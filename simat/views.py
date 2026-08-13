"""Vistas SIMAT: reporte de matrícula (exportación) y hub.

El exportador genera un Excel con el formato del "Reporte Plano" del SIMAT
(54 columnas) a partir de los aspirantes MATRICULADOS de la institución. Es
multi-institución: solo exporta la institución del usuario (el superusuario
puede pasar ?institucion=<id>). Los campos que HALU aún no captura salen en
blanco (se completan con la Fase 2 de captura).
"""
import logging

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.translation import gettext as _


# Orden EXACTO de columnas del Reporte Plano del SIMAT.
COLUMNAS_REPORTE = [
    'ANO', 'ETC', 'ESTADO', 'JERARQUIA', 'INSTITUCION', 'DANE', 'CALENDARIO',
    'SECTOR', 'SEDE', 'CODIGO_DANE_SEDE', 'CONSECUTIVO', 'ZONA_SEDE', 'JORNADA',
    'GRADO_COD', 'GRUPO', 'RENOMBRE', 'MODELO', 'MOTIVO', 'FECHAINI', 'FECHAFIN',
    'NUI', 'ESTRATO', 'SISBEN IV', 'PER_ID', 'DOC', 'TIPODOC', 'APELLIDO1',
    'APELLIDO2', 'NOMBRE1', 'NOMBRE2', 'GENERO', 'FECHA_NACIMIENTO', 'BARRIO',
    'EPS', 'TIPO DE SANGRE', 'MATRICULACONTRATADA', 'FUENTE_RECURSOS', 'INTERNADO',
    'NUM_CONTRATO', 'HA_ESTADO_VINCULADO_SRPA', 'ESTA_ACTIVO_SRPA', 'DISCAPACIDAD',
    'PAIS_ORIGEN', 'CORREO', 'TELEFONO', 'ETNIA', 'TRA_ESP_APR_ESCOLAR',
    'APOYO_ACADEMICO_ESPECIAL', 'LIST_CAP_EXCEPCIONALES', 'CAMPESINO',
    'PAIS_NACIMIENTO', 'PAIS_NACIONALIDAD2', 'CATEGORIA_AULA', 'CONTADOR',
]


def _sn(valor):
    """Bool → 'S'/'N' (formato SIMAT)."""
    return 'S' if valor else 'N'


def _si_no(valor):
    """Bool → 'SI'/'NO' (formato SIMAT para campos SI/NO)."""
    return 'SI' if valor else 'NO'


def _txt(valor):
    return '' if valor is None else str(valor)


# ── Tablas de códigos oficiales SIMAT (para exportar CÓDIGO, no la etiqueta) ──
_TIPODOC_SIMAT = {'CC': '1', 'TI': '2', 'CE': '3', 'RC': '5', 'NES': '8', 'PEP': '10', 'VISA': '11', 'TMF': '12'}
_GENERO_SIMAT = {'M': 'M', 'F': 'F'}
_JORNADA_SIMAT = {'COMPLETA': '1', 'MANANA': '2', 'TARDE': '3', 'NOCHE': '4', 'FIN_DE_SEMANA': '5', 'UNICA': '6'}
_ZONA_SIMAT = {'URBANA': '1', 'RURAL': '2'}
_SECTOR_SIMAT = {'OFICIAL': '1001', 'NO_OFICIAL': '1002'}
_DISCAP_SIMAT = {
    'NINGUNA': '99', 'FISICA': '15', 'INTELECTUAL': '8', 'PSICOSOCIAL': '18',
    'VISUAL_BAJA': '3', 'VISUAL_CEGUERA': '4', 'AUDITIVA_HIPOACUSIA': '2',
    'AUDITIVA_SORDA': '1', 'SORDOCEGUERA': '14', 'MULTIPLE': '10', 'SISTEMICA': '17',
    'VOZ_Y_HABLA': '19', 'TEA': '7', 'OTRA': '11',
}
_CAPACID_SIMAT = {
    'NINGUNA': '9', 'GLOBAL': '1', 'TALENTO_CIENTIFICO': '3',
    'TALENTO_ARTISTICO': '4', 'TALENTO_DEPORTIVO': '5', 'OTRA': '',
}
_VICTIMA_SIMAT = {
    'DESPLAZADO': '1', 'DESVINCULADO': '2', 'HIJO_DESMOVILIZADO': '3',
    'VICTIMA_MINAS': '4', 'OTRA': '17',
}


def _cod(mapa, valor):
    """Devuelve el código SIMAT de un valor de choice; '' si no mapea."""
    return mapa.get(valor or '', '')


def _solo_digitos(valor):
    """Solo los dígitos de un texto (para validar longitudes de códigos DANE)."""
    return ''.join(ch for ch in (valor or '') if ch.isdigit())


def _jornada_efectiva(asp):
    """Jornada del estudiante para el reporte SIMAT (código de choice o '').

    Si el estudiante no tiene jornada propia, HEREDA la «Jornada principal»
    configurada en su sede. Así basta con fijar la jornada una sola vez en la
    sede (Configuración › Sedes) y aplica a todos sus estudiantes, sin tener
    que asignarla estudiante por estudiante.
    """
    propia = (getattr(asp, 'jornada', '') or '').strip()
    if propia:
        return propia
    sede = getattr(asp, 'sede', None)
    return (getattr(sede, 'jornada_principal', '') or '').strip() if sede else ''


def _institucion_de(request):
    """Institución objetivo: la del usuario; superusuario puede pasar ?institucion=."""
    from finanzas.models import InstitucionEducativa
    if request.user.is_superuser and request.GET.get('institucion'):
        return InstitucionEducativa.objects.filter(pk=request.GET['institucion']).first()
    return getattr(request.user, 'institucion_asociada', None)


def _puede_gestionar(user):
    rol = getattr(user, 'rol', '') or ''
    return user.is_superuser or rol in ('coordinador', 'rector', 'administrador', 'secretaria')


def _fila_aspirante(asp, institucion, anio, contador):
    """Arma el dict de una fila del reporte a partir de un Aspirante matriculado."""
    sede = asp.sede
    grado = asp.grado_aspira
    # Etnia: código oficial SIMAT (no la etiqueta). Sin etnia → '0' (NO APLICA).
    etnia = asp.etnia_simat.codigo if asp.etnia_simat_id else '0'
    eps = asp.eps_simat.nombre if asp.eps_simat_id else _txt(asp.eps)
    fnac = asp.fecha_nacimiento.strftime('%Y-%m-%d') if asp.fecha_nacimiento else ''
    etc = institucion.simat_municipio_etc.codigo if institucion.simat_municipio_etc_id else ''
    # Grupo/sección: prioriza el grupo estructurado del estudiante matriculado;
    # si no, cae al texto libre capturado en la admisión.
    estudiante = getattr(asp, 'estudiante_creado', None)
    if estudiante is not None and getattr(estudiante, 'grupo_id', None):
        grupo_nombre = estudiante.grupo.nombre
    else:
        grupo_nombre = asp.grupo
    return {
        'ANO': anio,
        'ETC': _txt(etc),
        'ESTADO': 'MATRICULADO',
        'JERARQUIA': _txt(etc),
        'INSTITUCION': _txt(institucion.nombre),
        'DANE': _txt(institucion.codigo_dane),
        'CALENDARIO': _txt(institucion.simat_calendario),
        'SECTOR': _cod(_SECTOR_SIMAT, institucion.simat_sector),
        'SEDE': _txt(sede.nombre) if sede else _txt(institucion.nombre),
        'CODIGO_DANE_SEDE': _txt(sede.codigo_dane_sede) if sede else '',
        'CONSECUTIVO': _txt(sede.consecutivo) if sede else '',
        'ZONA_SEDE': _cod(_ZONA_SIMAT, sede.zona) if sede else '',
        'JORNADA': _cod(_JORNADA_SIMAT, _jornada_efectiva(asp)),
        'GRADO_COD': _txt(grado.nombre) if grado else '',
        'GRUPO': _txt(grupo_nombre),
        'RENOMBRE': '',
        'MODELO': _txt(asp.modelo_educativo),
        'MOTIVO': '',
        'FECHAINI': '',
        'FECHAFIN': '',
        'NUI': _txt(asp.simat_nui),
        'ESTRATO': ('NO APLICA' if asp.estrato == '0' else _txt(asp.estrato)),
        'SISBEN IV': _txt(asp.sisben_grupo),
        'PER_ID': _txt(asp.simat_per_id),
        'DOC': _txt(asp.numero_documento),
        'TIPODOC': _cod(_TIPODOC_SIMAT, asp.tipo_documento),
        'APELLIDO1': _txt(asp.primer_apellido),
        'APELLIDO2': _txt(asp.segundo_apellido),
        'NOMBRE1': _txt(asp.primer_nombre),
        'NOMBRE2': _txt(asp.segundo_nombre),
        'GENERO': _cod(_GENERO_SIMAT, asp.sexo),
        'FECHA_NACIMIENTO': fnac,
        'BARRIO': _txt(asp.barrio),
        'EPS': eps,
        'TIPO DE SANGRE': _txt(asp.grupo_sanguineo),
        'MATRICULACONTRATADA': _sn(asp.matricula_contratada),
        'FUENTE_RECURSOS': _txt(asp.fuente_recursos),
        'INTERNADO': _txt(asp.internado),
        'NUM_CONTRATO': '',
        'HA_ESTADO_VINCULADO_SRPA': _sn(asp.srpa),
        'ESTA_ACTIVO_SRPA': _sn(asp.srpa),
        'DISCAPACIDAD': _cod(_DISCAP_SIMAT, asp.discapacidad_categoria),
        'PAIS_ORIGEN': _txt(asp.pais_origen),
        'CORREO': _txt(asp.email_contacto),
        'TELEFONO': _txt(asp.telefono_contacto),
        'ETNIA': etnia,
        'TRA_ESP_APR_ESCOLAR': '',
        'APOYO_ACADEMICO_ESPECIAL': _sn(asp.apoyo_academico_especial),
        'LIST_CAP_EXCEPCIONALES': _cod(_CAPACID_SIMAT, getattr(asp, 'capacidad_excepcional', '')),
        'CAMPESINO': _sn(asp.campesino),
        'PAIS_NACIMIENTO': _txt(asp.pais_nacimiento),
        'PAIS_NACIONALIDAD2': _txt(asp.nacionalidad),
        'CATEGORIA_AULA': '',
        'CONTADOR': contador,
    }


# ── Encabezado EXACTO del reporte plano del SIMAT (55 columnas, orden oficial) ──
# El SIMAT identifica al alumno por men_per_id / simat_anexo_id (por eso el plano
# NO lleva documento ni nombres: es el anexo que se cruza por ID interno). Lo que
# HALU aún no captura sale vacío (ver auditoría). El Excel conserva los nombres
# legibles para revisión humana.
OFICIAL_COLUMNAS = [
    'simat_anexo_id', 'anio', 'municipio_id', 'dane', 'dane_sede', 'consecutivo_sede', 'sede',
    'prestacion_servicio', 'expedicion_departamento_id', 'expedicion_municipio_id',
    'dir_departamento_id', 'dir_municipio_id', 'estrato_id', 'sisben',
    'nacimiento_departamento_id', 'nacimiento_municipio_id', 'genero_id', 'tipo_victima_id',
    'expulsor_departamento_id', 'expulsor_municipio_id', 'proviene_sector_privado',
    'proviene_otro_municipio', 'discapacidad_id', 'capacidad_id', 'etnia_id', 'resguardo_id',
    'institucion_bienestar', 'jornada_id', 'caracter_id', 'especialidad_id', 'grado_id', 'grupo',
    'metodologia_id', 'subsidiado', 'repitente', 'nuevo', 'situacion_academica_va_id',
    'condicion_alumno_va_id', 'recurso_id', 'zona_id', 'madre_cf', 'hijo_mcf',
    'beneficiario_veterano', 'beneficiario_heroe', 'codigo_internado', 'codigo_valoracion_1',
    'codigo_valoracion_2', 'numero_convenio', 'men_per_id', 'apoyo_acad_esp', 'sist_resp_penal',
    'pais_origen', 'trastorno_id', 'fecha_anexo', 'tipo_anexo_id',
]


def _fk_cod(obj):
    return obj.codigo if obj else ''


def _fila_oficial(asp, institucion, anio, contador):
    """Fila con los nombres/orden EXACTOS del reporte plano del SIMAT. Los campos
    que HALU aún no captura salen vacíos (ver auditoría)."""
    sede = asp.sede
    est = getattr(asp, 'estudiante_creado', None)
    if est is not None and getattr(est, 'grupo_id', None):
        grupo_nombre = est.grupo.nombre
    else:
        grupo_nombre = asp.grupo
    etc = institucion.simat_municipio_etc.codigo if institucion.simat_municipio_etc_id else ''
    tipo_victima = _cod(_VICTIMA_SIMAT, asp.tipo_poblacion_victima) if asp.victima_conflicto else '99'
    return {
        'simat_anexo_id': '',
        'anio': anio,
        'municipio_id': _txt(etc),
        'dane': _txt(institucion.codigo_dane),
        'dane_sede': _txt(sede.codigo_dane_sede) if sede else '',
        'consecutivo_sede': _txt(sede.consecutivo) if sede else '',
        'sede': _txt(sede.nombre) if sede else '',
        'prestacion_servicio': _txt(getattr(institucion, 'simat_prestacion_servicio', '')),
        'expedicion_departamento_id': _fk_cod(asp.lugar_expedicion_departamento),
        'expedicion_municipio_id': _fk_cod(asp.lugar_expedicion_municipio),
        'dir_departamento_id': _fk_cod(asp.departamento_residencia),
        'dir_municipio_id': _fk_cod(asp.municipio_residencia),
        'estrato_id': (_txt(asp.estrato) if asp.estrato and asp.estrato != '0' else ''),
        'sisben': _txt(asp.sisben_simat),
        'nacimiento_departamento_id': _fk_cod(asp.departamento_nacimiento),
        'nacimiento_municipio_id': _fk_cod(asp.municipio_nacimiento),
        'genero_id': _cod(_GENERO_SIMAT, asp.sexo),
        'tipo_victima_id': tipo_victima,
        'expulsor_departamento_id': _fk_cod(asp.expulsor_departamento),
        'expulsor_municipio_id': _fk_cod(asp.expulsor_municipio),
        'proviene_sector_privado': _txt(asp.proviene_sector_privado),
        'proviene_otro_municipio': _txt(asp.proviene_otro_municipio),
        'discapacidad_id': _cod(_DISCAP_SIMAT, asp.discapacidad_categoria),
        'capacidad_id': _cod(_CAPACID_SIMAT, getattr(asp, 'capacidad_excepcional', '')),
        'etnia_id': asp.etnia_simat.codigo if asp.etnia_simat_id else '0',
        'resguardo_id': _fk_cod(asp.resguardo),
        'institucion_bienestar': _txt(asp.institucion_bienestar),
        'jornada_id': _cod(_JORNADA_SIMAT, _jornada_efectiva(asp)),
        'caracter_id': _txt(asp.caracter),
        'especialidad_id': _txt(asp.especialidad),
        'grado_id': _txt(getattr(asp.grado_aspira, 'simat_grado_id', '') if asp.grado_aspira_id else ''),
        'grupo': _txt(grupo_nombre),
        'metodologia_id': _txt(asp.metodologia),
        'subsidiado': _txt(asp.subsidiado),
        'repitente': _si_no(asp.repitente),
        'nuevo': _txt(asp.es_nuevo),
        'situacion_academica_va_id': _txt(asp.situacion_va),
        'condicion_alumno_va_id': _txt(asp.condicion_va),
        'recurso_id': _txt(asp.fuente_recurso),
        'zona_id': _cod(_ZONA_SIMAT, asp.zona_residencia),
        'madre_cf': _txt(asp.madre_cabeza_familia),
        'hijo_mcf': _txt(asp.hijo_madre_cabeza_familia),
        'beneficiario_veterano': _txt(asp.beneficiario_veterano),
        'beneficiario_heroe': _txt(asp.beneficiario_heroe),
        'codigo_internado': _txt(asp.tipo_internado),
        'codigo_valoracion_1': _txt(asp.valoracion_p1),
        'codigo_valoracion_2': _txt(asp.valoracion_p2),
        'numero_convenio': _txt(asp.numero_convenio),
        'men_per_id': _txt(asp.simat_per_id),
        'apoyo_acad_esp': _si_no(asp.apoyo_academico_especial),
        'sist_resp_penal': _si_no(asp.srpa),
        'pais_origen': _txt(asp.pais_origen),
        'trastorno_id': '',
        'fecha_anexo': '',
        'tipo_anexo_id': '',
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  ANEXO 6A — Archivo plano de MATRÍCULA (40 columnas, delimitado por «|»)
#  Orden y formato según la estructura nacional estándar (base inamovible):
#  sin fila de encabezados, una línea por estudiante, UTF-8, valores saneados.
#  El orden sigue el desglose/ejemplo oficial (la FECHA va en la col. 12, la
#  DIRECCIÓN en la 16). Los códigos que la plataforma aún no captura salen vacíos.
# ═══════════════════════════════════════════════════════════════════════════════

def _san(texto):
    """Sanitiza texto libre para el plano: MAYÚSCULAS, sin tildes, Ñ→N, sin
    saltos de línea/tabs/delimitadores/comas. Todo en una sola línea."""
    import unicodedata
    s = '' if texto is None else str(texto)
    # Descompone y elimina marcas diacríticas (tildes); la ñ → n.
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = s.upper()
    for ch in ('\n', '\r', '\t', '|', ';', ','):
        s = s.replace(ch, ' ')
    return ' '.join(s.split()).strip()


def _dmy(fecha):
    """date → DD/MM/YYYY (formato de fecha del Anexo 6A)."""
    return fecha.strftime('%d/%m/%Y') if fecha else ''


def _bin(valor):
    """Cualquier verdad → '1', si no '0' (booleanos numéricos del Anexo 6A)."""
    if isinstance(valor, str):
        return '1' if valor.strip().upper() in ('1', 'S', 'SI', 'SÍ', 'TRUE', 'X') else '0'
    return '1' if valor else '0'


def _depto2(dep_obj):
    """Código DANE de departamento a 2 dígitos (relleno de ceros)."""
    cod = (dep_obj.codigo if dep_obj else '') or ''
    return cod.zfill(2) if cod.isdigit() else cod


def _mpio3(mpio_obj):
    """Código DANE de municipio a 3 dígitos (los últimos 3 del código DANE)."""
    cod = (mpio_obj.codigo if mpio_obj else '') or ''
    if cod.isdigit():
        return cod[-3:].zfill(3)   # DANE municipio suele venir como 5 díg. (depto+mpio)
    return cod


def _fila_anexo6a(asp, institucion):
    """Lista ORDENADA de las 40 columnas del Anexo 6A para un aspirante matriculado.
    Reusa los mismos datos codificados que el reporte oficial."""
    sede = asp.sede
    grado = asp.grado_aspira
    est = getattr(asp, 'estudiante_creado', None)
    if est is not None and getattr(est, 'grupo_id', None):
        grupo_nombre = est.grupo.nombre
    else:
        grupo_nombre = asp.grupo
    # Inclusión: evitar nulos (99 = No Aplica) según la norma.
    victima = _cod(_VICTIMA_SIMAT, asp.tipo_poblacion_victima) if asp.victima_conflicto else '99'
    discap = _cod(_DISCAP_SIMAT, asp.discapacidad_categoria) or '99'
    capac = _cod(_CAPACID_SIMAT, getattr(asp, 'capacidad_excepcional', '')) or '99'
    etnia = asp.etnia_simat.codigo if asp.etnia_simat_id else '0'
    resguardo = (asp.resguardo.codigo if asp.resguardo_id else '') or '0'
    veterano_heroe = _bin(getattr(asp, 'beneficiario_veterano', False)) == '1' \
        or _bin(getattr(asp, 'beneficiario_heroe', False)) == '1'
    grado_id = _txt(getattr(grado, 'simat_grado_id', '') if grado else '')
    return [
        _txt(institucion.codigo_dane),                                   # 1  DANE_ESTABLECIMIENTO
        _txt(sede.codigo_dane_sede) if sede else '',                     # 2  DANE_SEDE
        _txt(sede.consecutivo) if sede else '1',                         # 3  CONSECUTIVO_SEDE
        _cod(_TIPODOC_SIMAT, asp.tipo_documento),                        # 4  TIPO_DOCUMENTO
        _txt(asp.numero_documento),                                      # 5  NUMERO_DOCUMENTO
        _depto2(asp.lugar_expedicion_departamento),                      # 6  EXPEDICION_DEPTO
        _mpio3(asp.lugar_expedicion_municipio),                          # 7  EXPEDICION_MUNICIPIO
        _san(asp.primer_nombre),                                         # 8  PRIMER_NOMBRE
        _san(asp.segundo_nombre),                                        # 9  SEGUNDO_NOMBRE
        _san(asp.primer_apellido),                                       # 10 PRIMER_APELLIDO
        _san(asp.segundo_apellido),                                      # 11 SEGUNDO_APELLIDO
        _dmy(asp.fecha_nacimiento),                                      # 12 FECHA_NACIMIENTO
        _depto2(asp.departamento_nacimiento),                            # 13 NACIMIENTO_DEPTO
        _mpio3(asp.municipio_nacimiento),                                # 14 NACIMIENTO_MUNICIPIO
        _cod(_GENERO_SIMAT, asp.sexo),                                   # 15 GENERO
        _san(asp.direccion),                                             # 16 DIRECCION_RESIDENCIA
        _txt(asp.telefono_contacto),                                     # 17 TELEFONO
        ('' if (asp.estrato in (None, '')) else _txt(asp.estrato)),      # 18 ESTRATO (0-6)
        _txt(asp.sisben_simat or asp.sisben_grupo),                      # 19 SISBEN
        _cod(_JORNADA_SIMAT, _jornada_efectiva(asp)),                    # 20 JORNADA
        _txt(asp.caracter),                                              # 21 CARACTER
        _txt(asp.especialidad),                                          # 22 ESPECIALIDAD
        grado_id,                                                        # 23 GRADO
        _txt(grupo_nombre),                                              # 24 GRUPO
        _txt(asp.metodologia),                                           # 25 METODOLOGIA
        _bin(asp.repitente),                                             # 26 REPITENTE
        _bin(getattr(asp, 'es_nuevo', '')),                             # 27 NUEVO
        _bin(asp.subsidiado),                                            # 28 SUBSIDIADO
        victima,                                                         # 29 TIPO_VICTIMA
        discap,                                                          # 30 DISCAPACIDAD
        capac,                                                           # 31 CAPACIDAD_EXCEPCIONAL
        etnia,                                                           # 32 ETNIA
        resguardo,                                                       # 33 RESGUARDO
        _bin(asp.srpa),                                                  # 34 SIST_RESPONSABILIDAD_PENAL
        _bin(asp.apoyo_academico_especial),                             # 35 APOYO_ACADEMICO_ESPECIAL
        _txt(asp.pais_origen or '170'),                                 # 36 PAIS_ORIGEN (170=Colombia)
        '99',                                                            # 37 TRASTORNO (no capturado → 99)
        _bin(asp.campesino),                                             # 38 POBLACION_CAMPESINA
        _bin(getattr(asp, 'hijo_madre_cabeza_familia', False)),         # 39 HIJO_MADRE_CABEZA_FAMILIA
        '1' if veterano_heroe else '0',                                 # 40 VETERANO_HEROE_FUERZA_PUBLICA
    ]


@login_required
def exportar_anexo6a_txt(request):
    """Descarga el Anexo 6A de matrícula: texto plano, 40 columnas separadas por
    «|», sin encabezado, UTF-8. Solo estudiantes MATRICULADOS de la institución."""
    if not _puede_gestionar(request.user):
        raise PermissionDenied
    from admisiones.models import Aspirante

    institucion = _institucion_de(request)
    if institucion is None:
        raise PermissionDenied("Sin institución asociada.")

    aspirantes = (
        Aspirante.objects
        .filter(institucion=institucion, estado=Aspirante.EstadoAdmision.MATRICULADO)
        .select_related('sede', 'grado_aspira', 'etnia_simat', 'resguardo',
                        'lugar_expedicion_departamento', 'lugar_expedicion_municipio',
                        'departamento_nacimiento', 'municipio_nacimiento',
                        'estudiante_creado', 'estudiante_creado__grupo')
        .order_by('primer_apellido', 'primer_nombre')
    )

    resp = HttpResponse(content_type='text/plain; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="Anexo6A_{institucion.codigo_dane or "matricula"}.txt"'
    lineas = []
    for asp in aspirantes.iterator():
        lineas.append('|'.join(_fila_anexo6a(asp, institucion)))
    resp.write('\r\n'.join(lineas))
    return resp


@login_required
def exportar_reporte_simat(request):
    """Descarga el Reporte Plano SIMAT (.xlsx) de los matriculados."""
    if not _puede_gestionar(request.user):
        raise PermissionDenied
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from admisiones.models import Aspirante

    institucion = _institucion_de(request)
    if institucion is None:
        raise PermissionDenied("Sin institución asociada.")
    anio = timezone.now().year

    aspirantes = (
        Aspirante.objects
        .filter(institucion=institucion, estado=Aspirante.EstadoAdmision.MATRICULADO)
        .select_related('sede', 'grado_aspira', 'etnia_simat', 'eps_simat',
                        'estudiante_creado', 'estudiante_creado__grupo')
        .order_by('primer_apellido', 'primer_nombre', 'apellidos')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "REPORTE SIMAT"
    ws.append(COLUMNAS_REPORTE)
    for i, cell in enumerate(ws[1], start=1):
        cell.fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal='center')

    for idx, asp in enumerate(aspirantes.iterator(), start=1):
        fila = _fila_aspirante(asp, institucion, anio, idx)
        ws.append([fila.get(col, '') for col in COLUMNAS_REPORTE])

    for col_cells in ws.columns:
        w = max((len(str(c.value or '')) for c in col_cells), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(w + 2, 10), 40)
    ws.freeze_panes = "A2"

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="reporte_simat_{anio}.xlsx"'
    wb.save(resp)
    return resp


def _escribir_reporte_delimitado(resp, institucion, delimiter=','):
    """Escribe el reporte plano con la estructura OFICIAL del SIMAT (55 columnas,
    orden y nombres exactos). Compartido por .csv (coma) y .txt (tabulador).
    Comillas en valores con el separador (QUOTE_MINIMAL), fin de línea CRLF,
    encoding UTF-8."""
    import csv
    from admisiones.models import Aspirante
    anio = timezone.now().year
    aspirantes = (
        Aspirante.objects
        .filter(institucion=institucion, estado=Aspirante.EstadoAdmision.MATRICULADO)
        .select_related(
            'sede', 'grado_aspira', 'etnia_simat', 'resguardo',
            'estudiante_creado', 'estudiante_creado__grupo',
            'departamento_residencia', 'municipio_residencia',
            'departamento_nacimiento', 'municipio_nacimiento',
            'lugar_expedicion_departamento', 'lugar_expedicion_municipio',
            'expulsor_departamento', 'expulsor_municipio',
        )
        .order_by('primer_apellido', 'primer_nombre', 'apellidos')
    )
    writer = csv.writer(resp, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL, lineterminator='\r\n')
    writer.writerow(OFICIAL_COLUMNAS)
    for idx, asp in enumerate(aspirantes.iterator(), start=1):
        fila = _fila_oficial(asp, institucion, anio, idx)
        writer.writerow([fila.get(col, '') for col in OFICIAL_COLUMNAS])


@login_required
def exportar_reporte_simat_csv(request):
    """Reporte plano SIMAT en CSV (separado por comas). El SIMAT ingesta la
    matrícula como archivo plano; el Excel queda para revisión humana."""
    if not _puede_gestionar(request.user):
        raise PermissionDenied
    institucion = _institucion_de(request)
    if institucion is None:
        raise PermissionDenied("Sin institución asociada.")
    anio = timezone.now().year
    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="reporte_simat_{anio}.csv"'
    _escribir_reporte_delimitado(resp, institucion)
    return resp


@login_required
def exportar_reporte_simat_txt(request):
    """Reporte plano SIMAT en .txt (formato oficial: el SIMAT genera el reporte
    'en CSV' pero lo entrega como archivo de texto .txt). Mismo contenido que el
    CSV."""
    if not _puede_gestionar(request.user):
        raise PermissionDenied
    institucion = _institucion_de(request)
    if institucion is None:
        raise PermissionDenied("Sin institución asociada.")
    anio = timezone.now().year
    resp = HttpResponse(content_type='text/plain; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="reporte_simat_{anio}.txt"'
    _escribir_reporte_delimitado(resp, institucion, delimiter='\t')
    return resp


# ─────────────────────────────────────────────────────────────────────────────
#  men_per_id (PER_ID del SIMAT) — carga masiva por Excel (ida y vuelta)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def descargar_plantilla_men_per_id(request):
    """Genera un Excel con los estudiantes matriculados (documento, nombres,
    sede, grado, grupo) y una columna men_per_id para que el usuario la llene con
    el PER_ID que devuelve el SIMAT y la vuelva a subir."""
    if not _puede_gestionar(request.user):
        raise PermissionDenied
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from admisiones.models import Aspirante

    institucion = _institucion_de(request)
    if institucion is None:
        raise PermissionDenied("Sin institución asociada.")

    aspirantes = (
        Aspirante.objects
        .filter(institucion=institucion, estado=Aspirante.EstadoAdmision.MATRICULADO)
        .select_related('sede', 'grado_aspira', 'estudiante_creado', 'estudiante_creado__grupo')
        .order_by('primer_apellido', 'primer_nombre', 'apellidos')
    )
    cols = ['documento', 'apellidos', 'nombres', 'sede', 'grado', 'grupo', 'men_per_id']
    wb = Workbook()
    ws = wb.active
    ws.title = "MEN_PER_ID"
    ws.append(cols)
    for i, cell in enumerate(ws[1], start=1):
        cell.fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal='center')
    for asp in aspirantes.iterator():
        est = getattr(asp, 'estudiante_creado', None)
        grupo = est.grupo.nombre if (est and getattr(est, 'grupo_id', None)) else _txt(asp.grupo)
        apellidos = ' '.join(p for p in [asp.primer_apellido, asp.segundo_apellido] if p) or _txt(asp.apellidos)
        nombres = ' '.join(p for p in [asp.primer_nombre, asp.segundo_nombre] if p) or _txt(asp.nombres)
        ws.append([
            _txt(asp.numero_documento), apellidos, nombres,
            _txt(asp.sede.nombre) if asp.sede else '',
            _txt(asp.grado_aspira.nombre) if asp.grado_aspira_id else '',
            _txt(grupo), _txt(asp.simat_per_id),
        ])
    for col_cells in ws.columns:
        w = max((len(str(c.value or '')) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(w + 2, 12), 40)
    ws.freeze_panes = "A2"
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="plantilla_men_per_id.xlsx"'
    wb.save(resp)
    return resp


@login_required
def cargar_men_per_id(request):
    """Recibe el Excel de plantilla_men_per_id con la columna men_per_id llena y
    asigna el PER_ID a cada estudiante (cruzando por documento, en la institución)."""
    if not _puede_gestionar(request.user):
        raise PermissionDenied
    institucion = _institucion_de(request)
    if institucion is None:
        raise PermissionDenied("Sin institución asociada.")
    if request.method == 'POST' and request.FILES.get('archivo'):
        from openpyxl import load_workbook
        from admisiones.models import Aspirante
        archivo = request.FILES['archivo']
        if not archivo.name.lower().endswith('.xlsx'):
            messages.error(request, _("El archivo debe ser un Excel (.xlsx)."))
            return redirect('simat:hub')
        try:
            wb = load_workbook(archivo, read_only=True, data_only=True)
            ws = wb.active
            filas = ws.iter_rows(values_only=True)
            encabezados = [str(h).strip().lower() if h is not None else '' for h in next(filas)]
            if 'documento' not in encabezados or 'men_per_id' not in encabezados:
                messages.error(request, _("El archivo debe tener las columnas 'documento' y 'men_per_id'."))
                return redirect('simat:hub')
            i_doc = encabezados.index('documento')
            i_men = encabezados.index('men_per_id')
            actualizados, sin_match = 0, 0
            for fila in filas:
                if fila is None or len(fila) <= max(i_doc, i_men):
                    continue
                doc = str(fila[i_doc]).strip() if fila[i_doc] is not None else ''
                men = str(fila[i_men]).strip() if fila[i_men] is not None else ''
                if not doc or not men:
                    continue
                asp = Aspirante.objects.filter(institucion=institucion, numero_documento=doc).first()
                if not asp:
                    sin_match += 1
                    continue
                asp.simat_per_id = men[:20]
                asp.save(update_fields=['simat_per_id'])
                est = getattr(asp, 'estudiante_creado', None)
                if est is not None:
                    car = getattr(est, 'caracterizacion', None)
                    if car is not None:
                        car.simat_per_id = men[:20]
                        car.save(update_fields=['simat_per_id'])
                actualizados += 1
            msg = _("Se asignaron %(n)s PER_ID (men_per_id).") % {'n': actualizados}
            if sin_match:
                msg += _(" %(s)s documento(s) no coincidieron con ningún estudiante.") % {'s': sin_match}
            messages.success(request, msg)
        except Exception:
            logging.getLogger(__name__).exception("Error cargando men_per_id (institución=%s)", getattr(institucion, 'pk', None))
            messages.error(request, _("No se pudo leer el archivo. Verifica que sea el Excel de la plantilla."))
    return redirect('simat:hub')


# ─────────────────────────────────────────────────────────────────────────────
#  Pre-validador de calidad SIMAT (reglas de auditoría del MEN, revisadas en local)
# ─────────────────────────────────────────────────────────────────────────────

# Edad típica esperada por grado (ID SIMAT). Preescolar (-) y adultos (21-25) se
# omiten del chequeo de edad.
_EDAD_ESPERADA = {'0': 5, '1': 6, '2': 7, '3': 8, '4': 9, '5': 10, '6': 11,
                  '7': 12, '8': 13, '9': 14, '10': 15, '11': 16, '12': 17, '13': 18}


def _validar_matricula_simat(institucion, anio):
    """Corre localmente las reglas de calidad del SIMAT sobre los matriculados.
    Devuelve dict con 'config' (errores de establecimiento), 'errores' y
    'advertencias' por estudiante, y contadores."""
    from admisiones.models import Aspirante

    config, errores, advertencias = [], [], []

    # ── Configuración del establecimiento ──
    if not (institucion.codigo_dane or '').strip():
        config.append(_("Falta el Código DANE de la institución."))
    else:
        _d = _solo_digitos(institucion.codigo_dane)
        if len(_d) != 12:
            config.append(_("El Código DANE de la institución debe tener 12 dígitos (actualmente tiene %(n)s). Corrígelo en la configuración de la institución.") % {'n': len(_d)})
    if not institucion.simat_municipio_etc_id:
        config.append(_("Falta el municipio (ETC) del SIMAT."))
    if not (institucion.simat_calendario or '').strip():
        config.append(_("Falta el calendario (A/B)."))
    if not (institucion.simat_sector or '').strip():
        config.append(_("Falta el sector (oficial/no oficial)."))
    sedes = list(institucion.sedes.filter(activa=True))
    if not sedes:
        config.append(_("No hay sedes activas."))
    for s in sedes:
        if not (s.codigo_dane_sede or '').strip():
            config.append(_("La sede «%(n)s» no tiene Código DANE de sede.") % {'n': s.nombre})
        else:
            _ds = _solo_digitos(s.codigo_dane_sede)
            if len(_ds) != 12:
                config.append(_("El Código DANE de la sede «%(n)s» debe tener 12 dígitos (actualmente tiene %(d)s). Corrígelo en Configuración › Sedes.") % {'n': s.nombre, 'd': len(_ds)})
        if not (s.consecutivo or '').strip():
            config.append(_("La sede «%(n)s» no tiene consecutivo.") % {'n': s.nombre})

    # Departamento de la institución (para la Regla 4) = 2 primeros dígitos del ETC.
    depto_inst = ''
    if institucion.simat_municipio_etc_id:
        depto_inst = (institucion.simat_municipio_etc.codigo or '')[:2]

    aspirantes = list(
        Aspirante.objects
        .filter(institucion=institucion, estado=Aspirante.EstadoAdmision.MATRICULADO)
        .select_related('sede', 'grado_aspira', 'municipio_residencia', 'departamento_residencia',
                        'estudiante_creado', 'estudiante_creado__grupo')
    )

    # Índices para duplicados (Regla 1 y 1.2).
    docs, perids = {}, {}
    for a in aspirantes:
        d = (a.numero_documento or '').strip()
        if d:
            docs.setdefault(d, []).append(a)
        p = (a.simat_per_id or '').strip()
        if p:
            perids.setdefault(p, []).append(a)

    def _nombre(a):
        n = ' '.join(x for x in [a.primer_nombre, a.primer_apellido] if x) or (a.nombres + ' ' + a.apellidos)
        return f"{n.strip()} ({a.numero_documento})"

    for a in aspirantes:
        nom = _nombre(a)
        # Regla 1 — documento faltante o duplicado
        if not (a.numero_documento or '').strip():
            errores.append((_("Regla 1 · Documento"), _("%(e)s no tiene número de documento.") % {'e': nom}))
        elif len(docs.get(a.numero_documento.strip(), [])) > 1:
            errores.append((_("Regla 1 · Duplicado"), _("Documento repetido: %(e)s.") % {'e': nom}))
        if not (a.tipo_documento or '').strip():
            advertencias.append((_("Identificación"), _("%(e)s no tiene tipo de documento.") % {'e': nom}))
        # Regla 1.2 — PER_ID duplicado
        p = (a.simat_per_id or '').strip()
        if p and len(perids.get(p, [])) > 1:
            errores.append((_("Regla 1.2 · PER_ID"), _("PER_ID repetido (%(p)s): %(e)s.") % {'p': p, 'e': nom}))
        # Género válido (F/M)
        if a.sexo not in ('F', 'M'):
            errores.append((_("Género"), _("%(e)s: el género debe ser Femenino o Masculino.") % {'e': nom}))
        # Fecha de nacimiento
        if not a.fecha_nacimiento:
            errores.append((_("Fecha de nacimiento"), _("%(e)s no tiene fecha de nacimiento.") % {'e': nom}))
        # Regla 2/grado — grado sin ID SIMAT
        gid = getattr(a.grado_aspira, 'simat_grado_id', '') if a.grado_aspira_id else ''
        if not a.grado_aspira_id:
            errores.append((_("Regla 2 · Grado"), _("%(e)s no tiene grado.") % {'e': nom}))
        elif not gid:
            errores.append((_("Regla 2 · Grado"), _("El grado «%(g)s» no tiene ID SIMAT asignado.") % {'g': a.grado_aspira.nombre}))
        # Jornada — obligatoria; el MEN rechaza el registro si va vacía. Se
        # valida la jornada EFECTIVA: si el estudiante no tiene una propia,
        # hereda la «Jornada principal» de su sede.
        if not _jornada_efectiva(a):
            errores.append((_("Jornada"), _("%(e)s no tiene jornada. Asígnala en la ficha del estudiante o define la «Jornada principal» de su sede (Configuración › Sedes).") % {'e': nom}))
        # Estrato — obligatorio (0 a 6) para el reporte.
        if a.estrato in (None, ''):
            errores.append((_("Estrato"), _("%(e)s no tiene estrato socioeconómico (0 a 6).") % {'e': nom}))
        # DIVIPOLA de nacimiento y de expedición del documento — obligatorios
        # para estudiantes de nacionalidad colombiana (país 170).
        es_colombiano = (a.pais_origen or '170') == '170'
        if es_colombiano:
            if not (a.departamento_nacimiento_id and a.municipio_nacimiento_id):
                errores.append((_("Lugar de nacimiento"), _("%(e)s no tiene departamento/municipio de nacimiento (DANE).") % {'e': nom}))
            if not (a.lugar_expedicion_departamento_id and a.lugar_expedicion_municipio_id):
                errores.append((_("Lugar de expedición"), _("%(e)s no tiene departamento/municipio de expedición del documento (DANE).") % {'e': nom}))
        # Regla 3 — edad atípica para el grado
        if a.fecha_nacimiento and gid in _EDAD_ESPERADA:
            edad = anio - a.fecha_nacimiento.year
            esp = _EDAD_ESPERADA[gid]
            if edad < esp - 2:
                advertencias.append((_("Regla 3 · Edad"), _("%(e)s: edad %(a)s años parece baja para el grado (esperado ~%(x)s).") % {'e': nom, 'a': edad, 'x': esp}))
            elif edad > esp + 3:
                advertencias.append((_("Regla 3 · Edad"), _("%(e)s: extraedad (%(a)s años; esperado ~%(x)s).") % {'e': nom, 'a': edad, 'x': esp}))
        # Regla 4 — departamento de residencia distante
        if depto_inst and a.departamento_residencia_id:
            dep_est = (a.departamento_residencia.codigo or '')[:2]
            if dep_est and dep_est != depto_inst:
                advertencias.append((_("Regla 4 · Residencia"), _("%(e)s reside en otro departamento (%(d)s) distinto al de la institución.") % {'e': nom, 'd': a.departamento_residencia.nombre}))

    return {
        'total': len(aspirantes),
        'config': config,
        'errores': errores,
        'advertencias': advertencias,
        'ok': not config and not errores and not advertencias,
    }


@login_required
def validar_simat(request):
    """Muestra el resultado de la pre-validación de calidad antes de exportar."""
    if not _puede_gestionar(request.user):
        raise PermissionDenied
    institucion = _institucion_de(request)
    if institucion is None:
        raise PermissionDenied("Sin institución asociada.")
    anio = timezone.now().year
    resultado = _validar_matricula_simat(institucion, anio)
    logging.getLogger(__name__).info(
        "Validación SIMAT institución=%s: %s estudiantes, %s errores config, %s errores, %s advertencias",
        institucion.pk, resultado['total'], len(resultado['config']),
        len(resultado['errores']), len(resultado['advertencias']),
    )
    return render(request, 'simat/validacion.html', {
        'titulo_pagina': _('Validación SIMAT'),
        'institucion': institucion,
        'r': resultado,
    })


@login_required
def hub_simat(request):
    """Página con el estado de configuración SIMAT y el botón de exportación."""
    if not _puede_gestionar(request.user):
        raise PermissionDenied
    from admisiones.models import Aspirante

    institucion = _institucion_de(request)
    matriculados = 0
    faltan_config = []
    if institucion is not None:
        matriculados = Aspirante.objects.filter(
            institucion=institucion, estado=Aspirante.EstadoAdmision.MATRICULADO
        ).count()
        if not institucion.codigo_dane:
            faltan_config.append("Código DANE de la institución")
        if not institucion.simat_municipio_etc_id:
            faltan_config.append("Municipio (ETC)")
        if not institucion.simat_calendario:
            faltan_config.append("Calendario (A/B)")
        if not institucion.simat_sector:
            faltan_config.append("Sector (oficial/no oficial)")
        if not institucion.sedes.exists():
            faltan_config.append("Al menos una Sede")

    return render(request, 'simat/hub.html', {
        'titulo_pagina': 'Reporte SIMAT',
        'institucion': institucion,
        'matriculados': matriculados,
        'faltan_config': faltan_config,
    })


# ─────────────────────────────────────────────────────────────────────────────
#  CRUD de Sedes (institución-scoped)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def lista_sedes(request):
    if not _puede_gestionar(request.user):
        raise PermissionDenied
    from .models import Sede
    institucion = _institucion_de(request)
    if institucion is None:
        raise PermissionDenied("Sin institución asociada.")
    # Garantiza al menos la Sede Principal (por si es una institución antigua).
    Sede.asegurar_principal(institucion)
    sedes = Sede.objects.filter(institucion=institucion).order_by('-es_principal', 'nombre')
    return render(request, 'simat/sedes_lista.html', {
        'titulo_pagina': _('Sedes'),
        'institucion': institucion,
        'sedes': sedes,
    })


@login_required
def crear_sede(request):
    if not _puede_gestionar(request.user):
        raise PermissionDenied
    from .forms import SedeForm
    from .models import Sede
    institucion = _institucion_de(request)
    if institucion is None:
        raise PermissionDenied("Sin institución asociada.")
    if request.method == 'POST':
        form = SedeForm(request.POST)
        if form.is_valid():
            sede = form.save(commit=False)
            sede.institucion = institucion
            # Consecutivo automático (si la institución lo tiene activado y quedó vacío).
            if getattr(institucion, 'simat_consecutivo_sede_automatico', True) and not (sede.consecutivo or '').strip():
                sede.consecutivo = Sede.siguiente_consecutivo(institucion)
            sede.save()
            messages.success(request, _("Sede «%(n)s» creada.") % {'n': sede.nombre})
            return redirect('simat:lista_sedes')
    else:
        initial = {}
        if getattr(institucion, 'simat_consecutivo_sede_automatico', True):
            initial['consecutivo'] = Sede.siguiente_consecutivo(institucion)
        form = SedeForm(initial=initial)
    return render(request, 'simat/sede_form.html', {
        'titulo_pagina': _('Nueva sede'), 'form': form, 'institucion': institucion,
    })


@login_required
def editar_sede(request, pk):
    if not _puede_gestionar(request.user):
        raise PermissionDenied
    from .forms import SedeForm
    from .models import Sede
    institucion = _institucion_de(request)
    sede = get_object_or_404(Sede, pk=pk, institucion=institucion)
    if request.method == 'POST':
        form = SedeForm(request.POST, instance=sede)
        if form.is_valid():
            form.save()
            messages.success(request, _("Sede «%(n)s» actualizada.") % {'n': sede.nombre})
            return redirect('simat:lista_sedes')
    else:
        form = SedeForm(instance=sede)
    return render(request, 'simat/sede_form.html', {
        'titulo_pagina': _('Editar sede'), 'form': form, 'institucion': institucion, 'sede': sede,
    })


@login_required
def eliminar_sede(request, pk):
    if not _puede_gestionar(request.user):
        raise PermissionDenied
    from .models import Sede
    institucion = _institucion_de(request)
    sede = get_object_or_404(Sede, pk=pk, institucion=institucion)
    if request.method == 'POST':
        if sede.es_principal:
            messages.error(request, _("No puedes eliminar la Sede Principal."))
        else:
            nombre = sede.nombre
            sede.delete()
            messages.success(request, _("Sede «%(n)s» eliminada.") % {'n': nombre})
    return redirect('simat:lista_sedes')


# ─────────────────────────────────────────────────────────────────────────────
#  CRUD de Grupos / Secciones (institución-scoped)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def lista_grupos(request):
    if not _puede_gestionar(request.user):
        raise PermissionDenied
    from gestion_academica.models import Grupo
    institucion = _institucion_de(request)
    if institucion is None:
        raise PermissionDenied("Sin institución asociada.")
    grupos = (
        Grupo.objects.filter(institucion=institucion)
        .select_related('grado', 'sede')
        .order_by('grado__orden', 'grado__nombre', 'nombre')
    )
    return render(request, 'simat/grupos_lista.html', {
        'titulo_pagina': _('Grupos'),
        'institucion': institucion,
        'grupos': grupos,
    })


@login_required
def crear_grupo(request):
    if not _puede_gestionar(request.user):
        raise PermissionDenied
    from .forms import GrupoForm
    institucion = _institucion_de(request)
    if institucion is None:
        raise PermissionDenied("Sin institución asociada.")
    if request.method == 'POST':
        form = GrupoForm(request.POST, institucion=institucion)
        if form.is_valid():
            grupo = form.save(commit=False)
            grupo.institucion = institucion
            grupo.save()
            messages.success(request, _("Grupo «%(n)s» creado.") % {'n': str(grupo)})
            return redirect('simat:lista_grupos')
    else:
        form = GrupoForm(institucion=institucion)
    return render(request, 'simat/grupo_form.html', {
        'titulo_pagina': _('Nuevo grupo'), 'form': form, 'institucion': institucion,
    })


@login_required
def editar_grupo(request, pk):
    if not _puede_gestionar(request.user):
        raise PermissionDenied
    from .forms import GrupoForm
    from gestion_academica.models import Grupo
    institucion = _institucion_de(request)
    grupo = get_object_or_404(Grupo, pk=pk, institucion=institucion)
    if request.method == 'POST':
        form = GrupoForm(request.POST, instance=grupo, institucion=institucion)
        if form.is_valid():
            grupo = form.save(commit=False)
            grupo.institucion = institucion
            grupo.save()
            messages.success(request, _("Grupo «%(n)s» actualizado.") % {'n': str(grupo)})
            return redirect('simat:lista_grupos')
    else:
        form = GrupoForm(instance=grupo, institucion=institucion)
    return render(request, 'simat/grupo_form.html', {
        'titulo_pagina': _('Editar grupo'), 'form': form, 'institucion': institucion, 'grupo': grupo,
    })


@login_required
def eliminar_grupo(request, pk):
    if not _puede_gestionar(request.user):
        raise PermissionDenied
    from gestion_academica.models import Grupo
    institucion = _institucion_de(request)
    grupo = get_object_or_404(Grupo, pk=pk, institucion=institucion)
    if request.method == 'POST':
        if grupo.estudiantes.exists():
            messages.error(
                request,
                _("No puedes eliminar «%(n)s»: tiene estudiantes asignados. "
                  "Muévelos a otro grupo primero.") % {'n': str(grupo)},
            )
        else:
            nombre = str(grupo)
            grupo.delete()
            messages.success(request, _("Grupo «%(n)s» eliminado.") % {'n': nombre})
    return redirect('simat:lista_grupos')
