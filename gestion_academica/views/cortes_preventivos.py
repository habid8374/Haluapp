"""
gestion_academica/views/cortes_preventivos.py
=============================================
Módulo de Corte Preventivo — HALU Platform
Informes académicos intermedios de alerta temprana.
"""
import io
import json
from decimal import Decimal
from datetime import date

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from ..models import (
    CortePreventivo, ConfiguracionCortePreventivo, ResultadoCorteEstudiante,
    DetalleMateriaCortePrev, Grado, PeriodoAcademico, Notificacion,
    Estudiante, Curso,
)

# ──────────────────────────────────────────────────────────────────────────────
#  Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _get_institucion(request):
    return getattr(request.user, 'institucion_asociada', None)


def _puede_gestionar_cortes(user):
    """Coordinador, admin institución o superusuario pueden gestionar cortes."""
    return user.is_superuser or getattr(user, 'cargo', '') in (
        'coordinador', 'admin_institucion', 'rector',
    )


def _color_riesgo(nivel):
    return {
        'ALTO':       ('danger',  'bi-exclamation-triangle-fill', '#dc2626'),
        'MEDIO':      ('warning', 'bi-exclamation-circle-fill',   '#d97706'),
        'BAJO':       ('info',    'bi-info-circle-fill',           '#2563eb'),
        'SIN_RIESGO': ('success', 'bi-check-circle-fill',          '#16a34a'),
    }.get(nivel, ('secondary', 'bi-dash-circle', '#6b7280'))


# ──────────────────────────────────────────────────────────────────────────────
#  1. LISTADO DE CORTES
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def lista_cortes_preventivos(request):
    institucion = _get_institucion(request)
    if not institucion:
        messages.error(request, "No tienes una institución asignada.")
        return redirect('gestion_academica:inicio_academico')

    qs = CortePreventivo.objects.filter(
        institucion=institucion
    ).select_related('grado', 'periodo_academico', 'generado_por').order_by(
        '-fecha_corte', 'grado__nombre'
    )

    # Filtros opcionales
    periodo_id = request.GET.get('periodo')
    grado_id   = request.GET.get('grado')
    estado     = request.GET.get('estado')
    if periodo_id:
        qs = qs.filter(periodo_academico_id=periodo_id)
    if grado_id:
        qs = qs.filter(grado_id=grado_id)
    if estado:
        qs = qs.filter(estado=estado)

    periodos = PeriodoAcademico.objects.filter(institucion=institucion).order_by('-año_escolar', '-fecha_inicio')
    grados   = Grado.objects.filter(institucion=institucion).order_by('nombre')

    context = {
        'cortes':    qs,
        'periodos':  periodos,
        'grados':    grados,
        'filtro_periodo': periodo_id,
        'filtro_grado':   grado_id,
        'filtro_estado':  estado,
        'puede_gestionar': _puede_gestionar_cortes(request.user),
    }
    return render(request, 'gestion_academica/cortes_preventivos/lista.html', context)


# ──────────────────────────────────────────────────────────────────────────────
#  2. CREAR CORTE
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def crear_corte_preventivo(request):
    institucion = _get_institucion(request)
    if not institucion or not _puede_gestionar_cortes(request.user):
        messages.error(request, "No tienes permiso para crear cortes preventivos.")
        return redirect('gestion_academica:lista_cortes_preventivos')

    periodos = PeriodoAcademico.objects.filter(institucion=institucion).order_by('-año_escolar', '-fecha_inicio')
    grados   = Grado.objects.filter(institucion=institucion).order_by('nombre')

    if request.method == 'POST':
        nombre      = request.POST.get('nombre_corte', '').strip()
        periodo_id  = request.POST.get('periodo_academico')
        fecha_str   = request.POST.get('fecha_corte')
        todos_grados = request.POST.get('todos_grados') == '1'
        grado_ids    = request.POST.getlist('grados') if not todos_grados else []

        # Validaciones básicas
        if not nombre:
            messages.error(request, "El nombre del corte es obligatorio.")
        elif not periodo_id:
            messages.error(request, "Debes seleccionar un período académico.")
        elif not fecha_str:
            messages.error(request, "La fecha de corte es obligatoria.")
        else:
            try:
                fecha_corte = date.fromisoformat(fecha_str)
            except ValueError:
                messages.error(request, "Fecha inválida.")
                fecha_corte = None

            if fecha_corte:
                periodo = get_object_or_404(PeriodoAcademico, pk=periodo_id, institucion=institucion)

                if todos_grados:
                    grados_a_crear = list(Grado.objects.filter(institucion=institucion))
                else:
                    if not grado_ids:
                        messages.error(request, "Debes seleccionar al menos un grado.")
                        grados_a_crear = []
                    else:
                        grados_a_crear = list(Grado.objects.filter(pk__in=grado_ids, institucion=institucion))

                creados = 0
                errores_dup = []
                cortes_nuevos = []
                for grado in grados_a_crear:
                    corte_nombre = f"{nombre}" if len(grados_a_crear) == 1 else f"{nombre} — {grado.nombre}"
                    try:
                        corte, created = CortePreventivo.objects.get_or_create(
                            institucion=institucion,
                            periodo_academico=periodo,
                            grado=grado,
                            fecha_corte=fecha_corte,
                            defaults={
                                'nombre_corte': corte_nombre,
                                'generado_por': request.user,
                            }
                        )
                        if created:
                            creados += 1
                            cortes_nuevos.append(corte)
                        else:
                            errores_dup.append(grado.nombre)
                    except Exception as e:
                        messages.error(request, f"Error al crear corte para {grado.nombre}: {e}")

                if errores_dup:
                    messages.warning(
                        request,
                        f"Ya existía un corte para esa fecha en: {', '.join(errores_dup)}. Se omitieron."
                    )

                if creados > 0:
                    messages.success(
                        request,
                        f"✅ {creados} corte(s) creado(s) correctamente. "
                        "Ahora puedes lanzar el cálculo desde cada corte."
                    )
                    if creados == 1:
                        return redirect('gestion_academica:detalle_corte_preventivo', pk=cortes_nuevos[0].pk)
                    return redirect('gestion_academica:lista_cortes_preventivos')

    context = {
        'periodos':     periodos,
        'grados':       grados,
        'fecha_hoy':    date.today().isoformat(),
    }
    return render(request, 'gestion_academica/cortes_preventivos/crear.html', context)


# ──────────────────────────────────────────────────────────────────────────────
#  3. DETALLE / DASHBOARD DEL CORTE
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def detalle_corte_preventivo(request, pk):
    institucion = _get_institucion(request)
    corte = get_object_or_404(
        CortePreventivo, pk=pk,
        institucion=institucion if not request.user.is_superuser else CortePreventivo.objects.get(pk=pk).institucion
    )

    resultados_qs = ResultadoCorteEstudiante.objects.filter(
        corte=corte
    ).select_related('estudiante__usuario').prefetch_related('detalles_materias__curso__materia')

    # Filtros del dashboard
    filtro_riesgo = request.GET.get('riesgo', '')
    filtro_citacion = request.GET.get('citacion', '')
    if filtro_riesgo:
        resultados_qs = resultados_qs.filter(nivel_riesgo=filtro_riesgo)
    if filtro_citacion == '1':
        resultados_qs = resultados_qs.filter(requiere_citacion_padres=True)
    elif filtro_citacion == '0':
        resultados_qs = resultados_qs.filter(requiere_citacion_padres=False)

    # Estadísticas
    todos_resultados = ResultadoCorteEstudiante.objects.filter(corte=corte)
    stats = {
        'total':       todos_resultados.count(),
        'alto':        todos_resultados.filter(nivel_riesgo='ALTO').count(),
        'medio':       todos_resultados.filter(nivel_riesgo='MEDIO').count(),
        'bajo':        todos_resultados.filter(nivel_riesgo='BAJO').count(),
        'sin_riesgo':  todos_resultados.filter(nivel_riesgo='SIN_RIESGO').count(),
        'con_citacion': todos_resultados.filter(requiere_citacion_padres=True).count(),
    }

    # Cursos del grado para el header
    cursos = Curso.objects.filter(
        grado=corte.grado, periodo_academico=corte.periodo_academico,
        institucion=corte.institucion
    ).select_related('materia').order_by('materia__nombre_materia')

    context = {
        'corte':          corte,
        'resultados':     resultados_qs,
        'cursos':         cursos,
        'stats':          stats,
        'filtro_riesgo':  filtro_riesgo,
        'filtro_citacion': filtro_citacion,
        'puede_gestionar': _puede_gestionar_cortes(request.user),
        'color_riesgo':    {r[0]: _color_riesgo(r[0]) for r in ResultadoCorteEstudiante.RIESGO_CHOICES},
    }
    return render(request, 'gestion_academica/cortes_preventivos/detalle.html', context)


# ──────────────────────────────────────────────────────────────────────────────
#  4. LANZAR CÁLCULO (async Celery)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@require_POST
def calcular_corte(request, pk):
    institucion = _get_institucion(request)
    corte = get_object_or_404(CortePreventivo, pk=pk, institucion=institucion)

    if not _puede_gestionar_cortes(request.user):
        messages.error(request, "No tienes permiso para lanzar el cálculo.")
        return redirect('gestion_academica:detalle_corte_preventivo', pk=pk)

    from ..tasks import calcular_corte_preventivo_task
    calcular_corte_preventivo_task.delay(corte.pk, user_id=request.user.pk)

    corte.estado = 'CALCULANDO'
    corte.save(update_fields=['estado'])

    messages.success(
        request,
        "⏳ Cálculo iniciado en segundo plano. "
        "Recibirás una notificación cuando esté listo."
    )
    return redirect('gestion_academica:detalle_corte_preventivo', pk=pk)


# ──────────────────────────────────────────────────────────────────────────────
#  5. PUBLICAR CORTE
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@require_POST
def publicar_corte(request, pk):
    institucion = _get_institucion(request)
    corte = get_object_or_404(CortePreventivo, pk=pk, institucion=institucion)

    if not _puede_gestionar_cortes(request.user):
        messages.error(request, "No tienes permiso para publicar cortes.")
        return redirect('gestion_academica:detalle_corte_preventivo', pk=pk)

    if corte.total_estudiantes_evaluados == 0:
        messages.warning(request, "⚠️ El corte no tiene resultados calculados todavía. Lanza el cálculo primero.")
        return redirect('gestion_academica:detalle_corte_preventivo', pk=pk)

    corte.estado = 'PUBLICADO'
    corte.fecha_publicacion = timezone.now()
    corte.save(update_fields=['estado', 'fecha_publicacion'])

    messages.success(request, f"✅ Corte '{corte.nombre_corte}' publicado correctamente.")
    return redirect('gestion_academica:detalle_corte_preventivo', pk=pk)


# ──────────────────────────────────────────────────────────────────────────────
#  6. ARCHIVAR CORTE
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@require_POST
def archivar_corte(request, pk):
    institucion = _get_institucion(request)
    corte = get_object_or_404(CortePreventivo, pk=pk, institucion=institucion)

    if not _puede_gestionar_cortes(request.user):
        messages.error(request, "No tienes permiso.")
        return redirect('gestion_academica:detalle_corte_preventivo', pk=pk)

    corte.estado = 'ARCHIVADO'
    corte.save(update_fields=['estado'])
    messages.success(request, f"Corte '{corte.nombre_corte}' archivado.")
    return redirect('gestion_academica:lista_cortes_preventivos')


# ──────────────────────────────────────────────────────────────────────────────
#  7. ELIMINAR CORTE
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@require_POST
def eliminar_corte(request, pk):
    institucion = _get_institucion(request)
    corte = get_object_or_404(CortePreventivo, pk=pk, institucion=institucion)

    if not _puede_gestionar_cortes(request.user):
        messages.error(request, "No tienes permiso.")
        return redirect('gestion_academica:lista_cortes_preventivos')

    nombre = corte.nombre_corte
    corte.delete()
    messages.success(request, f"Corte '{nombre}' eliminado correctamente.")
    return redirect('gestion_academica:lista_cortes_preventivos')


# ──────────────────────────────────────────────────────────────────────────────
#  8. GUARDAR OBSERVACIÓN DEL COORDINADOR/DIRECTOR
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@require_POST
def guardar_observacion_corte(request, pk):
    """Guarda la observación general del coordinador en el corte."""
    institucion = _get_institucion(request)
    corte = get_object_or_404(CortePreventivo, pk=pk, institucion=institucion)
    obs = request.POST.get('observacion_general', '').strip()
    corte.observacion_general = obs
    corte.save(update_fields=['observacion_general'])
    messages.success(request, "Observación guardada correctamente.")
    return redirect('gestion_academica:detalle_corte_preventivo', pk=pk)


@login_required
@require_POST
def guardar_observacion_estudiante(request, pk, resultado_pk):
    """Guarda la observación del director de curso para un estudiante."""
    institucion = _get_institucion(request)
    corte    = get_object_or_404(CortePreventivo, pk=pk, institucion=institucion)
    resultado = get_object_or_404(ResultadoCorteEstudiante, pk=resultado_pk, corte=corte, institucion=institucion)

    obs = request.POST.get('observacion', '').strip()
    citar = request.POST.get('requiere_citacion') == '1'

    resultado.observacion_director_curso = obs
    resultado.requiere_citacion_padres   = citar
    resultado.save(update_fields=['observacion_director_curso', 'requiere_citacion_padres'])

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'ok': True})
    messages.success(request, "Observación del estudiante guardada.")
    return redirect('gestion_academica:detalle_corte_preventivo', pk=pk)


@login_required
@require_POST
def guardar_observacion_materia(request, pk, detalle_pk):
    """Guarda la observación del docente en un DetalleMateriaCortePrev."""
    institucion = _get_institucion(request)
    corte   = get_object_or_404(CortePreventivo, pk=pk, institucion=institucion)
    detalle = get_object_or_404(DetalleMateriaCortePrev, pk=detalle_pk, institucion=institucion)

    obs = request.POST.get('observacion', '').strip()
    detalle.observacion_docente = obs
    detalle.save(update_fields=['observacion_docente'])

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'ok': True})
    messages.success(request, "Observación del docente guardada.")
    return redirect('gestion_academica:detalle_corte_preventivo', pk=corte.pk)


# ──────────────────────────────────────────────────────────────────────────────
#  9. ENVIAR NOTIFICACIONES MASIVAS
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@require_POST
def notificar_familias_corte(request, pk):
    institucion = _get_institucion(request)
    corte = get_object_or_404(CortePreventivo, pk=pk, institucion=institucion)

    if not _puede_gestionar_cortes(request.user):
        messages.error(request, "No tienes permiso para enviar notificaciones.")
        return redirect('gestion_academica:detalle_corte_preventivo', pk=pk)

    solo_riesgo = request.POST.get('solo_riesgo') == '1'
    resultados  = ResultadoCorteEstudiante.objects.filter(corte=corte, notificacion_enviada=False)
    if solo_riesgo:
        resultados = resultados.filter(nivel_riesgo__in=['ALTO', 'MEDIO'])

    enviadas = 0
    for resultado in resultados.select_related('estudiante__usuario'):
        # Buscar familiar del estudiante
        familiares_qs = resultado.estudiante.usuario.familiares_asociados.all() \
            if hasattr(resultado.estudiante.usuario, 'familiares_asociados') else []

        # Notificación interna al propio estudiante
        Notificacion.objects.create(
            destinatario=resultado.estudiante.usuario,
            mensaje=(
                f"Se ha generado el Corte Preventivo «{corte.nombre_corte}». "
                f"Tu nivel de desempeño actual es: {resultado.get_nivel_desempeno_general_display()}. "
                f"Ingresa a tu portal para más detalles."
            ),
            institucion=institucion,
        )
        resultado.notificacion_enviada = True
        resultado.fecha_notificacion   = timezone.now()
        resultado.save(update_fields=['notificacion_enviada', 'fecha_notificacion'])
        enviadas += 1

    # Actualizar total en riesgo
    corte.total_en_riesgo = ResultadoCorteEstudiante.objects.filter(
        corte=corte, nivel_riesgo__in=['ALTO', 'MEDIO']
    ).count()
    corte.save(update_fields=['total_en_riesgo'])

    messages.success(request, f"✅ {enviadas} notificación(es) enviada(s) correctamente.")
    return redirect('gestion_academica:detalle_corte_preventivo', pk=pk)


# ──────────────────────────────────────────────────────────────────────────────
#  10. EXPORTAR PDF POR GRADO
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def exportar_pdf_grado(request, pk):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, KeepTogether,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    institucion = _get_institucion(request)
    corte = get_object_or_404(CortePreventivo, pk=pk, institucion=institucion)

    resultados = ResultadoCorteEstudiante.objects.filter(
        corte=corte
    ).select_related('estudiante__usuario').prefetch_related(
        'detalles_materias__curso__materia'
    ).order_by('estudiante__usuario__last_name', 'estudiante__usuario__first_name')

    cursos = Curso.objects.filter(
        grado=corte.grado, periodo_academico=corte.periodo_academico,
        institucion=corte.institucion
    ).select_related('materia').order_by('materia__nombre_materia')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    COLOR_PRIMARIO  = colors.HexColor('#1e3a5f')
    COLOR_ACENTO    = colors.HexColor('#2563eb')
    COLOR_RIESGO_A  = colors.HexColor('#fee2e2')
    COLOR_RIESGO_M  = colors.HexColor('#fef9c3')
    COLOR_OK        = colors.HexColor('#f0fdf4')
    COLOR_CABECERA  = colors.HexColor('#1e3a5f')

    estilo_titulo   = ParagraphStyle('titulo',   fontName='Helvetica-Bold', fontSize=16,
                                      textColor=COLOR_PRIMARIO, alignment=TA_CENTER, spaceAfter=4)
    estilo_sub      = ParagraphStyle('sub',      fontName='Helvetica',      fontSize=10,
                                      textColor=colors.HexColor('#374151'), alignment=TA_CENTER, spaceAfter=2)
    estilo_alerta   = ParagraphStyle('alerta',   fontName='Helvetica-Oblique', fontSize=8,
                                      textColor=colors.HexColor('#6b7280'), alignment=TA_CENTER, spaceAfter=6)
    estilo_seccion  = ParagraphStyle('seccion',  fontName='Helvetica-Bold', fontSize=11,
                                      textColor=COLOR_ACENTO, spaceBefore=10, spaceAfter=4)
    estilo_celda    = ParagraphStyle('celda',    fontName='Helvetica',      fontSize=7,
                                      leading=9)
    estilo_celda_b  = ParagraphStyle('celdab',   fontName='Helvetica-Bold', fontSize=7,
                                      leading=9)

    elements = []

    # ── Encabezado del documento
    elements.append(Paragraph(
        f"CORTE PREVENTIVO ACADÉMICO — {corte.nombre_corte.upper()}",
        estilo_titulo
    ))
    elements.append(Paragraph(
        f"{institucion.nombre}  |  Grado: {corte.grado.nombre}  "
        f"|  Período: {corte.periodo_academico}  |  Fecha de Corte: {corte.fecha_corte.strftime('%d/%m/%Y')}",
        estilo_sub
    ))
    elements.append(Paragraph(
        "★ INFORME DE CARÁCTER PREVENTIVO — No constituye boletín oficial de calificaciones ★",
        estilo_alerta
    ))
    elements.append(HRFlowable(width="100%", thickness=2, color=COLOR_ACENTO, spaceAfter=8))

    # ── Resumen estadístico
    total       = resultados.count()
    alto        = resultados.filter(nivel_riesgo='ALTO').count()
    medio       = resultados.filter(nivel_riesgo='MEDIO').count()
    sin_riesgo  = resultados.filter(nivel_riesgo='SIN_RIESGO').count() + resultados.filter(nivel_riesgo='BAJO').count()
    promedios   = [r.promedio_general for r in resultados if r.promedio_general]
    prom_grado  = (sum(promedios) / len(promedios)).quantize(Decimal('0.01')) if promedios else '—'

    resumen_data = [
        ['Total Evaluados', 'En Riesgo Alto', 'En Riesgo Medio', 'Sin Riesgo', 'Promedio General'],
        [str(total), str(alto), str(medio), str(sin_riesgo), str(prom_grado)],
    ]
    t_resumen = Table(resumen_data, colWidths=[5*cm]*5, hAlign='CENTER')
    t_resumen.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), COLOR_CABECERA),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 9),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white]),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#d1d5db')),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    elements.append(t_resumen)
    elements.append(Spacer(1, 10))

    if corte.observacion_general:
        elements.append(Paragraph(f"<b>Observación del Coordinador:</b> {corte.observacion_general}", estilo_celda))
        elements.append(Spacer(1, 6))

    # ── Tabla consolidada de estudiantes
    elements.append(Paragraph("RESUMEN POR ESTUDIANTE", estilo_seccion))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb'), spaceAfter=4))

    nombres_materias = [c.materia.nombre_materia[:8] for c in cursos]
    cabecera_tabla = (
        ['#', 'Apellidos y Nombres', 'Prom.', 'Desempeño', 'Asistencia%', 'M.Riesgo', 'Riesgo'] +
        nombres_materias +
        ['Obs. Director']
    )

    col_widths = [0.5*cm, 4.5*cm, 1.3*cm, 2*cm, 1.8*cm, 1.3*cm, 1.8*cm]
    col_widths += [1.3*cm] * len(cursos)
    col_widths += [4*cm]

    tabla_data = [cabecera_tabla]
    for i, res in enumerate(resultados, 1):
        nombre = (
            f"{res.estudiante.usuario.last_name.upper()}, "
            f"{res.estudiante.usuario.first_name}"
        )
        asist = f"{res.porcentaje_asistencia}%" if res.porcentaje_asistencia else '—'
        prom  = str(res.promedio_general) if res.promedio_general else '—'
        nivel = {
            'SUPERIOR':'Superior','ALTO':'Alto','BASICO':'Básico','BAJO':'Bajo','SIN_DATOS':'—'
        }.get(res.nivel_desempeno_general, '—')
        riesgo_txt = {
            'ALTO':'🔴 ALTO','MEDIO':'🟡 MEDIO','BAJO':'🔵 BAJO','SIN_RIESGO':'🟢 OK'
        }.get(res.nivel_riesgo, '—')

        notas_por_curso = {d.curso_id: d.promedio_materia for d in res.detalles_materias.all()}
        notas_fila = [
            str(notas_por_curso.get(c.pk, '—')) if notas_por_curso.get(c.pk) else '—'
            for c in cursos
        ]

        obs = res.observacion_director_curso[:60] + '...' if len(res.observacion_director_curso) > 60 else res.observacion_director_curso

        fila = [str(i), nombre, prom, nivel, asist, str(res.materias_en_riesgo_count), riesgo_txt] + notas_fila + [obs]
        tabla_data.append(fila)

    t = Table(tabla_data, colWidths=col_widths, repeatRows=1)
    # Estilos base
    estilo_tabla = [
        ('BACKGROUND',  (0,0), (-1,0), COLOR_CABECERA),
        ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,-1), 7),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
        ('GRID',        (0,0), (-1,-1), 0.3, colors.HexColor('#d1d5db')),
        ('TOPPADDING',    (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
    ]
    # Color por nivel de riesgo
    for i, res in enumerate(resultados, 1):
        if res.nivel_riesgo == 'ALTO':
            estilo_tabla.append(('BACKGROUND', (0,i), (6,i), COLOR_RIESGO_A))
        elif res.nivel_riesgo == 'MEDIO':
            estilo_tabla.append(('BACKGROUND', (0,i), (6,i), COLOR_RIESGO_M))
        elif res.nivel_riesgo in ('SIN_RIESGO', 'BAJO'):
            estilo_tabla.append(('BACKGROUND', (0,i), (6,i), COLOR_OK))

    t.setStyle(TableStyle(estilo_tabla))
    elements.append(t)

    # ── Pie de página con texto de la institución
    elements.append(Spacer(1, 12))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb')))
    try:
        cfg = ConfiguracionCortePreventivo.objects.get(institucion=institucion)
        pie_texto = cfg.texto_pie_pagina
    except ConfiguracionCortePreventivo.DoesNotExist:
        pie_texto = "Este informe es de carácter preventivo y no constituye el boletín oficial de calificaciones."
    elements.append(Paragraph(pie_texto, ParagraphStyle('pie', fontName='Helvetica-Oblique', fontSize=7,
                                                          textColor=colors.HexColor('#9ca3af'), alignment=TA_CENTER)))
    elements.append(Paragraph(
        f"Generado: {date.today().strftime('%d/%m/%Y')}  |  {institucion.nombre}",
        ParagraphStyle('pie2', fontName='Helvetica', fontSize=7,
                        textColor=colors.HexColor('#9ca3af'), alignment=TA_CENTER)
    ))

    doc.build(elements)
    buffer.seek(0)
    nombre_archivo = f"corte_preventivo_{corte.grado.nombre.replace(' ','_')}_{corte.fecha_corte}.pdf"
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


# ──────────────────────────────────────────────────────────────────────────────
#  11. EXPORTAR PDF INDIVIDUAL DE ESTUDIANTE
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def exportar_pdf_estudiante(request, pk, resultado_pk):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    institucion = _get_institucion(request)
    corte    = get_object_or_404(CortePreventivo, pk=pk, institucion=institucion)
    resultado = get_object_or_404(
        ResultadoCorteEstudiante, pk=resultado_pk, corte=corte, institucion=institucion
    )
    detalles  = resultado.detalles_materias.select_related(
        'curso__materia', 'curso__periodo_academico'
    ).prefetch_related('curso__docentes_asignados__usuario').order_by('curso__materia__nombre_materia')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                             leftMargin=2*cm, rightMargin=2*cm,
                             topMargin=2*cm, bottomMargin=2*cm)

    COLOR_PRIMARIO = colors.HexColor('#1e3a5f')
    COLOR_ACENTO   = colors.HexColor('#2563eb')
    RIESGO_COLORES = {
        'ALTO':       colors.HexColor('#fee2e2'),
        'MEDIO':      colors.HexColor('#fef9c3'),
        'BAJO':       colors.HexColor('#dbeafe'),
        'SIN_RIESGO': colors.HexColor('#dcfce7'),
    }
    RIESGO_TEXTO = {
        'ALTO': '🔴 RIESGO ALTO', 'MEDIO': '🟡 RIESGO MEDIO',
        'BAJO': '🔵 RIESGO BAJO', 'SIN_RIESGO': '🟢 SIN RIESGO',
    }

    s = getSampleStyleSheet()
    est = lambda name, **kw: ParagraphStyle(name, **kw)
    titulo  = est('tit', fontName='Helvetica-Bold', fontSize=15, textColor=COLOR_PRIMARIO, alignment=TA_CENTER, spaceAfter=2)
    sub     = est('sub', fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#374151'), alignment=TA_CENTER, spaceAfter=2)
    alerta  = est('alt', fontName='Helvetica-Oblique', fontSize=7, textColor=colors.HexColor('#6b7280'), alignment=TA_CENTER, spaceAfter=8)
    seccion = est('sec', fontName='Helvetica-Bold', fontSize=10, textColor=COLOR_ACENTO, spaceBefore=8, spaceAfter=4)
    normal  = est('nor', fontName='Helvetica', fontSize=9, leading=13)
    pie_s   = est('pie', fontName='Helvetica-Oblique', fontSize=7, textColor=colors.HexColor('#9ca3af'), alignment=TA_CENTER)

    elems = []
    estudiante = resultado.estudiante
    nombre_completo = f"{estudiante.usuario.last_name.upper()} {estudiante.usuario.first_name}"

    # ── Encabezado
    elems.append(Paragraph(f"INFORME DE CORTE PREVENTIVO", titulo))
    elems.append(Paragraph(f"{institucion.nombre}", sub))
    elems.append(Paragraph(
        f"Grado: {corte.grado.nombre}  |  Período: {corte.periodo_academico}  "
        f"|  Fecha de Corte: {corte.fecha_corte.strftime('%d/%m/%Y')}",
        sub
    ))
    elems.append(Paragraph(
        "★ Informe preventivo — no constituye boletín oficial ★", alerta
    ))
    elems.append(HRFlowable(width="100%", thickness=2, color=COLOR_ACENTO, spaceAfter=8))

    # ── Datos del estudiante
    elems.append(Paragraph("DATOS DEL ESTUDIANTE", seccion))
    datos_est = [
        ['Nombre Completo:', nombre_completo,
         'Documento:', estudiante.documento_identidad or '—'],
        ['Grado Actual:', corte.grado.nombre,
         'Código:', estudiante.codigo_estudiante or '—'],
    ]
    t_est = Table(datos_est, colWidths=[3.5*cm, 7*cm, 3*cm, 4*cm])
    t_est.setStyle(TableStyle([
        ('FONTNAME',  (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME',  (2,0), (2,-1), 'Helvetica-Bold'),
        ('FONTSIZE',  (0,0), (-1,-1), 9),
        ('VALIGN',    (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#f1f5f9')),
        ('BACKGROUND', (2,0), (2,-1), colors.HexColor('#f1f5f9')),
    ]))
    elems.append(t_est)
    elems.append(Spacer(1, 8))

    # ── Banner de nivel de riesgo
    color_banner = RIESGO_COLORES.get(resultado.nivel_riesgo, colors.white)
    riesgo_txt   = RIESGO_TEXTO.get(resultado.nivel_riesgo, '—')
    prom_txt     = str(resultado.promedio_general) if resultado.promedio_general else '—'
    nivel_txt    = {'SUPERIOR':'Superior','ALTO':'Alto','BASICO':'Básico','BAJO':'Bajo','SIN_DATOS':'—'}.get(
        resultado.nivel_desempeno_general, '—'
    )
    asist_txt = f"{resultado.porcentaje_asistencia}%" if resultado.porcentaje_asistencia else '—'

    banner_data = [
        [riesgo_txt, f"Promedio: {prom_txt}/5.0", f"Desempeño: {nivel_txt}", f"Asistencia: {asist_txt}"]
    ]
    t_banner = Table(banner_data, colWidths=[4.5*cm, 4*cm, 4*cm, 5*cm], hAlign='CENTER')
    t_banner.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), color_banner),
        ('FONTNAME',   (0,0), (0,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 10),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('BOX',        (0,0), (-1,-1), 1.5, COLOR_ACENTO),
    ]))
    elems.append(t_banner)
    elems.append(Spacer(1, 10))

    # ── Tabla de notas por materia
    elems.append(Paragraph("NOTAS POR MATERIA", seccion))
    elems.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb'), spaceAfter=4))

    cab_mat = ['Materia', 'Docente', 'Promedio', 'Desempeño', 'Actvs.Total', 'Calificadas', 'Pendientes', 'Estado']
    mat_data = [cab_mat]
    for d in detalles:
        docentes_txt = ', '.join(
            doc.usuario.get_full_name() or doc.usuario.username
            for doc in d.curso.docentes_asignados.all()
        ) or '—'
        prom_m  = str(d.promedio_materia) if d.promedio_materia else '—'
        nivel_m = {'SUPERIOR':'Superior','ALTO':'Alto','BASICO':'Básico','BAJO':'Bajo','SIN_DATOS':'—'}.get(d.nivel_desempeno, '—')
        estado_m = '⚠️ RIESGO' if d.en_riesgo else '✅ OK'
        mat_data.append([
            d.curso.materia.nombre_materia, docentes_txt, prom_m, nivel_m,
            str(d.actividades_registradas), str(d.actividades_calificadas),
            str(d.actividades_pendientes), estado_m,
        ])

    col_w_mat = [4*cm, 3.5*cm, 1.8*cm, 2.2*cm, 2*cm, 2.2*cm, 2.2*cm, 2*cm]
    t_mat = Table(mat_data, colWidths=col_w_mat, repeatRows=1)
    est_mat = [
        ('BACKGROUND', (0,0), (-1,0), COLOR_PRIMARIO),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('ALIGN',      (1,0), (-1,-1), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('GRID',       (0,0), (-1,-1), 0.3, colors.HexColor('#d1d5db')),
        ('TOPPADDING',    (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
    ]
    for i, d in enumerate(detalles, 1):
        if d.en_riesgo:
            est_mat.append(('BACKGROUND', (0,i), (3,i), colors.HexColor('#fee2e2')))
    t_mat.setStyle(TableStyle(est_mat))
    elems.append(t_mat)

    # ── Observaciones
    if resultado.observacion_director_curso:
        elems.append(Spacer(1, 8))
        elems.append(Paragraph("OBSERVACIÓN DEL DIRECTOR DE CURSO", seccion))
        elems.append(Paragraph(resultado.observacion_director_curso, normal))

    obs_docs = [(d.curso.materia.nombre_materia, d.observacion_docente) for d in detalles if d.observacion_docente]
    if obs_docs:
        elems.append(Spacer(1, 6))
        elems.append(Paragraph("OBSERVACIONES DE DOCENTES", seccion))
        for materia, obs in obs_docs:
            elems.append(Paragraph(f"<b>{materia}:</b> {obs}", normal))

    # ── Sección de firmas
    elems.append(Spacer(1, 16))
    elems.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb')))
    firmas_data = [
        ['Director de Curso', '', 'Padre/Madre o Acudiente'],
        ['________________________', 'Fecha de enterado: ________', '________________________'],
        ['Firma y sello', '', 'Firma'],
    ]
    t_firmas = Table(firmas_data, colWidths=[6*cm, 5*cm, 6*cm], hAlign='CENTER')
    t_firmas.setStyle(TableStyle([
        ('FONTNAME',  (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',  (0,0), (-1,-1), 8),
        ('ALIGN',     (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',    (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
    ]))
    elems.append(t_firmas)

    # ── Pie
    elems.append(Spacer(1, 8))
    try:
        cfg = ConfiguracionCortePreventivo.objects.get(institucion=institucion)
        pie_texto = cfg.texto_pie_pagina
    except ConfiguracionCortePreventivo.DoesNotExist:
        pie_texto = "Este informe es de carácter preventivo y no constituye el boletín oficial."
    elems.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb')))
    elems.append(Paragraph(pie_texto, pie_s))
    elems.append(Paragraph(f"Generado: {date.today().strftime('%d/%m/%Y')}  |  {institucion.nombre}", pie_s))

    doc.build(elems)
    buffer.seek(0)
    nombre_archivo = f"corte_prev_{nombre_completo.replace(' ','_')}_{corte.fecha_corte}.pdf"
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


# ──────────────────────────────────────────────────────────────────────────────
#  12. EXPORTAR EXCEL
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def exportar_excel_corte(request, pk):
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    institucion = _get_institucion(request)
    corte = get_object_or_404(CortePreventivo, pk=pk, institucion=institucion)

    resultados = list(
        ResultadoCorteEstudiante.objects.filter(corte=corte)
        .select_related('estudiante__usuario')
        .prefetch_related('detalles_materias__curso__materia')
        .order_by('estudiante__usuario__last_name', 'estudiante__usuario__first_name')
    )
    cursos = list(
        Curso.objects.filter(
            grado=corte.grado, periodo_academico=corte.periodo_academico, institucion=corte.institucion
        ).select_related('materia').order_by('materia__nombre_materia')
    )

    wb = openpyxl.Workbook()

    # ── Estilos
    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    thin = Side(style='thin', color='D1D5DB')
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    F_CABECERA  = fill('1E3A5F'); F_RIESGO_A = fill('FEE2E2'); F_RIESGO_M = fill('FEF9C3')
    F_OK        = fill('F0FDF4'); F_GRIS     = fill('F8FAFC'); F_ACENTO   = fill('2563EB')
    FONT_BLANCO = Font(color='FFFFFF', bold=True, size=10)
    FONT_BOLD   = Font(bold=True, size=10)
    FONT_NORMAL = Font(size=9)

    def set_header(ws, row, col, value, bg=None, font=None, wrap=True):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = bg or F_CABECERA
        c.font = font or FONT_BLANCO
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
        c.border = borde
        return c

    # ── HOJA 1: Datos del grado
    ws = wb.active
    ws.title = f"Grado {corte.grado.nombre}"[:31]
    ws.freeze_panes = 'A9'

    # Encabezado institucional (filas 1-6)
    ws.merge_cells('A1:Z1')
    c = ws['A1']
    c.value = institucion.nombre.upper()
    c.font  = Font(bold=True, size=13, color='1E3A5F')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    ws.merge_cells('A2:Z2')
    ws['A2'].value = f"CORTE PREVENTIVO — {corte.nombre_corte}"
    ws['A2'].font  = Font(bold=True, size=11, color='2563EB')
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:Z3')
    ws['A3'].value = (
        f"Grado: {corte.grado.nombre}  |  Período: {corte.periodo_academico}  "
        f"|  Fecha de Corte: {corte.fecha_corte.strftime('%d/%m/%Y')}  "
        f"|  Generado: {date.today().strftime('%d/%m/%Y')}"
    )
    ws['A3'].font      = Font(size=9, color='374151')
    ws['A3'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A4:Z4')
    ws['A4'].value     = "★ Informe preventivo — no constituye boletín oficial de calificaciones ★"
    ws['A4'].font      = Font(italic=True, size=8, color='9CA3AF')
    ws['A4'].alignment = Alignment(horizontal='center')

    # Estadísticas en fila 6-7
    total_e  = len(resultados)
    alto_c   = sum(1 for r in resultados if r.nivel_riesgo == 'ALTO')
    medio_c  = sum(1 for r in resultados if r.nivel_riesgo == 'MEDIO')
    ok_c     = total_e - alto_c - medio_c
    promedios_l = [r.promedio_general for r in resultados if r.promedio_general]
    prom_g   = round(sum(promedios_l) / len(promedios_l), 2) if promedios_l else '—'

    stats_cabeceras = ['Total Evaluados', 'Riesgo Alto', 'Riesgo Medio', 'Sin Riesgo', 'Prom. General']
    stats_valores   = [total_e, alto_c, medio_c, ok_c, prom_g]
    for i, (cab, val) in enumerate(zip(stats_cabeceras, stats_valores), 1):
        ws.cell(row=6, column=i, value=cab).font  = Font(bold=True, size=9, color='FFFFFF')
        ws.cell(row=6, column=i).fill             = F_CABECERA
        ws.cell(row=6, column=i).alignment        = Alignment(horizontal='center')
        ws.cell(row=7, column=i, value=val).font  = Font(bold=True, size=11)
        ws.cell(row=7, column=i).alignment        = Alignment(horizontal='center')
    ws.row_dimensions[6].height = 16
    ws.row_dimensions[7].height = 20

    # ── Cabecera de tabla (fila 8)
    col = 1
    cabeceras_fijas = ['#', 'Código', 'Apellidos', 'Nombres', 'Prom. Gral.',
                        'Desempeño', '% Asistencia', 'Nivel Riesgo', 'Mats. Riesgo',
                        'Citación', 'Obs. Director']
    for cab in cabeceras_fijas:
        set_header(ws, 8, col, cab)
        col += 1
    # Columnas por materia
    for curso in cursos:
        set_header(ws, 8, col, curso.materia.nombre_materia)
        col += 1
    ws.row_dimensions[8].height = 30

    # ── Filas de datos (desde fila 9)
    RIESGO_LABEL = {'ALTO':'🔴 ALTO','MEDIO':'🟡 MEDIO','BAJO':'🔵 BAJO','SIN_RIESGO':'🟢 OK'}
    NIVEL_LABEL  = {'SUPERIOR':'Superior','ALTO':'Alto','BASICO':'Básico','BAJO':'Bajo','SIN_DATOS':'—'}

    for idx, res in enumerate(resultados, 1):
        fila = 8 + idx
        est  = res.estudiante
        notas_dict = {d.curso_id: d.promedio_materia for d in res.detalles_materias.all()}

        valores = [
            idx,
            est.codigo_estudiante or '—',
            est.usuario.last_name.upper(),
            est.usuario.first_name,
            float(res.promedio_general) if res.promedio_general else None,
            NIVEL_LABEL.get(res.nivel_desempeno_general, '—'),
            float(res.porcentaje_asistencia) if res.porcentaje_asistencia else None,
            RIESGO_LABEL.get(res.nivel_riesgo, '—'),
            res.materias_en_riesgo_count,
            '✅ Sí' if res.requiere_citacion_padres else 'No',
            res.observacion_director_curso,
        ]
        for curso in cursos:
            nota = notas_dict.get(curso.pk)
            valores.append(float(nota) if nota else None)

        for col_i, val in enumerate(valores, 1):
            c = ws.cell(row=fila, column=col_i, value=val)
            c.font      = FONT_NORMAL
            c.border    = borde
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=(col_i == 11))

        # Colorear fila según riesgo
        f_riesgo = F_RIESGO_A if res.nivel_riesgo == 'ALTO' else (
                    F_RIESGO_M if res.nivel_riesgo == 'MEDIO' else (
                    F_GRIS if idx % 2 == 0 else None))
        if f_riesgo:
            for col_i in range(1, 9):
                ws.cell(row=fila, column=col_i).fill = f_riesgo

        # Colorear celdas de notas por materia
        for j, curso in enumerate(cursos):
            col_nota = 12 + j
            nota_val = notas_dict.get(curso.pk)
            if nota_val is not None:
                if nota_val < Decimal('3.0'):
                    ws.cell(row=fila, column=col_nota).fill = F_RIESGO_A
                elif nota_val < Decimal('3.5'):
                    ws.cell(row=fila, column=col_nota).fill = F_RIESGO_M
                else:
                    ws.cell(row=fila, column=col_nota).fill = F_OK

        ws.row_dimensions[fila].height = 18

    # Ajustar anchos
    anchos = [4, 10, 22, 18, 10, 12, 13, 12, 12, 10, 35]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for j in range(len(cursos)):
        ws.column_dimensions[get_column_letter(12 + j)].width = 12

    # ── HOJA 2: Estudiantes en Riesgo Alto
    ws2 = wb.create_sheet("En Riesgo Alto")
    ws2.merge_cells('A1:G1')
    ws2['A1'].value = f"ESTUDIANTES EN RIESGO ALTO — {corte.nombre_corte}"
    ws2['A1'].font  = Font(bold=True, size=12, color='DC2626')
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2.row_dimensions[1].height = 20

    for i, cab in enumerate(['Apellidos', 'Nombres', 'Promedio', '% Asistencia', 'Materias en Riesgo', 'Citación', 'Obs. Director'], 1):
        set_header(ws2, 2, i, cab)
    ws2.row_dimensions[2].height = 20

    fila2 = 3
    for res in resultados:
        if res.nivel_riesgo != 'ALTO':
            continue
        est = res.estudiante
        datos = [
            est.usuario.last_name.upper(),
            est.usuario.first_name,
            float(res.promedio_general) if res.promedio_general else None,
            float(res.porcentaje_asistencia) if res.porcentaje_asistencia else None,
            res.materias_en_riesgo_count,
            'Sí' if res.requiere_citacion_padres else 'No',
            res.observacion_director_curso,
        ]
        for ci, val in enumerate(datos, 1):
            c = ws2.cell(row=fila2, column=ci, value=val)
            c.fill   = F_RIESGO_A
            c.border = borde
            c.font   = FONT_NORMAL
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=(ci == 7))
        ws2.row_dimensions[fila2].height = 18
        fila2 += 1

    for i, w in enumerate([22, 18, 10, 13, 16, 10, 40], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Guardar y responder
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre_archivo = f"corte_preventivo_{corte.grado.nombre.replace(' ','_')}_{corte.fecha_corte}.xlsx"
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


# ──────────────────────────────────────────────────────────────────────────────
#  13. CONFIGURACIÓN DEL MÓDULO
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def configuracion_corte_preventivo(request):
    institucion = _get_institucion(request)
    if not institucion or not _puede_gestionar_cortes(request.user):
        messages.error(request, "No tienes permiso para cambiar esta configuración.")
        return redirect('gestion_academica:lista_cortes_preventivos')

    cfg, _ = ConfiguracionCortePreventivo.objects.get_or_create(
        institucion=institucion,
        defaults={
            'umbral_riesgo_bajo':  Decimal('2.9'),
            'umbral_riesgo_medio': Decimal('3.4'),
        }
    )

    if request.method == 'POST':

        def _to_decimal(key, default):
            """Convierte el valor del POST a Decimal, tolerando coma o punto."""
            raw = request.POST.get(key, '').strip().replace(',', '.')
            try:
                return Decimal(raw) if raw else Decimal(default)
            except Exception:
                return Decimal(default)

        def _to_int(key, default):
            raw = request.POST.get(key, '').strip()
            try:
                return int(raw) if raw else default
            except Exception:
                return default

        # Usar update() directo en el queryset → genera UPDATE SQL sin pasar
        # por validaciones del modelo que podrían bloquear silenciosamente.
        try:
            ConfiguracionCortePreventivo.objects.filter(pk=cfg.pk).update(
                umbral_riesgo_bajo            = _to_decimal('umbral_riesgo_bajo',  '2.9'),
                umbral_riesgo_medio           = _to_decimal('umbral_riesgo_medio', '3.4'),
                porcentaje_inasistencia_alerta = _to_int('porcentaje_inasistencia_alerta', 20),
                mostrar_promedio_parcial       = 'mostrar_promedio_parcial'      in request.POST,
                mostrar_asistencia             = 'mostrar_asistencia'            in request.POST,
                mostrar_observaciones_docente  = 'mostrar_observaciones_docente' in request.POST,
                firma_rector_en_reporte        = 'firma_rector_en_reporte'       in request.POST,
                permitir_descarga_familiar     = 'permitir_descarga_familiar'    in request.POST,
                texto_pie_pagina               = request.POST.get('texto_pie_pagina', '').strip(),
            )
            messages.success(request, "✅ Configuración guardada correctamente.")
        except Exception as e:
            messages.error(request, f"Error al guardar: {e}")
        return redirect('gestion_academica:configuracion_corte_preventivo')

    return render(request, 'gestion_academica/cortes_preventivos/configuracion.html', {'cfg': cfg})
