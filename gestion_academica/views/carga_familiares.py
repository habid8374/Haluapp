"""Carga masiva de Familiares / Acudientes (con enlace a estudiantes matriculados).

Imita la carga de estudiantes: una plantilla Excel descargable que trae —desde la
BD— la lista de estudiantes matriculados de la institución (hoja oculta + lista
desplegable + autocompletado), para enlazar Acudiente ↔ Estudiante.

Multi-institución: todo se filtra por la institución del usuario.
"""
import logging

import openpyxl
import pandas as pd
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from django.db.models import Q
from gestion_academica.models import Estudiante, Familiar, TIPO_DOCUMENTO_CHOICES

logger = logging.getLogger(__name__)
Usuario = get_user_model()

COLUMNAS = [
    "TIPO_DOC", "DOCUMENTO", "NOMBRES", "APELLIDOS", "PARENTESCO",
    "TELEFONO", "EMAIL", "OCUPACION", "LUGAR_TRABAJO", "DIRECCION",
    "DOCUMENTO_ESTUDIANTE", "NOMBRE_ESTUDIANTE", "RESPONSABLE_FACTURACION",
]
# A=TIPO_DOC  B=DOCUMENTO  C=NOMBRES  D=APELLIDOS  E=PARENTESCO  F=TELEFONO
# G=EMAIL  H=OCUPACION  I=LUGAR_TRABAJO  J=DIRECCION  K=DOCUMENTO_ESTUDIANTE
# L=NOMBRE_ESTUDIANTE (VLOOKUP automático)  M=RESPONSABLE_FACTURACION


def _institucion(request):
    return getattr(request.user, "institucion_asociada", None)


def _estudiantes_para_familiares(institucion):
    """Devuelve todos los estudiantes vinculables: activos + preliminares de admisiones.

    Los estudiantes creados durante la importación de aspirantes nacen con
    activo=False hasta que completan la matrícula. Igualmente necesitan acudientes
    para generar las facturas electrónicas de inscripción, así que los incluimos.
    """
    return (
        Estudiante.objects.filter(institucion=institucion)
        .filter(Q(activo=True) | Q(aspirante_origen__isnull=False))
        .select_related("usuario", "grado_actual")
        .order_by("usuario__last_name", "usuario__first_name")
        .distinct()
    )


@login_required
def descargar_plantilla_familiares(request):
    """Genera la plantilla Excel con los estudiantes matriculados embebidos."""
    institucion = _institucion(request)
    if not institucion and not request.user.is_superuser:
        messages.error(request, "Tu usuario no está asociado a ninguna institución.")
        return redirect("gestion_academica:inicio_academico")

    estudiantes = _estudiantes_para_familiares(institucion)

    wb = Workbook()

    # Hoja oculta con los estudiantes (documento | nombre | grado)
    data = wb.create_sheet(title="Estudiantes")
    data.append(["DOCUMENTO", "NOMBRE", "GRADO"])
    for e in estudiantes:
        data.append([
            e.documento_identidad or "",
            e.usuario.get_full_name() or e.usuario.username,
            str(e.grado_actual) if e.grado_actual else "",
        ])
    n = estudiantes.count()

    # Hoja de tipos de documento
    tdoc = wb.create_sheet(title="TiposDoc")
    for code, _label in TIPO_DOCUMENTO_CHOICES:
        tdoc.append([code])

    # Hoja principal
    ws = wb.active
    ws.title = "ACUDIENTES"
    ws.append(COLUMNAS)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4F46E5")

    # Validaciones
    dv_td = DataValidation(type="list", formula1="=TiposDoc!$A$1:$A$6", allow_blank=False)
    ws.add_data_validation(dv_td); dv_td.add("A2:A2000")

    if n:
        dv_est = DataValidation(type="list", formula1=f"=Estudiantes!$A$2:$A${n + 1}", allow_blank=False)
        ws.add_data_validation(dv_est); dv_est.add("K2:K2000")

    dv_resp = DataValidation(type="list", formula1='"SI,NO"', allow_blank=True)
    ws.add_data_validation(dv_resp); dv_resp.add("M2:M2000")

    # Autocompletado del nombre del estudiante (col L) desde la hoja oculta
    for r in range(2, 2001):
        ws[f"L{r}"] = f'=IFERROR(VLOOKUP(K{r},Estudiantes!A:B,2,FALSE),"")'

    data.sheet_state = "hidden"
    tdoc.sheet_state = "hidden"

    # Anchos
    for col, w in {
        "A": 10, "B": 16, "C": 18, "D": 18, "E": 14, "F": 14, "G": 24,
        "H": 18, "I": 22, "J": 26, "K": 18, "L": 26, "M": 12,
    }.items():
        ws.column_dimensions[col].width = w

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="plantilla_acudientes.xlsx"'
    wb.save(resp)
    return resp


@login_required
def cargar_familiares(request):
    """GET: muestra la página (descargar plantilla + subir). POST: procesa el archivo."""
    institucion = _institucion(request)
    if not institucion and not request.user.is_superuser:
        messages.error(request, "Tu usuario no está asociado a ninguna institución.")
        return redirect("gestion_academica:inicio_academico")

    if request.method == "POST" and request.FILES.get("archivo"):
        resultado = _procesar_archivo(request, institucion, request.FILES["archivo"])
        creados, vinculos, errores = resultado
        if creados or vinculos:
            messages.success(
                request,
                f"Proceso completado: {creados} acudiente(s) creados/actualizados, "
                f"{vinculos} vínculo(s) con estudiantes.",
            )
        for err in errores[:30]:
            messages.warning(request, err)
        if len(errores) > 30:
            messages.warning(request, f"... y {len(errores) - 30} error(es) más.")
        return redirect("gestion_academica:cargar_familiares")

    return render(request, "gestion_academica/carga_familiares.html", {
        "titulo_pagina": "Carga de Acudientes / Familiares",
        "total_estudiantes": _estudiantes_para_familiares(institucion).count(),
        "total_familiares": Familiar.objects.filter(institucion=institucion).count(),
    })


def _procesar_archivo(request, institucion, archivo):
    creados = 0
    vinculos = 0
    errores = []
    try:
        df = pd.read_excel(archivo, sheet_name="ACUDIENTES", dtype=str).fillna("")
    except Exception as exc:
        return 0, 0, [f"No se pudo leer el archivo (use la plantilla 'ACUDIENTES'): {exc}"]

    requeridas = {"TIPO_DOC", "DOCUMENTO", "NOMBRES", "APELLIDOS", "EMAIL", "DOCUMENTO_ESTUDIANTE"}
    if not requeridas.issubset(set(df.columns)):
        return 0, 0, [f"Faltan columnas requeridas: {', '.join(sorted(requeridas - set(df.columns)))}"]

    for idx, row in df.iterrows():
        fila = idx + 2
        doc = str(row.get("DOCUMENTO", "")).strip()
        doc_est = str(row.get("DOCUMENTO_ESTUDIANTE", "")).strip()
        email = str(row.get("EMAIL", "")).strip()
        if not doc or not doc_est:
            errores.append(f"Fila {fila}: falta documento del acudiente o del estudiante.")
            continue
        if not email:
            errores.append(f"Fila {fila}: el acudiente {doc} no tiene email (requerido para login y factura).")
            continue

        estudiante = Estudiante.objects.filter(
            documento_identidad=doc_est, institucion=institucion
        ).select_related("usuario").first()
        if not estudiante:
            errores.append(f"Fila {fila}: no existe estudiante matriculado con documento {doc_est}.")
            continue

        try:
            with transaction.atomic():
                familiar = Familiar.objects.filter(
                    documento_identidad=doc, institucion=institucion
                ).select_related("usuario").first()

                if not familiar:
                    # Crear Usuario (login) rol=familiar
                    username = email or doc
                    if Usuario.objects.filter(username=username).exists():
                        username = f"fam_{doc}"
                    user = Usuario.objects.create_user(
                        username=username,
                        email=email,
                        password=doc,  # contraseña temporal = documento (cambiar luego)
                        first_name=str(row.get("NOMBRES", "")).strip(),
                        last_name=str(row.get("APELLIDOS", "")).strip(),
                    )
                    user.rol = "familiar"
                    user.institucion_asociada = institucion
                    user.save()
                    familiar = Familiar.objects.create(
                        usuario=user,
                        parentesco=str(row.get("PARENTESCO", "")).strip() or "Acudiente",
                        telefono=str(row.get("TELEFONO", "")).strip(),
                        documento_identidad=doc,
                        tipo_documento=str(row.get("TIPO_DOC", "")).strip()[:2] or "CC",
                        ocupacion=str(row.get("OCUPACION", "")).strip(),
                        lugar_trabajo=str(row.get("LUGAR_TRABAJO", "")).strip(),
                        direccion=str(row.get("DIRECCION", "")).strip(),
                        institucion=institucion,
                    )
                    creados += 1
                else:
                    creados += 1  # actualización

                familiar.estudiantes_asociados.add(estudiante)
                vinculos += 1

                if str(row.get("RESPONSABLE_FACTURACION", "")).strip().upper() in ("SI", "SÍ", "X", "1", "TRUE"):
                    estudiante.acudiente_responsable = familiar
                    estudiante.save(update_fields=["acudiente_responsable"])
        except Exception as exc:
            errores.append(f"Fila {fila}: error al procesar ({exc}).")
            logger.error("Carga familiar fila %s: %s", fila, exc, exc_info=True)

    return creados, vinculos, errores
