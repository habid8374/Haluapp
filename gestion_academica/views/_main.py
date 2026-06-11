# gestion_academica/views.py
# Importaciones estándar de Django
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.template.loader import get_template, render_to_string
from django.http import HttpResponse, HttpResponseNotFound, HttpResponseRedirect, Http404, HttpResponseForbidden
from django.db.models import Q, Prefetch, Func, Count
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin 
from django.contrib import messages
from django.utils.decorators import method_decorator
from django.contrib.staticfiles import finders
from django.db import transaction, IntegrityError
from django.utils.timezone import make_aware, datetime, localdate
from gestion_academica.models import DirectorCurso, PeriodoAcademico
from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.core.exceptions import ObjectDoesNotExist
from django.conf import settings
from django.utils.crypto import get_random_string
from django.core.mail import EmailMessage, EmailMultiAlternatives, get_connection
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.cache import never_cache
import os
from gestion_academica.utils import registrar_inasistencias_docentes
import openpyxl
from openpyxl.styles import Font
from collections import OrderedDict
from django.views.generic import View 
import unicodedata
from email.header import Header
from django.utils.encoding import force_str
from itertools import chain
from collections import defaultdict
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth.models import Group
from ..serializers import EventoCalendarioSerializer
from rest_framework.generics import RetrieveAPIView
from ..serializers import ActividadInteractivaSerializer, EnviarRespuestaSerializer
from itertools import groupby
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference
from io import BytesIO
from openpyxl.worksheet.datavalidation import DataValidation
from django.core.exceptions import PermissionDenied
from collections import OrderedDict
import random
from celery.result import AsyncResult
import qrcode
import io
import base64
import logging
from xhtml2pdf import pisa
from ..utils import (
    calcular_estado_academico_curso,
    obtener_desempeno,
    analizar_riesgo_academico_curso,
    contar_pares_estudiante_curso_en_riesgo_academico,
    ESTADOS_RIESGO_ACADEMICO_CURSO,
    generar_boletin_pdf_en_memoria,
    crear_lecciones_diarias_desde_planeacion,
)
import pandas as pd 
import datetime
from datetime import date, datetime, time
from decimal import Decimal, Inexact, ROUND_HALF_UP
from django.utils import timezone
from django.utils.formats import date_format
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from datetime import timedelta
import json
from django.core.management import call_command
from io import StringIO
import google.generativeai as genai
from utils.mensajes import mensaje_exito, mensaje_error, mostrar_mensaje
import google.generativeai as genai
from admisiones.models import Aspirante
from gestion_academica.decorators import requiere_pagos_al_dia, EstaAlDiaPermission
from ..utils import (
    obtener_promedio_materia_por_grado, 
    obtener_conteo_estudiantes_por_grado,
    get_absent_students_by_grade,
    obtener_resumen_financiero_estudiantes, # <-- CORRECCIÓN: Nombre nuevo
    get_top_student_in_school,
    get_observation_count_for_student,
    estudiante_en_curso_actividad,
    actividades_calificables_accesibles_para_usuario,
)
from allauth.socialaccount.models import SocialToken, SocialApp
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from ..tasks import (
    sincronizar_horario_google_calendar_task, 
    generar_contenido_planeacion_task, 
    analizar_propuesta_candidato_task, 
    analizar_comportamiento_task, 
    generar_propuesta_horario_task,
    analizar_plagio_tarea_task
)  
from cuestionarios.models import Cuestionario  

from ..models import (
    Usuario, Grado, Estudiante, Docente, Familiar,
    Materia, PeriodoAcademico, Curso, DirectorCurso,
    EsquemaCalificacion, TipoActividad, ActividadCalificable, Calificacion,
    PlanCurricular, Deber, EntregaDeber, MencionReconocimiento, ArchivoPlanAcademico,
    ConfiguracionInstitucion, Noticia, RegistroAsistencia, BloqueHorario, LeccionDiaria,
    Pregunta, Opcion, RespuestaEstudiante, IntentoActividad, DescriptorLogro, ObservacionBoletin,
    PeriodoAcademico, EscalaValorativa, AnotacionObservador, AnalisisRiesgo, PrediccionRiesgoEstudiante,
    Notificacion, DisponibilidadDocente, CitaReunion, Eleccion, Candidato, Voto, Estudiante, RegistroAsistenciaDocente,
    Aula, AreaAcademica, Egresado, ArchivoHistorico, SolicitudDocumento, Logro, EvaluacionLogroPreescolar, EscalaCualitativa,
    DimensionDesarrollo, LogroPreescolar, TicketSoporte, RespuestaTicket, PlaneacionClase, DetalleClase, NivelEscolaridad,
    AnalisisComportamientoIA,
    MallaCurricular,
    PlanSemanal,
)

from finanzas.models import InstitucionEducativa 
from finanzas.models import ConceptoPago, CuentaPorCobrarEstudiante, PagoRegistrado 
from finanzas.models import NOMBRES_MESES_ESPANOL, TipoConceptoPago, ConceptoPago
from finanzas.institucion_credentials import google_api_key as institucion_google_api_key
      

from ..forms import (
    GradoForm,
    EstudianteForm, 
    CustomUserCreationForm, CustomUserUpdateForm,
    DocenteForm, EstudianteForm,
    MateriaForm,
    PeriodoAcademicoForm,
    CursoForm,
    DirectorCursoForm,
    EsquemaCalificacionForm,
    TipoActividadForm,
    ActividadCalificableForm,
    CalificacionForm,
    DeberForm,
    EntregaDeberForm,
    PlanCurricularForm,
    MencionReconocimientoForm,
    ArchivoPlanAcademicoForm,
    RegistroInicialForm,
    DescriptorLogroForm, 
    LeccionDiariaForm,
    UploadFileForm,
    NoticiaForm,
    ObservacionBoletinForm,
    AnotacionObservadorForm,
    DocenteActividadForm,
    CalificarEntregaForm, 
    FamiliarForm,
    DisponibilidadDocenteForm,
    GestionCitaForm,
    EleccionForm,
    PreguntaForm,
    OpcionFormSet,
    ActividadConfigForm,
    AulaForm,
    AreaAcademicaForm,
    LogroForm,
    DimensionDesarrolloForm,
    EscalaCualitativaForm, LogroPreescolarForm,
    TicketSoporteForm, 
    RespuestaTicketForm,
    PlaneacionClaseForm,
    LeccionDiariaIaForm,
    CandidatoForm,
    UserEditForm, UserPasswordChangeForm

)
from cuestionarios.models import IntentoCuestionario

from ..tasks import generar_ranking_institucional_task
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

logger = logging.getLogger(__name__)

# Para cálculos en vistas
from django.db.models import Sum, Avg, F, ExpressionWrapper, DecimalField

import google.ai.generativelanguage as glm



# --- Funciones Auxiliares ---
def get_filtered_queryset(model, request_user, base_queryset=None):
    if base_queryset is None:
        queryset = model.objects.all()
    else:
        queryset = base_queryset
    if request_user.is_superuser:
        return queryset
    if hasattr(request_user, 'institucion_asociada') and request_user.institucion_asociada:
        return queryset.filter(institucion=request_user.institucion_asociada)
    return model.objects.none()

def get_current_institution(request_user):
    if request_user.is_superuser:
        return InstitucionEducativa.objects.first()
    if hasattr(request_user, 'institucion_asociada') and request_user.institucion_asociada:
        return request_user.institucion_asociada
    return None


def link_callback(uri, rel):
    """Resuelve URIs de recursos para xhtml2pdf con protección contra path traversal."""
    if uri.startswith(settings.MEDIA_URL):
        path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, "", 1))
        allowed_root = os.path.realpath(settings.MEDIA_ROOT)
    elif uri.startswith(settings.STATIC_URL):
        path = os.path.join(settings.STATICFILES_DIRS[0], uri.replace(settings.STATIC_URL, "", 1))
        allowed_root = os.path.realpath(settings.STATICFILES_DIRS[0])
    else:
        return uri
    real_path = os.path.realpath(path)
    if not real_path.startswith(allowed_root + os.sep) and real_path != allowed_root:
        logger.warning("link_callback: path traversal bloqueado para URI: %s", uri)
        return None
    if not os.path.isfile(real_path):
        return None
    return real_path

# Vista de inicio
@login_required
def inicio_academico(request):
    """
    Redirige al usuario a su dashboard o a la página de espera si aún no tiene un rol.
    VERSIÓN CORREGIDA PARA INTEGRARSE CON GOOGLE SIGN-IN.
    """
    user = request.user

    # 1. Redirección por rol (Tu lógica original se mantiene)
    if hasattr(user, 'rol'):
        if user.rol == 'estudiante':
            return redirect('gestion_academica:dashboard_estudiante')
        elif user.rol == 'docente':
            return redirect('gestion_academica:dashboard_docente')
        elif user.rol == 'coordinador':
            return redirect('gestion_academica:dashboard_coordinador')
        elif user.rol == 'familiar':
            return redirect('gestion_academica:portal_familiar_inicio')

    # 2. Acceso para staff/admin — HALU PULSE con KPIs en vivo
    if user.is_staff:
        institucion = getattr(user, 'institucion_asociada', None)
        today = timezone.localdate()

        # ── KPIs ─────────────────────────────────────────────────────
        total_alumnos = Estudiante.objects.filter(activo=True, institucion=institucion).count()
        periodo_activo = PeriodoAcademico.objects.filter(
            institucion=institucion, activo=True
        ).first()

        # Promedio institucional del período activo
        qs_cal = Calificacion.objects.filter(institucion=institucion)
        if periodo_activo:
            qs_cal = qs_cal.filter(
                actividad_calificable__curso__periodo_academico=periodo_activo
            )
        from django.db.models import Avg as _Avg
        promedio_raw = qs_cal.aggregate(prom=_Avg('valor_numerico'))['prom']
        promedio_institucional = round(promedio_raw, 1) if promedio_raw else None

        # Asistencia hoy
        total_hoy = RegistroAsistencia.objects.filter(
            fecha_solo=today, institucion=institucion
        ).count()
        presentes_hoy = RegistroAsistencia.objects.filter(
            fecha_solo=today, estado='PRESENTE', institucion=institucion
        ).count()
        asistencia_hoy_pct = round(presentes_hoy / total_hoy * 100) if total_hoy > 0 else None

        # Alumnos en riesgo alto (IA)
        try:
            alumnos_riesgo = PrediccionRiesgoEstudiante.objects.filter(
                nivel_riesgo='ALTO', institucion=institucion
            ).count()
        except Exception:
            alumnos_riesgo = 0

        # Casos Sentinel abiertos
        try:
            from gestion_academica.models import CasoConvivencia
            casos_sentinel = CasoConvivencia.objects.filter(
                estado__in=[CasoConvivencia.Estado.ABIERTO, CasoConvivencia.Estado.EN_SEGUIMIENTO],
                institucion=institucion
            ).count()
        except Exception:
            casos_sentinel = 0

        context = {
            'nombre_usuario': user.get_full_name() or user.username,
            'rol_usuario': getattr(user, 'rol', 'Administrador'),
            'total_alumnos': total_alumnos,
            'periodo_activo': periodo_activo,
            'promedio_institucional': promedio_institucional,
            'asistencia_hoy_pct': asistencia_hoy_pct,
            'alumnos_riesgo': alumnos_riesgo,
            'casos_sentinel': casos_sentinel,
        }
        return render(request, 'gestion_academica/inicio_academico.html', context)

    # --- 3. NUEVA LÓGICA PARA USUARIOS REGISTRADOS CON GOOGLE ---
    # Si el usuario está autenticado pero no es staff y no tiene ningún perfil
    # (Estudiante, Docente, Familiar), es un usuario "huérfano" esperando activación.
    if not hasattr(user, 'estudiante') and not hasattr(user, 'docente') and not hasattr(user, 'familiar'):
        return redirect('gestion_academica:espera_activacion')

    # 4. Fallback final (Tu lógica original se mantiene como última opción)
    messages.warning(request, "Tu cuenta no tiene un rol válido para acceder al sistema.")
    return redirect('logout')



# --- Vistas para Grados ---
class GradoListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Grado; template_name = 'gestion_academica/grado_lista.html'; context_object_name = 'grados'; permission_required = 'gestion_academica.view_grado'
    def get_queryset(self): return get_filtered_queryset(self.model, self.request.user).order_by('orden', 'nombre')
    def get_context_data(self, **kwargs): context = super().get_context_data(**kwargs); context['titulo_pagina'] = "Listado de Grados"; return context

class GradoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Grado
    form_class = GradoForm
    template_name = 'gestion_academica/grado_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_grados')
    permission_required = 'gestion_academica.add_grado'

    def get_context_data(self, **kwargs):
        """Añade el título al contexto para la plantilla."""
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Crear Nuevo Grado"
        return context

    def get_form_kwargs(self):
        """Pasa el objeto 'request' al formulario."""
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        """
        Gestiona la asignación de la institución antes de guardar.
        Esta es la versión final, segura y corregida.
        """
        # 1. Lógica para asignar la institución
        if not self.request.user.is_superuser:
            # Para un usuario normal, FORZAMOS la asignación de su institución.
            # Esto previene que pueda crear objetos en otro lugar.
            form.instance.institucion = self.request.user.institucion_asociada
        
        # 2. Lógica de validación para el superusuario
        # Si el usuario ES superusuario Y no seleccionó una institución en el formulario...
        elif self.request.user.is_superuser and not form.instance.institucion:
            # ...le mostramos un error y detenemos el proceso.
            messages.error(self.request, "Como superusuario, debes seleccionar una institución.")
            return self.form_invalid(form)

        # 3. Si todo lo anterior pasó sin problemas, mostramos el mensaje de éxito
        messages.success(self.request, "Grado creado exitosamente.")
        
        # 4. Y finalmente, guardamos el objeto en la base de datos y redirigimos.
        return super().form_valid(form)

class GradoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Grado; form_class = GradoForm; template_name = 'gestion_academica/grado_formulario.html'; success_url = reverse_lazy('gestion_academica:lista_grados'); permission_required = 'gestion_academica.change_grado'
    def get_queryset(self): return get_filtered_queryset(self.model, self.request.user)
    def get_context_data(self, **kwargs): context = super().get_context_data(**kwargs); context['titulo_formulario'] = "Editar Grado"; return context
    def get_form_kwargs(self): kwargs = super().get_form_kwargs(); kwargs['request'] = self.request; return kwargs
    def form_valid(self, form): messages.success(self.request, "Grado actualizado exitosamente."); return super().form_valid(form)

class GradoDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Grado; template_name = 'gestion_academica/grado_confirmar_eliminar.html'; success_url = reverse_lazy('gestion_academica:lista_grados'); context_object_name = 'grado'; permission_required = 'gestion_academica.delete_grado'
    def get_queryset(self): return get_filtered_queryset(self.model, self.request.user)
    def delete(self, request, *args, **kwargs): messages.success(request, f"El grado '{self.get_object().nombre}' ha sido eliminado."); return super().delete(request, *args, **kwargs)
    def get_context_data(self, **kwargs): context = super().get_context_data(**kwargs); context['titulo_pagina'] = "Confirmar Eliminación de Grado"; return context

class AulaListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Aula
    template_name = 'gestion_academica/aula_lista.html'
    context_object_name = 'aulas'
    # Debes crear este permiso o usar uno existente de administrador
    permission_required = 'gestion_academica.view_aula'
    paginate_by = 15

    def get_queryset(self):
        # Reutilizamos tu función para filtrar por la institución del usuario. ¡Perfecto!
        base_queryset = super().get_queryset().order_by('nombre')
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Gestión de Aulas y Salones"
        return context

class AulaCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Aula
    form_class = AulaForm
    template_name = 'gestion_academica/aula_formulario.html' # Reutilizaremos esta plantilla
    success_url = reverse_lazy('gestion_academica:lista_aulas')
    permission_required = 'gestion_academica.add_aula'

    def get_form_kwargs(self):
        # Es crucial pasar el 'request' al formulario para la lógica multi-institución
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        # Asignamos la institución automáticamente para usuarios no-superadmin
        if not self.request.user.is_superuser:
            form.instance.institucion = self.request.user.institucion_asociada
        elif self.request.user.is_superuser and not form.cleaned_data.get('institucion'):
            messages.error(self.request, "Como superusuario, debes seleccionar una institución.")
            return self.form_invalid(form)
            
        messages.success(self.request, f"Aula '{form.cleaned_data['nombre']}' creada exitosamente.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Crear Nueva Aula"
        return context

class AulaUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Aula
    form_class = AulaForm
    template_name = 'gestion_academica/aula_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_aulas')
    permission_required = 'gestion_academica.change_aula'

    def get_queryset(self):
        # Filtramos para que un admin solo pueda editar aulas de su institución
        return get_filtered_queryset(self.model, self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, f"Aula '{self.object.nombre}' actualizada exitosamente.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = f"Editar Aula: {self.object.nombre}"
        return context

class AulaDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Aula
    template_name = 'gestion_academica/confirmar_eliminar_generico.html' # Puedes usar una plantilla genérica
    success_url = reverse_lazy('gestion_academica:lista_aulas')
    context_object_name = 'object'
    permission_required = 'gestion_academica.delete_aula'

    def get_queryset(self):
        return get_filtered_queryset(self.model, self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Aula"
        context['mensaje_confirmacion'] = f"¿Estás seguro de que deseas eliminar el aula '{self.object.nombre}'? Esta acción no se puede deshacer."
        context['url_cancelar'] = reverse('gestion_academica:lista_aulas')
        return context

    def form_valid(self, form):
        # Usamos form_valid para el mensaje de éxito, ya que es más seguro que sobreescribir delete()
        messages.success(self.request, f"El aula '{self.object.nombre}' ha sido eliminada.")
        return super().form_valid(form)

class GradoParaEstudiantesListView(LoginRequiredMixin, ListView):
    model = Grado
    template_name = 'gestion_academica/grado_para_estudiantes_lista.html'
    context_object_name = 'grados'

    def dispatch(self, request, *args, **kwargs):
        # Tu lógica de permisos original se mantiene, es correcta.
        if not (request.user.is_staff or hasattr(request.user, 'docente')):
            raise PermissionDenied("No tienes los permisos necesarios para ver esta lista.")
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        """
        Filtra los grados de forma segura, manejando todos los casos.
        """
        user = self.request.user
        institucion = getattr(user, 'institucion_asociada', None)

        # Si el usuario no tiene institución y no es superusuario, no debe ver nada.
        if not institucion and not user.is_superuser:
            messages.warning(self.request, "Tu cuenta no está asociada a ninguna institución.")
            return Grado.objects.none()

        # --- Lógica para staff/admin/superusuario ---
        if user.is_staff:
            # Empezamos con todos los grados
            base_qs = Grado.objects.all()
            # Si no es superusuario, filtramos por su institución
            if not user.is_superuser:
                base_qs = base_qs.filter(institucion=institucion)
            
            # Devolvemos solo los grados de esa institución que tienen estudiantes
            return base_qs.annotate(
                num_estudiantes=Count('estudiantes_actuales')
            ).filter(num_estudiantes__gt=0).order_by('nombre')

        # --- Lógica para docentes ---
        if hasattr(user, 'docente'):
            docente = user.docente
            periodo_activo = PeriodoAcademico.objects.filter(institucion=institucion, activo=True).first()
            
            if periodo_activo:
                cursos_docente = Curso.objects.filter(docentes_asignados=docente, periodo_academico=periodo_activo)
                grados_ids = cursos_docente.values_list('grado_id', flat=True)
                
                # Devolvemos los grados donde el docente da clase y que tienen estudiantes
                return Grado.objects.filter(
                    id__in=set(grados_ids),
                    institucion=institucion
                ).annotate(num_estudiantes=Count('estudiantes_actuales')).order_by('nombre')

        # Si no se cumple ninguna condición, devolvemos una lista vacía.
        return Grado.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Seleccionar Grado para ver Estudiantes"
        return context

class EstudiantesPorGradoListView(LoginRequiredMixin, ListView):
    model = Estudiante
    template_name = 'gestion_academica/estudiante_lista_por_grado.html' # Nueva plantilla
    context_object_name = 'estudiantes'
    paginate_by = 20

    def dispatch(self, request, *args, **kwargs):
        # Lógica de seguridad: Un docente solo puede ver la lista de estudiantes
        # de un grado si tiene relación con ese grado.
        user = request.user
        grado_pk = self.kwargs['grado_pk']
        
        if not user.is_staff: # Si no es admin, debe ser docente
            try:
                docente = user.docente
                grado = Grado.objects.get(pk=grado_pk)
                periodo_activo = PeriodoAcademico.objects.filter(institucion=docente.institucion, activo=True).first()
                
                if periodo_activo:
                    es_director = DirectorCurso.objects.filter(docente=docente, grado=grado, periodo_academico=periodo_activo).exists()
                    da_clases_en_grado = Curso.objects.filter(docentes_asignados=docente, grado=grado, periodo_academico=periodo_activo).exists()
                    
                    if not (es_director or da_clases_en_grado):
                        raise PermissionDenied("No tienes permiso para ver los estudiantes de este grado.")
                else:
                    raise PermissionDenied("No hay un periodo académico activo.")
            except (Docente.DoesNotExist, Grado.DoesNotExist):
                raise PermissionDenied("Acceso no válido.")

        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        """ Filtra los estudiantes por el grado seleccionado. """
        self.grado = get_object_or_404(Grado, pk=self.kwargs['grado_pk'])
        
        # Filtro simple y seguro gracias a la validación en dispatch
        return Estudiante.objects.filter(
            grado_actual=self.grado
        ).select_related('usuario').order_by('usuario__last_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['grado'] = self.grado
        context['titulo_pagina'] = f"Estudiantes de {self.grado.nombre}"
        return context        

class EstudianteDetailView(LoginRequiredMixin, DetailView):
    model = Estudiante
    template_name = 'gestion_academica/estudiante_detalle.html'
    context_object_name = 'estudiante'

    def get_object(self, queryset=None):
        # Tu método get_object está perfecto, no necesita cambios.
        obj = super().get_object(queryset)
        user = self.request.user
        tiene_permiso = False
        
        if user.is_superuser:
            tiene_permiso = True
        elif user == obj.usuario:
            tiene_permiso = True
        elif user.is_staff and user.rol in ['administrador', 'coordinador'] and getattr(user, 'institucion_asociada', None) == obj.institucion:
            tiene_permiso = True
        elif hasattr(user, 'familiar') and obj in user.familiar.estudiantes_asociados.all():
            tiene_permiso = True
        elif hasattr(user, 'docente'):
            periodo_activo = PeriodoAcademico.objects.filter(institucion=obj.institucion, activo=True).first()
            if periodo_activo:
                if DirectorCurso.objects.filter(docente=user.docente, grado=obj.grado_actual, periodo_academico=periodo_activo).exists():
                    tiene_permiso = True
                elif Curso.objects.filter(grado=obj.grado_actual, periodo_academico=periodo_activo, docentes_asignados=user.docente).exists():
                    tiene_permiso = True
        
        if not tiene_permiso:
            raise PermissionDenied("No tienes los permisos necesarios para ver el perfil de este estudiante.")
            
        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Detalle de Estudiante: {self.object.usuario.get_full_name()}"
        
        # --- INICIO DE LA MODIFICACIÓN ---
        # Añadimos el periodo activo al contexto para que la plantilla lo pueda usar.
        periodo_activo = PeriodoAcademico.objects.filter(
            activo=True, institucion=self.object.institucion
        ).first()
        context['periodo_activo'] = periodo_activo
        # --- FIN DE LA MODIFICACIÓN ---

        # Tu lógica para las alertas de riesgo se mantiene igual
        ultimo_analisis = AnalisisRiesgo.objects.order_by('-fecha_analisis').first()
        if ultimo_analisis:
            context['alertas_academicas'] = ultimo_analisis.predicciones.filter(estudiante=self.object)
        else:
            context['alertas_academicas'] = []
            
        return context


@login_required
def calendario_academico_view(request):
    """
    Muestra la página del calendario académico principal.
    """
    context = {
        'titulo_pagina': "Calendario Académico Institucional",
        'institucion_actual': get_current_institution(request.user)
    }
    return render(request, 'gestion_academica/calendario_academico.html', context)



@mensaje_exito("Estudiante creado exitosamente")
@mensaje_error("Hubo un error creando el estudiante")
@transaction.atomic
@login_required
@permission_required('gestion_academica.add_estudiante')
def crear_estudiante(request):
    institution = get_current_institution(request.user)
    if not institution and not request.user.is_superuser:
        messages.error(request, "No se puede crear un estudiante sin una institución asignada.")
        return redirect('gestion_academica:inicio_academico')

    if request.method == 'POST':
        usuario_form = CustomUserCreationForm(request.POST, prefix="usr", request=request)
        estudiante_form = EstudianteForm(request.POST, request.FILES, prefix="est", request=request)
        if usuario_form.is_valid() and estudiante_form.is_valid():
            usuario = usuario_form.save(commit=False)
            usuario.rol = 'estudiante'
            usuario.save()
            estudiante = estudiante_form.save(commit=False)
            estudiante.usuario = usuario
            if not request.user.is_superuser and institution:
                estudiante.institucion = institution
            estudiante.save()
            messages.success(request, f"Estudiante '{usuario.get_full_name()}' creado exitosamente.")
            return redirect('gestion_academica:lista_estudiantes')
        else:
            messages.error(request, "Por favor, corrige los errores en el formulario.")
    else:
        usuario_form = CustomUserCreationForm(prefix="usr", request=request)
        estudiante_form = EstudianteForm(prefix="est", request=request)
    context = {'usuario_form': usuario_form, 'estudiante_form': estudiante_form, 'titulo': 'Registrar Nuevo Estudiante'}
    return render(request, 'gestion_academica/estudiante_formulario.html', context)


@login_required
@permission_required('gestion_academica.change_estudiante')
@mensaje_exito("Datos del estudiante actualizados exitosamente.")
def editar_estudiante(request, pk):
    estudiante = get_object_or_404(Estudiante, pk=pk)
    usuario = estudiante.usuario
    if request.method == 'POST':
        usuario_form = CustomUserUpdateForm(request.POST, instance=usuario, prefix="usr", request=request)
        estudiante_form = EstudianteForm(request.POST, request.FILES or None, instance=estudiante, prefix="est", request=request)
        if usuario_form.is_valid() and estudiante_form.is_valid():
            usuario_form.save()
            estudiante_form.save()
            
            # --- CORRECCIÓN DE LA REDIRECCIÓN ---
            if estudiante.grado_actual:
                return redirect('gestion_academica:lista_estudiantes_por_grado', grado_pk=estudiante.grado_actual.pk)
            else:
                return redirect('gestion_academica:lista_grados_para_estudiantes')
            
    else:
        usuario_form = CustomUserUpdateForm(instance=usuario, prefix="usr", request=request)
        estudiante_form = EstudianteForm(instance=estudiante, prefix="est", request=request)
    
    context = {
        'usuario_form': usuario_form, 
        'estudiante_form': estudiante_form, 
        'titulo': f'Editar Estudiante: {usuario.get_full_name()}', 
        'estudiante_obj': estudiante
    }
    return render(request, 'gestion_academica/estudiante_formulario.html', context)


@login_required
def ayuda_soporte_view(request):
    """
    Muestra el portal de ayuda, permite crear nuevos tickets de soporte
    y lista los tickets existentes del usuario.
    """
    institucion = getattr(request.user, 'institucion_asociada', None)
    
    if request.method == 'POST':
        form = TicketSoporteForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.usuario_reporta = request.user
            ticket.institucion = institucion
            ticket.save()
            
            # (Aquí añadiremos la notificación por correo al superadmin en el siguiente paso)
            
            messages.success(request, f"¡Ticket [{ticket.ticket_id}] creado exitosamente! Nuestro equipo de soporte lo revisará pronto.")
            return redirect('gestion_academica:ayuda_soporte')
    else:
        form = TicketSoporteForm()

    # Obtenemos el historial de tickets del usuario que ha iniciado sesión
    tickets_del_usuario = TicketSoporte.objects.filter(usuario_reporta=request.user)

    context = {
        'titulo_pagina': "Centro de Ayuda y Soporte",
        'form': form,
        'tickets_del_usuario': tickets_del_usuario
    }
    return render(request, 'gestion_academica/ayuda_soporte.html', context)

@login_required
def ticket_detail_view(request, ticket_id):
    """
    Muestra la conversación completa de un ticket para el usuario que lo creó
    y le permite añadir nuevas respuestas.
    """
    # Buscamos el ticket, asegurándonos de que pertenezca al usuario logueado
    ticket = get_object_or_404(TicketSoporte, ticket_id=ticket_id, usuario_reporta=request.user)
    
    if request.method == 'POST':
        form = RespuestaTicketForm(request.POST, request.FILES)
        if form.is_valid():
            respuesta = form.save(commit=False)
            respuesta.ticket = ticket
            respuesta.autor = request.user
            respuesta.save()
            
            # Actualizamos el ticket para que aparezca primero en la lista del admin
            ticket.save(update_fields=['ultima_actualizacion'])
            
            messages.success(request, "Tu respuesta ha sido añadida al ticket.")
            return redirect('gestion_academica:ticket_detail', ticket_id=ticket.ticket_id)
    else:
        form = RespuestaTicketForm()

    context = {
        'titulo_pagina': f"Detalle del Ticket [{ticket.ticket_id}]",
        'ticket': ticket,
        'respuestas': ticket.respuestas.select_related('autor').order_by('fecha_creacion'),
        'form': form
    }
    return render(request, 'gestion_academica/ticket_detail.html', context)    


class EstudianteDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Estudiante
    template_name = 'gestion_academica/estudiante_confirmar_eliminar.html'
    
    # ✅ MEJORA 1: Redirigir a la nueva lista de grados.
    # La antigua 'lista_estudiantes' ya no existe.
    success_url = reverse_lazy('gestion_academica:lista_grados_para_estudiantes')
    
    context_object_name = 'estudiante'
    permission_required = 'gestion_academica.delete_estudiante'

    def get_queryset(self):
        # Tu lógica de seguridad multi-institución está perfecta y se mantiene.
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    # ✅ MEJORA 2: Usar form_valid para el mensaje de éxito.
    # Esta es la forma recomendada por Django para añadir acciones tras una eliminación.
    def form_valid(self, form):
        # Guardamos el nombre antes de que el objeto se elimine.
        nombre_completo = self.object.usuario.get_full_name() or self.object.usuario.username
        
        # El método de la clase padre se encarga de la eliminación.
        response = super().form_valid(form)
        
        # Añadimos el mensaje de éxito después de la eliminación.
        messages.success(
            self.request,
            f"✅ El estudiante <strong>{nombre_completo}</strong> y su cuenta de usuario han sido eliminados correctamente."
        )
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Estudiante"
        return context

# --- Vistas para Docentes (Gestión por Admin) ---
class DocenteListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Docente; template_name = 'gestion_academica/docente_lista.html'; context_object_name = 'docentes'; permission_required = 'gestion_academica.view_docente'; paginate_by = 10
    def get_queryset(self):
        base_queryset = Docente.objects.select_related('usuario').order_by('usuario__last_name', 'usuario__first_name')
        return get_filtered_queryset(self.model, self.request.user, base_queryset)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs); context['titulo_pagina'] = "Listado de Docentes"; return context

class DocenteDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Docente; template_name = 'gestion_academica/docente_detalle.html'; context_object_name = 'docente'; permission_required = 'gestion_academica.view_docente'
    def get_queryset(self):
        base_queryset = super().get_queryset().select_related('usuario')
        return get_filtered_queryset(self.model, self.request.user, base_queryset)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs); context['titulo_pagina'] = f"Detalle de Docente: {self.object.usuario.get_full_name() or self.object.usuario.username}"; return context

class DocenteDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Docente; template_name = 'gestion_academica/docente_confirmar_eliminar.html'; success_url = reverse_lazy('gestion_academica:lista_docentes'); context_object_name = 'docente'; permission_required = 'gestion_academica.delete_docente'
    def get_queryset(self): return get_filtered_queryset(self.model, self.request.user)
    def delete(self, request, *args, **kwargs):
        nombre_completo = self.get_object().usuario.get_full_name() or self.get_object().usuario.username
        messages.success(request, f"El docente '{nombre_completo}' y su cuenta de usuario han sido eliminados."); return super().delete(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs); context['titulo_pagina'] = "Confirmar Eliminación de Docente"; return context


@mensaje_exito("Docente creado exitosamente")
@mensaje_error("Hubo un error creando el docente")
@login_required
@permission_required('gestion_academica.add_docente')
def crear_docente(request):
    institution = get_current_institution(request.user)
    if request.method == 'POST':
        usuario_form = CustomUserCreationForm(request.POST, prefix="usr", request=request)
        docente_form = DocenteForm(request.POST, request.FILES or None, prefix="doc", request=request)
        if usuario_form.is_valid() and docente_form.is_valid():
            usuario = usuario_form.save(commit=False)
            usuario.rol = 'docente'
            usuario.save()
            docente = docente_form.save(commit=False)
            docente.usuario = usuario
            if not request.user.is_superuser and institution:
                docente.institucion = institution
            docente.save()
            messages.success(request, f"Docente '{usuario.get_full_name()}' creado exitosamente.")
            return redirect('gestion_academica:lista_docentes')
    else:
        usuario_form = CustomUserCreationForm(prefix="usr", request=request)
        docente_form = DocenteForm(prefix="doc", request=request)
    context = {'usuario_form': usuario_form, 'docente_form': docente_form, 'titulo': 'Registrar Nuevo Docente'}
    return render(request, 'gestion_academica/docente_formulario.html', context)

@login_required
@permission_required('gestion_academica.change_docente')
@mensaje_exito("Datos del docente actualizados exitosamente.")
def editar_docente(request, pk):
    docente = get_object_or_404(Docente, pk=pk)
    usuario = docente.usuario

    if request.method == 'POST':
        usuario_form = CustomUserUpdateForm(request.POST, instance=usuario, prefix="usr", request=request)
        docente_form = DocenteForm(request.POST, request.FILES or None, instance=docente, prefix="doc", request=request)
        if usuario_form.is_valid() and docente_form.is_valid():
            usuario_form.save()
            docente_form.save()
            return redirect('gestion_academica:lista_docentes')
    else:
        usuario_form = CustomUserUpdateForm(instance=usuario, prefix="usr", request=request)
        docente_form = DocenteForm(instance=docente, prefix="doc", request=request)

    context = {
        'usuario_form': usuario_form,
        'docente_form': docente_form,
        'titulo': f'Editar Docente: {usuario.get_full_name()}',
        'docente_obj': docente
    }
    return render(request, 'gestion_academica/docente_formulario.html', context)

# --- Vistas para Materias ---
class MateriaListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Materia
    template_name = 'gestion_academica/materia_lista.html'
    context_object_name = 'materias'
    permission_required = 'gestion_academica.view_materia'
    paginate_by = 15

    def get_queryset(self):
        base_queryset = super().get_queryset().order_by('nombre_materia')
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Listado de Materias"
        inst = getattr(self.request.user, 'institucion_asociada', None)
        context['es_bilingue'] = getattr(inst, 'es_bilingue', False)
        context['idioma_secundario'] = getattr(inst, 'get_idioma_secundario_display', lambda: 'Inglés')()
        return context

class MateriaCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Materia
    form_class = MateriaForm
    template_name = 'gestion_academica/materia_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_materias')
    permission_required = 'gestion_academica.add_materia'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Crear Nueva Materia"
        inst = getattr(self.request.user, 'institucion_asociada', None)
        context['es_bilingue'] = getattr(inst, 'es_bilingue', False)
        context['idioma_secundario'] = getattr(inst, 'get_idioma_secundario_display', lambda: 'Inglés')()
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        if not self.request.user.is_superuser and hasattr(self.request.user, 'institucion_asociada') and self.request.user.institucion_asociada:
            form.instance.institucion = self.request.user.institucion_asociada
        elif self.request.user.is_superuser and not form.instance.institucion:
            messages.error(self.request, "Como superusuario, debes seleccionar una institución para la materia.")
            return self.form_invalid(form)

        messages.success(self.request, f"Materia '{form.cleaned_data['nombre_materia']}' creada exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

class MateriaUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Materia
    form_class = MateriaForm
    template_name = 'gestion_academica/materia_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_materias')
    permission_required = 'gestion_academica.change_materia'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Materia"
        inst = getattr(self.request.user, 'institucion_asociada', None)
        context['es_bilingue'] = getattr(inst, 'es_bilingue', False)
        context['idioma_secundario'] = getattr(inst, 'get_idioma_secundario_display', lambda: 'Inglés')()
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, f"Materia '{form.cleaned_data['nombre_materia']}' actualizada exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

class MateriaDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Materia
    template_name = 'gestion_academica/materia_confirmar_eliminar.html'
    success_url = reverse_lazy('gestion_academica:lista_materias')
    context_object_name = 'materia'
    permission_required = 'gestion_academica.delete_materia'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def delete(self, request, *args, **kwargs):
        materia_eliminada = self.get_object()
        messages.success(request, f"La materia '{materia_eliminada.nombre_materia}' ha sido eliminada exitosamente.") 
        return super().delete(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Materia"
        return context

# --- Vistas para Periodos Académicos ---
class PeriodoAcademicoListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = PeriodoAcademico
    template_name = 'gestion_academica/periodo_lista.html'
    context_object_name = 'periodos'
    permission_required = 'gestion_academica.view_periodoacademico'
    paginate_by = 10

    def get_queryset(self):
        base_queryset = super().get_queryset().order_by('-año_escolar', '-fecha_inicio')
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Listado de Periodos Académicos"
        return context

class PeriodoAcademicoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = PeriodoAcademico
    form_class = PeriodoAcademicoForm
    template_name = 'gestion_academica/periodo_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_periodos')
    permission_required = 'gestion_academica.add_periodoacademico'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Crear Nuevo Periodo Académico"
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        # Asigna la institución antes de hacer cualquier otra cosa
        if not self.request.user.is_superuser and hasattr(self.request.user, 'institucion_asociada'):
            form.instance.institucion = self.request.user.institucion_asociada
        elif self.request.user.is_superuser and not form.instance.institucion:
            messages.error(self.request, "Como superusuario, debes seleccionar una institución.")
            return self.form_invalid(form)

        # Si el checkbox 'activo' está marcado en el formulario
        if form.cleaned_data.get('activo'):
            # Desactiva TODOS los otros periodos de ESA MISMA institución.
            # No necesitamos excluir nada, porque el objeto nuevo aún no existe.
            PeriodoAcademico.objects.filter(
                institucion=form.instance.institucion, 
                activo=True
            ).update(activo=False)
        
        messages.success(self.request, f"Periodo Académico '{form.cleaned_data['nombre']}' creado exitosamente.")
        
        # La llamada a super().form_valid(form) se encarga de guardar el nuevo objeto
        return super().form_valid(form)


class PeriodoAcademicoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = PeriodoAcademico
    form_class = PeriodoAcademicoForm
    template_name = 'gestion_academica/periodo_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_periodos')
    permission_required = 'gestion_academica.change_periodoacademico'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Periodo Académico"
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        if form.cleaned_data.get('activo'):
            if form.instance.institucion: 
                PeriodoAcademico.objects.exclude(pk=self.object.pk).filter(activo=True, institucion=form.instance.institucion).update(activo=False)
            else: 
                PeriodoAcademico.objects.exclude(pk=self.object.pk).filter(activo=True).update(activo=False)
        messages.success(self.request, f"Periodo Académico '{form.cleaned_data['nombre']}' actualizado exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

class PeriodoAcademicoDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = PeriodoAcademico
    template_name = 'gestion_academica/periodo_confirmar_eliminar.html'
    success_url = reverse_lazy('gestion_academica:lista_periodos')
    context_object_name = 'periodo'
    permission_required = 'gestion_academica.delete_periodoacademico'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def delete(self, request, *args, **kwargs):
        periodo_eliminado = self.get_object()
        messages.success(request, f"El Periodo Académico '{periodo_eliminado.nombre}' ha sido eliminado exitosamente.")
        return super().delete(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Periodo Académico"
        return context

# --- Vistas para Cursos ---
class CursoListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Curso
    template_name = 'gestion_academica/curso_lista.html'
    context_object_name = 'cursos'
    permission_required = 'gestion_academica.view_curso'
    paginate_by = 10

    def get_queryset(self):
        base_queryset = Curso.objects.select_related('materia', 'grado', 'periodo_academico').prefetch_related('docentes_asignados__usuario').all().order_by(
            '-periodo_academico__año_escolar', '-periodo_academico__fecha_inicio', 'grado__nombre', 'materia__nombre_materia'
        )
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Listado de Cursos"
        return context

class CursoDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Curso
    template_name = 'gestion_academica/curso_detalle.html'
    context_object_name = 'curso'
    permission_required = 'gestion_academica.view_curso'

    def get_queryset(self):
        base_queryset = super().get_queryset().select_related('materia', 'grado', 'periodo_academico').prefetch_related('docentes_asignados__usuario')
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Detalle del Curso: {self.object}"
        return context

class CursoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Curso
    form_class = CursoForm
    template_name = 'gestion_academica/curso_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_cursos')
    permission_required = 'gestion_academica.add_curso'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Crear Nuevo Curso"
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        if not self.request.user.is_superuser and hasattr(self.request.user, 'institucion_asociada') and self.request.user.institucion_asociada:
            form.instance.institucion = self.request.user.institucion_asociada
        elif self.request.user.is_superuser and not form.instance.institucion:
            messages.error(self.request, "Como superusuario, debes seleccionar una institución para el curso.")
            return self.form_invalid(form)

        messages.success(self.request, f"Curso '{form.instance}' creado exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

class CursoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Curso
    form_class = CursoForm
    template_name = 'gestion_academica/curso_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_cursos')
    permission_required = 'gestion_academica.change_curso'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Curso"
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, f"Curso '{form.instance}' actualizado exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

class CursoDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Curso
    template_name = 'gestion_academica/curso_confirmar_eliminar.html'
    success_url = reverse_lazy('gestion_academica:lista_cursos')
    context_object_name = 'curso'
    permission_required = 'gestion_academica.delete_curso'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def delete(self, request, *args, **kwargs):
        curso_eliminado = self.get_object()
        messages.success(request, f"El curso '{curso_eliminado}' ha sido eliminado exitosamente.")
        return super().delete(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Curso"
        return context

# --- Vistas para Directores de Curso ---
class DirectorCursoListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = DirectorCurso
    template_name = 'gestion_academica/director_curso_lista.html'
    context_object_name = 'directores_curso'
    permission_required = 'gestion_academica.view_directorcurso'
    paginate_by = 10

    def get_queryset(self):
        base_queryset = DirectorCurso.objects.select_related(
            'docente__usuario', 'grado', 'periodo_academico'
        ).order_by(
            '-periodo_academico__año_escolar', '-periodo_academico__fecha_inicio', 'grado__nombre'
        )
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Listado de Directores de Curso"
        return context

class DirectorCursoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = DirectorCurso
    form_class = DirectorCursoForm
    template_name = 'gestion_academica/director_curso_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_directores_curso')
    permission_required = 'gestion_academica.add_directorcurso'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Asignar Nuevo Director de Curso"
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        if not self.request.user.is_superuser and hasattr(self.request.user, 'institucion_asociada') and self.request.user.institucion_asociada:
            form.instance.institucion = self.request.user.institucion_asociada
        elif self.request.user.is_superuser and not form.instance.institucion:
            messages.error(self.request, "Como superusuario, debes seleccionar una institución para la asignación.")
            return self.form_invalid(form)

        messages.success(self.request, f"Director de curso asignado exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

class DirectorCursoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = DirectorCurso
    form_class = DirectorCursoForm
    template_name = 'gestion_academica/director_curso_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_directores_curso')
    permission_required = 'gestion_academica.change_directorcurso'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Asignación de Director de Curso"
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, "Asignación de director de curso actualizada exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

class DirectorCursoDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = DirectorCurso
    template_name = 'gestion_academica/director_curso_confirmar_eliminar.html'
    success_url = reverse_lazy('gestion_academica:lista_directores_curso')
    context_object_name = 'director_curso'
    permission_required = 'gestion_academica.delete_directorcurso'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def delete(self, request, *args, **kwargs):
        director_eliminado = self.get_object()
        messages.success(request, f"La asignación del director '{director_eliminado.docente}' para el grado '{director_eliminado.grado}' ha sido eliminada.")
        return super().delete(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Asignación"
        return context

# --- Vistas para Esquemas de Calificación ---
class EsquemaCalificacionListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = EsquemaCalificacion
    template_name = 'gestion_academica/esquema_calificacion_lista.html'
    context_object_name = 'esquemas'
    permission_required = 'gestion_academica.view_esquemacalificacion'
    paginate_by = 10

    def get_queryset(self):
        base_queryset = super().get_queryset().order_by('nombre')
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Esquemas de Calificación"
        return context

class EsquemaCalificacionCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = EsquemaCalificacion
    form_class = EsquemaCalificacionForm
    template_name = 'gestion_academica/esquema_calificacion_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_esquemas_calificacion')
    permission_required = 'gestion_academica.add_esquemacalificacion'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Crear Nuevo Esquema de Calificación"
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        if not self.request.user.is_superuser and hasattr(self.request.user, 'institucion_asociada') and self.request.user.institucion_asociada:
            form.instance.institucion = self.request.user.institucion_asociada
        elif self.request.user.is_superuser and not form.instance.institucion:
            messages.error(self.request, "Como superusuario, debes seleccionar una institución para el esquema de calificación.")
            return self.form_invalid(form)

        messages.success(self.request, f"Esquema '{form.cleaned_data['nombre']}' creado exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

class EsquemaCalificacionUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = EsquemaCalificacion
    form_class = EsquemaCalificacionForm
    template_name = 'gestion_academica/esquema_calificacion_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_esquemas_calificacion')
    permission_required = 'gestion_academica.change_esquemacalificacion'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Esquema de Calificación"
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, f"Esquema '{form.cleaned_data['nombre']}' actualizada exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

class EsquemaCalificacionDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = EsquemaCalificacion
    template_name = 'gestion_academica/esquema_calificacion_confirmar_eliminar.html'
    success_url = reverse_lazy('gestion_academica:lista_esquemas_calificacion')
    context_object_name = 'esquema'
    permission_required = 'gestion_academica.delete_esquemacalificacion'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def delete(self, request, *args, **kwargs):
        esquema_eliminado = self.get_object()
        messages.success(request, f"El esquema '{esquema_eliminado.nombre}' ha sido eliminado.") 
        return super().delete(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Esquema"
        return context

# --- Vistas para Tipos de Actividad ---
class TipoActividadListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = TipoActividad
    template_name = 'gestion_academica/tipo_actividad_lista.html'
    context_object_name = 'tipos_actividad'
    permission_required = 'gestion_academica.view_tipoactividad'
    paginate_by = 10

    def get_queryset(self):
        base_queryset = super().get_queryset().order_by('nombre')
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Categorías de Actividad"
        qs = context['object_list']
        context['total_porcentaje'] = sum(c.porcentaje for c in qs if c.porcentaje) if qs else 0
        return context

class TipoActividadCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = TipoActividad
    form_class = TipoActividadForm
    template_name = 'gestion_academica/tipo_actividad_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_tipos_actividad')
    permission_required = 'gestion_academica.add_tipoactividad'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Crear Nuevo Tipo de Actividad"
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        if not self.request.user.is_superuser and hasattr(self.request.user, 'institucion_asociada') and self.request.user.institucion_asociada:
            form.instance.institucion = self.request.user.institucion_asociada
        elif self.request.user.is_superuser and not form.instance.institucion:
            messages.error(self.request, "Como superusuario, debes seleccionar una institución para el tipo de actividad.")
            return self.form_invalid(form)

        messages.success(self.request, f"Tipo de actividad '{form.cleaned_data['nombre']}' creado exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

class TipoActividadUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = TipoActividad
    form_class = TipoActividadForm
    template_name = 'gestion_academica/tipo_actividad_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_tipos_actividad')
    permission_required = 'gestion_academica.change_tipoactividad'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Tipo de Actividad"
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, f"Tipo de actividad '{form.cleaned_data['nombre']}' actualizado exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

class TipoActividadDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = TipoActividad
    template_name = 'gestion_academica/tipo_actividad_confirmar_eliminar.html'
    success_url = reverse_lazy('gestion_academica:lista_tipos_actividad')
    context_object_name = 'tipo_actividad'
    permission_required = 'gestion_academica.delete_tipoactividad'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def delete(self, request, *args, **kwargs):
        tipo_eliminado = self.get_object()
        messages.success(request, f"El tipo de actividad '{tipo_eliminado.nombre}' ha sido eliminado.") 
        return super().delete(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Tipo de Actividad"
        return context

# --- Vistas para Actividades Calificables ---
class ActividadCalificableListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = ActividadCalificable
    template_name = 'gestion_academica/actividad_calificable_lista.html'
    context_object_name = 'actividades'
    permission_required = 'gestion_academica.view_actividadcalificable'
    paginate_by = 10

    def get_queryset(self):
        base_queryset = ActividadCalificable.objects.select_related(
            'curso__materia', 'curso__grado', 'curso__periodo_academico', 'tipo_actividad'
        ).prefetch_related('cuestionario').order_by(
            '-curso__periodo_academico__año_escolar', '-curso__periodo_academico__fecha_inicio',
            'curso__grado', 'curso__materia', '-fecha_publicacion'
        )
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Listado de Actividades Calificables"
        return context

class ActividadCalificableDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = ActividadCalificable
    template_name = 'gestion_academica/actividad_calificable_detalle.html'
    context_object_name = 'actividad'
    permission_required = 'gestion_academica.view_actividadcalificable'

    def get_queryset(self):
        base_queryset = super().get_queryset().select_related(
            'curso__materia', 'curso__grado', 'curso__periodo_academico', 'tipo_actividad'
        )
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Detalle: {self.object.titulo}"
        return context

class ActividadCalificableCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = ActividadCalificable
    form_class = ActividadCalificableForm
    template_name = 'gestion_academica/actividad_calificable_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_actividades_calificables')
    permission_required = 'gestion_academica.add_actividadcalificable'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Crear Nueva Actividad Calificable"
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        if not self.request.user.is_superuser and hasattr(self.request.user, 'institucion_asociada') and self.request.user.institucion_asociada:
            form.instance.institucion = self.request.user.institucion_asociada
        elif self.request.user.is_superuser and not form.instance.institucion:
            messages.error(self.request, "Como superusuario, debes seleccionar una institución para la actividad.")
            return self.form_invalid(form)

        messages.success(self.request, f"Actividad '{form.cleaned_data['titulo']}' creada exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

class ActividadCalificableUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = ActividadCalificable
    form_class = ActividadCalificableForm
    template_name = 'gestion_academica/actividad_calificable_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_actividades_calificables')
    permission_required = 'gestion_academica.change_actividadcalificable'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Actividad Calificable"
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, f"Actividad '{form.cleaned_data['titulo']}' actualizada exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

class ActividadCalificableDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = ActividadCalificable
    template_name = 'gestion_academica/actividad_calificable_confirmar_eliminar.html'
    success_url = reverse_lazy('gestion_academica:lista_actividades_calificables')
    context_object_name = 'actividad'
    permission_required = 'gestion_academica.delete_actividadcalificable'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def delete(self, request, *args, **kwargs):
        actividad_eliminada = self.get_object()
        messages.success(request, f"La actividad '{actividad_eliminada.titulo}' ha sido eliminada.") 
        return super().delete(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Actividad"
        return context

# --- Vistas para Deberes (Tareas) ---
class DeberListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Deber
    template_name = 'gestion_academica/deber_lista.html'
    context_object_name = 'deberes'
    permission_required = 'gestion_academica.view_deber'
    paginate_by = 10

    def get_queryset(self):
        base_queryset = Deber.objects.select_related(
            'curso__materia', 'curso__grado', 'curso__periodo_academico'
        ).order_by('-curso__periodo_academico__año_escolar', 'curso__grado', 'curso__materia', '-fecha_entrega')
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Listado de Deberes / Tareas"
        return context

class DeberDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Deber
    template_name = 'gestion_academica/deber_detalle.html'
    context_object_name = 'deber'
    permission_required = 'gestion_academica.view_deber'

    def get_queryset(self):
        base_queryset = super().get_queryset().select_related(
            'curso__materia', 'curso__grado', 'curso__periodo_academico'
        )
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Detalle del Deber: {self.object.titulo}"
        return context

class DeberCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Deber
    form_class = DeberForm
    template_name = 'gestion_academica/deber_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_deberes')
    permission_required = 'gestion_academica.add_deber'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        deber = form.save(commit=False)
        
        # Si el usuario no es superadmin, la institución se asigna automáticamente
        if not self.request.user.is_superuser:
            deber.institucion = self.request.user.institucion_asociada
        # Si es superadmin, la institución ya viene del formulario, así que no hacemos nada extra aquí
        
        deber.save()
        
        # --- INICIO DE LA CORRECCIÓN CLAVE ---
        # Asignamos el objeto recién guardado a self.object
        self.object = deber
        # --- FIN DE LA CORRECCIÓN CLAVE ---
        
        messages.success(self.request, f"Deber '{deber.titulo}' creado exitosamente.")
        
        # Ahora la redirección funcionará porque self.object ya no es nulo
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Crear Nuevo Deber / Tarea"
        return context
    

class DeberUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Deber
    form_class = DeberForm
    template_name = 'gestion_academica/deber_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_deberes')
    permission_required = 'gestion_academica.change_deber'

    def form_valid(self, form):
        messages.success(self.request, f"Deber '{form.cleaned_data['titulo']}' actualizado exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Deber / Tarea"
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

class DeberDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Deber
    template_name = 'gestion_academica/deber_confirmar_eliminar.html'
    success_url = reverse_lazy('gestion_academica:lista_deberes')
    context_object_name = 'deber'
    permission_required = 'gestion_academica.delete_deber'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def delete(self, request, *args, **kwargs):
        deber_eliminado = self.get_object()
        messages.success(request, f"El deber '{deber_eliminado.titulo}' ha sido eliminado.") 
        return super().delete(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Deber"
        return context

def calcular_nota_final_curso(curso, estudiante):
    """
    Función centralizada para calcular la nota final ponderada de un estudiante en un curso.
    Esta función será usada por el boletín, el libro de notas y el resumen del estudiante.
    """
    # Obtenemos todas las calificaciones del estudiante para este curso específico
    calificaciones = Calificacion.objects.filter(
        estudiante=estudiante,
        actividad_calificable__curso=curso
    )
    # Obtenemos todas las actividades de ese curso
    actividades = ActividadCalificable.objects.filter(curso=curso).select_related('tipo_actividad')

    # Creamos un mapa para acceder rápidamente a las notas ya obtenidas
    calificaciones_map = {cal.actividad_calificable_id: cal.valor_numerico for cal in calificaciones}
    
    # Agrupamos las actividades por su categoría (Ej: "Exámenes", "Tareas")
    actividades_por_categoria = defaultdict(list)
    for actividad in actividades:
        if actividad.tipo_actividad and actividad.tipo_actividad.porcentaje is not None:
            actividades_por_categoria[actividad.tipo_actividad].append(actividad)

    # Iniciamos el cálculo de la nota final
    nota_final_curso = Decimal('0.0')
    for categoria, actividades_en_categoria in actividades_por_categoria.items():
        notas_de_la_categoria = [
            calificaciones_map.get(act.id) 
            for act in actividades_en_categoria 
            if calificaciones_map.get(act.id) is not None
        ]
        
        if notas_de_la_categoria:
            # Calculamos el promedio simple dentro de la categoría
            promedio_categoria = sum(notas_de_la_categoria) / len(notas_de_la_categoria)
            
            # Aplicamos el porcentaje de la categoría a su promedio
            porcentaje_categoria = categoria.porcentaje
            nota_ponderada_categoria = promedio_categoria * (porcentaje_categoria / Decimal('100.0'))
            
            # Sumamos el resultado a la nota final del curso
            nota_final_curso += nota_ponderada_categoria
            
    return nota_final_curso if nota_final_curso > 0 else None


# ===================================================================
# ▼▼▼ 2. AHORA, REEMPLAZA TU VISTA DE RESUMEN CON ESTA ▼▼▼
# ===================================================================
@login_required
@permission_required('gestion_academica.ver_mis_calificaciones') 
@requiere_pagos_al_dia
def mis_cursos_y_calificaciones_resumen(request):
    """
    Muestra al estudiante un resumen de sus cursos y calificaciones,
    usando la misma lógica de cálculo que el boletín imprimible para
    garantizar la consistencia de los datos.
    """
    try:
        estudiante_actual = Estudiante.objects.get(usuario=request.user)
    except Estudiante.DoesNotExist:
        messages.error(request, "Tu perfil de estudiante no está configurado.")
        return redirect('gestion_academica:inicio_academico')

    periodo_activo = PeriodoAcademico.objects.filter(
        activo=True, institucion=estudiante_actual.institucion
    ).first()

    cursos_con_notas = []
    # Variables para el promedio general PONDERADO POR IHS
    total_puntos_ponderados_general = Decimal('0.0')
    total_ihs_general = 0

    if periodo_activo and estudiante_actual.grado_actual:
        # Hacemos la misma consulta optimizada que en el boletín
        cursos_del_estudiante = Curso.objects.filter(
            grado=estudiante_actual.grado_actual, periodo_academico=periodo_activo
        ).select_related('materia').prefetch_related(
            Prefetch('actividades_calificables', queryset=ActividadCalificable.objects.select_related('tipo_actividad').prefetch_related(
                Prefetch('calificaciones_recibidas', queryset=Calificacion.objects.filter(estudiante=estudiante_actual), to_attr='mi_calificacion')
            ))
        ).order_by('materia__nombre_materia')

        for curso in cursos_del_estudiante:
            # ▼▼▼ INICIO: LÓGICA DE CÁLCULO COPIADA EXACTAMENTE DEL BOLETÍN ▼▼▼
            nota_final_curso = Decimal('0.0')
            actividades_del_curso = curso.actividades_calificables.all()
            actividades_por_categoria = defaultdict(list)
            for actividad in actividades_del_curso:
                if actividad.tipo_actividad and actividad.tipo_actividad.porcentaje is not None:
                    actividades_por_categoria[actividad.tipo_actividad].append(actividad)

            for categoria, actividades_en_categoria in actividades_por_categoria.items():
                notas_de_la_categoria = []
                for actividad in actividades_en_categoria:
                    if hasattr(actividad, 'mi_calificacion') and actividad.mi_calificacion and actividad.mi_calificacion[0].valor_numerico is not None:
                        notas_de_la_categoria.append(actividad.mi_calificacion[0].valor_numerico)
                
                if notas_de_la_categoria:
                    promedio_categoria = sum(notas_de_la_categoria) / len(notas_de_la_categoria)
                    porcentaje_categoria = categoria.porcentaje
                    nota_ponderada_categoria = promedio_categoria * (porcentaje_categoria / Decimal('100.0'))
                    nota_final_curso += nota_ponderada_categoria
            
            nota_final_curso = nota_final_curso if nota_final_curso > 0 else None
            # ▲▲▲ FIN: LÓGICA DE CÁLCULO COPIADA EXACTAMENTE DEL BOLETÍN ▲▲▲

            # Acumulamos para el promedio general ponderado por IHS
            if nota_final_curso is not None and curso.materia.intensidad_horaria_semanal > 0:
                total_puntos_ponderados_general += nota_final_curso * curso.materia.intensidad_horaria_semanal
                total_ihs_general += curso.materia.intensidad_horaria_semanal
            
            cursos_con_notas.append({
                'curso': curso,
                'nota_final': nota_final_curso,
                'desempeno': obtener_desempeno(nota_final_curso, estudiante_actual.institucion) if nota_final_curso is not None else None,
            })

    # Cálculo final del promedio general
    promedio_general_periodo = total_puntos_ponderados_general / total_ihs_general if total_ihs_general > 0 else None

    context = {
        'titulo_pagina': "Mis Cursos y Calificaciones",
        'cursos_con_notas': cursos_con_notas,
        'promedio_general_periodo': promedio_general_periodo,
        'periodo_activo': periodo_activo,
        'estudiante': estudiante_actual,
    }
    return render(request, 'gestion_academica/estudiante_mis_cursos_calificaciones.html', context)

@login_required
@permission_required('gestion_academica.ver_mis_calificaciones') 
@requiere_pagos_al_dia
def detalle_mis_calificaciones_por_curso(request, curso_pk):
    try:
        estudiante_actual = get_filtered_queryset(Estudiante, request.user).get(usuario=request.user)
    except Estudiante.DoesNotExist:
        messages.error(request, "Tu perfil de estudiante no está configurado correctamente o no tienes permiso para acceder.")
        return redirect('gestion_academica:inicio_academico')
    
    curso = get_object_or_404(
        get_filtered_queryset(Curso, request.user), 
        pk=curso_pk
    )
    
    if estudiante_actual.grado_actual != curso.grado or estudiante_actual.institucion != curso.institucion:
        messages.error(request, "No tienes permiso para ver las calificaciones de este curso.")
        return redirect('gestion_academica:mis_cursos_calificaciones')
    
    actividades_del_curso = ActividadCalificable.objects.filter(curso=curso).order_by('fecha_publicacion', 'titulo')
    calificaciones_del_estudiante = Calificacion.objects.filter(estudiante=estudiante_actual, actividad_calificable__in=actividades_del_curso).select_related('actividad_calificable')
    calificaciones_por_actividad = {cal.actividad_calificable_id: cal for cal in calificaciones_del_estudiante}
    actividades_con_calificacion = [{'actividad': act, 'calificacion': calificaciones_por_actividad.get(act.pk)} for act in actividades_del_curso]
    
    context = {'titulo_pagina': f"Mis Calificaciones en: {curso.materia.nombre_materia}", 'curso': curso, 'actividades_con_calificacion': actividades_con_calificacion, 'estudiante': estudiante_actual}
    return render(request, 'gestion_academica/estudiante_detalle_calificaciones_curso.html', context)

# --- Vistas para Estudiantes - Entrega de Deberes ---
@login_required
@permission_required('gestion_academica.ver_mis_deberes') 
@requiere_pagos_al_dia
def mis_deberes_lista(request):
    try:
        estudiante_actual = get_filtered_queryset(Estudiante, request.user).get(usuario=request.user)
    except Estudiante.DoesNotExist:
        messages.error(request, "Tu perfil de estudiante no está configurado correctamente o no tienes permiso para acceder.")
        return redirect('gestion_academica:inicio_academico')
    
    if not estudiante_actual.grado_actual:
        messages.info(request, "Aún no estás asignado a un grado, por lo que no tienes deberes asignados.")
        context = {'titulo_pagina': "Mis Deberes", 'deberes_con_estado_entrega': []}
        return render(request, 'gestion_academica/estudiante_mis_deberes_lista.html', context)
    
    periodo_activo = PeriodoAcademico.objects.filter(
        activo=True,
        institucion=estudiante_actual.institucion
    ).first()

    if not periodo_activo:
        messages.warning(request, "No hay un periodo académico activo para tu institución. No se pueden mostrar deberes.")
        context = {'titulo_pagina': "Mis Deberes", 'deberes_con_estado_entrega': []}
        return render(request, 'gestion_academica/estudiante_mis_deberes_lista.html', context)
    
    cursos_del_estudiante = Curso.objects.filter(
        grado=estudiante_actual.grado_actual, 
        periodo_academico=periodo_activo,
        institucion=estudiante_actual.institucion 
    )
    deberes_asignados = Deber.objects.filter(
        curso__in=cursos_del_estudiante,
        institucion=estudiante_actual.institucion 
    ).select_related('curso__materia', 'curso__grado').order_by('-fecha_entrega')
    
    entregas_realizadas_ids = EntregaDeber.objects.filter(
        estudiante=estudiante_actual, deber__in=deberes_asignados
    ).values_list('deber_id', flat=True)
    
    deberes_con_estado_entrega = [
        {'deber': deber, 'entregado': deber.id in entregas_realizadas_ids} 
        for deber in deberes_asignados
    ]
    
    context = {'titulo_pagina': "Mis Deberes / Tareas Asignadas", 'deberes_con_estado_entrega': deberes_con_estado_entrega, 'periodo_activo': periodo_activo}
    return render(request, 'gestion_academica/estudiante_mis_deberes_lista.html', context)


@login_required
@permission_required('gestion_academica.puede_realizar_entrega_deber')
@requiere_pagos_al_dia
def realizar_entrega_deber(request, deber_pk):
    # (Esta parte de la vista no cambia)
    try:
        estudiante_actual = request.user.estudiante
    except Estudiante.DoesNotExist:
        messages.error(request, "Tu perfil de estudiante no está configurado.")
        return redirect('gestion_academica:inicio_academico')

    deber = get_object_or_404(Deber.objects.select_related('curso__grado'), pk=deber_pk)

    if estudiante_actual.grado_actual != deber.curso.grado:
        messages.error(request, "No tienes permiso para realizar una entrega para este deber.")
        return redirect('gestion_academica:dashboard_estudiante')

    entrega_obj, created = EntregaDeber.objects.get_or_create(
        deber=deber, estudiante=estudiante_actual,
        defaults={'institucion': estudiante_actual.institucion}
    )

    esta_vencido = timezone.now().date() > deber.fecha_entrega
    if esta_vencido and not entrega_obj.archivo_adjunto_estudiante:
        messages.warning(request, f"La fecha límite para '{deber.titulo}' ya ha pasado.")
        return redirect('gestion_academica:dashboard_estudiante')

    if request.method == 'POST':
        form = EntregaDeberForm(request.POST, request.FILES, instance=entrega_obj, request=request)
        if form.is_valid():
            entrega_guardada = form.save(commit=False)
            
            # --- INICIO DE LA MODIFICACIÓN ---
            # Guardamos una bandera para saber si se subió un archivo nuevo
            archivo_nuevo_subido = 'archivo_adjunto_estudiante' in request.FILES
            if archivo_nuevo_subido:
                entrega_guardada.fecha_entrega_real = timezone.now()
            
            entrega_guardada.save()
            
            # Si se subió un archivo, llamamos a la tarea de análisis de plagio.
            if archivo_nuevo_subido:
                analizar_plagio_tarea_task.delay(entrega_guardada.id)
            
            # Cambiamos el mensaje para notificar al estudiante
            messages.success(request, f"Entrega para '{deber.titulo}' guardada. HALU la analizará para asegurar su originalidad.")
            # --- FIN DE LA MODIFICACIÓN ---
            
            return redirect('gestion_academica:dashboard_estudiante')
    else:
        form = EntregaDeberForm(instance=entrega_obj, request=request)

    context = {
        'form': form, 'deber': deber, 'entrega_existente': entrega_obj,
        'titulo_formulario': f"{'Actualizar' if entrega_obj.archivo_adjunto_estudiante else 'Realizar'} Entrega para: {deber.titulo}"
    }
    return render(request, 'gestion_academica/estudiante_realizar_entrega_deber.html', context)


# --- Vistas para Registro de Calificaciones (por Docentes) ---
@login_required
@permission_required('gestion_academica.puede_calificar_estudiantes') 
def listar_estudiantes_para_calificar(request, actividad_pk):
    actividad = get_object_or_404(
        get_filtered_queryset(ActividadCalificable, request.user).select_related('curso__grado', 'curso__materia', 'curso__periodo_academico'), 
        pk=actividad_pk
    )

    if not request.user.is_superuser and not (hasattr(request.user, 'docente') and actividad.curso.docentes_asignados.filter(pk=request.user.docente.pk).exists()):
        messages.error(request, "No tienes permiso para calificar estudiantes en esta actividad.")
        return redirect('gestion_academica:docente_seleccionar_curso_libro_notas') 

    estudiantes_del_grado = Estudiante.objects.filter(
        grado_actual=actividad.curso.grado,
        institucion=actividad.institucion 
    ).select_related('usuario').order_by('usuario__last_name', 'usuario__first_name')
    
    calificaciones_existentes = Calificacion.objects.filter(
        actividad_calificable=actividad,
        estudiante__institucion=actividad.institucion 
    ).select_related('estudiante')
    
    calificaciones_por_estudiante = {cal.estudiante_id: cal for cal in calificaciones_existentes}
    
    estudiantes_con_calificacion = [
        {'estudiante': est, 'calificacion': calificaciones_por_estudiante.get(est.pk)} 
        for est in estudiantes_del_grado
    ]
    
    context = {'actividad': actividad, 'estudiantes_con_calificacion': estudiantes_con_calificacion, 'titulo_pagina': f"Calificar: {actividad.titulo}"}
    return render(request, 'gestion_academica/actividad_calificar_estudiantes.html', context)

@login_required
@permission_required('gestion_academica.puede_calificar_estudiantes') 
def registrar_editar_calificacion(request, actividad_pk, estudiante_pk):
    actividad = get_object_or_404(
        get_filtered_queryset(ActividadCalificable, request.user), 
        pk=actividad_pk
    )
    estudiante = get_object_or_404(
        get_filtered_queryset(Estudiante, request.user), 
        pk=estudiante_pk
    )

    if not request.user.is_superuser and not (hasattr(request.user, 'docente') and actividad.curso.docentes_asignados.filter(pk=request.user.docente.pk).exists()):
        messages.error(request, "No tienes permiso para calificar a estudiantes en esta actividad.")
        return redirect('gestion_academica:docente_seleccionar_curso_libro_notas') 

    if estudiante.grado_actual != actividad.curso.grado or estudiante.institucion != actividad.institucion:
        messages.error(request, "El estudiante no pertenece a este curso o institución.")
        return redirect('gestion_academica:listar_estudiantes_para_calificar', actividad_pk=actividad.pk)
    
    calificacion_obj, created = Calificacion.objects.get_or_create(
        actividad_calificable=actividad,
        estudiante=estudiante,
        defaults={'registrada_por': request.user.docente if hasattr(request.user, 'docente') else None, 'institucion': actividad.institucion} 
    )

    if request.method == 'POST':
        form = CalificacionForm(request.POST, instance=calificacion_obj)
        if form.is_valid():
            calificacion_guardada = form.save(commit=False)
            calificacion_guardada.actividad_calificable = actividad
            calificacion_guardada.estudiante = estudiante
            if hasattr(request.user, 'docente'):
                calificacion_guardada.registrada_por = request.user.docente
            if not calificacion_guardada.institucion:
                calificacion_guardada.institucion = actividad.institucion
            calificacion_guardada.save()
            messages.success(request, f"Calificación para {estudiante.usuario.get_full_name()} en '{actividad.titulo}' guardada exitosamente.")
            return redirect('gestion_academica:listar_estudiantes_para_calificar', actividad_pk=actividad.pk)
        else:
            messages.error(request, "Error al guardar la calificación. Por favor, revisa el formulario.")
    else:
        form = CalificacionForm(instance=calificacion_obj)
    
    context = {'form': form, 'actividad': actividad, 'estudiante': estudiante, 'titulo_formulario': f"Calificar a {estudiante.usuario.get_full_name()} en '{actividad.titulo}'"}
    return render(request, 'gestion_academica/calificacion_formulario.html', context)

# --- Vistas para Planes Curriculares ---
class PlanCurricularListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = PlanCurricular
    template_name = 'gestion_academica/plan_curricular_lista.html'
    context_object_name = 'planes'
    permission_required = 'gestion_academica.view_plancurricular'
    paginate_by = 10

    def get_queryset(self):
        base_queryset = PlanCurricular.objects.select_related(
            'grado_asociado', 'materia_asociada', 'periodo_academico_asociado', 'creado_por'
        ).order_by('-fecha_publicacion', 'nombre')
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Planes Curriculares"
        return context

class PlanCurricularDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = PlanCurricular
    template_name = 'gestion_academica/plan_curricular_detalle.html'
    context_object_name = 'plan'
    permission_required = 'gestion_academica.view_plancurricular'

    def get_queryset(self):
        base_queryset = super().get_queryset().select_related(
            'grado_asociado', 'materia_asociada', 'periodo_academico_asociado', 'creado_por'
        )
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Detalle: {self.object.nombre}"
        return context

class PlanCurricularCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = PlanCurricular
    form_class = PlanCurricularForm
    template_name = 'gestion_academica/plan_curricular_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_planes_curriculares')
    permission_required = 'gestion_academica.add_plancurricular'

    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        if not self.request.user.is_superuser and hasattr(self.request.user, 'institucion_asociada') and self.request.user.institucion_asociada:
            form.instance.institucion = self.request.user.institucion_asociada
        elif self.request.user.is_superuser and not form.instance.institucion:
            messages.error(self.request, "Como superusuario, debes seleccionar una institución para el plan curricular.")
            return self.form_invalid(form)

        messages.success(self.request, f"Plan Curricular '{form.cleaned_data['nombre']}' creado exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Crear Nuevo Plan Curricular"
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

class PlanCurricularUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = PlanCurricular
    form_class = PlanCurricularForm
    template_name = 'gestion_academica/plan_curricular_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_planes_curriculares')
    permission_required = 'gestion_academica.change_plancurricular'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def form_valid(self, form):
        messages.success(self.request, f"Plan Curricular '{form.cleaned_data['nombre']}' actualizado exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Plan Curricular"
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

class PlanCurricularDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = PlanCurricular
    template_name = 'gestion_academica/plan_curricular_confirmar_eliminar.html'
    success_url = reverse_lazy('gestion_academica:lista_planes_curriculares')
    context_object_name = 'plan'
    permission_required = 'gestion_academica.delete_plancurricular'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def delete(self, request, *args, **kwargs):
        plan_eliminado = self.get_object()
        messages.success(request, f"El Plan Curricular '{plan_eliminado.nombre}' ha sido eliminado.") 
        return super().delete(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Plan Curricular"
        return context

# --- Vistas para Menciones y Reconocimientos ---
class MencionReconocimientoListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = MencionReconocimiento
    template_name = 'gestion_academica/mencion_lista.html'
    context_object_name = 'menciones'
    permission_required = 'gestion_academica.view_mencionreconocimiento'
    paginate_by = 10

    def get_queryset(self):
        base_queryset = MencionReconocimiento.objects.select_related(
            'estudiante__usuario', 'curso__materia', 'periodo', 'otorgado_por__usuario'
        ).order_by('-fecha_otorgamiento', 'estudiante__usuario__last_name')
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Menciones y Reconocimientos"
        return context

class MencionReconocimientoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = MencionReconocimiento
    form_class = MencionReconocimientoForm
    template_name = 'gestion_academica/mencion_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_menciones')
    permission_required = 'gestion_academica.add_mencionreconocimiento'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user 
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        if not form.cleaned_data.get('otorgado_por') and hasattr(self.request.user, 'docente'):
            form.instance.otorgado_por = self.request.user.docente
        
        if not self.request.user.is_superuser and hasattr(self.request.user, 'institucion_asociada') and self.request.user.institucion_asociada:
            form.instance.institucion = self.request.user.institucion_asociada
        elif self.request.user.is_superuser and not form.instance.institucion:
            messages.error(self.request, "Como superusuario, debes seleccionar una institución para la mención.")
            return self.form_invalid(form)

        messages.success(self.request, f"Mención/Reconocimiento para '{form.cleaned_data['estudiante']}' creado exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Registrar Nueva Mención o Reconocimiento"
        return context

class MencionReconocimientoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = MencionReconocimiento
    form_class = MencionReconocimientoForm
    template_name = 'gestion_academica/mencion_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_menciones')
    permission_required = 'gestion_academica.change_mencionreconocimiento'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        kwargs['request'] = self.request 
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, f"Mención/Reconocimiento para '{form.cleaned_data['estudiante']}' actualizado exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Mención o Reconocimiento"
        return context

class MencionReconocimientoDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = MencionReconocimiento
    template_name = 'gestion_academica/mencion_confirmar_eliminar.html'
    success_url = reverse_lazy('gestion_academica:lista_menciones')
    context_object_name = 'mencion'
    permission_required = 'gestion_academica.delete_mencionreconocimiento'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def delete(self, request, *args, **kwargs):
        mencion_eliminada = self.get_object()
        messages.success(request, f"La mención/reconocimiento para '{mencion_eliminada.estudiante}' ha sido eliminada.") 
        return super().delete(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Mención/Reconocimiento"
        return context

# --- Vistas para Archivos de Planes Académicos y Materiales ---
class ArchivoPlanAcademicoListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = ArchivoPlanAcademico
    template_name = 'gestion_academica/archivo_plan_lista.html'
    context_object_name = 'archivos'
    permission_required = 'gestion_academica.view_archivoplanacademico'
    paginate_by = 10

    def get_queryset(self):
        base_queryset = ArchivoPlanAcademico.objects.select_related(
            'curso_asociado__materia', 'curso_asociado__grado', 
            'materia_asociada', 'subido_por'
        ).order_by('-fecha_subida')
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Materiales y Planes Académicos"
        return context

class ArchivoPlanAcademicoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = ArchivoPlanAcademico
    form_class = ArchivoPlanAcademicoForm
    template_name = 'gestion_academica/archivo_plan_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_archivos_plan')
    permission_required = 'gestion_academica.add_archivoplanacademico'

    def form_valid(self, form):
        form.instance.subido_por = self.request.user
        if not self.request.user.is_superuser and hasattr(self.request.user, 'institucion_asociada') and self.request.user.institucion_asociada:
            form.instance.institucion = self.request.user.institucion_asociada
        elif self.request.user.is_superuser and not form.instance.institucion:
            messages.error(self.request, "Como superusuario, debes seleccionar una institución para el archivo.")
            return self.form_invalid(form)

        messages.success(self.request, f"Archivo '{form.cleaned_data['nombre_archivo_descriptivo']}' subido exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Subir Nuevo Archivo/Material"
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

class ArchivoPlanAcademicoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = ArchivoPlanAcademico
    form_class = ArchivoPlanAcademicoForm
    template_name = 'gestion_academica/archivo_plan_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_archivos_plan')
    permission_required = 'gestion_academica.change_archivoplanacademico'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def form_valid(self, form):
        messages.success(self.request, f"Archivo '{form.cleaned_data['nombre_archivo_descriptivo']}' actualizado exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Archivo/Material"
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

class ArchivoPlanAcademicoDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = ArchivoPlanAcademico
    template_name = 'gestion_academica/archivo_plan_confirmar_eliminar.html'
    success_url = reverse_lazy('gestion_academica:lista_archivos_plan')
    context_object_name = 'archivo_plan'
    permission_required = 'gestion_academica.delete_archivoplanacademico'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def delete(self, request, *args, **kwargs):
        archivo_eliminado = self.get_object()
        messages.success(request, f"El archivo '{archivo_eliminado.nombre_archivo_descriptivo}' ha sido eliminado.") 
        return super().delete(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Archivo"
        return context

# --- Vistas para Estudiantes - Mi Boletín ---
@login_required
@permission_required('gestion_academica.ver_mi_boletin')
@requiere_pagos_al_dia
def mi_boletin_periodo_actual(request):
    try:
        estudiante_actual = Estudiante.objects.select_related('usuario', 'grado_actual', 'institucion').get(usuario=request.user)
    except Estudiante.DoesNotExist:
        messages.error(request, "Tu perfil de estudiante no está configurado.")
        return redirect('gestion_academica:inicio_academico')

    if not estudiante_actual.grado_actual:
        messages.info(request, "Aún no estás asignado a un grado para generar el boletín.")
        return render(request, 'gestion_academica/estudiante_mi_boletin.html', {'estudiante': estudiante_actual, 'cursos_con_detalle': []})

    periodo_activo = PeriodoAcademico.objects.filter(activo=True, institucion=estudiante_actual.institucion).first()

    cursos_con_detalle = []
    total_puntos_ponderados_general = Decimal('0.0')
    total_ihs_general = 0

    if periodo_activo:
        cursos_del_estudiante = Curso.objects.filter(
            grado=estudiante_actual.grado_actual, periodo_academico=periodo_activo
        ).select_related('materia').order_by('materia__nombre_materia')

        for curso in cursos_del_estudiante:
            estado_academico = calcular_estado_academico_curso(curso, estudiante_actual)
            nota_final_curso = estado_academico.get('nota_final_ponderada')

            if nota_final_curso is not None and curso.materia.intensidad_horaria_semanal > 0:
                total_puntos_ponderados_general += nota_final_curso * curso.materia.intensidad_horaria_semanal
                total_ihs_general += curso.materia.intensidad_horaria_semanal
            
            cursos_con_detalle.append({
                'curso': curso,
                'nota_final_curso': nota_final_curso,
                'desempeno': obtener_desempeno(nota_final_curso, estudiante_actual.institucion) if nota_final_curso is not None else None,
            })
    
    promedio_general_periodo = total_puntos_ponderados_general / total_ihs_general if total_ihs_general > 0 else None

    context = {
        'estudiante': estudiante_actual,
        'periodo_activo': periodo_activo,
        'cursos_con_detalle': cursos_con_detalle,
        'promedio_general_periodo': promedio_general_periodo,
        'titulo_pagina': 'Resumen de Calificaciones'
    }
    return render(request, 'gestion_academica/estudiante_mi_boletin.html', context)

# =========================================================================
# FIN: VISTA DEL BOLETÍN MODIFICADA
# =========================================================================

def obtener_desempeno(nota):
    """
    Función auxiliar para convertir una nota numérica a una escala cualitativa.
    """
    if nota is None:
        return ""
    nota = Decimal(str(nota)) # Asegurar que es Decimal
    # Ajusta estos rangos según la escala de tu institución
    if nota >= 4.6: return "Sup."
    if nota >= 4.0: return "Alt."
    if nota >= 3.0: return "Bas."
    return "Baj."

@login_required
@permission_required('gestion_academica.ver_mi_boletin', raise_exception=True)
@requiere_pagos_al_dia
def boletin_imprimible(request, estudiante_pk, periodo_pk):
    """
    Genera un boletín en PDF para un estudiante y periodo específicos,
    con tamaño de papel dinámico y agrupación por áreas académicas.
    """
    # 1. Obtiene el formato de papel desde la URL (ej: ?formato=oficio)
    formato_papel = request.GET.get('formato', 'carta').lower()

    # 2. Obtiene los objetos principales de la base de datos
    try:
        estudiante_actual = Estudiante.objects.select_related(
            'usuario', 'grado_actual', 'institucion'
        ).get(pk=estudiante_pk)
        periodo = PeriodoAcademico.objects.get(
            pk=periodo_pk, institucion=estudiante_actual.institucion
        )
    except (Estudiante.DoesNotExist, PeriodoAcademico.DoesNotExist):
        messages.error(request, "El estudiante o periodo académico solicitado no es válido.")
        return redirect('gestion_academica:inicio_academico') # Ajusta esta URL si es necesario

    # 3. Verifica los permisos de acceso al boletín
    es_el_mismo_estudiante = (request.user.pk == estudiante_actual.usuario.pk)
    es_staff = request.user.is_staff
    es_familiar = hasattr(request.user, 'familiar') and request.user.familiar.estudiantes_asociados.filter(pk=estudiante_pk).exists()
    
    if not (es_el_mismo_estudiante or es_staff or es_familiar):
        messages.error(request, "No tienes permiso para ver este boletín.")
        return redirect('gestion_academica:inicio_academico') # Ajusta esta URL si es necesario

    # 4. Consulta principal para obtener los cursos y pre-cargar datos relacionados
    cursos = Curso.objects.filter(
        grado=estudiante_actual.grado_actual, periodo_academico=periodo
    ).select_related('materia').prefetch_related(
        'materia__areaacademica_set',
        Prefetch('materia__descriptores', queryset=DescriptorLogro.objects.filter(periodo_academico=periodo), to_attr='descriptores_del_periodo')
    )
    
    # 5. Procesa los datos para agruparlos por área y calcular promedios
    areas_data = defaultdict(lambda: {'materias': [], 'nombre': ''})
    total_puntos_ponderados_general = Decimal('0.0')
    total_ihs_general = 0

    for curso in cursos:
        # Lógica de agrupación por área
        area_obj = curso.materia.areaacademica_set.first()
        area_nombre = area_obj.nombre if area_obj else "Otras Materias"
        
        # Lógica de cálculo de notas
        estado_academico = calcular_estado_academico_curso(curso, estudiante_actual)
        nota_final_curso = estado_academico.get('nota_final_ponderada')

        if nota_final_curso is not None and curso.materia.intensidad_horaria_semanal > 0:
            total_puntos_ponderados_general += nota_final_curso * curso.materia.intensidad_horaria_semanal
            total_ihs_general += curso.materia.intensidad_horaria_semanal

        # Construye el diccionario de datos para cada materia
        materia_data = {
            'nombre': curso.materia.nombre_materia,
            'nombre_L2': curso.materia.nombre_idioma_secundario or '',
            'ihs': curso.materia.intensidad_horaria_semanal,
            'calificacion': nota_final_curso,
            'desempeno': obtener_desempeno(nota_final_curso, estudiante_actual.institucion),
            'inasistencias': RegistroAsistencia.objects.filter(estudiante=estudiante_actual, curso=curso, estado='AUSENTE').count(),
            'descriptores': [d.descripcion for d in curso.materia.descriptores_del_periodo]
        }
        
        areas_data[area_nombre]['materias'].append(materia_data)
        if not areas_data[area_nombre]['nombre']:
            areas_data[area_nombre]['nombre'] = area_nombre

    # 6. Realiza cálculos finales y obtiene datos adicionales
    boletin_data_ordenado = sorted(areas_data.values(), key=lambda x: x['nombre'])
    promedio_general = total_puntos_ponderados_general / total_ihs_general if total_ihs_general > 0 else None
    
    director_de_grupo = DirectorCurso.objects.select_related('docente__usuario').filter(
        grado=estudiante_actual.grado_actual, periodo_academico=periodo
    ).first()
    
    observacion_obj = ObservacionBoletin.objects.filter(estudiante=estudiante_actual, periodo=periodo).first()
    observaciones_texto = observacion_obj.observacion if observacion_obj else "No hay observaciones registradas para este periodo."
    
    # 7. Construye el contexto final para la plantilla
    context = {
        'institucion': estudiante_actual.institucion,
        'periodo': periodo,
        'estudiante': estudiante_actual,
        'boletin_data': boletin_data_ordenado,
        'promedio_general': promedio_general,
        'director_de_grupo': director_de_grupo,
        'observaciones': observaciones_texto,
        'fecha_emision': timezone.now(),
        'formato_papel': formato_papel,
        'es_bilingue': getattr(estudiante_actual.institucion, 'es_bilingue', False),
    }

    # 8. Renderiza el HTML y lo convierte a PDF
    template_path = 'gestion_academica/boletin_imprimible_formal.html'
    template = get_template(template_path)
    html = template.render(context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Boletin_{estudiante_actual.usuario.username}_{periodo.nombre}.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
    
    if pisa_status.err:
        logger.error("Error al generar PDF con xhtml2pdf (pisa.err=%s)", pisa_status.err)
        return HttpResponse('Error al generar el PDF. Por favor, inténtelo de nuevo.', status=500)
        
    return response


def _resumen_por_estudiante_portal_familiar(estudiante):
    """
    Datos ligeros para la portada del portal familiar (y API móvil): compromisos
    próximos (14 días) y aviso si el estudiante no está al día en pagos.
    """
    hoy = timezone.now().date()
    limite = hoy + timedelta(days=14)
    out = {
        "proximos_compromisos": 0,
        "alerta_finanzas": False,
        "hay_periodo_y_grado": False,
    }
    periodo = PeriodoAcademico.objects.filter(
        activo=True, institucion_id=estudiante.institucion_id
    ).first()
    if not periodo or not estudiante.grado_actual_id:
        return out
    out["hay_periodo_y_grado"] = True
    curso_ids = list(
        Curso.objects.filter(
            grado_id=estudiante.grado_actual_id,
            periodo_academico_id=periodo.pk,
        ).values_list("pk", flat=True)
    )
    if not curso_ids:
        return out
    out["proximos_compromisos"] = Deber.objects.filter(
        curso_id__in=curso_ids,
        fecha_entrega__range=(hoy, limite),
    ).count() + ActividadCalificable.objects.filter(
        curso_id__in=curso_ids,
        fecha_entrega_limite__range=(hoy, limite),
    ).count()
    try:
        out["alerta_finanzas"] = not estudiante.esta_al_dia()
    except Exception:
        out["alerta_finanzas"] = False
    return out


# --- Vistas para Portal de Familiares ---
@login_required
@permission_required('gestion_academica.acceso_portal_familiar') 
def portal_familiar_inicio(request):
    """
    Página principal para el rol de Familiar. Muestra los estudiantes
    asociados y las acciones principales que puede realizar.
    """
    try:
        # Usamos el método directo y claro para obtener el familiar
        familiar_actual = request.user.familiar
    except (AttributeError, Familiar.DoesNotExist):
        # Si el usuario no tiene un perfil de familiar, se le niega el acceso
        messages.error(request, "Acceso denegado o perfil de familiar no configurado.")
        return redirect('gestion_academica:inicio_academico')
    
    # Obtenemos los estudiantes asociados a este familiar y los datos relacionados
    estudiantes_qs = familiar_actual.estudiantes_asociados.select_related(
        'usuario',
        'grado_actual',
        'institucion',
    ).order_by('usuario__last_name', 'usuario__first_name')

    estudiantes_con_resumen = [
        {'estudiante': est, 'resumen': _resumen_por_estudiante_portal_familiar(est)}
        for est in estudiantes_qs
    ]

    context = {
        'titulo_pagina': "Portal de Familiares",
        'familiar': familiar_actual,
        'estudiantes_con_resumen': estudiantes_con_resumen,
    }
    return render(request, 'gestion_academica/portal_familiar_inicio.html', context)

@login_required
@permission_required('gestion_academica.ver_calificaciones_estudiante_familiar') 
def familiar_ver_calificaciones_estudiante(request, estudiante_pk):
    try:
        # Aseguramos que el familiar solo pueda acceder a sus estudiantes
        familiar_actual = request.user.familiar
        estudiante_seleccionado = familiar_actual.estudiantes_asociados.select_related(
            'usuario', 'grado_actual', 'institucion'
        ).get(pk=estudiante_pk)
    except (Familiar.DoesNotExist, Estudiante.DoesNotExist):
        messages.error(request, "No tienes permiso para ver la información de este estudiante.")
        return redirect('gestion_academica:portal_familiar_inicio')

    periodo_activo = PeriodoAcademico.objects.filter(
        activo=True, 
        institucion=estudiante_seleccionado.institucion
    ).first()
    
    cursos_del_estudiante = []
    if periodo_activo and estudiante_seleccionado.grado_actual:
        cursos_del_estudiante = Curso.objects.filter(
            grado=estudiante_seleccionado.grado_actual,
            periodo_academico=periodo_activo,
            institucion=estudiante_seleccionado.institucion,
        ).select_related('materia')
    
    context = {
        'titulo_pagina': f"Resumen de Calificaciones de {estudiante_seleccionado.usuario.get_full_name()}",
        'estudiante_seleccionado': estudiante_seleccionado,
        'cursos': cursos_del_estudiante, # Ahora la variable 'cursos' tendrá contenido
        'periodo_activo': periodo_activo,
        'es_vista_familiar': True,
    }
    # Asegúrate que la plantilla se llame así o ajusta el nombre
    return render(request, 'gestion_academica/familiar_ver_calificaciones.html', context)

@login_required
@permission_required('gestion_academica.ver_calificaciones_estudiante_familiar') 
def familiar_ver_detalle_calificaciones_curso_estudiante(request, estudiante_pk, curso_pk):
    familiar_actual = None
    try:
        familiar_actual = get_filtered_queryset(Familiar, request.user).get(usuario=request.user)
    except Familiar.DoesNotExist:
        messages.error(request, "Acceso denegado o perfil de familiar no configurado o no tienes permiso para acceder.")
        return redirect('gestion_academica:inicio_academico')

    estudiante = get_object_or_404(
        get_filtered_queryset(Estudiante, request.user), 
        pk=estudiante_pk
    )
    curso = get_object_or_404(
        get_filtered_queryset(Curso, request.user).select_related('materia', 'grado', 'periodo_academico'), 
        pk=curso_pk
    )

    if not familiar_actual.estudiantes_asociados.filter(pk=estudiante.pk).exists() or \
       (estudiante.grado_actual and estudiante.grado_actual != curso.grado) or \
       estudiante.institucion != curso.institucion: 
        messages.error(request, "No tienes permiso para ver esta información o el estudiante no pertenece a este curso.")
        return redirect('gestion_academica:portal_familiar_inicio')

    actividades_del_curso = ActividadCalificable.objects.filter(curso=curso).order_by('fecha_publicacion', 'titulo')
    calificaciones_del_estudiante = Calificacion.objects.filter(
        estudiante=estudiante,
        actividad_calificable__in=actividades_del_curso
    ).select_related('actividad_calificable')
    calificaciones_por_actividad = {cal.actividad_calificable_id: cal for cal in calificaciones_del_estudiante}
    actividades_con_calificacion = []
    for act in actividades_del_curso:
        actividades_con_calificacion.append({
            'actividad': act,
            'calificacion': calificaciones_por_actividad.get(act.pk)
        })
    
    context = {
        'titulo_pagina': f"Calificaciones de {estudiante.usuario.get_full_name()} en {curso.materia.nombre_materia}",
        'curso': curso,
        'actividades_con_calificacion': actividades_con_calificacion,
        'estudiante': estudiante,
        'es_vista_familiar': True,
    }
    return render(request, 'gestion_academica/estudiante_detalle_calificaciones_curso.html', context)

@login_required
@permission_required('gestion_academica.ver_boletin_estudiante_familiar') 
def familiar_ver_boletin_estudiante(request, estudiante_pk):
    familiar_actual = None
    try:
        familiar_actual = get_filtered_queryset(Familiar, request.user).get(usuario=request.user)
    except Familiar.DoesNotExist:
        messages.error(request, "Acceso denegado o perfil de familiar no configurado o no tienes permiso para acceder.")
        return redirect('gestion_academica:inicio_academico')
        
    estudiante_seleccionado = get_object_or_404(
        get_filtered_queryset(Estudiante, request.user).select_related('usuario', 'grado_actual'), 
        pk=estudiante_pk
    )

    if not familiar_actual.estudiantes_asociados.filter(pk=estudiante_seleccionado.pk).exists():
        messages.error(request, "No tienes permiso para ver el boletín de este estudiante.")
        return redirect('gestion_academica:portal_familiar_inicio')

    periodo_activo = PeriodoAcademico.objects.filter(
        activo=True,
        institucion=estudiante_seleccionado.institucion 
    ).first()
    
    cursos_con_detalle = []
    promedio_general_periodo = None
    
    if estudiante_seleccionado.grado_actual and periodo_activo:
        cursos_del_estudiante = Curso.objects.filter(
            grado=estudiante_seleccionado.grado_actual,
            periodo_academico=periodo_activo,
            institucion=estudiante_seleccionado.institucion 
        ).select_related('materia', 'grado', 'periodo_academico').prefetch_related('docentes_asignados__usuario').order_by('materia__nombre_materia')

        suma_ponderada_total = Decimal('0.00')
        suma_porcentajes_total = Decimal('0.00') 
        
        promedio_general_periodo_numerador = Decimal('0.00')
        promedio_general_periodo_denominador = 0

        for curso_iter in cursos_del_estudiante:
            actividades_del_curso = ActividadCalificable.objects.filter(curso=curso_iter).order_by('fecha_publicacion', 'titulo')
            calificaciones_del_estudiante = Calificacion.objects.filter(
                estudiante=estudiante_seleccionado,
                actividad_calificable__in=actividades_del_curso,
                valor_numerico__isnull=False
            ).select_related('actividad_calificable')
            
            calificaciones_por_actividad = {cal.actividad_calificable_id: cal for cal in calificaciones_del_estudiante}
            actividades_para_boletin = []
            suma_notas_ponderadas_curso = Decimal('0.00')
            suma_porcentajes_curso = Decimal('0.00')

            for act in actividades_del_curso:
                calificacion_actual = calificaciones_por_actividad.get(act.pk)
                if calificacion_actual and calificacion_actual.valor_numerico is not None and act.porcentaje_en_periodo is not None:
                    suma_notas_ponderadas_curso += (Decimal(str(calificacion_actual.valor_numerico)) * Decimal(str(act.porcentaje_en_periodo))) 
                    suma_porcentajes_curso += Decimal(str(act.porcentaje_en_periodo))
                actividades_para_boletin.append({'actividad': act, 'calificacion': calificacion_actual})
            
            nota_final_curso = None
            if suma_porcentajes_curso > Decimal('0.00'):
                nota_final_curso = suma_notas_ponderadas_curso / suma_porcentajes_curso
                
                if nota_final_curso is not None:
                    promedio_general_periodo_numerador += nota_final_curso
                    promedio_general_periodo_denominador += 1
            
            cursos_con_detalle.append({'curso': curso_iter, 'actividades_con_calificacion': actividades_para_boletin, 'nota_final_curso': nota_final_curso})

        if promedio_general_periodo_denominador > 0:
            promedio_general_periodo = promedio_general_periodo_numerador / Decimal(str(promedio_general_periodo_denominador)) 
        
    context = {
        'titulo_pagina': f"Boletín de {estudiante_seleccionado.usuario.get_full_name()}",
        'estudiante': estudiante_seleccionado,
        'periodo_activo': periodo_activo,
        'cursos_con_detalle': cursos_con_detalle,
        'promedio_general_periodo': promedio_general_periodo,
        'es_vista_familiar': True,
    }
    return render(request, 'gestion_academica/estudiante_mi_boletin.html', context)

@login_required
@permission_required('gestion_academica.ver_deberes_estudiante_familiar') 
def familiar_ver_deberes_estudiante(request, estudiante_pk):
    familiar_actual = None
    try:
        familiar_actual = get_filtered_queryset(Familiar, request.user).get(usuario=request.user)
    except Familiar.DoesNotExist:
        messages.error(request, "Acceso denegado o perfil de familiar no configurado o no tienes permiso para acceder.")
        return redirect('gestion_academica:inicio_academico') # CORRECCIÓN: redirigir a inicio_academico
        
    estudiante_seleccionado = get_object_or_404(
        get_filtered_queryset(Estudiante, request.user).select_related('usuario', 'grado_actual'), 
        pk=estudiante_pk
    )

    if not familiar_actual.estudiantes_asociados.filter(pk=estudiante_seleccionado.pk).exists():
        messages.error(request, "No tienes permiso para ver los deberes de este estudiante.")
        return redirect('gestion_academica:portal_familiar_inicio')

    periodo_activo = PeriodoAcademico.objects.filter(
        activo=True,
        institucion=estudiante_seleccionado.institucion 
    ).first()
    
    deberes_con_estado_entrega = []

    if estudiante_seleccionado.grado_actual and periodo_activo:
        cursos_del_estudiante = Curso.objects.filter(
            grado=estudiante_seleccionado.grado_actual,
            periodo_academico=periodo_activo,
            institucion=estudiante_seleccionado.institucion 
        )
        deberes_asignados = Deber.objects.filter(
            curso__in=cursos_del_estudiante,
            institucion=estudiante_seleccionado.institucion 
        ).select_related('curso__materia', 'curso__grado').order_by('-fecha_entrega')
        
        entregas_realizadas_ids = EntregaDeber.objects.filter(
            estudiante=estudiante_seleccionado,
            deber__in=deberes_asignados
        ).values_list('deber_id', flat=True)
        
        for deber in deberes_asignados:
            deberes_con_estado_entrega.append({
                'deber': deber,
                'entregado': deber.id in entregas_realizadas_ids,
                'puede_entregar': False 
            })
        
    context = {
        'titulo_pagina': f"Deberes de {estudiante_seleccionado.usuario.get_full_name()}",
        'deberes_con_estado_entrega': deberes_con_estado_entrega,
        'periodo_activo': periodo_activo,
        'estudiante_seleccionado': estudiante_seleccionado,
        'es_vista_familiar': True,
    }
    return render(request, 'gestion_academica/estudiante_mis_deberes_lista.html', context)

# --- Vistas para Docentes - Libro de Notas ---
@login_required
def docente_libro_de_notas_por_curso(request, curso_pk):
    # Determinar si quien accede es coordinador/admin (acceso de sólo lectura)
    es_coordinador = request.user.is_staff and getattr(request.user, 'rol', None) in ['administrador', 'coordinador']

    # Verificar que el usuario es docente, coordinador o superusuario
    if not (hasattr(request.user, 'docente') or es_coordinador or request.user.is_superuser):
        messages.error(request, "Acceso denegado.")
        return redirect('gestion_academica:inicio_academico')

    curso = get_object_or_404(Curso.objects.select_related('materia', 'grado', 'periodo_academico'), pk=curso_pk)

    # Los coordinadores de la misma institución pueden ver cualquier curso; los docentes sólo los suyos
    if not es_coordinador and not request.user.is_superuser:
        if not hasattr(request.user, 'docente') or not curso.docentes_asignados.filter(pk=request.user.docente.pk).exists():
            messages.error(request, "No tienes permiso para ver el libro de notas de este curso.")
            return redirect('gestion_academica:dashboard_docente')

    estudiantes_del_curso = Estudiante.objects.filter(
        grado_actual=curso.grado, institucion=curso.institucion
    ).select_related('usuario').order_by('usuario__last_name', 'usuario__first_name')

    actividades_del_curso = ActividadCalificable.objects.filter(
    curso=curso
    ).select_related('tipo_actividad').order_by('tipo_actividad__orden', 'titulo')

    if request.method == 'POST':
        # Los coordinadores acceden en modo sólo lectura — no pueden guardar notas
        if es_coordinador and not request.user.is_superuser:
            messages.warning(request, "Los coordinadores pueden consultar el libro de notas pero no modificar calificaciones.")
            return redirect('gestion_academica:docente_libro_de_notas_por_curso', curso_pk=curso.pk)
        with transaction.atomic():
            calificaciones_a_actualizar = []
            calificaciones_a_crear_temp = []
            calificaciones_ids_para_revisar = []

            for estudiante in estudiantes_del_curso:
                for actividad in actividades_del_curso:
                    nota_str = request.POST.get(f'nota-E{estudiante.pk}-A{actividad.pk}')
                    if nota_str is not None and nota_str.strip() != '':
                        try:
                            valor_nota = Decimal(nota_str.replace(',', '.'))
                            calificacion_existente = Calificacion.objects.filter(estudiante=estudiante, actividad_calificable=actividad).first()
                            
                            if calificacion_existente:
                                if calificacion_existente.valor_numerico != valor_nota:
                                    calificacion_existente.valor_numerico = valor_nota
                                    calificaciones_a_actualizar.append(calificacion_existente)
                                    calificaciones_ids_para_revisar.append(calificacion_existente.id)
                            else:
                                calificaciones_a_crear_temp.append(
                                    Calificacion(
                                        estudiante=estudiante, actividad_calificable=actividad,
                                        valor_numerico=valor_nota, registrada_por=request.user.docente,
                                        institucion=curso.institucion
                                    )
                                )
                        except (ValueError, Inexact):
                            messages.warning(request, f"El valor '{nota_str}' para {estudiante} no es una nota válida.")

            if calificaciones_a_actualizar:
                Calificacion.objects.bulk_update(calificaciones_a_actualizar, ['valor_numerico'])
            
            if calificaciones_a_crear_temp:
                calificaciones_creadas = Calificacion.objects.bulk_create(calificaciones_a_crear_temp)
                # Recopilamos los IDs de las nuevas calificaciones creadas
                for cal in calificaciones_creadas:
                    calificaciones_ids_para_revisar.append(cal.id)
            
            messages.success(request, "¡Notas guardadas exitosamente!")

        if calificaciones_ids_para_revisar:
            api_key = institucion_google_api_key(curso.institucion)
            ia_disponible = False
            model = None
            if api_key:
                try:
                    genai.configure(api_key=api_key)
                    # Usamos el modelo estable actual: gemini-2.5-flash
                    model = genai.GenerativeModel('gemini-2.5-flash')
                    ia_disponible = True
                except Exception as e:
                    print(f"--- ERROR FATAL DE IA: No se pudo configurar la API de Google: {e} ---")
            else:
                print("--- IA omitida: la institución no tiene google_api_key (Gemini). ---")

            if ia_disponible:
                calificaciones_procesadas = Calificacion.objects.filter(pk__in=calificaciones_ids_para_revisar)
                for calificacion in calificaciones_procesadas:
                    estudiante = calificacion.estudiante
                    nota_minima_aprobacion = getattr(estudiante.institucion, 'nota_minima_aprobacion', Decimal('3.0'))

                    if calificacion.valor_numerico is not None and calificacion.valor_numerico < nota_minima_aprobacion:
                        actividad = calificacion.actividad_calificable
                        
                        prompt = (
                            f"Actúa como un tutor amigable llamado HALU. Un estudiante de '{estudiante.grado_actual.nombre}' "
                            f"obtuvo una calificación baja de '{calificacion.valor_numerico}' en la materia de '{actividad.curso.materia.nombre_materia}' "
                            f"sobre el tema '{actividad.titulo}'. "
                            "Genera un consejo corto en español con 3 pasos de estudio concretos y accionables. "
                            "El tono debe ser alentador y positivo."
                        )
                        
                        try:
                            response = model.generate_content(prompt)
                            consejo_generado = response.text
                            
                            Notificacion.objects.create(
                                destinatario=estudiante.usuario, 
                                mensaje=f"HALU te ha generado un plan de estudio para '{actividad.titulo}'.",
                                consejo_ia=consejo_generado,
                                institucion=estudiante.institucion # <-- LÍNEA AÑADIDA
                            )
                        except Exception as e:
                            print(f">>> ERROR al llamar a la API de Google: {e} !!!")

        messages.success(request, "¡Notas guardadas y notificaciones de refuerzo enviadas!")
        return redirect('gestion_academica:docente_libro_de_notas_por_curso', curso_pk=curso.pk)

    # 2. Preparación de datos para mostrar la planilla (la parte que ya tenías, pero mejorada)
    calificaciones_existentes = Calificacion.objects.filter(actividad_calificable__in=actividades_del_curso)
    calificaciones_map = defaultdict(dict)
    for cal in calificaciones_existentes:
        calificaciones_map[cal.estudiante_id][cal.actividad_calificable_id] = cal.valor_numerico

    actividades_agrupadas = defaultdict(list)
    for actividad in actividades_del_curso:
        actividades_agrupadas[actividad.tipo_actividad].append(actividad)

    # 3. Lógica para calcular y mostrar la nota final en la planilla
    libro_notas_data = []
    for estudiante in estudiantes_del_curso:
        nota_final_curso = Decimal('0.0')
        for categoria, actividades_en_categoria in actividades_agrupadas.items():
            notas_de_la_categoria = [
                calificaciones_map.get(estudiante.pk, {}).get(act.pk) 
                for act in actividades_en_categoria 
                if calificaciones_map.get(estudiante.pk, {}).get(act.pk) is not None
            ]
            if notas_de_la_categoria:
                promedio_categoria = sum(notas_de_la_categoria) / len(notas_de_la_categoria)
                if categoria and categoria.porcentaje:
                    nota_final_curso += promedio_categoria * (categoria.porcentaje / Decimal('100.0'))
        
        libro_notas_data.append({
            'estudiante': estudiante,
            'nota_final_curso': nota_final_curso if nota_final_curso > 0 else None
        })

    # ── VISTA ANUAL: todos los períodos del mismo año escolar ──────────────
    año_escolar = curso.periodo_academico.año_escolar
    periodos_del_año = PeriodoAcademico.objects.filter(
        año_escolar=año_escolar,
        institucion=curso.institucion
    ).order_by('fecha_inicio')

    # Mapa periodo.pk → curso de la misma materia+grado en ese período
    cursos_por_periodo = {}
    for p in periodos_del_año:
        cursos_por_periodo[p.pk] = Curso.objects.filter(
            materia=curso.materia,
            grado=curso.grado,
            periodo_academico=p,
            institucion=curso.institucion
        ).first()

    nota_minima = getattr(curso.institucion, 'nota_minima_aprobacion', Decimal('3.0'))

    datos_anuales = []
    for estudiante in estudiantes_del_curso:
        notas_periodos = []
        notas_validas = []
        for periodo in periodos_del_año:
            curso_p = cursos_por_periodo.get(periodo.pk)
            if curso_p:
                estado = calcular_estado_academico_curso(curso_p, estudiante)
                nota = estado.get('nota_final_ponderada')
            else:
                nota = None
            es_actual = (periodo.pk == curso.periodo_academico.pk)
            notas_periodos.append({
                'periodo': periodo,
                'nota': nota,
                'es_actual': es_actual,
            })
            if nota is not None:
                notas_validas.append(nota)

        promedio_anual = (sum(notas_validas) / Decimal(len(notas_validas))) if notas_validas else None
        datos_anuales.append({
            'estudiante': estudiante,
            'notas_periodos': notas_periodos,
            'promedio_anual': promedio_anual,
            'aprueba': (promedio_anual >= nota_minima) if promedio_anual is not None else None,
        })
    # ──────────────────────────────────────────────────────────────────────

    context = {
        'curso': curso,
        'estudiantes': estudiantes_del_curso,
        'actividades_agrupadas': dict(actividades_agrupadas),
        'notas_map': calificaciones_map,
        'libro_notas_data': libro_notas_data,
        'datos_anuales': datos_anuales,
        'periodos_del_año': periodos_del_año,
        'año_escolar': año_escolar,
        'nota_minima': nota_minima,
        'titulo_pagina': f"Libro de Notas: {curso}",
        'es_coordinador': es_coordinador,
    }
    return render(request, 'gestion_academica/docente_libro_de_notas_por_curso.html', context)

@login_required
@permission_required('gestion_academica.acceso_libro_notas_docente')
def docente_seleccionar_curso_libro_notas(request):
    """
    Muestra al docente una lista de sus cursos, permitiendo filtrar por
    periodo académico.
    """
    if not hasattr(request.user, 'docente'):
        messages.error(request, "Acceso denegado.")
        return redirect('gestion_academica:inicio_academico')

    docente = request.user.docente
    institucion = request.user.institucion_asociada
    
    # Obtenemos todos los periodos de la institución para el filtro
    periodos_disponibles = PeriodoAcademico.objects.filter(institucion=institucion).order_by('-año_escolar', '-fecha_inicio')
    
    # Determinamos el periodo a mostrar: el que viene por GET o el activo por defecto
    periodo_seleccionado_id = request.GET.get('periodo')
    if periodo_seleccionado_id:
        periodo_a_mostrar = get_object_or_404(PeriodoAcademico, pk=periodo_seleccionado_id, institucion=institucion)
    else:
        periodo_a_mostrar = periodos_disponibles.filter(activo=True).first()

    # Filtramos los cursos del docente basados en el periodo seleccionado
    cursos_del_docente = []
    if periodo_a_mostrar:
        cursos_del_docente = Curso.objects.filter(
            docentes_asignados=docente,
            periodo_academico=periodo_a_mostrar
        ).select_related('materia', 'grado')

    context = {
        'cursos': cursos_del_docente,
        'periodos_disponibles': periodos_disponibles,
        'periodo_seleccionado': periodo_a_mostrar,
        'titulo_pagina': "Seleccionar Curso para Calificar"
    }
    return render(request, 'gestion_academica/seleccionar_curso_libro_notas.html', context)


@login_required
def coordinador_seleccionar_curso_libro_notas(request):
    """
    Libro de notas del coordinador — navegación jerárquica de solo lectura:
      Nivel 1 (default): lista de grados con cursos en el período seleccionado
      Nivel 2 (?grado_pk=X): materias/cursos de ese grado
    Solo lectura; los docentes son los únicos que pueden modificar notas.
    """
    institucion = getattr(request.user, 'institucion_asociada', None)
    es_coordinador = request.user.is_staff and getattr(request.user, 'rol', None) in ['administrador', 'coordinador']

    if not (es_coordinador or request.user.is_superuser):
        messages.error(request, "Acceso denegado. Esta sección es exclusiva del coordinador.")
        return redirect('gestion_academica:inicio_academico')

    periodos_disponibles = PeriodoAcademico.objects.filter(
        institucion=institucion
    ).order_by('-año_escolar', '-fecha_inicio') if institucion else PeriodoAcademico.objects.none()

    periodo_seleccionado_id = request.GET.get('periodo')
    if periodo_seleccionado_id:
        periodo_a_mostrar = get_object_or_404(PeriodoAcademico, pk=periodo_seleccionado_id, institucion=institucion)
    else:
        periodo_a_mostrar = periodos_disponibles.filter(activo=True).first() or periodos_disponibles.first()

    grado_pk   = request.GET.get('grado_pk')
    grado_actual = None
    grados     = []
    cursos     = []
    nivel      = 'grados'

    if periodo_a_mostrar:
        if grado_pk:
            # Nivel 2: materias del grado seleccionado
            grado_actual = get_object_or_404(Grado, pk=grado_pk, institucion=institucion)
            cursos = Curso.objects.filter(
                periodo_academico=periodo_a_mostrar,
                institucion=institucion,
                grado=grado_actual,
            ).select_related('materia', 'grado').prefetch_related(
                'docentes_asignados__usuario'
            ).order_by('materia__nombre_materia')
            nivel = 'materias'
        else:
            # Nivel 1: grados que tienen al menos un curso en este período
            grados = Grado.objects.filter(
                institucion=institucion,
                cursos__periodo_academico=periodo_a_mostrar,
                cursos__institucion=institucion,
            ).distinct().annotate(
                num_materias=Count(
                    'cursos',
                    filter=Q(
                        cursos__periodo_academico=periodo_a_mostrar,
                        cursos__institucion=institucion,
                    )
                )
            ).order_by('orden', 'nombre')
            nivel = 'grados'

    context = {
        'nivel': nivel,
        'grados': grados,
        'grado_actual': grado_actual,
        'cursos': cursos,
        'periodos_disponibles': periodos_disponibles,
        'periodo_seleccionado': periodo_a_mostrar,
        'titulo_pagina': "Supervisión — Libro de Notas",
        'es_coordinador': True,
    }
    return render(request, 'gestion_academica/coordinador_seleccionar_curso_libro_notas.html', context)


#@login_required
#@permission_required('gestion_academica.puede_realizar_registro_inicial')
def registro_inicial(request):
    if get_user_model().objects.exists():
        messages.warning(request, "El sistema ya ha sido configurado. Inicia sesión para continuar.")
        return redirect('login') 

    if request.method == 'POST':
        form = RegistroInicialForm(request.POST, request.FILES) 
        if form.is_valid():
            institucion = form.save(commit=False)
            institucion.save() 
            
            user = get_user_model().objects.create_user(
                username=form.cleaned_data['username'],
                email=form.cleaned_data['email'], 
                password=form.cleaned_data['password'],
            )
            user.is_staff = True 
            user.is_superuser = True 
            user.institucion_asociada = institucion 
            user.rol = 'administrativo' 
            user.save()

            messages.success(request, "Configuración inicial y superusuario creados exitosamente. Por favor, inicia sesión.")
            return redirect('login')
        else:
            messages.error(request, "Hubo un error al registrar la información inicial. Por favor, revisa los datos.")
    else:
        form = RegistroInicialForm()

    return render(request, 'gestion_academica/registro_inicial.html', {'form': form}) 

class NoticiaListView(ListView):
    model = Noticia
    template_name = 'gestion_academica/noticia_lista.html' 
    context_object_name = 'noticias'
    paginate_by = 5

    def get_queryset(self):
        base_queryset = Noticia.objects.all().order_by('-fecha_publicacion')
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Noticias y Anuncios"
        return context

class NoticiaDetailView(DetailView):
    model = Noticia
    template_name = 'gestion_academica/noticia_detalle.html' 
    context_object_name = 'noticia'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = self.object.titulo
        return context

class NoticiaCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Noticia
    form_class = NoticiaForm
    template_name = 'gestion_academica/noticia_formulario.html' 
    success_url = reverse_lazy('gestion_academica:lista_noticias_gestion') 
    permission_required = 'gestion_academica.add_noticia'

    def form_valid(self, form):
        form.instance.publicado_por = self.request.user
        if not self.request.user.is_superuser and hasattr(self.request.user, 'institucion_asociada') and self.request.user.institucion_asociada:
            form.instance.institucion = self.request.user.institucion_asociada
        elif self.request.user.is_superuser and not form.instance.institucion:
            messages.error(self.request, "Como superusuario, debes seleccionar una institución para la noticia.")
            return self.form_invalid(form)

        messages.success(self.request, f"Noticia '{form.cleaned_data['titulo']}' creada exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Crear Nueva Noticia/Anuncio"
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

class NoticiaUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Noticia
    form_class = NoticiaForm
    template_name = 'gestion_academica/noticia_formulario.html' 
    success_url = reverse_lazy('gestion_academica:lista_noticias_gestion')
    permission_required = 'gestion_academica.change_noticia'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Noticia/Anuncio"
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Noticia '{form.cleaned_data['titulo']}' actualizada exitosamente.") # CORRECCIÓN: self.request usado para messages
        return super().form_valid(form)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

class NoticiaDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Noticia
    template_name = 'gestion_academica/noticia_confirmar_eliminar.html' 
    success_url = reverse_lazy('gestion_academica:lista_noticias_gestion')
    context_object_name = 'noticia'
    permission_required = 'gestion_academica.delete_noticia'

    def get_queryset(self):
        base_queryset = super().get_queryset()
        return get_filtered_queryset(self.model, self.request.user, base_queryset)

    def delete(self, request, *args, **kwargs):
        noticia_eliminada = self.get_object()
        messages.success(request, f"La noticia '{noticia_eliminada.titulo}' ha sido eliminada.") 
        return super().delete(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Noticia"
        return context

class NoticiaGestionListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Noticia
    template_name = 'gestion_academica/noticia_gestion_lista.html' 
    context_object_name = 'noticias'
    permission_required = 'gestion_academica.view_noticia' 
    paginate_by = 10

    def get_queryset(self):
        base_queryset = Noticia.objects.all().order_by('-fecha_publicacion')
        filtered_by_institution = get_filtered_queryset(self.model, self.request.user, base_queryset)

        if hasattr(self.request.user, 'docente') and not self.request.user.is_superuser and not (hasattr(self.request.user, 'rol') and self.request.user.rol == 'administrativo'):
            return filtered_by_institution.filter(publicado_por=self.request.user)
        
        return filtered_by_institution

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Gestión de Noticias y Anuncios"
        return context
    
@login_required
# Podrías añadir un permiso específico para ver carnets si quieres
def generar_carnet_estudiante(request, estudiante_pk):
    """
    Genera una vista de carnet digital para un estudiante específico,
    incluyendo un código QR único.
    """
    try:
        estudiante = Estudiante.objects.select_related('usuario', 'grado_actual').get(pk=estudiante_pk)
    except Estudiante.DoesNotExist:
        return HttpResponseNotFound("<h1>Estudiante no encontrado</h1>")

    # --- Generación del Código QR ---
    # El dato que vamos a codificar en el QR es el UUID único del estudiante.
    qr_data = str(estudiante.qr_identifier)
    
    # Creamos el objeto QR
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)

    # Creamos la imagen del QR en memoria, sin guardarla en el disco
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Convertimos la imagen a bytes para poder mostrarla en el HTML
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    img_str = base64.b64encode(buffer.getvalue()).decode()
    
    qr_code_url = f"data:image/png;base64,{img_str}"
    
    # --- Preparación del Contexto para la Plantilla ---
    context = {
        'estudiante': estudiante,
        'qr_code_url': qr_code_url,
        'titulo_pagina': f"Carnet de {estudiante}",
    }
    
    return render(request, 'gestion_academica/carnet_estudiante.html', context)
# =========================================================================
# FIN: VISTA NUEVA
# =========================================================================
 
@require_POST
@login_required
def registrar_asistencia_api(request):
    """
    Endpoint de API para registrar la asistencia.
    VERSIÓN DEFINITIVA: Usa una búsqueda manual por rango para evitar errores de SQLite.
    """
    try:
        data = json.loads(request.body)
        qr_identifier = data.get('qr_identifier')
        curso_id = data.get('curso_id')
        if not qr_identifier or not curso_id:
            return JsonResponse({'status': 'error', 'message': 'Faltan datos.'}, status=400)
        
        estudiante = Estudiante.objects.get(qr_identifier=qr_identifier)
        curso = Curso.objects.get(pk=curso_id)

        if estudiante.grado_actual != curso.grado:
            return JsonResponse({'status': 'error', 'message': 'Estudiante no pertenece a este curso.'}, status=403)

        # --- LÓGICA DE GUARDADO DEFINITIVA ---
        hoy = timezone.localdate()
        hoy_inicio = make_aware(datetime.combine(hoy, datetime.min.time()))
        hoy_fin = make_aware(datetime.combine(hoy, datetime.max.time()))

        registro_existente = RegistroAsistencia.objects.filter(
            estudiante=estudiante,
            curso=curso,
            fecha__range=(hoy_inicio, hoy_fin)
        ).first()

        # 2. Decidimos si actualizar o crear
        if registro_existente:
            registro_existente.estado = 'PRESENTE'
            registro_existente.registrado_por = request.user
            registro_existente.fecha = timezone.now() # Actualizamos la hora
            registro_existente.save()
            message = f"Estado de {estudiante.usuario.get_full_name()} actualizado a PRESENTE."
        else:
            RegistroAsistencia.objects.create(
                estudiante=estudiante,
                curso=curso,
                estado='PRESENTE',
                registrado_por=request.user,
                fecha=timezone.now(),
                institucion=estudiante.institucion  # ✅ Corrección
            )
            message = f"Asistencia de {estudiante.usuario.get_full_name()} registrada."

        return JsonResponse({'status': 'success', 'message': message})

    except Exception as e:
        logger.error(f"Error en API de asistencia: {e}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': 'Error interno del servidor.'}, status=500)

# =========================================================================
# FIN: VISTA API
# =========================================================================    

@login_required
@permission_required('gestion_academica.add_registroasistencia')
def escaner_asistencia(request, curso_pk): # <--- CAMBIO: de curso_id a curso_pk
    """
    Muestra la página con el escáner de QR para un curso específico.
    """
    curso = get_object_or_404(Curso.objects.select_related('materia', 'grado'), pk=curso_pk) # <--- CAMBIO: usa curso_pk
    
    # Lógica de seguridad
    if not (request.user.is_superuser or hasattr(request.user, 'docente')):
        messages.error(request, "No tienes permiso para pasar lista en este curso.")
        return redirect('gestion_academica:dashboard_docente')

    context = {
        'curso': curso,
        'titulo_pagina': f"Pasar Lista: {curso}"
    }
    return render(request, 'gestion_academica/escaner_asistencia.html', context)

@login_required
@permission_required('gestion_academica.add_registroasistencia')
def seleccionar_curso_asistencia(request):
    """
    Muestra TODOS los cursos activos de la institución para que un docente
    pueda seleccionar uno y gestionar su asistencia (ideal para cubrir ausencias).
    """
    if not hasattr(request.user, 'docente'):
        messages.error(request, "Acceso denegado. Solo para docentes.")
        return redirect('gestion_academica:inicio_academico')

    # Obtenemos la institución del docente logueado
    institucion_docente = request.user.institucion_asociada
    periodo_activo = PeriodoAcademico.objects.filter(activo=True, institucion=institucion_docente).first()
    
    cursos_institucion = []
    if periodo_activo and institucion_docente:
        # Buscamos TODOS los cursos de la institución en el periodo activo
        cursos_institucion = Curso.objects.filter(
            institucion=institucion_docente,
            periodo_academico=periodo_activo
        ).select_related('materia', 'grado').prefetch_related('docentes_asignados__usuario').order_by('grado__nombre', 'materia__nombre_materia')

    context = {
        'cursos': cursos_institucion,
        'periodo_activo': periodo_activo,
        'titulo_pagina': "Seleccionar un Curso para Asistir"
    }
    # La plantilla ahora debe mostrar la lista de todos los cursos.
    # El enlace de cada curso debe apuntar a la vista 'gestionar_asistencia'.
    return render(request, 'gestion_academica/seleccionar_curso_asistencia.html', context)

@login_required
@requiere_pagos_al_dia
def mi_historial_asistencia(request):
    """
    Muestra al estudiante su historial de asistencias, tardanzas y ausencias
    para el periodo académico activo.
    """
    try:
        estudiante = request.user.estudiante
    except Estudiante.DoesNotExist:
        # Esto previene errores si un usuario no estudiante intenta acceder.
        messages.error(request, "Tu perfil de estudiante no está configurado.")
        return redirect('gestion_academica: ')

    periodo_activo = PeriodoAcademico.objects.filter(
        activo=True, institucion=estudiante.institucion
    ).first()
    
    historial_asistencia = []
    if periodo_activo:
        # Buscamos todos los registros del estudiante en el periodo activo y los ordenamos por fecha.
        historial_asistencia = RegistroAsistencia.objects.filter(
            estudiante=estudiante,
            curso__periodo_academico=periodo_activo
        ).select_related('curso__materia').order_by('-fecha')

    context = {
        'titulo_pagina': "Mi Historial de Asistencia",
        'periodo_activo': periodo_activo,
        'historial': historial_asistencia,
    }
    
    # Esta vista usará la plantilla 'mi_historial_asistencia.html' que ya creamos.
    return render(request, 'gestion_academica/mi_historial_asistencia.html', context)

@login_required
@permission_required('gestion_academica.view_registroasistencia')
def exportar_asistencia_excel(request, curso_pk):  # 👈 Asegúrate de que URL también lo llame curso_pk
    hoy = localdate()

    # Seguridad multi-institución
    curso = get_object_or_404(Curso, pk=curso_pk)
    if not request.user.is_superuser and curso.institucion != request.user.institucion_asociada:
        return HttpResponse("No autorizado para acceder a este curso.", status=403)

    # Obtiene registros filtrando por curso, fecha e institución
    registros = RegistroAsistencia.objects.filter(
        curso=curso,
        fecha__date=hoy,
        institucion=curso.institucion  # 👈 Garantiza coherencia
    ).select_related('estudiante__usuario')

    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Asistencia_{curso}"

    # Encabezados
    ws.append(['Nombre del Estudiante', 'Documento', 'Estado', 'Fecha'])

    for r in registros:
        ws.append([
            r.estudiante.usuario.get_full_name(),
            r.estudiante.documento_identidad,
            r.estado,
            r.fecha.strftime("%Y-%m-%d %H:%M:%S")
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Asistencia_{curso}_{hoy}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

@login_required
def ver_mi_perfil(request):
    """
    Muestra una página de perfil enriquecida para el usuario logueado,
    adaptando el contenido y los datos según su rol.
    Acepta POST para actualizar foto de perfil.
    """
    user = request.user

    # ── Manejo de subida/eliminación de foto ────────────────────────
    if request.method == 'POST' and request.POST.get('eliminar_foto') == 'si':
        if user.foto_perfil:
            try:
                user.foto_perfil.delete(save=False)
            except Exception:
                pass
            user.foto_perfil = None
            user.save(update_fields=['foto_perfil'])
            messages.success(request, "Foto de perfil eliminada.")
        return redirect('gestion_academica:ver_mi_perfil')

    if request.method == 'POST' and request.FILES.get('foto_perfil'):
        foto = request.FILES['foto_perfil']
        # Validar tipo de archivo
        tipos_validos = ['image/jpeg', 'image/png', 'image/webp', 'image/gif']
        if foto.content_type in tipos_validos and foto.size <= 5 * 1024 * 1024:  # max 5 MB
            # Borrar foto anterior si existe
            if user.foto_perfil:
                try:
                    user.foto_perfil.delete(save=False)
                except Exception:
                    pass
            user.foto_perfil = foto
            user.save(update_fields=['foto_perfil'])
            messages.success(request, "Foto de perfil actualizada correctamente.")
        else:
            messages.error(request, "Archivo no válido. Usa JPG, PNG o WebP (máx. 5 MB).")
        return redirect('gestion_academica:ver_mi_perfil')

    context = {
        'user': user,
        'titulo_pagina': "Mi Perfil",
        'estudiante_profile': None,
        'docente_profile': None,
        'familiar_profile': None,
    }

    # Intentamos obtener cada tipo de perfil
    estudiante_profile = getattr(user, 'estudiante', None)
    docente_profile = getattr(user, 'docente', None)
    familiar_profile = getattr(user, 'familiar', None)

    if estudiante_profile:
        context['estudiante_profile'] = estudiante_profile
        # Calculamos datos adicionales para el estudiante
        cuentas_pendientes = CuentaPorCobrarEstudiante.objects.filter(
            estudiante=estudiante_profile
        ).exclude(estado='PAGADO')
        
        saldo_total = cuentas_pendientes.aggregate(
            total=Sum('monto_asignado') - Sum('pagos__valor_pagado')
        )['total'] or 0
        
        context['saldo_pendiente'] = saldo_total
        context['esta_en_mora'] = cuentas_pendientes.filter(estado='VENCIDO').exists()

    elif docente_profile:
        context['docente_profile'] = docente_profile
        # Buscamos los cursos que el docente imparte en el periodo activo
        periodo_activo = PeriodoAcademico.objects.filter(institucion=docente_profile.institucion, activo=True).first()
        if periodo_activo:
            context['cursos_del_docente'] = Curso.objects.filter(
                docentes_asignados=docente_profile,
                periodo_academico=periodo_activo
            ).select_related('materia', 'grado')

    elif familiar_profile:
        context['familiar_profile'] = familiar_profile
    
    return render(request, 'gestion_academica/mi_perfil.html', context)

    
@login_required
@requiere_pagos_al_dia
def detalle_curso_aula_virtual(request, curso_pk):
    curso = get_object_or_404(
        Curso.objects.select_related(
            'materia', 'grado', 'periodo_academico', 'aula', 'institucion'
        ).prefetch_related(
            'deberes',
            'actividades_calificables',
            'enlaces_videollamada',
            'archivos_material_apoyo_curso'
        ),
        pk=curso_pk
    )

    usuario = request.user
    institucion_curso = curso.institucion
    acceso_permitido = False

    # ✅ 1. Estudiante
    if hasattr(usuario, 'estudiante'):
        estudiante = usuario.estudiante
        if estudiante.grado_actual == curso.grado and estudiante.institucion == institucion_curso:
            acceso_permitido = True

    # ✅ 2. Docente
    elif hasattr(usuario, 'docente'):
        docente = usuario.docente
        if docente in curso.docentes_asignados.all() and docente.institucion == institucion_curso:
            acceso_permitido = True

    # ✅ 3. Familiar
    elif hasattr(usuario, 'familiar'):
        familiar = usuario.familiar
        if familiar.institucion == institucion_curso:
            for estudiante in familiar.estudiantes_asociados.all():
                if estudiante.grado_actual == curso.grado:
                    acceso_permitido = True
                    break

    # ✅ 4. Staff (rector, coordinador, etc.)
    elif usuario.is_staff and getattr(usuario, 'institucion_asociada', None):
        if usuario.institucion_asociada == institucion_curso:
            acceso_permitido = True

    if not acceso_permitido:
        messages.error(request, "No tienes permiso para acceder a esta aula virtual.")
        return redirect('gestion_academica:inicio_academico')

    context = {
        'curso': curso,
        'deberes': curso.deberes.all(),
        'actividades': curso.actividades_calificables.all(),
        'enlaces_video': curso.enlaces_videollamada.all(),
        'recursos': curso.archivos_material_apoyo_curso.all(),
        'aula': curso.aula,
        'titulo_pagina': f"Aula Virtual: {curso.materia.nombre_materia}"
    }

    return render(request, 'gestion_academica/aula_virtual.html', context)

# gestion_academica/views.py (COMPLETO y FINAL)

@login_required
def dashboard_estudiante(request):
    try:
        estudiante = Estudiante.objects.select_related('usuario', 'grado_actual', 'institucion').get(usuario=request.user)
    except Estudiante.DoesNotExist:
        messages.error(request, "Tu perfil de estudiante no está configurado. Contacta a la administración.")
        return redirect('gestion_academica:inicio_academico')

    # --- LÓGICA FINANCIERA ---
    # Se mantiene la consulta histórica para los KPIs (vencidas, próximas a vencer),
    # pero ahora delegamos la decisión de "bloqueo real" a Estudiante.esta_al_dia(),
    # que respeta el toggle por institución (bloquear_portal_por_mora) y los días
    # de gracia (dias_gracia_mora). Ver Fase C — gestion_academica/decorators.py.
    hoy = timezone.now().date()
    cuentas_no_pagadas_qs = estudiante.cuentas_por_cobrar.exclude(estado='PAGADO')
    cuentas_pendientes_count = cuentas_no_pagadas_qs.count()
    cuentas_vencidas_qs = cuentas_no_pagadas_qs.filter(fecha_vencimiento_especifica__lt=hoy)
    cuentas_vencidas_count = cuentas_vencidas_qs.count()
    cuentas_vencidas_lista = list(
        cuentas_vencidas_qs
        .select_related('concepto_pago')
        .order_by('fecha_vencimiento_especifica')[:10]
    )
    limite_proximo_pago = hoy + timedelta(days=30)
    cuentas_proximas_a_vencer_count = cuentas_no_pagadas_qs.filter(
        fecha_vencimiento_especifica__gte=hoy,
        fecha_vencimiento_especifica__lte=limite_proximo_pago
    ).count()
    portal_bloqueado_por_mora = not estudiante.esta_al_dia()
    estudiante_moroso = portal_bloqueado_por_mora
    dias_atraso_max = estudiante.dias_de_atraso_max if portal_bloqueado_por_mora else 0
    saldo_total_vencido = sum(
        (c.saldo_pendiente for c in cuentas_vencidas_lista),
        Decimal('0.00'),
    )

    # --- LÓGICA DE NOTIFICACIONES Y PERIODO ---
    notificaciones_sin_leer = Notificacion.objects.filter(
        destinatario=request.user, leido=False, institucion=estudiante.institucion
    )
    periodo_activo = PeriodoAcademico.objects.filter(activo=True, institucion=estudiante.institucion).first()

    # --- LÓGICA PARA ELECCIONES (se mantiene intacta) ---
    now = timezone.now()
    elecciones_activas = Eleccion.objects.filter(
        institucion=estudiante.institucion,
        fecha_inicio__lte=now,
        fecha_fin__gte=now
    ).annotate(num_candidatos=Count('candidatos')).order_by('-fecha_fin')
    votos_emitidos = Voto.objects.filter(
        votante=estudiante,
        eleccion__in=elecciones_activas
    ).values_list('eleccion_id', flat=True)

    # --- NUEVA LÓGICA PARA PREPARAR EL HORARIO ---
    horario_agrupado = defaultdict(list)
    dia_semana_hoy = hoy.weekday()

    # --- CONTEXTO PRINCIPAL ---
    context = {
        'estudiante': estudiante,
        'periodo_activo': periodo_activo,
        'titulo_pagina': "Mi Panel de Inicio",
        'estudiante_moroso': estudiante_moroso,
        'portal_bloqueado_por_mora': portal_bloqueado_por_mora,
        'cuentas_vencidas_count': cuentas_vencidas_count,
        'cuentas_vencidas_lista': cuentas_vencidas_lista,
        'cuentas_pendientes_count': cuentas_pendientes_count,
        'cuentas_proximas_a_vencer_count': cuentas_proximas_a_vencer_count,
        'dias_atraso_max': dias_atraso_max,
        'saldo_total_vencido': saldo_total_vencido,
        'notificaciones_sin_leer': notificaciones_sin_leer,
        'elecciones_activas': elecciones_activas,
        'votos_emitidos_ids': list(votos_emitidos),

        # Variables necesarias para el nuevo horario interactivo
        'horario_agrupado': {},
        'dias_semana': BloqueHorario.DIA_SEMANA_CHOICES,
        'dia_semana_hoy': dia_semana_hoy,
        
        # Inicialización de otras variables del contexto
        'cursos_info': [], 
        'materias_con_calificaciones': {},
        'ultimas_noticias': [], 
        'proximos_eventos_agenda': [],
        'inasistencias_periodo': 0, 
        'ultimas_lecciones': [],
    }

    if periodo_activo and estudiante.grado_actual:
        cursos_del_estudiante = Curso.objects.filter(
            grado=estudiante.grado_actual,
            periodo_academico=periodo_activo,
            institucion=estudiante.institucion,
        ).select_related('materia').prefetch_related(
            'actividades_calificables__cuestionario'
        )
        
        context.update({
            'ultimas_menciones': MencionReconocimiento.objects.filter(estudiante=estudiante, periodo=periodo_activo).order_by('-fecha_otorgamiento')[:3],
            'ultimas_noticias': Noticia.objects.filter(institucion=estudiante.institucion).order_by('-fecha_publicacion')[:3],
            'cursos_info': list(cursos_del_estudiante),
            'inasistencias_periodo': RegistroAsistencia.objects.filter(estudiante=estudiante, curso__periodo_academico=periodo_activo, estado='AUSENTE').count(),
            'ultimas_lecciones': LeccionDiaria.objects.filter(curso__in=cursos_del_estudiante).select_related('curso__materia').order_by('-fecha')[:3]
        })

        calificaciones_recientes = Calificacion.objects.filter(estudiante=estudiante, actividad_calificable__curso__in=cursos_del_estudiante).select_related('actividad_calificable__curso__materia').order_by('-fecha_registro')[:20]
        materias_agrupadas = defaultdict(list)
        for cal in calificaciones_recientes:
            materias_agrupadas[cal.actividad_calificable.curso.materia].append(cal)
        context['materias_con_calificaciones'] = dict(materias_agrupadas)

        # --- LÓGICA DEL HORARIO (LLENADO DE DATOS) ---
        bloques_horario = BloqueHorario.objects.filter(curso__in=cursos_del_estudiante).select_related('curso__materia', 'aula').order_by('hora_inicio')
        for bloque in bloques_horario:
            horario_agrupado[bloque.dia_semana].append(bloque)
        context['horario_agrupado'] = dict(horario_agrupado)
        
        # --- LÓGICA DE LA AGENDA (se mantiene intacta) ---
        limite_agenda = hoy + timedelta(days=14)
        eventos_agenda = []
        for deber in Deber.objects.filter(curso__in=cursos_del_estudiante, fecha_entrega__range=(hoy, limite_agenda)):
            eventos_agenda.append({'fecha': deber.fecha_entrega, 'titulo': f"Entrega: {deber.titulo}", 'tipo': 'tarea', 'url': reverse('gestion_academica:realizar_entrega_deber', args=[deber.pk]), 'target_blank': False})
        acts_agenda = ActividadCalificable.objects.filter(
            curso__in=cursos_del_estudiante,
            fecha_entrega_limite__range=(hoy, limite_agenda),
        ).prefetch_related('preguntas')
        agenda_act_pks = list(acts_agenda.values_list('pk', flat=True))
        agenda_con_cuestionario = set(
            Cuestionario.objects.filter(actividad_calificable_id__in=agenda_act_pks).values_list(
                'actividad_calificable_id', flat=True
            )
        )
        for act in acts_agenda:
            if act.pk in agenda_con_cuestionario:
                url = reverse('cuestionarios:iniciar_cuestionario', args=[act.pk])
                target_blank = False
            elif act.preguntas.exists():
                url = reverse('gestion_academica:resolver_actividad', args=[act.pk])
                target_blank = False
            elif act.material_adjunto:
                url = act.material_adjunto.url
                target_blank = True
            else:
                url = '#'
                target_blank = False
            eventos_agenda.append({
                'fecha': act.fecha_entrega_limite,
                'titulo': f"Actividad: {act.titulo}",
                'tipo': 'evaluacion',
                'url': url,
                'target_blank': target_blank,
            })
        context['proximos_eventos_agenda'] = sorted(eventos_agenda, key=lambda e: e['fecha'])[:5]

        # Citas agendadas por el familiar del estudiante (lectura pura)
        ahora_est = timezone.now()
        context['citas_estudiante'] = list(
            CitaReunion.objects.filter(
                estudiante=estudiante,
                institucion=estudiante.institucion,
                estado__in=['PENDIENTE', 'CONFIRMADA'],
                fecha_hora_inicio__gte=ahora_est,
            ).select_related('docente__usuario', 'familiar__usuario')
            .order_by('fecha_hora_inicio')[:5]
        )

    return render(request, 'gestion_academica/dashboard_estudiante.html', context)


class CalendarioEventosAPIView(APIView):
    """
    API que devuelve todos los eventos relevantes para el estudiante logueado
    en el formato que espera FullCalendar. VERSIÓN CORREGIDA.
    """
    permission_classes = [IsAuthenticated, EstaAlDiaPermission]

    def get(self, request, *args, **kwargs):
        try:
            estudiante = request.user.estudiante
        except AttributeError: # 'estudiante' no existe en el usuario, ej: admin
             # Usamos AttributeError en lugar de Estudiante.DoesNotExist para mayor precisión
            return Response([], status=403) # 403 Forbidden

        periodo_activo = PeriodoAcademico.objects.filter(
            institucion=estudiante.institucion, activo=True
        ).first()

        if not periodo_activo or not estudiante.grado_actual:
            return Response([])

        eventos = []

        # 1. Bloques de horario (eventos recurrentes)
        bloques_horario = BloqueHorario.objects.filter(
            curso__grado=estudiante.grado_actual, curso__periodo_academico=periodo_activo
        ).select_related('curso__materia', 'aula')
        
        for bloque in bloques_horario:
            eventos.append({
                'title': bloque.curso.materia.nombre_materia,
                'daysOfWeek': [(bloque.dia_semana + 1) % 7], # Ajuste para FullCalendar (Domingo=0)
                'startTime': bloque.hora_inicio.strftime('%H:%M:%S'),
                'endTime': bloque.hora_fin.strftime('%H:%M:%S'),
                'color': '#0d6efd',
                'description': f"Aula: {bloque.aula.nombre}" if bloque.aula else ""
            })

        # 2. Deberes (eventos de día completo)
        deberes = Deber.objects.filter(curso__grado=estudiante.grado_actual, curso__periodo_academico=periodo_activo)
        for deber in deberes:
            eventos.append({
                'title': f"Entrega: {deber.titulo}",
                'start': deber.fecha_entrega.isoformat(), # formato YYYY-MM-DD
                'allDay': True,
                'color': '#dc3545',
                'url': reverse('gestion_academica:realizar_entrega_deber', kwargs={'deber_pk': deber.pk})
            })
            
        # 3. Actividades Calificables (eventos de día completo)
        actividades = ActividadCalificable.objects.filter(
            curso__grado=estudiante.grado_actual,
            curso__periodo_academico=periodo_activo,
            fecha_entrega_limite__isnull=False,
        ).prefetch_related('preguntas')
        cal_act_pks = list(actividades.values_list('pk', flat=True))
        cal_ids_cuestionario = set(
            Cuestionario.objects.filter(actividad_calificable_id__in=cal_act_pks).values_list(
                'actividad_calificable_id', flat=True
            )
        )
        for actividad in actividades:
            if actividad.pk in cal_ids_cuestionario:
                act_url = reverse('cuestionarios:iniciar_cuestionario', args=[actividad.pk])
            elif actividad.preguntas.exists():
                act_url = reverse('gestion_academica:resolver_actividad', args=[actividad.pk])
            elif actividad.material_adjunto:
                act_url = actividad.material_adjunto.url
            else:
                act_url = None
            ev = {
                'title': f"Evaluación: {actividad.titulo}",
                'start': actividad.fecha_entrega_limite.isoformat(),  # formato YYYY-MM-DD
                'allDay': True,
                'color': '#ffc107',
            }
            if act_url:
                ev['url'] = act_url
            eventos.append(ev)

        # Serializamos la lista de diccionarios que hemos construido
        serializer = EventoCalendarioSerializer(instance=eventos, many=True)
        return Response(serializer.data)

    
@login_required
@permission_required('gestion_academica.add_lecciondiaria') # Asegúrate de que los docentes tengan este permiso
def seleccionar_curso_para_leccion(request):
    """
    Muestra al docente una lista de sus cursos para que elija en cuál
    registrar una lección diaria.
    """
    if not hasattr(request.user, 'docente'):
        messages.error(request, "Acceso denegado. Solo para docentes.")
        return redirect('gestion_academica:inicio_academico')

    docente = request.user.docente
    periodo_activo = PeriodoAcademico.objects.filter(activo=True, institucion=docente.institucion).first()
    
    cursos_del_docente = []
    if periodo_activo:
        cursos_del_docente = Curso.objects.filter(
            docentes_asignados=docente,
            periodo_academico=periodo_activo
        ).select_related('materia', 'grado')

    context = {
        'cursos': cursos_del_docente,
        'periodo_activo': periodo_activo,
        'titulo_pagina': "Seleccionar Curso para Registrar Lección"
    }
    return render(request, 'gestion_academica/seleccionar_curso_leccion.html', context)


@login_required
@permission_required('gestion_academica.add_lecciondiaria')
def registrar_leccion_diaria(request, curso_pk):
    curso = get_object_or_404(Curso, pk=curso_pk)
    
    if not hasattr(request.user, 'docente') or not curso.docentes_asignados.filter(pk=request.user.docente.pk).exists():
        messages.error(request, "No tienes permiso para registrar lecciones en este curso.")
        return redirect('gestion_academica:seleccionar_curso_para_leccion')

    if request.method == 'POST':
        # Pasamos el usuario para que el queryset del formulario se filtre correctamente
        form = LeccionDiariaForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            leccion = form.save(commit=False)
            leccion.creado_por = request.user
            # Aseguramos que la institución se guarde si no viene en el form
            if not leccion.institucion:
                leccion.institucion = curso.institucion
            leccion.save()
            messages.success(request, f"Lección para '{curso.materia}' del {leccion.fecha} registrada exitosamente.")
            return redirect('gestion_academica:seleccionar_curso_para_leccion')
    else:
        # ================================================================
        #   INICIO: CORRECCIÓN CLAVE
        # ================================================================
        # Usamos timezone.localdate() para obtener la fecha correcta según settings.py
        form = LeccionDiariaForm(user=request.user, initial={'curso': curso, 'fecha': timezone.localdate()})
        # ================================================================
        #   FIN: CORRECCIÓN
        # ================================================================
    
    context = {
        'form': form,
        'curso': curso,
        'titulo_pagina': f"Registrar Lección para {curso.materia}"
    }
    return render(request, 'gestion_academica/registrar_leccion_diaria_form.html', context)  

class DetalleActividadAPIView(RetrieveAPIView):
    """
    API View para obtener los detalles de UNA actividad calificable interactiva.
    Incluye todas sus preguntas y opciones. VERSIÓN OPTIMIZADA.
    """
    serializer_class = ActividadInteractivaSerializer
    permission_classes = [IsAuthenticated, EstaAlDiaPermission]
    lookup_field = 'pk'

    def get_queryset(self):
        return (
            actividades_calificables_accesibles_para_usuario(self.request.user).prefetch_related(
                Prefetch(
                    'preguntas',
                    queryset=Pregunta.objects.order_by('orden').prefetch_related(
                        Prefetch('opciones', queryset=Opcion.objects.order_by('?'))
                    ),
                )
            )
        )
# =========

@login_required
@requiere_pagos_al_dia
def resolver_actividad_page(request, actividad_pk):
    if not hasattr(request.user, 'estudiante'):
        messages.error(request, "Solo los estudiantes pueden resolver esta actividad.")
        return redirect('gestion_academica:inicio_academico')

    estudiante = request.user.estudiante
    actividad = get_object_or_404(
        ActividadCalificable.objects.select_related(
            'curso__grado', 'curso__periodo_academico', 'institucion'
        ),
        pk=actividad_pk,
    )

    if not estudiante_en_curso_actividad(estudiante, actividad):
        messages.error(request, "No tienes permiso para esta actividad.")
        return redirect('gestion_academica:dashboard_estudiante')

    if Cuestionario.objects.filter(actividad_calificable_id=actividad.pk).exists():
        return redirect('cuestionarios:iniciar_cuestionario', actividad_pk)

    if not Pregunta.objects.filter(actividad=actividad).exists():
        messages.info(
            request,
            "Esta actividad no tiene preguntas interactivas configuradas aquí.",
        )
        return redirect('gestion_academica:dashboard_estudiante')

    institucion = estudiante.institucion

    intento = IntentoActividad.objects.filter(
        estudiante=estudiante,
        actividad=actividad,
        estado='en_progreso',
        institucion=institucion,
    ).first()

    if intento and actividad.duracion_minutos:
        tiempo_transcurrido = timezone.now() - intento.inicio
        if tiempo_transcurrido.total_seconds() >= actividad.duracion_minutos * 60:
            intento.estado = 'tiempo_agotado'
            intento.fin = timezone.now()
            intento.save()
            intento = None

    if not intento:
        total_intentos = IntentoActividad.objects.filter(
            estudiante=estudiante,
            actividad=actividad,
            institucion=institucion,
        ).count()
        if total_intentos >= actividad.numero_intentos_permitidos:
            messages.error(
                request,
                f"Has alcanzado el número máximo de {actividad.numero_intentos_permitidos} intentos para esta actividad.",
            )
            return redirect('gestion_academica:dashboard_estudiante')
        intento = IntentoActividad.objects.create(
            estudiante=estudiante,
            actividad=actividad,
            institucion=institucion,
        )

    tiempo_restante_segundos = None
    if actividad.duracion_minutos:
        tiempo_transcurrido = timezone.now() - intento.inicio
        tiempo_total_segundos = actividad.duracion_minutos * 60
        tiempo_restante_segundos = tiempo_total_segundos - tiempo_transcurrido.total_seconds()

        if tiempo_restante_segundos <= 0:
            intento.estado = 'tiempo_agotado'
            intento.fin = timezone.now()
            intento.save()
            messages.error(
                request,
                f"Se ha agotado el tiempo para la actividad '{actividad.titulo}'.",
            )
            return redirect('gestion_academica:dashboard_estudiante')

    if intento.estado == 'completado':
        messages.warning(
            request,
            f"Ya has completado la actividad '{actividad.titulo}'.",
        )
        return redirect('gestion_academica:dashboard_estudiante')

    return render(
        request,
        'gestion_academica/resolver_actividad.html',
        {
            'actividad': actividad,
            'tiempo_restante': tiempo_restante_segundos,
        },
    )


class EnviarRespuestasAPIView(APIView):
    """
    API para recibir, guardar y CALIFICAR AUTOMÁTICAMENTE las respuestas de un estudiante.
    VERSIÓN FINAL Y CORREGIDA.
    """
    permission_classes = [IsAuthenticated, EstaAlDiaPermission]

    def post(self, request, actividad_pk, *args, **kwargs):
        # 1. Verificar que el usuario es un estudiante
        try:
            estudiante = request.user.estudiante
        except AttributeError:
            return Response({'error': 'El usuario no tiene un perfil de estudiante asociado.'}, status=403)

        actividad = get_object_or_404(
            ActividadCalificable.objects.select_related('curso'),
            pk=actividad_pk,
        )
        if not estudiante_en_curso_actividad(estudiante, actividad):
            return Response({'error': 'No autorizado para esta actividad.'}, status=403)
        if Cuestionario.objects.filter(actividad_calificable_id=actividad_pk).exists():
            return Response(
                {'error': 'Esta evaluación se realiza en el módulo de cuestionarios.'},
                status=400,
            )

        # 2. Verificar que los datos de respuesta tengan el formato correcto
        respuestas_data = request.data.get('respuestas', {})
        if not isinstance(respuestas_data, dict):
            return Response({'error': 'El formato de respuestas enviado no es válido.'}, status=400)

        # 3. Buscar el intento activo.
        intento = (
            IntentoActividad.objects.filter(
                estudiante=estudiante,
                actividad_id=actividad_pk,
                estado='en_progreso',
            )
            .order_by('-inicio')
            .first()
        )
        if intento is None:
            return Response({'error': 'No se encontró un intento activo para esta actividad o ya fue completado.'}, status=404)

        # 4. Usar una transacción para asegurar la integridad de los datos
        with transaction.atomic():
            # Optimizamos la consulta para traer las preguntas y sus opciones correctas de una vez
            preguntas_de_la_actividad = Pregunta.objects.filter(
                actividad_id=actividad_pk
            ).prefetch_related(
                Prefetch('opciones', queryset=Opcion.objects.filter(es_correcta=True), to_attr='opcion_correcta_prefetched')
            )
            
            # Limpiamos las respuestas anteriores para este intento
            RespuestaEstudiante.objects.filter(
                estudiante=estudiante, pregunta__in=preguntas_de_la_actividad
            ).delete()

            # ================================================================
            #   INICIO: BUCLE DE GUARDADO DE RESPUESTAS (CORREGIDO)
            # ================================================================
            for pregunta_id_str, respuesta_valor in respuestas_data.items():
                pregunta_id = int(pregunta_id_str)
                
                try:
                    # Obtenemos el objeto de la pregunta de la lista que ya consultamos
                    pregunta_obj = next(p for p in preguntas_de_la_actividad if p.pk == pregunta_id)
                except StopIteration:
                    # Si el frontend envía una pregunta que no es de esta actividad, es un error.
                    transaction.set_rollback(True)
                    return Response({'error': f'La pregunta con ID {pregunta_id} no pertenece a esta actividad.'}, status=400)
                
                # Preparamos los datos para validar con el serializer
                datos_a_validar = {'pregunta_id': pregunta_id}
                if pregunta_obj.tipo == 'respuesta_abierta':
                    datos_a_validar['texto_respuesta'] = respuesta_valor
                else: # Opción múltiple, Verdadero/Falso, etc.
                    datos_a_validar['opcion_id'] = respuesta_valor

                serializer = EnviarRespuestaSerializer(data=datos_a_validar)
                
                if serializer.is_valid():
                    # Si los datos son válidos, creamos el objeto RespuestaEstudiante
                    data = serializer.validated_data
                    RespuestaEstudiante.objects.create(
                        estudiante=estudiante,
                        pregunta_id=data['pregunta_id'],
                        opcion_seleccionada_id=data.get('opcion_id'),
                        texto_respuesta=data.get('texto_respuesta', '')
                    )
                else:
                    # Si cualquier respuesta es inválida, revertimos toda la transacción
                    transaction.set_rollback(True)
                    return Response(serializer.errors, status=400)
            # ================================================================
            #   FIN: BUCLE DE GUARDADO DE RESPUESTAS (CORREGIDO)
            # ================================================================

            # --- Lógica de Calificación Automática (sin cambios) ---
            puntaje_total = 0
            total_preguntas_calificables = preguntas_de_la_actividad.exclude(tipo='respuesta_abierta').count()
            
            for pregunta in preguntas_de_la_actividad:
                if pregunta.tipo == 'opcion_multiple':
                    respuesta_dada_id = respuestas_data.get(str(pregunta.pk))
                    
                    opcion_correcta_lista = pregunta.opcion_correcta_prefetched
                    if opcion_correcta_lista:
                        opcion_correcta = opcion_correcta_lista[0]
                        if respuesta_dada_id == opcion_correcta.pk:
                            puntaje_total += 1
            
            if total_preguntas_calificables > 0:
                nota_final = (Decimal(puntaje_total) / Decimal(total_preguntas_calificables)) * Decimal('5.0')
            else:
                nota_final = None

            # Actualizamos el intento
            intento.estado = 'completado'
            intento.fin = timezone.now()
            intento.puntaje_obtenido = nota_final
            intento.save()

        # Devolvemos un mensaje de éxito informativo
        mensaje_exito = "Respuestas guardadas y actividad finalizada."
        if nota_final is not None:
            mensaje_exito += f" Tu calificación preliminar es: {nota_final:.2f}"

        return Response({'status': 'success', 'message': mensaje_exito}, status=201)


@login_required
def dashboard_docente(request):
    """
    Muestra el panel principal del docente, con la lógica de conteo de citas corregida.
    """
    try:
        docente = request.user.docente
    except (AttributeError, Docente.DoesNotExist):
        messages.error(request, "Acceso denegado. Esta sección es solo para docentes.")
        return redirect('gestion_academica:inicio_academico')

    institucion = request.user.institucion_asociada
    periodo_activo = PeriodoAcademico.objects.filter(activo=True, institucion=institucion).first()
    
    hoy_inicio = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    fin_rango = hoy_inicio + timedelta(days=7)

    citas_proxima_semana_count = CitaReunion.objects.filter(
        docente=docente,
        institucion=institucion,
        estado__in=['PENDIENTE', 'CONFIRMADA'],
        fecha_hora_inicio__range=(hoy_inicio, fin_rango)
    ).count()

    ahora_tz = timezone.now()
    ventana_citas_fin = ahora_tz + timedelta(days=28)
    citas_docente_horario = CitaReunion.objects.filter(
        docente=docente,
        institucion=institucion,
        estado__in=['PENDIENTE', 'CONFIRMADA'],
        fecha_hora_inicio__gte=ahora_tz - timedelta(hours=1),
        fecha_hora_inicio__lte=ventana_citas_fin,
    ).select_related('familiar__usuario', 'estudiante__usuario').order_by('fecha_hora_inicio')

    citas_por_dia_semana = {i: [] for i in range(7)}
    citas_reunion_hoy = []
    hoy_local = timezone.localtime(ahora_tz).date()
    for _cita in citas_docente_horario:
        local_inicio = timezone.localtime(_cita.fecha_hora_inicio)
        citas_por_dia_semana[local_inicio.weekday()].append(_cita)
        if local_inicio.date() == hoy_local:
            citas_reunion_hoy.append(_cita)
    for _k in citas_por_dia_semana:
        citas_por_dia_semana[_k].sort(key=lambda c: c.fecha_hora_inicio)
    # Buscamos los 5 cuestionarios más recientes creados por el usuario actual
    ultimos_cuestionarios = Cuestionario.objects.filter(
        creado_por=request.user
    ).select_related('actividad_calificable__curso__materia').order_by('-fecha_creacion')[:5]
    # --- FIN DE LA MODIFICACIÓN ---

    direccion_grupo = DirectorCurso.objects.filter(docente=docente, periodo_academico=periodo_activo, institucion=institucion).select_related('grado').first()
    notificaciones_sin_leer = Notificacion.objects.filter(destinatario=request.user, institucion=institucion, leido=False)

    horario_agrupado = defaultdict(list)
    hoy = timezone.now().date()
    dia_semana_hoy = hoy.weekday()
    cursos_de_hoy = Curso.objects.none()

    layout_guardado = docente.dashboard_layout or [] # Obtiene el layout o una lista vacía
    dashboard_layout_json = json.dumps(layout_guardado)

    
    context = {
        'docente': docente,
        'periodo_activo': periodo_activo,
        'es_director_de_grupo': bool(direccion_grupo),
        'titulo_pagina': "Panel de Docente",
        'cursos_asignados': [],
        'entregas_pendientes_calificar': 0,
        'lecciones_sin_registrar_hoy': [],
        'cantidad_estudiantes_riesgo': 0,
        'notificaciones_sin_leer': notificaciones_sin_leer,
        'citas_proxima_semana': citas_proxima_semana_count,
        'citas_por_dia_semana': citas_por_dia_semana,
        'citas_reunion_hoy': citas_reunion_hoy,
        'citas_proximas_lista': list(citas_docente_horario),  # lista plana cronológica para la sección dedicada
        'horario_agrupado': dict(horario_agrupado),
        'dias_semana': BloqueHorario.DIA_SEMANA_CHOICES,
        'cursos_de_hoy': cursos_de_hoy,
        'dia_semana_hoy': dia_semana_hoy,
        'ultimos_cuestionarios': ultimos_cuestionarios,
        'dashboard_layout_json': dashboard_layout_json,
        
    }

    if periodo_activo:
        cursos_del_docente = Curso.objects.filter(docentes_asignados=docente, periodo_academico=periodo_activo).select_related('materia', 'grado', 'aula')
        context['cursos_asignados'] = list(cursos_del_docente)

        # ── Mallas curriculares del docente ──────────────────────────────────
        año_actual_malla = date.today().year
        pares_docente = {(c.materia_id, c.grado_id) for c in cursos_del_docente}
        q_mallas = Q()
        for materia_id, grado_id in pares_docente:
            q_mallas |= Q(materia_id=materia_id, grado_id=grado_id)
        if direccion_grupo:
            q_mallas |= Q(grado=direccion_grupo.grado)
        if q_mallas:
            mallas_qs = (
                MallaCurricular.objects
                .filter(q_mallas, institucion=institucion, año_lectivo=año_actual_malla)
                .select_related('materia', 'grado')
                .prefetch_related('items')
                .order_by('grado__orden', 'materia__nombre_materia')
            )
        else:
            mallas_qs = MallaCurricular.objects.none()
        mallas_con_items = []
        for malla in mallas_qs:
            por_periodo = {1: [], 2: [], 3: [], 4: []}
            for item in malla.items.all().order_by('periodo', 'orden'):
                por_periodo.setdefault(item.periodo, []).append(item)
            mallas_con_items.append({'malla': malla, 'por_periodo': por_periodo})
        context['mallas_docente'] = mallas_con_items
        context['año_malla'] = año_actual_malla

        # ── Planes devueltos con observaciones (acción requerida) ─────────────
        context['planes_con_observaciones'] = list(
            PlanSemanal.objects
            .filter(docente=docente, institucion=institucion, estado=PlanSemanal.Estado.CON_OBSERVACIONES)
            .select_related('curso__materia', 'curso__grado')
            .order_by('-semana_inicio')
        )

        # --- INICIO DE LA MODIFICACIÓN ---
        # Contamos las entregas de deberes pendientes
        entregas_count = EntregaDeber.objects.filter(
            deber__curso__in=cursos_del_docente, 
            calificacion_obtenida__isnull=True, 
            archivo_adjunto_estudiante__isnull=False
        ).count()

        # Contamos los intentos de cuestionario que necesitan revisión manual
        intentos_count = IntentoCuestionario.objects.filter(
            cuestionario__actividad_calificable__curso__in=cursos_del_docente,
            estado='FINALIZADO',
            respuestas__pregunta__tipo='texto_libre'
        ).distinct().count()

        # Sumamos ambos para el total del widget
        context['entregas_pendientes_calificar'] = entregas_count + intentos_count
        # --- FIN DE LA MODIFICACIÓN ---

        context['lecciones_sin_registrar_hoy'] = cursos_del_docente.exclude(pk__in=LeccionDiaria.objects.filter(curso__in=cursos_del_docente, fecha=timezone.localdate()).values_list('curso_id', flat=True))

        bloques_horario = BloqueHorario.objects.filter(
            curso__in=cursos_del_docente
        ).select_related('curso__materia', 'curso__grado', 'aula').order_by('hora_inicio')

        for bloque in bloques_horario:
            horario_agrupado[bloque.dia_semana].append(bloque)
        context['horario_agrupado'] = dict(horario_agrupado)

        cursos_de_hoy = cursos_del_docente.filter(horarios__dia_semana=dia_semana_hoy).distinct()
        context['cursos_de_hoy'] = cursos_de_hoy

        if direccion_grupo:
            estudiantes_del_grupo = Estudiante.objects.filter(
                grado_actual=direccion_grupo.grado,
                institucion=docente.institucion,
            )
            cursos_del_grado = Curso.objects.filter(
                grado=direccion_grupo.grado,
                periodo_academico=periodo_activo,
            )
            context["cantidad_estudiantes_riesgo"] = contar_pares_estudiante_curso_en_riesgo_academico(
                estudiantes_del_grupo,
                cursos_del_grado,
            )

    return render(request, 'gestion_academica/dashboard_docente.html', context)

@require_POST
@login_required
def guardar_layout_dashboard(request):
    """
    Recibe la configuración del layout de GridStack vía AJAX y la guarda en el perfil del docente.
    """
    try:
        # Cargamos los datos JSON que envía el navegador
        data = json.loads(request.body)
        
        # Guardamos los datos en el campo que creamos en el modelo
        request.user.docente.dashboard_layout = data
        request.user.docente.save()
        
        return JsonResponse({'status': 'ok', 'message': 'Diseño guardado.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
        

@login_required
def exportar_reporte_riesgo_view(request):
    """
    Exporta a Excel el reporte COMPLETO de estudiantes en riesgo,
    ordenado alfabéticamente por estudiante.
    """
    # 1. Obtenemos los datos del docente y su curso (misma lógica que el reporte)
    try:
        docente = request.user.docente
        periodo_activo = PeriodoAcademico.objects.filter(activo=True, institucion=docente.institucion).first()
        direccion_grupo = DirectorCurso.objects.filter(docente=docente, periodo_academico=periodo_activo).select_related('grado').first()
    except (AttributeError, Docente.DoesNotExist):
        return HttpResponse("Acceso denegado.", status=403)

    if not direccion_grupo:
        return HttpResponse("No eres director de grupo en el periodo activo.", status=403)

    # 2. Calculamos los datos del reporte (misma lógica)
    estudiantes_del_grupo = Estudiante.objects.filter(grado_actual=direccion_grupo.grado, institucion=docente.institucion)
    cursos_del_grado = Curso.objects.filter(grado=direccion_grupo.grado, periodo_academico=periodo_activo)
    
    data_para_excel = []
    for estudiante in estudiantes_del_grupo:
        for curso in cursos_del_grado:
            resultado_riesgo = analizar_riesgo_academico_curso(curso, estudiante)
            if resultado_riesgo['estado'] in ESTADOS_RIESGO_ACADEMICO_CURSO:
                data_para_excel.append({
                    'Estudiante': estudiante.usuario.get_full_name(),
                    'Materia en Riesgo': curso.materia.nombre_materia,
                    'Estado Actual': resultado_riesgo['estado'],
                    'Nota Requerida': f"{resultado_riesgo['nota_requerida']:.2f}".replace('.', ',') if resultado_riesgo['nota_requerida'] is not None else ''
                })

    # ===============================================================
    # INICIO DEL CAMBIO: ORDENAMOS LOS DATOS
    # ===============================================================
    # Ordenamos la lista de resultados alfabéticamente por el nombre del estudiante
    data_ordenada = sorted(data_para_excel, key=lambda d: d['Estudiante'])
    # ===============================================================
    # FIN DEL CAMBIO
    # ===============================================================

    # 3. Generamos el archivo Excel con los datos ya ordenados
    df = pd.DataFrame(data_ordenada) # Usamos la lista ordenada
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Riesgo_Academico_{direccion_grupo.grado}.xlsx"'
    
    df.to_excel(response, index=False)
    return response   


# NUEVA VISTA PARA EL REPORTE DETALLADO
@login_required
def reporte_riesgo_academico_view(request):
    """
    Muestra el dashboard de 'HALU Sentinel' con las predicciones de riesgo.
    """
    if not (request.user.is_superuser or (request.user.is_staff and request.user.rol in ['administrador', 'coordinador'])):
        messages.error(request, "No tienes permiso para ver este reporte.")
        return redirect('gestion_academica:dashboard_coordinador')
        
    ultimo_analisis = AnalisisRiesgo.objects.order_by('-fecha_analisis').first()
    
    predicciones = []
    if ultimo_analisis:
        predicciones = ultimo_analisis.predicciones.select_related(
            'estudiante__usuario', 'materia'
        ).order_by('estudiante__usuario__last_name', 'materia__nombre_materia')

    context = {
        'titulo_pagina': "HALU Sentinel - Reporte de Riesgo Académico",
        'ultimo_analisis': ultimo_analisis,
        'predicciones': predicciones,
    }
    return render(request, 'gestion_academica/reporte_riesgo_academico.html', context)


# VISTA DE DETALLE (NUEVA) - Muestra el reporte de un solo estudiante
@login_required
def detalle_riesgo_estudiante_view(request, estudiante_pk):
    # La lógica de seguridad y obtención de datos es similar
    try:
        docente = request.user.docente
        periodo_activo = PeriodoAcademico.objects.filter(activo=True, institucion=docente.institucion).first()
        estudiante = Estudiante.objects.get(pk=estudiante_pk, institucion=docente.institucion)
        direccion_grupo = DirectorCurso.objects.get(docente=docente, periodo_academico=periodo_activo, grado=estudiante.grado_actual)
    except (Exception):
        messages.error(request, "Acceso denegado o datos no válidos.")
        return redirect('gestion_academica:reporte_riesgo_academico')

    cursos_del_grado = Curso.objects.filter(grado=direccion_grupo.grado, periodo_academico=periodo_activo)
    
    alertas_riesgo = []
    for curso in cursos_del_grado:
        resultado_riesgo = analizar_riesgo_academico_curso(curso, estudiante)
        if resultado_riesgo['estado'] in ESTADOS_RIESGO_ACADEMICO_CURSO:
            alertas_riesgo.append({
                'estudiante': estudiante,
                'curso': curso,
                'estado': resultado_riesgo['estado'],
                'nota_requerida': resultado_riesgo['nota_requerida']
            })
    
    # Ordenamos las alertas por el nombre de la materia
    alertas_ordenadas = sorted(alertas_riesgo, key=lambda a: a['curso'].materia.nombre_materia)

    context = {
        'titulo_pagina': f"Detalle de Riesgo: {estudiante.usuario.get_full_name()}",
        'estudiante': estudiante,
        'alertas_del_estudiante': alertas_ordenadas
    }
    return render(request, 'gestion_academica/detalle_riesgo_estudiante.html', context)

 

@login_required
@permission_required('gestion_academica.view_observacionboletin')
def gestionar_observaciones_curso(request, grado_pk, periodo_pk):
    """
    Paso 1: Muestra la lista de estudiantes para que el director elija a quién
    registrar una observación para el boletín.
    """
    grado = get_object_or_404(Grado, pk=grado_pk)
    periodo = get_object_or_404(PeriodoAcademico, pk=periodo_pk)

    # Lógica de permisos para asegurar que solo el director de grupo acceda
    try:
        director = DirectorCurso.objects.get(grado=grado, periodo_academico=periodo)
        if not request.user.is_superuser and director.docente != request.user.docente:
            messages.error(request, "No tienes permiso para gestionar las observaciones de este curso.")
            return redirect('gestion_academica:dashboard_docente')
    except DirectorCurso.DoesNotExist:
        messages.error(request, "No hay un director de grupo asignado para este curso y periodo.")
        return redirect('gestion_academica:dashboard_docente')

    # Obtenemos los estudiantes del curso
    estudiantes_del_curso = Estudiante.objects.filter(grado_actual=grado).select_related('usuario').order_by('usuario__last_name', 'usuario__first_name')
    
    context = {
        'grado': grado,
        'periodo': periodo,
        'estudiantes': estudiantes_del_curso,
        'titulo_pagina': f"Seleccionar Estudiante para Observaciones | {grado.nombre}"
    }
    # Renderizamos la NUEVA plantilla de lista
    return render(request, 'gestion_academica/gestionar_observaciones_lista.html', context)


# VISTA 2: La que muestra el formulario (esta es tu vista original, pero con un nuevo nombre)
@login_required
@permission_required('gestion_academica.add_observacionboletin')
def gestionar_observacion_estudiante_form(request, estudiante_pk, periodo_pk):
    """
    Paso 2: Muestra y procesa el formulario para un estudiante específico.
    """
    estudiante = get_object_or_404(Estudiante.objects.select_related('usuario', 'grado_actual'), pk=estudiante_pk)
    periodo = get_object_or_404(PeriodoAcademico, pk=periodo_pk)
    
    # Aquí puedes añadir la misma lógica de permisos de la vista anterior si quieres más seguridad

    observacion_obj, created = ObservacionBoletin.objects.get_or_create(
        estudiante=estudiante,
        periodo=periodo,
        defaults={
            'creado_por': request.user.docente if hasattr(request.user, 'docente') else None,
            'institucion': estudiante.institucion  # <-- CAMBIO CLAVE: Se añade la institución
        }
    )
    
    if request.method == 'POST':
        form = ObservacionBoletinForm(request.POST, instance=observacion_obj)
        if form.is_valid():
            form.save()
            messages.success(request, f"Observación para {estudiante.usuario.get_full_name()} guardada exitosamente.")
            # Regresa a la lista de estudiantes del mismo curso
            return redirect('gestion_academica:gestionar_observaciones', grado_pk=estudiante.grado_actual.pk, periodo_pk=periodo.pk)
    else:
        form = ObservacionBoletinForm(instance=observacion_obj)
    
    context = {
        'form': form,
        'estudiante': estudiante,
        'periodo': periodo,
        'titulo_pagina': f"Observación para {estudiante.usuario.get_full_name()}"
    }
    return render(request, 'gestion_academica/gestionar_observacion_formulario.html', context)

def obtener_desempeno(nota, institucion):
    """
    Busca en la base de datos la abreviatura del desempeño que corresponde
    a una nota específica para una institución dada.
    """
    if nota is None or institucion is None:
        return ""
    
    nota = Decimal(str(nota))
    
    escala = EscalaValorativa.objects.filter(
        institucion=institucion,
        nota_minima__lte=nota, # lte = Less Than or Equal (menor o igual que)
        nota_maxima__gte=nota  # gte = Greater Than or Equal (mayor o igual que)
    ).first()

    return escala.abreviatura if escala else "-"    

@login_required
@requiere_pagos_al_dia
def detalle_leccion_diaria(request, leccion_pk):
    """
    Muestra el detalle completo de una lección diaria para un estudiante.
    """
    leccion = get_object_or_404(LeccionDiaria.objects.select_related(
        'curso__materia', 'curso__grado', 'creado_por'
    ), pk=leccion_pk)

    # Lógica de seguridad
    try:
        estudiante = request.user.estudiante
        if estudiante.grado_actual != leccion.curso.grado:
            messages.error(request, "No tienes permiso para ver esta lección.")
            return redirect('gestion_academica:dashboard_estudiante')
    except AttributeError:
        if not request.user.is_staff:
            messages.error(request, "Acceso denegado.")
            return redirect('gestion_academica:inicio_academico')
            
    # ================================================================
    #   INICIO: CORRECCIÓN DEL FORMATEO DE FECHA
    # ================================================================
    # Formateamos la fecha usando Python, no la sintaxis de plantilla
    fecha_formateada = date_format(leccion.fecha, "d M, Y")
    
    context = {
        'leccion': leccion,
        'titulo_pagina': f"Clase del {fecha_formateada}"
    }
    # ================================================================
    #   FIN: CORRECCIÓN
    # ================================================================
    
    return render(request, 'gestion_academica/detalle_leccion.html', context)   

@login_required
@requiere_pagos_al_dia
def detalle_calificaciones_por_materia(request, materia_pk):
    """
    Muestra una página con el historial de calificaciones de un estudiante
    para una materia específica en el periodo activo.
    """
    try:
        estudiante = request.user.estudiante
    except Estudiante.DoesNotExist:
        messages.error(request, "Tu perfil de estudiante no está configurado.")
        return redirect('gestion_academica:dashboard_estudiante')

    periodo_activo = PeriodoAcademico.objects.filter(activo=True, institucion=estudiante.institucion).first()
    materia = get_object_or_404(Materia, pk=materia_pk, institucion=estudiante.institucion)

    calificaciones_materia = []
    if periodo_activo:
        calificaciones_materia = Calificacion.objects.filter(
            estudiante=estudiante,
            actividad_calificable__curso__materia=materia,
            actividad_calificable__curso__periodo_academico=periodo_activo
        ).select_related('actividad_calificable').order_by('-fecha_registro')

    context = {
        'titulo_pagina': f"Mis Notas en {materia.nombre_materia}",
        'materia': materia,
        'periodo_activo': periodo_activo,
        'calificaciones': calificaciones_materia,
        'estudiante': estudiante,
    }
    
    return render(request, 'gestion_academica/detalle_calificaciones_materia.html', context)   

class CentroGestionDocenteView(LoginRequiredMixin, TemplateView):
    """
    Página principal del Centro de Gestión del Docente, que servirá como
    un menú de acceso a las diferentes herramientas de gestión.
    """
    template_name = 'gestion_academica/centro_gestion_docente.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Centro de Gestión del Docente"
        return context

@login_required
def docente_hub_reportes(request):
    """
    Hub centralizado de reportes para el docente.
    Muestra reportes por materia siempre, y reportes de director de grupo
    únicamente si el docente es director en el periodo activo.
    """
    if not hasattr(request.user, 'docente'):
        messages.error(request, "Acceso denegado. Solo para docentes.")
        return redirect('gestion_academica:dashboard_docente')

    docente = request.user.docente
    periodo_activo = PeriodoAcademico.objects.filter(
        activo=True, institucion=docente.institucion
    ).first()

    direccion_grupo = None
    cursos_asignados = []
    if periodo_activo:
        direccion_grupo = DirectorCurso.objects.filter(
            docente=docente, periodo_academico=periodo_activo
        ).select_related('grado').first()
        cursos_asignados = list(
            Curso.objects.filter(
                docentes_asignados=docente,
                periodo_academico=periodo_activo
            ).select_related('materia', 'grado')
        )

    context = {
        'titulo_pagina': 'Centro de Reportes',
        'es_director_de_grupo': bool(direccion_grupo),
        'direccion_grupo': direccion_grupo,
        'cursos_asignados': cursos_asignados,
        'periodo_activo': periodo_activo,
    }
    return render(request, 'gestion_academica/docente_hub_reportes.html', context)


class DocenteMaterialListView(LoginRequiredMixin, ListView):
    """
    Muestra al docente una lista de los materiales que ÉL ha subido.
    """
    model = ArchivoPlanAcademico
    template_name = 'gestion_academica/docente_material_lista.html'
    context_object_name = 'archivos'
    paginate_by = 10

    def get_queryset(self):
        # Filtramos para mostrar solo los archivos subidos por el usuario actual
        return ArchivoPlanAcademico.objects.filter(subido_por=self.request.user).order_by('-fecha_subida')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Mis Archivos y Materiales"
        return context

class DocenteMaterialCreateView(LoginRequiredMixin, CreateView):
    """
    Permite al docente subir un nuevo archivo.
    """
    model = ArchivoPlanAcademico
    form_class = ArchivoPlanAcademicoForm
    template_name = 'gestion_academica/docente_material_formulario.html'
    success_url = reverse_lazy('gestion_academica:docente_lista_materiales')

    def get_form_kwargs(self):
        # Pasamos el objeto 'request' al formulario para poder filtrar los cursos
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        # Asignamos automáticamente el docente y la institución antes de guardar
        form.instance.subido_por = self.request.user
        form.instance.institucion = self.request.user.institucion_asociada
        messages.success(self.request, "¡Material subido exitosamente!")
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Subir Nuevo Material"
        return context

class DocenteMaterialUpdateView(LoginRequiredMixin, UpdateView):
    """
    Permite al docente editar la información de un material que ha subido.
    """
    model = ArchivoPlanAcademico
    form_class = ArchivoPlanAcademicoForm
    template_name = 'gestion_academica/docente_material_formulario.html'
    success_url = reverse_lazy('gestion_academica:docente_lista_materiales')

    def get_queryset(self):
        # Medida de seguridad: un docente solo puede editar sus propios archivos.
        return ArchivoPlanAcademico.objects.filter(subido_por=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, "Material actualizado exitosamente.")
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Material"
        return context

class DocenteMaterialDeleteView(LoginRequiredMixin, DeleteView):
    """
    Permite al docente eliminar un material que ha subido.
    """
    model = ArchivoPlanAcademico
    template_name = 'gestion_academica/docente_material_confirmar_eliminar.html'
    success_url = reverse_lazy('gestion_academica:docente_lista_materiales')
    context_object_name = 'archivo'

    def get_queryset(self):
        # Medida de seguridad: un docente solo puede eliminar sus propios archivos.
        return ArchivoPlanAcademico.objects.filter(subido_por=self.request.user)

    def form_valid(self, form):
        messages.success(self.request, f"El archivo '{self.object.nombre_archivo_descriptivo}' ha sido eliminado.")
        return super().form_valid(form)     
    
class DocenteDescriptorListView(LoginRequiredMixin, ListView):
    model = DescriptorLogro
    template_name = 'gestion_academica/docente_descriptor_lista.html'
    context_object_name = 'descriptores'

    def get_queryset(self):
        """
        Filtra los descriptores para mostrar solo los creados por el docente
        logueado y que pertenecen a su institución.
        """
        try:
            # Obtenemos la institución a través del perfil del docente
            institucion_actual = self.request.user.docente.institucion
            
            # --- LÍNEA CORREGIDA ---
            # Filtramos 'creado_por' usando el objeto de usuario directamente.
            return DescriptorLogro.objects.filter(
                institucion=institucion_actual, 
                creado_por=self.request.user  # <-- ¡AQUÍ ESTÁ EL CAMBIO!
            )
        except (AttributeError, Docente.DoesNotExist):
            return DescriptorLogro.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['upload_form'] = UploadFileForm()
        return context

    def post(self, request, *args, **kwargs):
        """
        Procesa el archivo Excel subido para la carga masiva.
        """
        form = UploadFileForm(request.POST, request.FILES)
        if not form.is_valid():
            messages.error(request, "Error en el formulario. Por favor, intenta de nuevo.")
            return redirect('gestion_academica:docente_lista_descriptores')

        archivo = request.FILES['file']
        
        try:
            df = pd.read_excel(archivo, sheet_name="PLANTILLA_BANCO_DE_LOGROS")
            required_cols = ['CODIGO_ASIGNATURA', 'PERIODO', 'TEXTO_DESCRIPTOR']
            if not all(col in df.columns for col in required_cols):
                messages.error(request, f"El archivo no tiene las columnas requeridas: {required_cols}.")
                return redirect('gestion_academica:docente_lista_descriptores')

            institucion_actual = request.user.docente.institucion
            
            logros_creados_con_exito = 0
            
            for index, row in df.iterrows():
                if pd.isna(row['CODIGO_ASIGNATURA']) or pd.isna(row['PERIODO']) or pd.isna(row['TEXTO_DESCRIPTOR']):
                    continue

                try:
                    materia_obj = Materia.objects.get(codigo_materia=str(row['CODIGO_ASIGNATURA']).strip(), institucion=institucion_actual)
                    periodo_str = str(row['PERIODO']).strip()
                    periodo_obj = PeriodoAcademico.objects.get(nombre=periodo_str, institucion=institucion_actual)
                    
                    # --- LÍNEA CORREGIDA ---
                    # Al crear, también usamos el objeto de usuario para 'creado_por'.
                    DescriptorLogro.objects.create(
                        materia=materia_obj,
                        periodo_academico=periodo_obj,
                        descripcion=str(row['TEXTO_DESCRIPTOR']).strip(),
                        creado_por=request.user, # <-- ¡AQUÍ ESTÁ EL SEGUNDO CAMBIO!
                        institucion=institucion_actual
                    )
                    logros_creados_con_exito += 1
                
                except Materia.DoesNotExist:
                    messages.warning(request, f"Fila {index + 2}: La materia con código '{row['CODIGO_ASIGNATURA']}' no existe en tu institución.")
                except PeriodoAcademico.DoesNotExist:
                    messages.warning(request, f"Fila {index + 2}: El periodo '{periodo_str}' no existe en tu institución.")
                except Exception as e:
                    messages.error(request, f"Fila {index + 2}: Ocurrió un error inesperado al procesar esta fila. Error: {e}")
            
            if logros_creados_con_exito > 0:
                messages.success(request, f"¡Éxito! Se han importado {logros_creados_con_exito} descriptores.")
            else:
                messages.info(request, "Proceso finalizado. No se importaron descriptores nuevos.")

        except Exception as e:
            messages.error(request, f"No se pudo procesar el archivo. Error: {e}")
        
        return redirect('gestion_academica:docente_lista_descriptores')

class DocenteDescriptorCreateView(LoginRequiredMixin, CreateView):
    model = DescriptorLogro
    form_class = DescriptorLogroForm
    template_name = 'gestion_academica/docente_descriptor_formulario.html'
    success_url = reverse_lazy('gestion_academica:docente_lista_descriptores')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        messages.success(self.request, "Descriptor de logro creado exitosamente.")
        return super().form_valid(form)

class DocenteDescriptorUpdateView(LoginRequiredMixin, UpdateView):
    model = DescriptorLogro
    form_class = DescriptorLogroForm
    template_name = 'gestion_academica/docente_descriptor_formulario.html'
    success_url = reverse_lazy('gestion_academica:docente_lista_descriptores')

    def get_queryset(self):
        # Seguridad: Un docente solo puede editar sus propios descriptores
        return DescriptorLogro.objects.filter(creado_por=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

class DocenteDescriptorDeleteView(LoginRequiredMixin, DeleteView):
    model = DescriptorLogro
    template_name = 'gestion_academica/docente_descriptor_confirmar_eliminar.html'
    success_url = reverse_lazy('gestion_academica:docente_lista_descriptores')
    context_object_name = 'descriptor'

    def get_queryset(self):
        return DescriptorLogro.objects.filter(creado_por=self.request.user) 

# --- Gestión de Menciones por Docente ---

class DocenteMencionListView(LoginRequiredMixin, ListView):
    model = MencionReconocimiento
    template_name = 'gestion_academica/docente_mencion_lista.html'
    context_object_name = 'menciones'

    def get_queryset(self):
        # Muestra solo las menciones creadas por el docente logueado
        return MencionReconocimiento.objects.filter(otorgado_por=self.request.user.docente).order_by('-fecha_otorgamiento')

class DocenteMencionCreateView(LoginRequiredMixin, CreateView):
    model = MencionReconocimiento
    form_class = MencionReconocimientoForm
    template_name = 'gestion_academica/docente_mencion_formulario.html'
    success_url = reverse_lazy('gestion_academica:docente_lista_menciones')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        # Asigna automáticamente el docente y la institución
        form.instance.otorgado_por = self.request.user.docente
        form.instance.institucion = self.request.user.institucion_asociada
        messages.success(self.request, "Mención creada exitosamente.")
        return super().form_valid(form)

class DocenteMencionUpdateView(LoginRequiredMixin, UpdateView):
    model = MencionReconocimiento
    form_class = MencionReconocimientoForm
    template_name = 'gestion_academica/docente_mencion_formulario.html'
    success_url = reverse_lazy('gestion_academica:docente_lista_menciones')

    def get_queryset(self):
        # Seguridad: Un docente solo puede editar sus propias menciones
        return MencionReconocimiento.objects.filter(otorgado_por=self.request.user.docente)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

class DocenteMencionDeleteView(LoginRequiredMixin, DeleteView):
    model = MencionReconocimiento
    template_name = 'gestion_academica/docente_mencion_confirmar_eliminar.html'
    success_url = reverse_lazy('gestion_academica:docente_lista_menciones')
    context_object_name = 'mencion'

    def get_queryset(self):
        return MencionReconocimiento.objects.filter(otorgado_por=self.request.user.docente)

def link_callback(uri, rel):  # noqa: F811
    """
    Convierte una URL de recurso (/media/... o /static/...) a una ruta de sistema
    de archivos absoluta que xhtml2pdf pueda encontrar. Protegida contra path traversal.
    """
    # Para archivos de MEDIA (logos de la institución, etc.)
    if uri.startswith(settings.MEDIA_URL):
        path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, "", 1))
        allowed_root = os.path.realpath(settings.MEDIA_ROOT)

    # Para archivos STATIC (marco del diploma, sello, etc.)
    elif uri.startswith(settings.STATIC_URL):
        path = os.path.join(settings.STATICFILES_DIRS[0], uri.replace(settings.STATIC_URL, "", 1))
        allowed_root = os.path.realpath(settings.STATICFILES_DIRS[0])

    else:
        return uri

    # Protección contra path traversal (ej: /../../../etc/passwd)
    real_path = os.path.realpath(path)
    if not real_path.startswith(allowed_root + os.sep) and real_path != allowed_root:
        logger.warning("link_callback: path traversal bloqueado para URI: %s", uri)
        return None

    if not os.path.isfile(real_path):
        return None

    return real_path

@login_required
def generar_mencion_pdf(request, mencion_pk):
    """
    Genera un certificado en PDF para una mención de honor específica.
    """
    try:
        # Obtenemos la mención con todos los datos relacionados para optimizar
        mencion = MencionReconocimiento.objects.select_related(
            'estudiante__usuario',
            'otorgado_por__usuario',
            'institucion'
        ).get(pk=mencion_pk)
    except MencionReconocimiento.DoesNotExist:
        return HttpResponse("La mención solicitada no existe.", status=404)

    # Lógica de seguridad: Solo el estudiante, su familiar o un staff pueden verla
    es_el_mismo_estudiante = (request.user.pk == mencion.estudiante.usuario.pk)
    es_familiar_asociado = hasattr(request.user, 'familiar') and request.user.familiar.estudiantes_asociados.filter(pk=mencion.estudiante.pk).exists()

    def _respuesta_no_disponible(titulo, mensaje, url_volver):
        return render(
            request,
            'gestion_academica/mencion_certificado_no_disponible.html',
            {'titulo': titulo, 'mensaje': mensaje, 'url_volver': url_volver},
            status=403,
        )

    if not (es_el_mismo_estudiante or es_familiar_asociado or request.user.is_staff):
        url = reverse('gestion_academica:inicio_academico')
        return _respuesta_no_disponible(
            'Acceso denegado',
            'No tienes permiso para ver este reconocimiento.',
            url,
        )

    # Misma política que el portal: sin diploma descargable si el estudiante está en mora
    # (el personal sí puede generarlo).
    if not request.user.is_staff:
        try:
            moroso = not mencion.estudiante.esta_al_dia()
        except Exception:
            moroso = False
        if moroso and (es_el_mismo_estudiante or es_familiar_asociado):
            if es_familiar_asociado:
                volver = reverse('gestion_academica:portal_familiar_inicio')
            else:
                volver = reverse('gestion_academica:dashboard_estudiante')
            return _respuesta_no_disponible(
                'Diploma no disponible por atrasos en pagos',
                'Mientras existan obligaciones vencidas sin regularizar, el certificado no se puede '
                'descargar desde el portal. Puedes pagar en «Pagos en línea» / estado de cuenta; '
                'si ya pagaste, espera unos minutos y vuelve a intentar.',
                volver,
            )

    context = {
        'mencion': mencion,
        'institucion': mencion.institucion,
    }

    template_path = 'gestion_academica/mencion_imprimible.html'
    template = get_template(template_path)
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    # El nombre del archivo será algo como "Mencion_Merito_Deportivo_Sebastian_Acuña.pdf"
    response['Content-Disposition'] = f'inline; filename="Mencion_{mencion.tipo.replace(" ", "_")}_{mencion.estudiante}.pdf"'

    # Creamos el PDF
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
    if pisa_status.err:
       return HttpResponse('Ocurrió un error al generar el PDF <pre>' + html + '</pre>')
    return response

# --- Gestión del Observador del Estudiante ---

@login_required
def seleccionar_estudiante_observador(request):
    """
    Muestra al docente una lista de sus estudiantes para seleccionar a quién
    le hará una anotación en el observador.
    """
    if not hasattr(request.user, 'docente'):
        messages.error(request, "Acceso denegado.")
        return redirect('gestion_academica:dashboard_docente')

    docente = request.user.docente
    periodo_activo = PeriodoAcademico.objects.filter(activo=True, institucion=docente.institucion).first()
    
    estudiantes = []
    if periodo_activo:
        cursos_docente = Curso.objects.filter(docentes_asignados=docente, periodo_academico=periodo_activo)
        grados_ids = cursos_docente.values_list('grado_id', flat=True).distinct()
        estudiantes = Estudiante.objects.filter(
            grado_actual_id__in=grados_ids,
            institucion=docente.institucion
        ).select_related('usuario', 'grado_actual').order_by('grado_actual__nombre', 'usuario__last_name')

    context = {
        'estudiantes': estudiantes,
        'titulo_pagina': 'Seleccionar Estudiante para Observador'
    }
    return render(request, 'gestion_academica/seleccionar_estudiante_observador.html', context)


@login_required
def historial_observador_estudiante(request, estudiante_pk):
    """
    Muestra el historial y guarda nuevas anotaciones, disparando el signal de Halu Sentinel.
    """
    estudiante = get_object_or_404(Estudiante.objects.select_related('usuario', 'institucion'), pk=estudiante_pk)
    
    if request.method == 'POST':
        form = AnotacionObservadorForm(request.POST, request=request)
        if form.is_valid():
            anotacion = form.save(commit=False)
            anotacion.estudiante = estudiante
            anotacion.registrado_por = request.user
            
            # Asignación clave para la integridad de la base de datos
            anotacion.institucion = estudiante.institucion
            
            # Al hacer .save(), el signal se disparará automáticamente
            anotacion.save()
            
            messages.success(request, f"Anotación para {estudiante} registrada. Halu Sentinel la está analizando.")
            return redirect('gestion_academica:historial_observador', estudiante_pk=estudiante.pk)
    else:
        form = AnotacionObservadorForm(request=request)

    anotaciones = AnotacionObservador.objects.filter(estudiante=estudiante).select_related('registrado_por', 'curso__materia').order_by('-fecha_hora')
    
    context = {
        'estudiante': estudiante,
        'form': form,
        'anotaciones': anotaciones,
        'titulo_pagina': f"Observador de {estudiante}"
    }
    return render(request, 'gestion_academica/historial_observador_estudiante.html', context)


@login_required
def exportar_observador_pdf(request, estudiante_pk):
    """
    Genera el Observador del Estudiante en PDF (formato oficial Decreto 1860/1994).
    Admite ?periodo=<id> para filtrar por período académico.
    """
    from gestion_academica.models import Familiar, PeriodoAcademico

    estudiante = get_object_or_404(
        Estudiante.objects.select_related('usuario', 'grado_actual', 'institucion'),
        pk=estudiante_pk,
    )
    institucion = estudiante.institucion

    # ── Familiar principal ──────────────────────────────────────────────────
    familiar = (
        Familiar.objects
        .filter(estudiantes_asociados=estudiante, institucion=institucion)
        .select_related('usuario')
        .first()
    )

    # ── Período académico ───────────────────────────────────────────────────
    periodo_id = request.GET.get('periodo')
    periodo_seleccionado = None
    if periodo_id:
        periodo_seleccionado = PeriodoAcademico.objects.filter(
            pk=periodo_id, institucion=institucion
        ).first()
    if not periodo_seleccionado:
        periodo_seleccionado = PeriodoAcademico.objects.filter(
            institucion=institucion, activo=True
        ).first()

    # ── Anotaciones (filtradas por período si se eligió uno) ────────────────
    qs_anotaciones = (
        AnotacionObservador.objects
        .filter(estudiante=estudiante)
        .select_related('registrado_por', 'registrado_por__docente', 'curso__materia')
        .order_by('fecha_hora')
    )
    if periodo_seleccionado:
        qs_anotaciones = qs_anotaciones.filter(
            fecha_hora__date__gte=periodo_seleccionado.fecha_inicio,
            fecha_hora__date__lte=periodo_seleccionado.fecha_fin,
        )

    # ── Firma del docente que genera (si es docente) ────────────────────────
    firma_docente = None
    docente_generador = None
    if hasattr(request.user, 'docente'):
        docente_generador = request.user.docente
        firma_docente = docente_generador.firma_docente or None

    # ── Todos los períodos para el selector en pantalla ─────────────────────
    periodos_disponibles = PeriodoAcademico.objects.filter(
        institucion=institucion
    ).order_by('-año_escolar', 'nombre')

    context = {
        'estudiante': estudiante,
        'familiar': familiar,
        'anotaciones': qs_anotaciones,
        'institucion': institucion,
        'fecha_generacion': timezone.now(),
        'periodo_seleccionado': periodo_seleccionado,
        'periodos_disponibles': periodos_disponibles,
        'firma_docente': firma_docente,
        'docente_generador': docente_generador,
        'generado_por': request.user.get_full_name() or request.user.username,
    }

    template_path = 'gestion_academica/observador_imprimible.html'
    template = get_template(template_path)
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    nombre_archivo = f"ObservadorEstudiante_{estudiante.usuario.get_full_name().replace(' ', '_')}.pdf"
    response['Content-Disposition'] = f'inline; filename="{nombre_archivo}"'

    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
    if pisa_status.err:
        return HttpResponse('Ocurrió un error al generar el PDF del Observador.')
    return response                
       

class DocenteActividadListView(LoginRequiredMixin, ListView):
    model = ActividadCalificable
    template_name = 'gestion_academica/docente_actividad_lista.html'
    context_object_name = 'actividades'

    def get_queryset(self):
        # Muestra solo las actividades de los cursos del docente en el periodo activo
        periodo_activo = PeriodoAcademico.objects.filter(activo=True, institucion=self.request.user.institucion_asociada).first()
        if periodo_activo:
            return ActividadCalificable.objects.filter(
                curso__docentes_asignados=self.request.user.docente,
                curso__periodo_academico=periodo_activo
            ).order_by('curso', 'titulo')
        return ActividadCalificable.objects.none()

class DocenteActividadCreateView(LoginRequiredMixin, CreateView):
    model = ActividadCalificable
    form_class = DocenteActividadForm
    template_name = 'gestion_academica/docente_actividad_formulario.html'
    success_url = reverse_lazy('gestion_academica:docente_lista_actividades')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        # Le pasamos el request al constructor del formulario
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.instance.institucion = self.request.user.institucion_asociada
        # El campo 'porcentaje_en_periodo' ya no se usa aquí, se usará el de la categoría.
        messages.success(self.request, "Actividad creada exitosamente.")
        return super().form_valid(form)

class DocenteActividadUpdateView(LoginRequiredMixin, UpdateView):
    model = ActividadCalificable
    form_class = DocenteActividadForm
    template_name = 'gestion_academica/docente_actividad_formulario.html'
    success_url = reverse_lazy('gestion_academica:docente_lista_actividades')

    def get_queryset(self):
        # Seguridad: un docente solo puede editar actividades de sus cursos
        return ActividadCalificable.objects.filter(curso__docentes_asignados=self.request.user.docente)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        # Le pasamos el request al constructor del formulario
        kwargs['request'] = self.request
        return kwargs

class DocenteActividadDeleteView(LoginRequiredMixin, DeleteView):
    model = ActividadCalificable
    template_name = 'gestion_academica/docente_actividad_confirmar_eliminar.html'
    success_url = reverse_lazy('gestion_academica:docente_lista_actividades')
    context_object_name = 'actividad'

    def get_queryset(self):
        return ActividadCalificable.objects.filter(curso__docentes_asignados=self.request.user.docente)

class DocenteTipoActividadListView(LoginRequiredMixin, ListView):
    model = TipoActividad
    template_name = 'gestion_academica/docente_tipo_actividad_lista.html'
    context_object_name = 'categorias'

    def get_queryset(self):
        # La consulta se mantiene, pero la ordenamos por nombre para consistencia
        return TipoActividad.objects.filter(institucion=self.request.user.institucion_asociada).order_by('nombre')

    def get_context_data(self, **kwargs):
        # Obtenemos el contexto original
        context = super().get_context_data(**kwargs)
        
        # --- INICIO DE LA NUEVA LÓGICA ---
        
        # Calculamos la suma de los porcentajes de las categorías listadas
        total_porcentaje = sum(cat.porcentaje for cat in context['categorias'] if cat.porcentaje)
        
        # Añadimos el total y una bandera de advertencia al contexto
        context['total_porcentaje'] = total_porcentaje
        context['porcentaje_excedido'] = total_porcentaje > 100
        
        # --- FIN DE LA NUEVA LÓGICA ---
        
        context['titulo_pagina'] = "Gestionar Categorías de Calificación"
        return context


class DocenteTipoActividadCreateView(LoginRequiredMixin, View):
    """
    Acceso bloqueado: la gestión de categorías es exclusiva del coordinador.
    Se mantiene la ruta para compatibilidad, pero redirige de inmediato.
    """
    def dispatch(self, request, *args, **kwargs):
        messages.warning(
            request,
            "La creación de categorías de calificación es exclusiva del coordinador académico."
        )
        return redirect('gestion_academica:docente_lista_tipos_actividad')


class DocenteTipoActividadUpdateView(LoginRequiredMixin, View):
    """
    Acceso bloqueado: la edición de categorías es exclusiva del coordinador.
    """
    def dispatch(self, request, *args, **kwargs):
        messages.warning(
            request,
            "La edición de categorías de calificación es exclusiva del coordinador académico."
        )
        return redirect('gestion_academica:docente_lista_tipos_actividad')


class DocenteTipoActividadDeleteView(LoginRequiredMixin, View):
    """
    Acceso bloqueado: la eliminación de categorías es exclusiva del coordinador.
    """
    def dispatch(self, request, *args, **kwargs):
        messages.warning(
            request,
            "La eliminación de categorías de calificación es exclusiva del coordinador académico."
        )
        return redirect('gestion_academica:docente_lista_tipos_actividad')

@login_required
def seleccionar_curso_reporte_nota_minima(request):
    """
    Muestra al docente una lista de sus cursos, permitiendo filtrar por
    periodo académico, para generar el reporte de nota mínima.
    """
    if not hasattr(request.user, 'docente'):
        messages.error(request, "Acceso denegado.")
        return redirect('gestion_academica:dashboard_docente')

    docente = request.user.docente
    institucion = request.user.institucion_asociada
    
    # Obtenemos todos los periodos de la institución para el filtro
    periodos_disponibles = PeriodoAcademico.objects.filter(institucion=institucion).order_by('-año_escolar', '-fecha_inicio')
    
    # Determinamos el periodo a mostrar: el que viene por GET o el activo por defecto
    periodo_seleccionado_id = request.GET.get('periodo')
    if periodo_seleccionado_id:
        periodo_a_mostrar = get_object_or_404(PeriodoAcademico, pk=periodo_seleccionado_id, institucion=institucion)
    else:
        periodo_a_mostrar = periodos_disponibles.filter(activo=True).first()

    # Filtramos los cursos del docente basados en el periodo seleccionado
    cursos_del_docente = []
    if periodo_a_mostrar:
        cursos_del_docente = Curso.objects.filter(
            docentes_asignados=docente,
            periodo_academico=periodo_a_mostrar
        ).select_related('materia', 'grado')

    context = {
        'cursos': cursos_del_docente,
        'periodos_disponibles': periodos_disponibles,
        'periodo_seleccionado': periodo_a_mostrar,
        'titulo_pagina': "Generar Reporte de Nota Mínima"
    }
    return render(request, 'gestion_academica/seleccionar_curso_reporte.html', context)


@login_required
def generar_reporte_nota_minima(request, curso_pk):
    curso = get_object_or_404(
        Curso.objects.select_related('grado', 'institucion'), pk=curso_pk
    )
    estudiantes = Estudiante.objects.filter(
        grado_actual=curso.grado
    ).select_related('usuario').order_by('usuario__last_name')

    reporte_data = []
    NOTA_OBJETIVO = getattr(curso.institucion, 'nota_minima_aprobacion', Decimal('3.0'))

    for estudiante in estudiantes:
        estado_academico = calcular_estado_academico_curso(curso, estudiante)
        nota_actual = estado_academico.get('nota_actual_promediada') or Decimal('0.0')
        porcentaje_evaluado = estado_academico.get('porcentaje_evaluado') or Decimal('0.0')
        puntos_acumulados = estado_academico.get('nota_final_ponderada') or Decimal('0.0')

        porcentaje_restante = Decimal('100.0') - porcentaje_evaluado
        estado = ""
        nota_requerida = None

        if porcentaje_restante <= 0:
            # ✅ Nuevo: Verificar si aprobó o reprobó al finalizar el curso
            if puntos_acumulados >= NOTA_OBJETIVO:
                estado = "Curso finalizado"
            else:
                estado = "Reprobado"
        else:
            puntos_necesarios = (NOTA_OBJETIVO - puntos_acumulados) * (Decimal('100.0') / porcentaje_restante)
            nota_requerida = puntos_necesarios.quantize(Decimal('0.01'))

            if nota_requerida <= 0:
                estado = "Aprobado (No necesita más puntos)"
            elif nota_requerida > 5.0:
                estado = "Situación Crítica (Imposible aprobar)"
            else:
                estado = "En Riesgo"

        reporte_data.append({
            'estudiante': estudiante,
            'promedio_actual': nota_actual,
            'porcentaje_evaluado': porcentaje_evaluado,
            'porcentaje_restante': porcentaje_restante,
            'nota_requerida': nota_requerida,
            'estado': estado
        })

    context = {
        'curso': curso,
        'reporte_data': reporte_data,
        'titulo_pagina': f"Reporte de Nota Mínima para {curso.materia}",
        'nota_objetivo': NOTA_OBJETIVO
    }
    return render(request, 'gestion_academica/reporte_nota_minima.html', context) 

def _get_panel_director_data(docente, periodo_activo):
    """
    Función auxiliar que calcula y prepara todos los datos para el panel del director.
    Devuelve un diccionario con los datos listos para ser usados.
    """
    direccion_grupo = DirectorCurso.objects.filter(docente=docente, periodo_academico=periodo_activo).select_related('grado').first()
    
    if not direccion_grupo:
        return None # Si no es director, no hay datos que procesar

    grado = direccion_grupo.grado
    estudiantes = Estudiante.objects.filter(grado_actual=grado).select_related('usuario').order_by('usuario__last_name')
    cursos = Curso.objects.filter(grado=grado, periodo_academico=periodo_activo).select_related('materia')
    
    panel_data = []
    for estudiante in estudiantes:
        notas_ordenadas = []
        for curso in cursos:
            estado_academico = calcular_estado_academico_curso(curso, estudiante)
            nota_final_materia = estado_academico.get('nota_final_ponderada')
            notas_ordenadas.append(nota_final_materia)

        notas_validas = [nota for nota in notas_ordenadas if nota is not None]
        promedio_estudiante = sum(notas_validas) / len(notas_validas) if notas_validas else None

        panel_data.append({
            'estudiante': estudiante,
            'notas_ordenadas': notas_ordenadas,
            'promedio_estudiante': promedio_estudiante,
        })

    return {
        'direccion_grupo': direccion_grupo,
        'cursos': cursos,
        'panel_data': panel_data,
        'titulo_pagina': f"Panel del Director de Grupo: {grado.nombre}"
    }        

@login_required
def panel_director_grupo(request):
    """
    Muestra el panel con la información de TODOS los grados de los que el docente es director.
    """
    try:
        docente = request.user.docente
    except (AttributeError, Docente.DoesNotExist):
        messages.error(request, "Acceso denegado. Solo para docentes.")
        return redirect('gestion_academica:inicio_academico')

    periodo_activo = PeriodoAcademico.objects.filter(activo=True, institucion=docente.institucion).first()
    
    if not periodo_activo:
        messages.error(request, "No hay un periodo académico activo configurado.")
        return redirect('gestion_academica:dashboard_docente')

    # 1. Buscamos TODAS las direcciones de grupo del docente, no solo la primera.
    direcciones_de_grupo = DirectorCurso.objects.filter(
        docente=docente, 
        periodo_academico=periodo_activo
    ).select_related('grado')

    if not direcciones_de_grupo.exists():
        messages.warning(request, "No eres director de ningún grupo en el periodo académico activo.")
        return redirect('gestion_academica:dashboard_docente')

    # 2. Preparamos una lista para guardar los datos de CADA panel de grado
    paneles_data = []
    
    # 3. Iteramos sobre cada grado que el docente dirige
    for direccion in direcciones_de_grupo:
        grado = direccion.grado
        estudiantes = Estudiante.objects.filter(grado_actual=grado, activo=True).select_related('usuario').order_by('usuario__last_name')
        cursos = Curso.objects.filter(grado=grado, periodo_academico=periodo_activo).select_related('materia').order_by('materia__nombre_materia')
        
        # Lógica de cálculo de notas para este grado específico
        data_estudiantes_grado = []
        for estudiante in estudiantes:
            notas_ordenadas = []
            for curso in cursos:
                # Reutilizamos tu función de cálculo para mantener la consistencia
                estado_academico = calcular_estado_academico_curso(curso, estudiante)
                nota_final_materia = estado_academico.get('nota_final_ponderada')
                notas_ordenadas.append(nota_final_materia)

            notas_validas = [nota for nota in notas_ordenadas if nota is not None]
            promedio_estudiante = sum(notas_validas) / len(notas_validas) if notas_validas else None

            data_estudiantes_grado.append({
                'estudiante': estudiante,
                'notas_ordenadas': notas_ordenadas,
                'promedio_estudiante': promedio_estudiante,
            })
        
        # Añadimos toda la información de este grado a nuestra lista de paneles
        paneles_data.append({
            'grado': grado,
            'cursos_header': cursos,
            'data_estudiantes': data_estudiantes_grado
        })

    context = {
        'paneles_data': paneles_data,
        'titulo_pagina': "Panel del Director de Grupo"
    }
    return render(request, 'gestion_academica/panel_director_grupo.html', context)

@login_required
def exportar_panel_director_excel(request):
    """
    Exporta los datos del panel del director a un archivo Excel,
    con una hoja de cálculo separada para cada grado que dirige.
    """
    try:
        docente = request.user.docente
        periodo_activo = PeriodoAcademico.objects.filter(activo=True, institucion=docente.institucion).first()
    except (AttributeError, Docente.DoesNotExist):
        return HttpResponse("Acceso denegado.", status=403)

    if not periodo_activo:
        messages.error(request, "No hay un periodo académico activo para exportar.")
        return redirect('gestion_academica:dashboard_docente')

    direcciones_de_grupo = DirectorCurso.objects.filter(docente=docente, periodo_academico=periodo_activo).select_related('grado')

    if not direcciones_de_grupo.exists():
        messages.warning(request, "No eres director de ningún grupo para exportar.")
        return redirect('gestion_academica:dashboard_docente')

    # Usamos BytesIO para crear el archivo Excel en memoria
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Iteramos sobre cada grado que el docente dirige
        for direccion in direcciones_de_grupo:
            grado = direccion.grado
            estudiantes = Estudiante.objects.filter(grado_actual=grado, activo=True).select_related('usuario').order_by('usuario__last_name')
            cursos = Curso.objects.filter(grado=grado, periodo_academico=periodo_activo).select_related('materia').order_by('materia__nombre_materia')
            
            # --- Lógica de cálculo de notas (idéntica a la de la vista del panel) ---
            panel_data = []
            for estudiante in estudiantes:
                notas_ordenadas = []
                for curso in cursos:
                    estado_academico = calcular_estado_academico_curso(curso, estudiante)
                    nota_final_materia = estado_academico.get('nota_final_ponderada')
                    notas_ordenadas.append(nota_final_materia)

                notas_validas = [nota for nota in notas_ordenadas if nota is not None]
                promedio_estudiante = sum(notas_validas) / len(notas_validas) if notas_validas else None

                panel_data.append({
                    'estudiante': estudiante,
                    'notas_ordenadas': notas_ordenadas,
                    'promedio_estudiante': promedio_estudiante,
                })
            # --- Fin de la lógica de cálculo ---

            # Preparamos los datos para el DataFrame de esta hoja
            headers = ['Estudiante'] + [c.materia.nombre_materia for c in cursos] + ['Promedio Final']
            data_para_excel = []
            for data_estudiante in panel_data:
                fila = {'Estudiante': data_estudiante['estudiante'].usuario.get_full_name()}
                for i, curso in enumerate(cursos):
                    nota = data_estudiante['notas_ordenadas'][i]
                    fila[curso.materia.nombre_materia] = f"{nota:.2f}".replace('.', ',') if nota is not None else ''
                
                promedio = data_estudiante['promedio_estudiante']
                fila['Promedio Final'] = f"{promedio:.2f}".replace('.', ',') if promedio is not None else ''
                data_para_excel.append(fila)

            df = pd.DataFrame(data_para_excel, columns=headers)
            
            # Creamos una hoja con el nombre del grado (acortado si es muy largo)
            sheet_name = str(grado.nombre)[:30]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    output.seek(0)
    
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Panel_Director_{docente.usuario.username}.xlsx"'
    return response 

@login_required
def exportar_reporte_nota_minima_excel(request, curso_pk):
    """
    Exporta el reporte de nota mínima a un archivo Excel.
    """
    # Reutilizamos la misma lógica de cálculo de la vista original
    curso = get_object_or_404(Curso.objects.select_related('grado'), pk=curso_pk)
    estudiantes = Estudiante.objects.filter(grado_actual=curso.grado).select_related('usuario').order_by('usuario__last_name')
    actividades = ActividadCalificable.objects.filter(curso=curso).select_related('tipo_actividad')
    calificaciones = Calificacion.objects.filter(actividad_calificable__in=actividades)
    calificaciones_map = defaultdict(dict)
    for cal in calificaciones:
        calificaciones_map[cal.estudiante_id][cal.actividad_calificable_id] = cal.valor_numerico
    
    actividades_por_categoria = defaultdict(list)
    for actividad in actividades:
        if actividad.tipo_actividad and actividad.tipo_actividad.porcentaje:
            actividades_por_categoria[actividad.tipo_actividad].append(actividad)

    reporte_data = []
    NOTA_OBJETIVO = Decimal('3.0')

    for estudiante in estudiantes:
        puntos_acumulados = Decimal('0.0')
        porcentaje_evaluado = Decimal('0.0')
        for categoria, lista_actividades in actividades_por_categoria.items():
            notas_categoria_estudiante = [calificaciones_map.get(estudiante.pk, {}).get(act.pk) for act in lista_actividades if calificaciones_map.get(estudiante.pk, {}).get(act.pk) is not None]
            if notas_categoria_estudiante:
                promedio_categoria = sum(notas_categoria_estudiante) / len(notas_categoria_estudiante)
                porcentaje_categoria = categoria.porcentaje
                puntos_acumulados += promedio_categoria * (porcentaje_categoria / Decimal('100.0'))
                porcentaje_evaluado += porcentaje_categoria
        
        porcentaje_restante = Decimal('100.0') - porcentaje_evaluado
        estado = ""
        nota_requerida = None
        if porcentaje_restante <= 0:
            estado = "Curso finalizado"
        else:
            puntos_necesarios = (NOTA_OBJETIVO * 100 - puntos_acumulados * 100) / porcentaje_restante
            nota_requerida = puntos_necesarios.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            if nota_requerida <= 0: estado = "Aprobado"
            elif nota_requerida > 5.0: estado = "Situación Crítica"
            else: estado = "En Riesgo"

        reporte_data.append({
            'Estudiante': estudiante.usuario.get_full_name(),
            'Promedio Actual': f"{(puntos_acumulados / (porcentaje_evaluado / 100)) if porcentaje_evaluado > 0 else 0:.2f}",
            '% Evaluado': f"{porcentaje_evaluado:.2f}%",
            'Nota Requerida en % Restante': f"{nota_requerida:.2f}" if nota_requerida is not None and nota_requerida > 0 else "-",
            'Estado': estado
        })

    # Crear DataFrame y archivo Excel
    df = pd.DataFrame(reporte_data)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Nota_Minima_{curso}.xlsx"'
    df.to_excel(response, index=False)
    return response    

@login_required
def exportar_libro_de_notas_excel(request, curso_pk):
    """
    Exporta el libro de notas completo de un curso a un archivo Excel,
    INCLUYENDO LA NOTA DEFINITIVA CALCULADA.
    """
    # 1. Obtener todos los datos necesarios (igual que antes)
    curso = get_object_or_404(Curso, pk=curso_pk)
    estudiantes = Estudiante.objects.filter(grado_actual=curso.grado).select_related('usuario').order_by('usuario__last_name')
    actividades = ActividadCalificable.objects.filter(curso=curso).select_related('tipo_actividad').order_by('tipo_actividad__nombre', 'titulo')
    
    calificaciones = Calificacion.objects.filter(actividad_calificable__in=actividades)
    calificaciones_map = defaultdict(dict)
    for cal in calificaciones:
        calificaciones_map[cal.estudiante_id][cal.actividad_calificable_id] = cal.valor_numerico

    actividades_por_categoria = defaultdict(list)
    for actividad in actividades:
        if actividad.tipo_actividad and actividad.tipo_actividad.porcentaje:
            actividades_por_categoria[actividad.tipo_actividad].append(actividad)

    # 2. Preparamos los datos para el DataFrame
    data_para_excel = []
    # Añadimos 'Definitiva' a la lista de columnas
    columnas = ['Estudiante'] + [act.titulo for act in actividades] + ['Definitiva']

    for estudiante in estudiantes:
        fila = {'Estudiante': estudiante.usuario.get_full_name()}
        
        # --- Lógica de cálculo de la nota final (la misma del boletín) ---
        nota_final_curso = Decimal('0.0')
        for categoria, lista_actividades in actividades_por_categoria.items():
            notas_categoria = [
                calificaciones_map.get(estudiante.pk, {}).get(act.pk) 
                for act in lista_actividades 
                if calificaciones_map.get(estudiante.pk, {}).get(act.pk) is not None
            ]
            if notas_categoria:
                promedio_cat = sum(notas_categoria) / len(notas_categoria)
                nota_final_curso += promedio_cat * (categoria.porcentaje / Decimal('100.0'))
        
        # Llenamos la fila con las notas de cada actividad
        for actividad in actividades:
            nota = calificaciones_map.get(estudiante.pk, {}).get(actividad.pk)
            fila[actividad.titulo] = f"{nota:.2f}" if nota is not None else ""
        
        # Añadimos la nota definitiva calculada a la fila
        fila['Definitiva'] = f"{nota_final_curso:.2f}" if nota_final_curso > 0 else ""
        
        data_para_excel.append(fila)

    # 3. Crear el DataFrame y generar el archivo Excel
    df = pd.DataFrame(data_para_excel, columns=columnas)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Libro_de_Notas_{curso}.xlsx"'
    
    df.to_excel(response, index=False)
    return response


class TareasPorCalificarView(LoginRequiredMixin, View):
    template_name = 'gestion_academica/tareas_por_calificar_lista.html'

    def get(self, request, *args, **kwargs):
        docente = request.user.docente
        institucion = request.user.institucion_asociada
        periodo_activo = PeriodoAcademico.objects.filter(activo=True, institucion=institucion).first()

        entregas_pendientes = EntregaDeber.objects.none()
        intentos_pendientes = IntentoCuestionario.objects.none()

        if periodo_activo:
            cursos_del_docente = Curso.objects.filter(docentes_asignados=docente, periodo_academico=periodo_activo)
            
            # 1. Obtenemos las entregas de deberes pendientes (sin cambios)
            entregas_pendientes = EntregaDeber.objects.filter(
                deber__curso__in=cursos_del_docente,
                calificacion_obtenida__isnull=True,
                archivo_adjunto_estudiante__isnull=False
            ).select_related('deber__curso', 'estudiante__usuario').order_by('-deber__fecha_entrega')

            # 2. Obtenemos los intentos de cuestionario que requieren revisión
            # --- INICIO DE LA MODIFICACIÓN CLAVE ---
            # Ahora buscamos los intentos de cuestionarios que TENGAN preguntas de texto libre,
            # sin importar si el estudiante las contestó o no.
            intentos_pendientes = IntentoCuestionario.objects.filter(
                cuestionario__actividad_calificable__curso__in=cursos_del_docente,
                estado='FINALIZADO',
                cuestionario__preguntas__tipo='texto_libre' # <-- LÍNEA CORREGIDA
            ).select_related(
                'estudiante__usuario', 
                'cuestionario__actividad_calificable__curso'
            ).distinct().order_by('-fecha_fin')
            # --- FIN DE LA MODIFICACIÓN CLAVE ---

        context = {
            'titulo_pagina': 'Actividades Pendientes por Calificar',
            'entregas_pendientes': entregas_pendientes,
            'intentos_pendientes': intentos_pendientes
        }
        
        return render(request, self.template_name, context)
      

class CalificarEntregaView(LoginRequiredMixin, UpdateView):
    model = EntregaDeber
    form_class = CalificarEntregaForm
    template_name = 'gestion_academica/calificar_entrega_formulario.html'
    context_object_name = 'entrega'
    success_url = reverse_lazy('gestion_academica:tareas_por_calificar')

    def get_queryset(self):
        """Seguridad: Solo permite calificar entregas de los cursos del docente."""
        return EntregaDeber.objects.filter(deber__curso__docentes_asignados=self.request.user.docente)

    # El método get_form_kwargs ha sido ELIMINADO porque CalificarEntregaForm no lo necesita.

    def form_valid(self, form):
        """Se ejecuta cuando el formulario es válido, antes de guardar."""
        # Tu lógica de sincronización se mantiene aquí
        entrega = form.save(commit=False)
        entrega.fecha_calificacion = timezone.now()
        entrega.save()

        tipo_tarea, _ = TipoActividad.objects.get_or_create(
            nombre='Tareas', 
            institucion=self.request.user.institucion_asociada,
            defaults={'porcentaje': 0}
        )

        actividad_calificable, _ = ActividadCalificable.objects.get_or_create(
            titulo=entrega.deber.titulo,
            curso=entrega.deber.curso,
            defaults={
                'tipo_actividad': tipo_tarea,
                'descripcion': entrega.deber.descripcion,
                'institucion': self.request.user.institucion_asociada
            }
        )

        Calificacion.objects.update_or_create(
            estudiante=entrega.estudiante,
            actividad_calificable=actividad_calificable,
            defaults={
                'valor_numerico': Decimal(entrega.calificacion_obtenida),
                'registrada_por': self.request.user.docente,
                'observaciones': entrega.comentarios_docente,
                'institucion': self.request.user.institucion_asociada
            }
        )
        
        messages.success(self.request, f"La tarea de {self.object.estudiante} ha sido calificada y sincronizada.")
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Calificar Tarea: {self.object.deber.titulo}"
        return context   

@login_required
def descargar_plantilla_view(request):
    """
    Genera y sirve el archivo Excel de plantilla personalizado para el docente.
    """
    try:
        docente = request.user.docente
        institucion_actual = docente.institucion
        # Filtramos las materias que el docente imparte en esa institución
        materias_docente = Materia.objects.filter(
            cursos__docentes_asignados=docente,
            institucion=institucion_actual
        ).distinct()

    except (AttributeError, Docente.DoesNotExist):
        messages.error(request, "Tu usuario no está configurado como docente o no tiene una institución asignada.")
        return redirect('gestion_academica:docente_lista_descriptores')

    workbook = Workbook()
    
    # Hoja de datos oculta para las materias
    sheet_data = workbook.create_sheet(title="Data_Asignaturas")
    sheet_data.append(['CODIGO_ASIGNATURA', 'NOMBRE_ASIGNATURA'])
    for materia in materias_docente:
        sheet_data.append([materia.codigo_materia, materia.nombre_materia])

    # Hoja principal
    sheet = workbook.active
    sheet.title = "PLANTILLA_BANCO_DE_LOGROS"
    sheet.append(['CODIGO_ASIGNATURA', 'NOMBRE_ASIGNATURA', 'PERIODO', 'TEXTO_DESCRIPTOR'])

    # Listas desplegables (Validación de datos)
    if materias_docente.exists():
        dv_materia = DataValidation(type="list", formula1=f"=Data_Asignaturas!$A$2:${chr(ord('A'))}{len(materias_docente) + 1}")
        sheet.add_data_validation(dv_materia)
        dv_materia.add('A2:A1000')

    dv_periodo = DataValidation(type="list", formula1='"Primer Periodo,Segundo Periodo,Tercer Periodo,Cuarto Periodo"')
    sheet.add_data_validation(dv_periodo)
    dv_periodo.add('C2:C1000')
    
    # Columna de autocompletado para el nombre de la materia
    for row_idx in range(2, 1001):
        sheet[f'B{row_idx}'] = f'=IFERROR(VLOOKUP(A{row_idx},Data_Asignaturas!A:B,2,FALSE),"")'

    sheet_data.sheet_state = 'hidden'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_descriptores.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def gestion_descriptores_view(request):
    """
    Maneja tanto la muestra del formulario de carga (GET) como el procesamiento del archivo subido (POST).
    """
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['file']
            try:
                # Usamos pandas para leer el archivo Excel
                df = pd.read_excel(archivo)
                
                # Verificar que las columnas necesarias existen
                required_cols = ['CODIGO_ASIGNATURA', 'PERIODO', 'TEXTO_DESCRIPTOR']
                if not all(col in df.columns for col in required_cols):
                    messages.error(request, "El archivo no tiene las columnas requeridas: CODIGO_ASIGNATURA, PERIODO, TEXTO_DESCRIPTOR.")
                    return redirect('gestion_descriptores')

                institucion_actual = request.user.institucion
                docente = request.user
                logros_creados = 0
                
                for index, row in df.iterrows():
                    try:
                        # Validar que los datos no estén vacíos
                        if pd.isna(row['CODIGO_ASIGNATURA']) or pd.isna(row['PERIODO']) or pd.isna(row['TEXTO_DESCRIPTOR']):
                            continue # Saltar filas vacías

                        # Buscar la materia asegurando que pertenezca a la institución correcta
                        materia = Materia.objects.get(
                            codigo_materia=str(row['CODIGO_ASIGNATURA']).strip(),
                            institucion=institucion_actual
                        )
                        
                        # Crear el descriptor y estampar los datos automáticos
                        DescriptorLogro.objects.create(
                            materia=materia,
                            periodo_academico=str(row['PERIODO']).strip(),
                            descripcion=str(row['TEXTO_DESCRIPTOR']).strip(),
                            creado_por=docente,
                            institucion=institucion_actual
                        )
                        logros_creados += 1
                        
                    except Materia.DoesNotExist:
                        messages.warning(request, f"Fila {index + 2}: La materia con código '{row['CODIGO_ASIGNATURA']}' no existe o no pertenece a tu institución.")
                    except Exception as e:
                        messages.error(request, f"Fila {index + 2}: Ocurrió un error inesperado. {e}")
                
                if logros_creados > 0:
                    messages.success(request, f"¡Éxito! Se han importado {logros_creados} descriptores de logro.")
                else:
                    messages.warning(request, "El proceso finalizó, pero no se importó ningún descriptor nuevo. Revisa los mensajes de advertencia.")

            except Exception as e:
                messages.error(request, f"No se pudo procesar el archivo. Error: {e}")
            
            return redirect('gestion_descriptores')
    else:
        form = UploadFileForm()
    
    return render(request, 'tu_app/gestion_descriptores.html', {'form': form})        

@login_required
def reporte_riesgo_global_view(request):
    user = request.user
    
    # --- Tu lógica de permisos originales (se mantiene intacta) ---
    if user.is_staff:
        pass # Acceso permitido
    elif hasattr(user, 'docente'):
        docente = user.docente
        periodo_activo = PeriodoAcademico.objects.filter(activo=True, institucion=user.institucion_asociada).first()
        if not periodo_activo or not DirectorCurso.objects.filter(docente=docente, periodo_academico=periodo_activo).exists():
            messages.error(request, "Acceso denegado. Esta sección es solo para personal directivo o directores de curso.")
            return redirect('gestion_academica:dashboard_docente')
    else:
        messages.error(request, "No tienes permiso para acceder a esta sección.")
        return redirect('gestion_academica:inicio_academico')

    # --- Obtención de filtros ---
    periodos = PeriodoAcademico.objects.filter(institucion=user.institucion_asociada).order_by('-año_escolar', 'nombre') if not user.is_superuser else PeriodoAcademico.objects.all().order_by('-año_escolar', 'nombre')
    grados = Grado.objects.filter(institucion=user.institucion_asociada).order_by('nombre') if not user.is_superuser else Grado.objects.all().order_by('nombre')
    
    periodo_seleccionado_id = request.GET.get('periodo')
    grado_seleccionado_id = request.GET.get('grado')

    # --- LÓGICA DE CONSULTA CORREGIDA ---
    # 1. Empezamos con el historial completo como base.
    predicciones_qs = PrediccionRiesgoEstudiante.objects.select_related(
        'estudiante__usuario', 
        'materia', 
        'estudiante__grado_actual', 
        'analisis__periodo_academico'
    ).order_by('estudiante__usuario__last_name', 'materia__nombre_materia')

    # 2. Aplicamos filtros si existen.
    if periodo_seleccionado_id:
        predicciones_qs = predicciones_qs.filter(analisis__periodo_academico_id=periodo_seleccionado_id)
    if grado_seleccionado_id:
        predicciones_qs = predicciones_qs.filter(estudiante__grado_actual_id=grado_seleccionado_id)

    # --- Contexto para la plantilla ---
    context = {
        'titulo_pagina': "HALU Sentinel - Reporte de Riesgo",
        'predicciones': predicciones_qs, # Pasamos la lista plana y filtrada
        'periodos': periodos,
        'grados': grados,
        'periodo_seleccionado_id': periodo_seleccionado_id,
        'grado_seleccionado_id': grado_seleccionado_id,
    }

    return render(request, 'gestion_academica/reporte_riesgo_global.html', context)


# VISTA PARA EXPORTAR EL REPORTE (CON PERMISOS CORREGIDOS Y FINALES)
@login_required
def exportar_reporte_riesgo_global_view(request):
    """
    Exporta a Excel el reporte global. Versión con permisos definitivos.
    """
    periodo_id = request.GET.get('periodo')
    grado_id = request.GET.get('grado')

    if not (periodo_id and grado_id):
        return HttpResponse("Faltan filtros.", status=400)
    
    try:
        periodo = get_object_or_404(PeriodoAcademico, pk=periodo_id)
        grado = get_object_or_404(Grado, pk=grado_id)
        
        # --- LÓGICA DE PERMISOS DEFINITIVA ---
        user = request.user
        user_inst = getattr(user, 'institucion_asociada', None)
        es_superusuario = user.is_superuser
        es_admin_o_coordinador = (user.is_staff and user_inst and user_inst.pk == periodo.institucion.pk and user.rol in ['administrador', 'coordinador'])
        es_director_del_grupo = False
        if hasattr(user, 'docente'):
            es_director_del_grupo = DirectorCurso.objects.filter(docente=user.docente, grado=grado, periodo_academico=periodo).exists()

        if not (es_superusuario or es_admin_o_coordinador or es_director_del_grupo):
            return HttpResponse("Acceso denegado.", status=403)
        # --- FIN DE LA LÓGICA DE PERMISOS ---
        
        # El resto del código de exportación...
        institucion = periodo.institucion
        estudiantes = Estudiante.objects.filter(grado_actual=grado, institucion=institucion)
        cursos = Curso.objects.filter(grado=grado, periodo_academico=periodo)
        
        alertas_riesgo = []
        for estudiante in estudiantes:
            for curso in cursos:
                resultado_riesgo = analizar_riesgo_academico_curso(curso, estudiante)
                if resultado_riesgo['estado'] in ESTADOS_RIESGO_ACADEMICO_CURSO:
                    alertas_riesgo.append({
                        'Estudiante': estudiante.usuario.get_full_name(),
                        'Materia en Riesgo': curso.materia.nombre_materia,
                        'Estado Actual': resultado_riesgo['estado'],
                        'Nota Requerida': f"{resultado_riesgo['nota_requerida']:.2f}".replace('.', ',') if resultado_riesgo['nota_requerida'] is not None else ''
                    })
        
        data_ordenada = sorted(alertas_riesgo, key=lambda x: x['Estudiante'])
        df = pd.DataFrame(data_ordenada)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Riesgo_{grado.nombre}_{periodo.nombre}.xlsx"'
        df.to_excel(response, index=False)
        return response

    except Exception as e:
        return HttpResponse(f"Ocurrió un error inesperado al generar el reporte: {e}", status=500)       

@login_required
def asistencia_diaria_data_api(request):
    """
    API que devuelve los datos de asistencia del día en formato JSON 
    para la actualización en tiempo real del dashboard.
    """
    if not request.user.is_staff:
        return JsonResponse({'error': 'Acceso denegado'}, status=403)

    fecha_str = request.GET.get('fecha', timezone.localdate().strftime('%Y-%m-%d'))
    try:
        fecha_a_consultar = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        fecha_a_consultar = timezone.localdate()

    institucion = getattr(request.user, 'institucion_asociada', InstitucionEducativa.objects.first())
    if not institucion:
        return JsonResponse({'error': 'Institución no encontrada'}, status=404)

    # Usamos el filtro de fecha robusto
    registros_del_dia = RegistroAsistencia.objects.filter(
        fecha__startswith=fecha_a_consultar.strftime('%Y-%m-%d'), 
        estudiante__institucion=institucion
    ).select_related('estudiante', 'curso__materia', 'registrado_por')

    # Creamos un diccionario con el estado de cada estudiante que SÍ tiene registro
    data = {
       registro.estudiante.pk: {
           "estado": registro.get_estado_display(),
           "info_registro": f"{registro.registrado_por.get_full_name() if registro.registrado_por else 'Sistema'} ({registro.curso.materia.nombre_materia if registro.curso else 'General'})",
           "hora_registro": timezone.localtime(registro.fecha).strftime('%I:%M %p'),
           "aula": registro.aula.nombre if registro.aula else (registro.curso.aula.nombre if registro.curso and registro.curso.aula else "Sin Aula")
       }
       for registro in registros_del_dia
    }
    
    return JsonResponse(data)


@login_required
@permission_required('gestion_academica.view_registroasistencia')
def asistencia_diaria_admin_view(request):
    """
    Muestra un dashboard con el estado de asistencia, usando un filtro de fecha robusto.
    """
    if not request.user.is_staff:
        messages.error(request, "Acceso denegado.")
        return redirect('gestion_academica:inicio_academico')

    fecha_str = request.GET.get('fecha', timezone.localdate().strftime('%Y-%m-%d'))
    try:
        fecha_a_consultar = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        fecha_a_consultar = timezone.localdate()

    institucion = getattr(request.user, 'institucion_asociada', InstitucionEducativa.objects.first())

    estudiantes_activos = Estudiante.objects.filter(
        institucion=institucion, usuario__is_active=True
    ).select_related('usuario', 'grado_actual').order_by('grado_actual__nombre', 'usuario__last_name')
    
    # --- FILTRO DE FECHA ROBUSTO Y DEFINITIVO ---
    inicio_del_dia = timezone.make_aware(datetime.combine(fecha_a_consultar, time.min))
    fin_del_dia = timezone.make_aware(datetime.combine(fecha_a_consultar, time.max))
    
    registros_del_dia = RegistroAsistencia.objects.filter(
        fecha__range=(inicio_del_dia, fin_del_dia), 
        estudiante__institucion=institucion
    ).select_related('estudiante', 'curso__materia', 'registrado_por')

    registros_por_estudiante = {registro.estudiante.pk: registro for registro in registros_del_dia}

    asistencia_por_grado = defaultdict(list)
    for estudiante in estudiantes_activos:
        if estudiante.grado_actual:
            asistencia_por_grado[estudiante.grado_actual.nombre].append({
                'estudiante': estudiante,
                'registro': registros_por_estudiante.get(estudiante.pk)
            })

    context = {
        'titulo_pagina': f"Control de Asistencia - {fecha_a_consultar.strftime('%d de %B de %Y')}",
        'fecha_consultada': fecha_a_consultar,
        'asistencia_data': dict(asistencia_por_grado),
    }
    
    return render(request, 'gestion_academica/admin_asistencia_diaria.html', context)

@login_required
@permission_required('gestion_academica.view_analisisriesgo') # O un permiso de admin/coordinador
def dashboard_riesgo_academico(request):
    """
    Muestra el dashboard principal de Analítica Predictiva, presentando
    un resumen del último análisis de riesgo realizado.
    """
    if not (request.user.is_superuser or (request.user.is_staff and request.user.rol in ['administrativo', 'coordinador'])):
        messages.error(request, "Acceso denegado. Esta sección es solo para personal directivo.")
        return redirect('gestion_academica:inicio_academico')

    # Buscamos el último análisis de riesgo completado
    ultimo_analisis = AnalisisRiesgo.objects.order_by('-fecha_analisis').first()

    predicciones = []
    riesgo_alto_count = 0
    riesgo_medio_count = 0

    if ultimo_analisis:
        # Obtenemos todas las predicciones asociadas a ese análisis
        predicciones = ultimo_analisis.predicciones.select_related(
            'estudiante__usuario', 'materia'
        ).order_by('nivel_riesgo', 'estudiante__usuario__last_name')
        
        riesgo_alto_count = predicciones.filter(nivel_riesgo='ALTO').count()
        riesgo_medio_count = predicciones.filter(nivel_riesgo='MEDIO').count()

    context = {
        'titulo_pagina': "HALU Sentinel - Riesgo Académico",
        'ultimo_analisis': ultimo_analisis,
        'predicciones': predicciones,
        'riesgo_alto_count': riesgo_alto_count,
        'riesgo_medio_count': riesgo_medio_count,
    }
    return render(request, 'gestion_academica/dashboard_riesgo_academico.html', context)    


@login_required
def citar_acudiente_view(request, prediccion_pk):
    """
    Envía un correo al acudiente, obteniendo dinámicamente las credenciales
    de correo desde el modelo de la Institución.
    """
    # La lógica de permisos y obtención de objetos se mantiene igual
    if not (request.user.is_superuser or (request.user.is_staff and request.user.rol in ['coordinador', 'administrador'])):
        messages.error(request, "No tienes permiso para realizar esta acción.")
        return redirect('gestion_academica:reporte_riesgo_academico')

    prediccion = get_object_or_404(PrediccionRiesgoEstudiante, pk=prediccion_pk)
    estudiante = prediccion.estudiante
    institucion = estudiante.institucion
    familiar = estudiante.familiares.select_related('usuario').first()

    if not (familiar and familiar.usuario.email):
        messages.error(request, f"El estudiante {estudiante} no tiene un familiar con correo electrónico registrado.")
        return redirect('gestion_academica:reporte_riesgo_academico')

    # --- INICIO DE LA LÓGICA CORREGIDA ---

    # 1. Verificamos si la institución tiene credenciales de correo configuradas
    if not (institucion.email_host_user and institucion.email_host_password):
        messages.error(request, f"La institución '{institucion.nombre}' no tiene configuradas las credenciales para enviar correos.")
        return redirect('gestion_academica:reporte_riesgo_academico')

    # 2. Renderizamos las plantillas como antes
    context_email = {'familiar': familiar, 'estudiante': estudiante, 'prediccion': prediccion, 'institucion': institucion}
    asunto = render_to_string('gestion_academica/email/citacion_asunto.txt', context_email).strip()
    cuerpo_texto = render_to_string('gestion_academica/email/citacion_cuerpo.txt', context_email)
    cuerpo_html = render_to_string('gestion_academica/email/citacion_cuerpo.html', context_email)
    
    try:
        # 3. Creamos una conexión SMTP personalizada con las credenciales de la institución
        connection = get_connection(
            host=institucion.email_host,
            port=institucion.email_port,
            username=institucion.email_host_user,
            password=institucion.email_host_password,
            use_tls=institucion.email_use_tls
        )

        # 4. Creamos el correo usando esas credenciales
        email = EmailMultiAlternatives(
            subject=asunto,
            body=cuerpo_texto,
            from_email=institucion.email_host_user, # Usamos el correo de la institución como remitente
            to=[familiar.usuario.email],
            connection=connection # ¡Le decimos al correo que use nuestra conexión personalizada!
        )
        email.attach_alternative(cuerpo_html, "text/html")
        email.send(fail_silently=False)
        
        messages.success(request, f"Citación enviada exitosamente al correo de {familiar.usuario.get_full_name()}.")

    except Exception as e:
        messages.error(request, f"No se pudo enviar el correo. Revisa la configuración SMTP de la institución. Error: {e}")

    return redirect('gestion_academica:reporte_riesgo_academico') 

@login_required
def ejecutar_analisis_riesgo_view(request):
    """
    Ejecuta el comando 'calcular_riesgo_academico' y redirige al reporte.
    """
    if not (request.user.is_superuser or (request.user.is_staff and request.user.rol in ['administrador', 'coordinador'])):
        messages.error(request, "No tienes permiso para ejecutar esta acción.")
        return redirect('gestion_academica:dashboard_coordinador')

    try:
        call_command('calcular_riesgo_academico')
        messages.success(request, "El análisis de riesgo se ha completado exitosamente.")
    except Exception as e:
        messages.error(request, f"Ocurrió un error al ejecutar el análisis: {e}")

    return redirect('gestion_academica:reporte_riesgo_academico')  

@login_required
def notificar_docente_view(request, prediccion_pk):
    """
    Crea notificaciones internas para el personal relevante, asegurando
    que la acción solo pueda ser ejecutada por personal de la misma institución.
    """
    if not (request.user.is_superuser or (request.user.is_staff and request.user.rol in ['coordinador', 'administrador'])):
        messages.error(request, "No tienes permiso para realizar esta acción.")
        return redirect('gestion_academica:reporte_riesgo_academico')

    # --- ✅ INICIO DE LA MEJORA DE SEGURIDAD ---
    # Obtenemos la predicción, pero nos aseguramos de que pertenezca a la
    # institución del usuario que está realizando la acción.
    prediccion = get_object_or_404(
        PrediccionRiesgoEstudiante.objects.select_related(
            'estudiante__usuario', 'materia', 'analisis__periodo_academico'
        ), 
        pk=prediccion_pk,
        institucion=request.user.institucion_asociada # <-- Filtro de seguridad
    )
    # --- FIN DE LA MEJORA DE SEGURIDAD ---

    estudiante = prediccion.estudiante
    periodo = prediccion.analisis.periodo_academico
    
    destinatarios = set()

    # (El resto de tu lógica para encontrar destinatarios y enviar la notificación no necesita cambios)
    director_de_grupo = DirectorCurso.objects.filter(
        grado=estudiante.grado_actual, 
        periodo_academico=periodo
    ).select_related('docente__usuario').first()
    
    if director_de_grupo:
        destinatarios.add(director_de_grupo.docente.usuario)

    if prediccion.materia:
        mensaje_notificacion = f"Alerta de Riesgo: {estudiante.usuario.get_full_name()} en {prediccion.materia.nombre_materia}."
        
        curso_especifico = Curso.objects.filter(
            materia=prediccion.materia, 
            grado=estudiante.grado_actual, 
            periodo_academico=periodo
        ).prefetch_related('docentes_asignados__usuario').first()
        
        if curso_especifico:
            for docente in curso_especifico.docentes_asignados.all():
                destinatarios.add(docente.usuario)
    else:
        mensaje_notificacion = f"Alerta de Seguimiento: {estudiante.usuario.get_full_name()} requiere atención. Motivo: {prediccion.factores_influyentes}"

    if not destinatarios:
        messages.warning(request, f"No se pudo enviar la notificación. No se encontró un director de grupo asignado para {estudiante.grado_actual}.")
        return redirect('gestion_academica:reporte_riesgo_academico')

    enlace_perfil_estudiante = request.build_absolute_uri(
        reverse('gestion_academica:detalle_estudiante', kwargs={'pk': estudiante.pk})
    )
    
    for usuario_destinatario in destinatarios:
        Notificacion.objects.create(
            destinatario=usuario_destinatario,
            mensaje=mensaje_notificacion,
            enlace=enlace_perfil_estudiante,
            institucion=estudiante.institucion
        )

    messages.success(request, f"Notificación enviada exitosamente a {len(destinatarios)} destinatario(s).")
    return redirect('gestion_academica:reporte_riesgo_academico')   

@login_required
@permission_required('gestion_academica.add_familiar') # Asegúrate de que los admins tengan este permiso
def crear_familiar(request):
    """
    Vista para crear un nuevo Familiar junto con su cuenta de Usuario.
    """
    if request.method == 'POST':
        # Pasamos el 'request' a los formularios para la lógica de multi-institución
        usuario_form = CustomUserCreationForm(request.POST, prefix="usr", request=request)
        familiar_form = FamiliarForm(request.POST, prefix="fam", request=request)
        
        if usuario_form.is_valid() and familiar_form.is_valid():
            # Creamos el usuario pero no lo guardamos todavía
            usuario = usuario_form.save(commit=False)
            usuario.rol = 'familiar' # Asignamos el rol correcto
            usuario.save() # Ahora guardamos el usuario

            # Creamos el perfil de familiar
            familiar = familiar_form.save(commit=False)
            familiar.usuario = usuario # Vinculamos el usuario recién creado
            familiar.institucion = usuario.institucion_asociada # Asignamos la misma institución
            familiar.save()
            
            # El método save() de un ModelMultipleChoiceField requiere que el objeto principal ya exista
            familiar_form.save_m2m()

            messages.success(request, f"Familiar '{usuario.get_full_name()}' creado exitosamente.")
            return redirect('admin:gestion_academica_familiar_changelist') # Redirige a la lista del admin
    else:
        usuario_form = CustomUserCreationForm(prefix="usr", request=request)
        familiar_form = FamiliarForm(prefix="fam", request=request)

    context = {
        'usuario_form': usuario_form,
        'familiar_form': familiar_form,
        'titulo_pagina': 'Registrar Nuevo Familiar'
    }
    return render(request, 'gestion_academica/familiar_formulario.html', context)    

@login_required
def dashboard_coordinador_view(request):
    """
    Muestra el panel principal para Coordinadores y Administradores.
    VERSIÓN ACTUALIZADA: Incluye un reporte de cursos filtrado por grado.
    """
    user = request.user
    user_inst = getattr(user, 'institucion_asociada', None)

    if not (user.is_superuser or (user.is_staff and user.rol in ['coordinador', 'administrador'])):
        messages.error(request, "Acceso denegado.")
        return redirect('gestion_academica:inicio_academico')
    
    # --- Inicialización de variables ---
    alertas_count = 0
    cantidad_convivencia = 0
    porcentaje_asistencia_hoy = 0
    grados_institucion = Grado.objects.none()
    cursos_filtrados = Curso.objects.none()
    planes_pendientes_count = 0
    grado_seleccionado_id = request.GET.get('grado') # Capturamos el ID del grado desde la URL

    if user_inst:
        # Lógica de KPIs existente (sin cambios)
        alertas_count = AnotacionObservador.objects.filter(estudiante__institucion=user_inst, requiere_revision=True).count()
        cantidad_convivencia = AnotacionObservador.objects.filter(estudiante__institucion=user_inst).exclude(Q(tipo_situacion_ia='NINGUNO') | Q(tipo_situacion_ia__isnull=True)).count()
        
        total_estudiantes_activos = Estudiante.objects.filter(institucion=user_inst, activo=True).count()
        presentes_hoy = RegistroAsistencia.objects.filter(institucion=user_inst, fecha__date=timezone.now().date(), estado='PRESENTE', curso__isnull=True).values('estudiante').distinct().count()
        if total_estudiantes_activos > 0:
            porcentaje_asistencia_hoy = round((presentes_hoy / total_estudiantes_activos) * 100)
        
        # --- Planes semanales pendientes de revisión ---
        planes_pendientes_count = PlanSemanal.objects.filter(
            institucion=user_inst, estado='ENVIADO'
        ).count()

        # --- Lógica añadida para el nuevo reporte ---
        grados_institucion = Grado.objects.filter(institucion=user_inst).order_by('orden')
        
        if grado_seleccionado_id:
            cursos_filtrados = Curso.objects.filter(
                periodo_academico__activo=True,
                grado_id=grado_seleccionado_id,
                institucion=user_inst 
            ).select_related('materia', 'grado', 'aula').prefetch_related('docentes_asignados__usuario').order_by('materia__nombre_materia')
        # --- Fin de la lógica añadida ---

    eleccion_mas_reciente = None
    votos_data = []
    total_votos = 0
    total_votantes_elegibles = 0
    eleccion_activa = False
    porcentaje_participacion = 0

    if user_inst:
        eleccion_mas_reciente = (
            Eleccion.objects.filter(institucion=user_inst).order_by('-fecha_inicio').first()
        )

    if eleccion_mas_reciente and user_inst:
        now = timezone.now()
        eleccion_activa = eleccion_mas_reciente.fecha_inicio <= now <= eleccion_mas_reciente.fecha_fin
        total_votos = Voto.objects.filter(eleccion=eleccion_mas_reciente).count()
        total_votantes_elegibles = Estudiante.objects.filter(
            institucion=user_inst, activo=True
        ).count()
        if total_votantes_elegibles > 0:
            porcentaje_participacion = round((total_votos / total_votantes_elegibles) * 100)

        for candidato in eleccion_mas_reciente.candidatos.select_related(
            'estudiante__usuario', 'estudiante__grado_actual'
        ).all():
            votos_count = candidato.votos_recibidos.count()
            porcentaje = round((votos_count / total_votos * 100)) if total_votos > 0 else 0
            votos_data.append({
                'candidato': candidato,
                'votos': votos_count,
                'porcentaje': porcentaje,
            })
        votos_data.sort(key=lambda x: x['votos'], reverse=True)

    # 3. Preparamos el contexto
    _ph_grado_api = 2_147_483_647
    context = {
        'titulo_pagina': 'Panel de Coordinación Académica',
        'usuario': user,
        'alertas_bienestar_count': alertas_count,
        'cantidad_convivencia': cantidad_convivencia,
        'porcentaje_asistencia_hoy': porcentaje_asistencia_hoy,
        'grados_institucion': grados_institucion,
        'cursos_filtrados': cursos_filtrados,
        'grado_seleccionado_id': grado_seleccionado_id,
        'eleccion_mas_reciente': eleccion_mas_reciente,
        'votos_data': votos_data,
        'total_votos': total_votos,
        'total_votantes_elegibles': total_votantes_elegibles,
        'eleccion_activa': eleccion_activa,
        'porcentaje_participacion': porcentaje_participacion,
        'api_cursos_por_grado_json': json.dumps({
            'pattern': reverse(
                'gestion_academica:api_get_cursos_por_grado',
                kwargs={'grado_id': _ph_grado_api},
            ),
            'ph': str(_ph_grado_api),
        }),
        'planes_pendientes_count': planes_pendientes_count,
    }
    
    return render(request, 'gestion_academica/dashboard_coordinador.html', context)

@login_required
def get_cursos_por_grado_partial(request, grado_id):
    """
    Vista parcial que devuelve solo el HTML de la tabla de cursos para un grado específico.
    Diseñada para ser llamada por AJAX.
    """
    user_inst = getattr(request.user, 'institucion_asociada', None)
    
    cursos_filtrados = Curso.objects.filter(
        periodo_academico__activo=True,
        grado_id=grado_id,
        institucion=user_inst 
    ).select_related('materia', 'grado', 'aula').prefetch_related('docentes_asignados__usuario').order_by('materia__nombre_materia')
    
    context = {
        'cursos_filtrados': cursos_filtrados,
    }
    # Apuntamos a una nueva plantilla que solo contiene la tabla
    return render(request, 'gestion_academica/partials/tabla_cursos_coordinador.html', context)


@login_required
def lista_notificaciones_view(request):
    """
    Muestra todas las notificaciones del usuario y las marca como leídas.
    """
    # Buscamos todas las notificaciones para el usuario logueado, de la más nueva a la más antigua.
    notificaciones = Notificacion.objects.filter(destinatario=request.user).order_by('-fecha_creacion')
    
    # Marcamos todas las notificaciones no leídas como leídas.
    # El método .update() es muy eficiente para actualizar muchos objetos a la vez.
    notificaciones.filter(leido=False).update(leido=True, fecha_leido=timezone.now())
    
    context = {
        'titulo_pagina': "Mis Notificaciones",
        'notificaciones': notificaciones
    }
    
    return render(request, 'gestion_academica/notificaciones_lista.html', context)        

def dashboard_bienestar_view(request):
    """
    HALU Sentinel — Dashboard principal de convivencia escolar.
    Muestra KPIs, casos activos por tipo/estado y análisis IA de comportamiento.
    Acceso: superusuario, coordinador, psicólogo.
    """
    from ..models import CasoConvivencia

    if not (request.user.is_superuser or (
        request.user.is_staff and request.user.rol in ['coordinador', 'psicologo', 'rector']
    )):
        messages.error(request, "Acceso denegado. Esta sección es confidencial.")
        return redirect('gestion_academica:inicio_academico')

    institucion = request.user.institucion_asociada

    # ── KPIs ──────────────────────────────────────────────────────────
    casos_qs = CasoConvivencia.objects.filter(institucion=institucion)
    kpi_abiertos       = casos_qs.filter(estado=CasoConvivencia.Estado.ABIERTO).count()
    kpi_seguimiento    = casos_qs.filter(estado=CasoConvivencia.Estado.EN_SEGUIMIENTO).count()
    kpi_vencidos       = casos_qs.filter(estado=CasoConvivencia.Estado.VENCIDO).count()
    kpi_tipo3_activos  = casos_qs.filter(
        tipo_situacion=CasoConvivencia.TipoSituacion.TIPO_III,
        estado__in=[CasoConvivencia.Estado.ABIERTO, CasoConvivencia.Estado.EN_SEGUIMIENTO, CasoConvivencia.Estado.VENCIDO]
    ).count()
    from django.utils import timezone as _tz
    inicio_mes = _tz.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    kpi_cerrados_mes   = casos_qs.filter(estado=CasoConvivencia.Estado.CERRADO, fecha_cierre__gte=inicio_mes).count()

    # ── Casos activos (tabla principal) ───────────────────────────────
    casos_activos = (
        casos_qs
        .filter(estado__in=[
            CasoConvivencia.Estado.ABIERTO,
            CasoConvivencia.Estado.EN_SEGUIMIENTO,
            CasoConvivencia.Estado.VENCIDO,
        ])
        .select_related('responsable', 'anotacion_origen__estudiante__usuario')
        .prefetch_related('involucrados__estudiante__usuario')
        .order_by('-tipo_situacion', 'fecha_limite')
    )

    # ── Alertas sin caso (anotaciones sin caso formal, Tipo II/III) ──
    alertas_sin_caso = list(
        AnotacionObservador.objects.filter(
            requiere_revision=True,
            institucion=institucion,
            caso_convivencia__isnull=True,
        ).select_related('estudiante__usuario', 'registrado_por').order_by('-fecha_hora')[:20]
    )

    # ── Análisis comportamental IA ────────────────────────────────────
    resumenes_ia = AnalisisComportamientoIA.objects.filter(
        institucion=institucion
    ).select_related('estudiante__usuario').order_by('-fecha_analisis')[:10]

    context = {
        'titulo_pagina': "HALU Sentinel — Ruta de Convivencia Escolar",
        # Lista estructurada para los KPI chips (móvil-friendly)
        'kpis': [
            {'valor': kpi_abiertos,      'label': 'Abiertos',      'color': 'warning'},
            {'valor': kpi_seguimiento,   'label': 'Seguimiento',   'color': 'info'},
            {'valor': kpi_vencidos,      'label': 'Vencidos',      'color': 'danger'},
            {'valor': kpi_tipo3_activos, 'label': 'Tipo III',      'color': 'dark'},
            {'valor': kpi_cerrados_mes,  'label': 'Cerrados/mes',  'color': 'success'},
            {'valor': len(alertas_sin_caso), 'label': 'Sin expediente', 'color': 'secondary'},
        ],
        'casos_activos':    casos_activos,
        'alertas_sin_caso': alertas_sin_caso,
        'resumenes_ia':     resumenes_ia,
        'CasoConvivencia':  CasoConvivencia,
    }
    return render(request, 'gestion_academica/dashboard_bienestar.html', context)


@login_required
def detalle_alerta_bienestar_view(request, pk):
    """Detalle de una anotación con requiere_revision=True (sin caso formal)."""
    if not (request.user.is_superuser or (
        request.user.is_staff and request.user.rol in ['coordinador', 'psicologo', 'rector']
    )):
        messages.error(request, "Acceso denegado a esta sección.")
        return redirect('gestion_academica:dashboard_coordinador')

    alerta = get_object_or_404(
        AnotacionObservador.objects.select_related('estudiante__usuario', 'registrado_por'),
        pk=pk, requiere_revision=True,
    )
    context = {
        'titulo_pagina': "Detalle de Alerta de Bienestar",
        'alerta': alerta,
    }
    return render(request, 'gestion_academica/detalle_alerta_bienestar.html', context)


# ── HALU SENTINEL: gestión de casos ────────────────────────────────────── #

def _sentinel_permiso(request):
    return request.user.is_superuser or (
        request.user.is_staff
        and request.user.rol in ['coordinador', 'psicologo', 'rector', 'administrador']
    )


@login_required
def detalle_caso_convivencia(request, pk):
    """
    Expediente completo de un CasoConvivencia:
    - Línea de tiempo de acciones
    - Formulario para registrar nueva acción
    - Cambio de estado
    - Enlace a generar acta PDF
    """
    from ..models import CasoConvivencia, AccionCaso, InvolucradoCaso
    from django import forms as django_forms

    if not _sentinel_permiso(request):
        messages.error(request, "Acceso denegado.")
        return redirect('gestion_academica:inicio_academico')

    caso = get_object_or_404(
        CasoConvivencia.objects.select_related(
            'responsable', 'anotacion_origen__estudiante__usuario',
            'anotacion_origen__registrado_por', 'institucion',
        ).prefetch_related(
            'acciones__ejecutado_por',
            'involucrados__estudiante__usuario',
        ),
        pk=pk,
        institucion=request.user.institucion_asociada,
    )

    # ── Formulario inline para nueva acción ───────────────────────────
    class AccionForm(django_forms.ModelForm):
        class Meta:
            model = AccionCaso
            fields = ['tipo_accion', 'descripcion', 'evidencia']
            widgets = {
                'tipo_accion': django_forms.Select(attrs={
                    'class': 'form-select form-select-sm w-100',
                }),
                'descripcion': django_forms.Textarea(attrs={
                    'rows': 3,
                    'class': 'form-control form-control-sm w-100',
                    'placeholder': 'Describe la acción realizada…',
                }),
                'evidencia': django_forms.ClearableFileInput(attrs={
                    'class': 'form-control form-control-sm w-100',
                }),
            }

    # ── Formulario de estado ──────────────────────────────────────────
    class EstadoForm(django_forms.ModelForm):
        class Meta:
            model = CasoConvivencia
            fields = ['estado', 'responsable', 'resolucion_final']
            widgets = {
                'estado': django_forms.Select(attrs={
                    'class': 'form-select form-select-sm w-100',
                }),
                'responsable': django_forms.Select(attrs={
                    'class': 'form-select form-select-sm w-100',
                }),
                'resolucion_final': django_forms.Textarea(attrs={
                    'rows': 3,
                    'class': 'form-control form-control-sm w-100',
                }),
            }

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            inst = caso.institucion
            self.fields['responsable'].queryset = (
                Usuario.objects.filter(
                    institucion_asociada=inst, is_staff=True,
                    rol__in=['coordinador', 'administrador', 'psicologo'],
                )
            )

    if request.method == 'POST':
        accion = request.POST.get('accion_form')

        if accion == 'nueva_accion':
            form_accion = AccionForm(request.POST, request.FILES)
            form_estado = EstadoForm(instance=caso)
            if form_accion.is_valid():
                nueva = form_accion.save(commit=False)
                nueva.caso = caso
                nueva.ejecutado_por = request.user
                nueva.save()
                # Cambiar estado a EN_SEGUIMIENTO si estaba ABIERTO
                if caso.estado == CasoConvivencia.Estado.ABIERTO:
                    caso.estado = CasoConvivencia.Estado.EN_SEGUIMIENTO
                    caso.save(update_fields=['estado'])
                messages.success(request, "Acción registrada en el expediente.")
                return redirect('gestion_academica:detalle_caso_convivencia', pk=pk)

        elif accion == 'cambiar_estado':
            form_estado = EstadoForm(request.POST, instance=caso)
            form_accion = AccionForm()
            if form_estado.is_valid():
                caso_actualizado = form_estado.save(commit=False)
                if caso_actualizado.estado == CasoConvivencia.Estado.CERRADO:
                    from django.utils import timezone as _tz
                    caso_actualizado.fecha_cierre = _tz.now()
                caso_actualizado.save()
                messages.success(request, f"Estado actualizado a: {caso.get_estado_display()}")
                return redirect('gestion_academica:detalle_caso_convivencia', pk=pk)
        else:
            form_accion = AccionForm()
            form_estado = EstadoForm(instance=caso)
    else:
        form_accion = AccionForm()
        form_estado = EstadoForm(instance=caso)

    context = {
        'titulo_pagina': f"Caso {caso.radicado}",
        'caso': caso,
        'form_accion': form_accion,
        'form_estado': form_estado,
        'CasoConvivencia': CasoConvivencia,
        'AccionCaso': AccionCaso,
    }
    return render(request, 'gestion_academica/detalle_caso_convivencia.html', context)


@login_required
def abrir_caso_manual(request, anotacion_pk):
    """
    Crea manualmente un CasoConvivencia desde una AnotacionObservador
    que no fue clasificada como Tipo II/III por la IA (ej: Tipo I explícito).
    """
    from ..models import CasoConvivencia, InvolucradoCaso
    from datetime import timedelta

    if not _sentinel_permiso(request):
        messages.error(request, "Acceso denegado.")
        return redirect('gestion_academica:inicio_academico')

    anotacion = get_object_or_404(
        AnotacionObservador,
        pk=anotacion_pk,
        institucion=request.user.institucion_asociada,
    )

    # Evitar duplicados
    if hasattr(anotacion, 'caso_convivencia') and anotacion.caso_convivencia:
        messages.info(request, f"Ya existe el caso {anotacion.caso_convivencia.radicado}.")
        return redirect('gestion_academica:detalle_caso_convivencia', pk=anotacion.caso_convivencia.pk)

    tipo = anotacion.tipo_situacion_ia or 'TIPO I'
    fecha_limite = timezone.now() + timedelta(days=7 if tipo != 'TIPO III' else 1)

    caso = CasoConvivencia.objects.create(
        institucion=anotacion.institucion,
        tipo_situacion=tipo,
        anotacion_origen=anotacion,
        descripcion_detalle=anotacion.descripcion,
        protocolo_ia=anotacion.acciones_protocolo_ia or '',
        fecha_limite=fecha_limite,
        responsable=request.user,
    )
    InvolucradoCaso.objects.create(
        caso=caso, estudiante=anotacion.estudiante,
        rol=CasoConvivencia.RolInvolucrado.VICTIMA,
    )
    messages.success(request, f"Caso {caso.radicado} abierto correctamente.")
    return redirect('gestion_academica:detalle_caso_convivencia', pk=caso.pk)


@login_required
def acta_caso_pdf(request, pk):
    """Genera el acta formal del caso en PDF usando WeasyPrint."""
    from ..models import CasoConvivencia
    try:
        from weasyprint import HTML
        from django.template.loader import render_to_string
        import io
        from django.http import HttpResponse
    except ImportError:
        messages.error(request, "WeasyPrint no está instalado.")
        return redirect('gestion_academica:detalle_caso_convivencia', pk=pk)

    if not _sentinel_permiso(request):
        messages.error(request, "Acceso denegado.")
        return redirect('gestion_academica:inicio_academico')

    caso = get_object_or_404(
        CasoConvivencia.objects.select_related('responsable', 'institucion')
        .prefetch_related('acciones__ejecutado_por', 'involucrados__estudiante__usuario'),
        pk=pk, institucion=request.user.institucion_asociada,
    )

    html_string = render_to_string('gestion_academica/acta_caso_convivencia_pdf.html', {
        'caso': caso,
        'institucion': caso.institucion,
        'fecha_generacion': timezone.now(),
        'generado_por': request.user,
    })
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="acta_{caso.radicado}.pdf"'
    return response

@login_required
def gestionar_disponibilidad_view(request):
    """
    Permite a un docente ver, añadir y eliminar sus bloques de
    disponibilidad para reuniones con padres.
    """
    try:
        docente = request.user.docente
    except (AttributeError, Docente.DoesNotExist):
        messages.error(request, "Acceso denegado. Esta sección es solo para docentes.")
        return redirect('gestion_academica:inicio_academico')

    if request.method == 'POST':
        form = DisponibilidadDocenteForm(request.POST)
        if form.is_valid():
            disponibilidad = form.save(commit=False)
            disponibilidad.docente = docente
            disponibilidad.institucion = docente.institucion
            disponibilidad.save()
            messages.success(request, "Nuevo bloque de disponibilidad añadido correctamente.")
            return redirect('gestion_academica:gestionar_disponibilidad')
        else:
            messages.error(request, "Hubo un error en el formulario. Por favor, revisa los datos.")
    else:
        form = DisponibilidadDocenteForm()

    # Obtenemos las disponibilidades existentes para mostrarlas en la tabla
    disponibilidades_actuales = DisponibilidadDocente.objects.filter(docente=docente).order_by('dia_semana', 'hora_inicio')
    
    context = {
        'titulo_pagina': "Gestionar mi Disponibilidad para Reuniones",
        'form': form,
        'disponibilidades': disponibilidades_actuales,
    }
    return render(request, 'gestion_academica/gestionar_disponibilidad.html', context)


@login_required
@require_POST # Para seguridad, esta vista solo acepta peticiones POST
def eliminar_disponibilidad_view(request, pk):
    """
    Elimina un bloque de disponibilidad específico.
    """
    try:
        docente = request.user.docente
        disponibilidad = get_object_or_404(DisponibilidadDocente, pk=pk, docente=docente)
        disponibilidad.delete()
        messages.success(request, "El bloque de disponibilidad ha sido eliminado.")
    except (AttributeError, Docente.DoesNotExist):
        messages.error(request, "Acción no permitida.")
    
    return redirect('gestion_academica:gestionar_disponibilidad')         

@login_required
def familiar_seleccionar_docente_cita(request):
    """
    Muestra al familiar una lista de los docentes de sus hijos
    para que pueda seleccionar con quién agendar una cita.
    """
    try:
        familiar = request.user.familiar
    except (AttributeError, Familiar.DoesNotExist):
        messages.error(request, "Acceso denegado. Esta sección es solo para familiares.")
        return redirect('gestion_academica:inicio_academico')

    # Obtenemos un listado único de docentes que enseñan a los hijos del familiar
    institucion_familiar = getattr(request.user, 'institucion_asociada', None)
    estudiantes_asociados = familiar.estudiantes_asociados.all()
    cursos_de_los_hijos = Curso.objects.filter(
        grado__in=[e.grado_actual for e in estudiantes_asociados],
        institucion=institucion_familiar,
    )
    docentes = Docente.objects.filter(
        cursos_impartidos__in=cursos_de_los_hijos,
        institucion=institucion_familiar,
    ).distinct().select_related('usuario')

    context = {
        'titulo_pagina': "Agendar Reunión - Seleccionar Docente",
        'docentes': docentes
    }
    return render(request, 'gestion_academica/familiar_seleccionar_docente.html', context)


def _cita_reunion_slot_minuto_utc(dt):
    """Minuto UTC del inicio; alinea cupos entre BD (UTC) y slots generados en TZ local."""
    from datetime import timezone as _dt_tz

    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return int(dt.astimezone(_dt_tz.utc).timestamp() // 60)


@login_required
def familiar_agendar_cita_view(request, docente_pk):
    """
    Permite a un familiar ver los horarios disponibles y reservar una cita,
    evitando que se puedan agendar citas en horarios ya ocupados.
    """
    try:
        familiar = request.user.familiar
    except (AttributeError, Familiar.DoesNotExist):
        messages.error(request, "Acceso denegado. Esta sección es solo para familiares.")
        return redirect('gestion_academica:inicio_academico')

    # Aislamiento multi-institución: el docente debe pertenecer a la misma institución del familiar
    institucion_familiar = getattr(request.user, 'institucion_asociada', None)
    docente = get_object_or_404(Docente, pk=docente_pk, institucion=institucion_familiar)

    citas_existentes = CitaReunion.objects.filter(
        docente=docente,
        fecha_hora_inicio__gte=timezone.now(),
    ).exclude(estado='CANCELADA')
    slots_ocupados_keys = {_cita_reunion_slot_minuto_utc(c.fecha_hora_inicio) for c in citas_existentes}

    disponibilidades = DisponibilidadDocente.objects.filter(docente=docente)
    slots_disponibles = []
    hoy = timezone.localdate()
    for i in range(14):
        fecha_actual = hoy + timedelta(days=i)
        dia_semana_actual = fecha_actual.weekday()

        for disp in disponibilidades.filter(dia_semana=dia_semana_actual):
            hora_actual_slot = timezone.make_aware(
                datetime.combine(fecha_actual, disp.hora_inicio),
                timezone.get_current_timezone(),
            )
            hora_fin_bloque = timezone.make_aware(
                datetime.combine(fecha_actual, disp.hora_fin),
                timezone.get_current_timezone(),
            )

            while hora_actual_slot < hora_fin_bloque:
                if _cita_reunion_slot_minuto_utc(hora_actual_slot) not in slots_ocupados_keys:
                    slots_disponibles.append(hora_actual_slot)
                hora_actual_slot += timedelta(minutes=15)

    if request.method == 'POST':
        fecha_hora_str = request.POST.get('fecha_hora')
        asunto = request.POST.get('asunto')
        estudiante_id = request.POST.get('estudiante_id')

        if not (fecha_hora_str and asunto and estudiante_id):
            messages.error(request, "Por favor, completa todos los campos del formulario.")
            return redirect('gestion_academica:familiar_agendar_cita', docente_pk=docente.pk)

        fecha_hora_cita = datetime.fromisoformat(fecha_hora_str)
        if timezone.is_naive(fecha_hora_cita):
            fecha_hora_cita = timezone.make_aware(
                fecha_hora_cita, timezone.get_current_timezone()
            )

        estudiante = get_object_or_404(
            Estudiante,
            pk=estudiante_id,
            pk__in=familiar.estudiantes_asociados.values_list("pk", flat=True),
        )

        try:
            with transaction.atomic():
                Docente.objects.select_for_update().get(pk=docente.pk)
                ocupadas = {
                    _cita_reunion_slot_minuto_utc(c.fecha_hora_inicio)
                    for c in CitaReunion.objects.filter(
                        docente_id=docente.pk,
                        fecha_hora_inicio__gte=timezone.now(),
                    ).exclude(estado='CANCELADA')
                }
                if _cita_reunion_slot_minuto_utc(fecha_hora_cita) in ocupadas:
                    messages.error(
                        request,
                        "Lo sentimos, ese horario acaba de ser reservado. Por favor, elige otro.",
                    )
                    return redirect('gestion_academica:familiar_agendar_cita', docente_pk=docente.pk)

                CitaReunion.objects.create(
                    docente=docente,
                    familiar=familiar,
                    estudiante=estudiante,
                    fecha_hora_inicio=fecha_hora_cita,
                    asunto=asunto,
                    institucion_id=estudiante.institucion_id,
                )
        except IntegrityError:
            messages.error(
                request,
                "Lo sentimos, ese horario acaba de ser reservado. Por favor, elige otro.",
            )
            return redirect('gestion_academica:familiar_agendar_cita', docente_pk=docente.pk)

        messages.success(
            request,
            f"Cita con {docente.usuario.get_full_name()} agendada exitosamente.",
        )
        return redirect('gestion_academica:portal_familiar_inicio')

    context = {
        'titulo_pagina': f"Agendar Cita con {docente.usuario.get_full_name()}",
        'docente': docente,
        'slots_disponibles': slots_disponibles,
        'estudiantes_del_familiar': familiar.estudiantes_asociados.all(),
    }
    return render(request, 'gestion_academica/familiar_agendar_cita.html', context)

@login_required
def mis_citas_view(request):
    """
    Muestra al docente una lista de todas sus citas agendadas.
    """
    try:
        docente = request.user.docente
    except (AttributeError, Docente.DoesNotExist):
        messages.error(request, "Acceso denegado. Esta sección es solo para docentes.")
        return redirect('gestion_academica:inicio_academico')

    # Buscamos todas las citas del docente que no estén canceladas o ya realizadas
    citas_agendadas = CitaReunion.objects.filter(
        docente=docente,
        estado__in=['PENDIENTE', 'CONFIRMADA']
    ).select_related(
        'familiar__usuario', 'estudiante__usuario'
    ).order_by('fecha_hora_inicio')

    context = {
        'titulo_pagina': "Mis Citas Agendadas",
        'citas': citas_agendadas,
    }
    return render(request, 'gestion_academica/docente_mis_citas.html', context)        

@login_required
def gestionar_cita_view(request, pk):
    """
    Permite a un docente actualizar el estado de una cita y añadir notas.
    """
    # Seguridad: Aseguramos que el docente solo pueda gestionar sus propias citas.
    cita = get_object_or_404(CitaReunion.objects.select_related(
        'familiar__usuario', 'estudiante__usuario'
    ), pk=pk, docente__usuario=request.user)

    if request.method == 'POST':
        form = GestionCitaForm(request.POST, instance=cita)
        if form.is_valid():
            form.save()
            messages.success(request, "La información de la cita ha sido actualizada.")
            return redirect('gestion_academica:docente_mis_citas')
    else:
        form = GestionCitaForm(instance=cita)

    context = {
        'titulo_pagina': "Gestionar Cita",
        'form': form,
        'cita': cita,
    }
    return render(request, 'gestion_academica/gestionar_cita.html', context)

@login_required
def supervisar_citas_view(request):
    """
    Permite al coordinador ver todas las citas agendadas en la institución,
    incluyendo las notas y acuerdos registrados por los docentes.
    """
    # 1. Lógica de permisos
    if not (request.user.is_superuser or (request.user.is_staff and request.user.rol in ['coordinador', 'administrador'])):
        messages.error(request, "Acceso denegado a esta sección.")
        return redirect('gestion_academica:inicio_academico')

    # 2. Obtenemos las citas de la institución del coordinador
    user_inst = getattr(request.user, 'institucion_asociada', None)
    citas = CitaReunion.objects.none() # Queryset vacío por defecto
    
    if user_inst:
        citas = CitaReunion.objects.filter(
            docente__institucion=user_inst
        ).select_related(
            'docente__usuario', 'familiar__usuario', 'estudiante__usuario'
        ).order_by('-fecha_hora_inicio')

    context = {
        'titulo_pagina': "Supervisión de Reuniones Agendadas",
        'citas': citas,
    }
    return render(request, 'gestion_academica/supervisar_citas.html', context)  

@login_required
def detalle_cita_supervision_view(request, pk):
    """
    Muestra al coordinador el detalle completo de una única cita agendada.
    """
    # Lógica de permisos para el coordinador/admin
    if not (request.user.is_superuser or (request.user.is_staff and request.user.rol in ['coordinador', 'administrador'])):
        messages.error(request, "Acceso denegado a esta sección.")
        return redirect('gestion_academica:inicio_academico')

    # Obtenemos la cita específica, asegurándonos de que pertenezca a la institución del coordinador
    cita = get_object_or_404(
        CitaReunion.objects.select_related(
            'docente__usuario', 'familiar__usuario', 'estudiante__usuario'
        ),
        pk=pk,
        docente__institucion=request.user.institucion_asociada
    )

    context = {
        'titulo_pagina': "Detalle de Reunión Agendada",
        'cita': cita,
    }
    return render(request, 'gestion_academica/detalle_cita_supervision.html', context)         

@login_required
@permission_required('gestion_academica.view_estudiante')
def seleccionar_estudiante_certificado_view(request):
    """
    Permite al administrador seleccionar primero un grado y luego un
    estudiante de ese grado para generar un certificado.
    """
    # Obtenemos los grados de la institución del usuario para el filtro
    grados = get_filtered_queryset(Grado, request.user).order_by('nombre')
    
    estudiantes = Estudiante.objects.none() # Por defecto, no mostramos ningún estudiante
    grado_seleccionado = None
    
    # Verificamos si el usuario ha seleccionado un grado desde el formulario
    grado_id = request.GET.get('grado')
    if grado_id:
        # Filtramos los estudiantes para mostrar solo los del grado seleccionado
        estudiantes = get_filtered_queryset(Estudiante, request.user).filter(
            grado_actual_id=grado_id
        ).select_related('usuario', 'grado_actual').order_by('usuario__last_name')
        
        # También obtenemos el objeto del grado para mostrar su nombre
        grado_seleccionado = get_object_or_404(Grado, pk=grado_id)

    context = {
        'titulo_pagina': "Generar Certificados",
        'grados': grados,
        'estudiantes': estudiantes,
        'grado_seleccionado': grado_seleccionado
    }
    return render(request, 'gestion_academica/seleccionar_estudiante_certificado.html', context)  

@login_required
@permission_required('gestion_academica.view_estudiante')
def generar_certificado_estudios_view(request, estudiante_pk):
    try:
        estudiante = Estudiante.objects.select_related(
            'usuario', 'grado_actual', 'institucion'
        ).get(pk=estudiante_pk)
    except Estudiante.DoesNotExist:
        return HttpResponse("Estudiante no encontrado.", status=404)

    institucion = estudiante.institucion
    configuracion_adicional = getattr(institucion, 'configuracioninstitucion', None)

    ultimo_periodo_cursado = PeriodoAcademico.objects.filter(
        cursos__grado=estudiante.grado_actual,
        institucion=institucion
    ).order_by('-fecha_fin').first()
    
    materias_con_notas = []
    if ultimo_periodo_cursado:
        cursos_del_periodo = Curso.objects.filter(
            grado=estudiante.grado_actual,
            periodo_academico=ultimo_periodo_cursado
        )
        for curso in cursos_del_periodo:
            estado_academico = calcular_estado_academico_curso(curso, estudiante)
            nota_final = estado_academico.get('nota_final_ponderada')
            
            if nota_final is not None:
                materias_con_notas.append({
                    'nombre': curso.materia.nombre_materia,
                    # --- CORRECCIÓN CLAVE ---
                    # Eliminamos la multiplicación y solo formateamos la nota a dos decimales.
                    'promedio': f"{nota_final:.2f}",
                    # -------------------------
                    'desempeno': obtener_desempeno(nota_final, institucion)
                })

    context = {
        'estudiante': estudiante,
        'institucion': institucion,
        'configuracion': configuracion_adicional,
        'fecha_generacion': timezone.now(),
        'año_cursado': ultimo_periodo_cursado.año_escolar if ultimo_periodo_cursado else "N/A",
        'materias': sorted(materias_con_notas, key=lambda x: x['nombre']),
    }

    template_path = 'gestion_academica/certificados/certificado_estudios.html'
    template = get_template(template_path)
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Certificado_{estudiante.usuario.last_name}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
    if pisa_status.err:
       return HttpResponse('Ocurrió un error al generar el PDF <pre>' + html + '</pre>')
    return response

# Normaliza texto (quita tildes, convierte a minúsculas)
def normalizar(texto):
    return ''.join(
        c for c in unicodedata.normalize('NFKD', texto.lower())
        if not unicodedata.combining(c)
    )

@login_required
@permission_required('gestion_academica.view_estudiante')
def generar_constancia_matricula_view(request, estudiante_pk):
    try:
        estudiante = Estudiante.objects.select_related('usuario', 'grado_actual', 'institucion').get(pk=estudiante_pk)
    except Estudiante.DoesNotExist:
        return HttpResponse("Estudiante no encontrado.", status=404)

    institucion = estudiante.institucion
    periodo_activo = PeriodoAcademico.objects.filter(institucion=institucion, activo=True).first()

    valor_matricula = 0
    valor_pension = 0

    # ✅ Buscar TODOS los conceptos de la institución
    conceptos = ConceptoPago.objects.filter(institucion=institucion)

    # ✅ Buscar matrícula por nombre
    concepto_matricula = None
    for concepto in conceptos:
        if "matricula" in normalizar(concepto.nombre_concepto):
            concepto_matricula = concepto
            break

    # ✅ Buscar pensión (mensualidad)
    concepto_pension = None
    for concepto in conceptos:
        if "mensualidad" in normalizar(concepto.nombre_concepto):
            concepto_pension = concepto
            break

    print(">>> MATRÍCULA:", concepto_matricula)
    print(">>> PENSIÓN:", concepto_pension)

    valor_matricula = float(concepto_matricula.valor) if concepto_matricula else 0
    valor_pension = float(concepto_pension.valor) if concepto_pension else 0

    print(">>> MATRÍCULA GENERAL:", valor_matricula)
    print(">>> PENSIÓN GENERAL:", valor_pension)

    context = {
        'estudiante': estudiante,
        'institucion': institucion,
        'periodo_activo': periodo_activo,
        'fecha_generacion': timezone.now(),
        'valor_matricula': valor_matricula,
        'valor_pension': valor_pension
    }

    template = get_template('gestion_academica/certificados/constancia_matricula.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Constancia_Matricula_{estudiante.usuario.username}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse("Error al generar PDF")
    return response

    
@login_required
@permission_required('gestion_academica.view_estudiante')
def generar_paz_y_salvo_view(request, estudiante_pk):
    """
    Genera un certificado de Paz y Salvo Financiero, con la lógica
    de verificación de deudas corregida.
    """
    try:
        estudiante = Estudiante.objects.select_related('usuario', 'institucion').get(pk=estudiante_pk)
    except Estudiante.DoesNotExist:
        return HttpResponse("Estudiante no encontrado.", status=404)

    # --- LÓGICA DE VERIFICACIÓN FINANCIERA CORREGIDA ---
    # Buscamos si existe CUALQUIER cuenta para este estudiante que NO esté en estado 'PAGADO'.
    cuentas_pendientes = CuentaPorCobrarEstudiante.objects.filter(
        estudiante=estudiante
    ).exclude(estado='PAGADO')
    # --- FIN DE LA CORRECCIÓN ---

    if cuentas_pendientes.exists():
        messages.error(request, f"No se puede generar el Paz y Salvo. El estudiante {estudiante.usuario.get_full_name()} tiene saldos pendientes.")
        return redirect('gestion_academica:seleccionar_estudiante_certificado')

    # El resto de la lógica para generar el PDF se mantiene igual
    context = {
        'estudiante': estudiante,
        'institucion': estudiante.institucion,
        'fecha_generacion': timezone.now(),
    }
    template_path = 'gestion_academica/certificados/paz_y_salvo.html'
    template = get_template(template_path)
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Paz_y_Salvo_{estudiante.usuario.last_name}.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
    if pisa_status.err:
       return HttpResponse('Ocurrió un error al generar el PDF.')
    return response

@login_required
@permission_required('gestion_academica.add_cuentaporcobrarestudiante') # O un permiso de admin
def promocion_anual_view(request):
    """
    Gestiona el proceso de fin de año: ver estado de estudiantes,
    promoverlos al siguiente grado y generar la matrícula del próximo año.
    """
    user_inst = getattr(request.user, 'institucion_asociada', None)
    
    # Lógica para poblar los filtros
    grados = Grado.objects.filter(institucion=user_inst) if user_inst else Grado.objects.all()
    periodos = PeriodoAcademico.objects.filter(institucion=user_inst) if user_inst else PeriodoAcademico.objects.all()

    grado_id = request.GET.get('grado')
    periodo_id = request.GET.get('periodo')
    
    lista_para_promocion = []

    # Si el admin ha filtrado, calculamos el estado de los estudiantes
    if grado_id and periodo_id:
        grado = get_object_or_404(Grado, pk=grado_id)
        periodo = get_object_or_404(PeriodoAcademico, pk=periodo_id)
        estudiantes = Estudiante.objects.filter(grado_actual=grado)

        for est in estudiantes:
            # Aquí necesitarías una función que calcule el promedio final del año
            # Por ahora, simularemos que todos aprueban
            estado_final = "Aprobado" 
            lista_para_promocion.append({'estudiante': est, 'estado': estado_final})

    # Lógica para procesar la promoción
    if request.method == 'POST':
        estudiantes_a_promover_ids = request.POST.getlist('estudiantes_a_promover')
        siguiente_periodo_id = request.POST.get('siguiente_periodo_id')
        
        siguiente_periodo = get_object_or_404(PeriodoAcademico, pk=siguiente_periodo_id)
        concepto_matricula = ConceptoPago.objects.filter(
            tipo_concepto__nombre__icontains='Matrícula',
            periodo_academico_aplicable=siguiente_periodo
        ).first()

        if not concepto_matricula:
            messages.error(request, "No se encontró un concepto de 'Matrícula' para el periodo de destino. Por favor, créalo primero en el módulo de Finanzas.")
            return redirect(request.path_info)

        estudiantes_promovidos = 0
        with transaction.atomic():
            for est_id in estudiantes_a_promover_ids:
                estudiante = Estudiante.objects.get(pk=est_id)
                if estudiante.grado_actual and estudiante.grado_actual.siguiente_grado:
                    estudiante.grado_actual = estudiante.grado_actual.siguiente_grado
                    estudiante.save(update_fields=['grado_actual'])
                    
                    # Creamos la cuenta de cobro para la matrícula
                    CuentaPorCobrarEstudiante.objects.get_or_create(
                        estudiante=estudiante,
                        concepto_pago=concepto_matricula,
                        institucion=user_inst,
                        defaults={
                            'monto_asignado': estudiante.valor_matricula,
                            'estado': 'PENDIENTE'
                        }
                    )
                    estudiantes_promovidos += 1

        messages.success(request, f"{estudiantes_promovidos} estudiantes han sido promovidos y sus matrículas generadas exitosamente.")
        return redirect('gestion_academica:promocion_anual')

    context = {
        'titulo_pagina': "Promoción Anual de Estudiantes",
        'grados': grados,
        'periodos': periodos,
        'lista_para_promocion': lista_para_promocion,
        'grado_seleccionado_id': grado_id,
        'periodo_seleccionado_id': periodo_id,
    }
    return render(request, 'gestion_academica/promocion_anual.html', context)

@login_required
@permission_required('gestion_academica.view_estudiante') # O un permiso más específico
def generar_paz_y_salvo_view(request, estudiante_pk):
    """
    Genera un certificado de Paz y Salvo Financiero, con la lógica de
    verificación de deudas corregida para usar el campo 'estado'.
    """
    try:
        estudiante = Estudiante.objects.select_related('usuario', 'institucion').get(pk=estudiante_pk)
    except Estudiante.DoesNotExist:
        return HttpResponse("Estudiante no encontrado.", status=404)

    # --- INICIO DE LA CORRECCIÓN: Usamos el campo 'estado' ---
    # Buscamos si existe CUALQUIER cuenta para este estudiante que NO esté en estado 'PAGADO'.
    # Usamos .exclude() para hacerlo más explícito y claro.
    cuentas_pendientes = CuentaPorCobrarEstudiante.objects.filter(
        estudiante=estudiante
    ).exclude(estado='PAGADO')
    # --- FIN DE LA CORRECCIÓN ---

    if cuentas_pendientes.exists():
        messages.error(request, f"No se puede generar el Paz y Salvo. El estudiante {estudiante.usuario.get_full_name()} tiene saldos pendientes.")
        # Redirigimos de vuelta a la página de selección de certificados
        return redirect('gestion_academica:seleccionar_estudiante_certificado')

    # El resto de la lógica para generar el PDF se mantiene igual
    context = {
        'estudiante': estudiante,
        'institucion': estudiante.institucion,
        'fecha_generacion': timezone.now(),
    }
    template_path = 'gestion_academica/certificados/paz_y_salvo.html'
    template = get_template(template_path)
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Paz_y_Salvo_{estudiante.usuario.last_name}.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
    if pisa_status.err:
       return HttpResponse('Ocurrió un error al generar el PDF.')
    return response  

@login_required
@permission_required('gestion_academica.change_grado') # Solo usuarios con permiso para cambiar grados pueden acceder
def gestionar_promocion_grados_view(request):
    """
    Permite al administrador asignar el 'siguiente_grado' a cada grado
    para automatizar el proceso de promoción.
    """
    user_inst = getattr(request.user, 'institucion_asociada', None)
    
    if not request.user.is_staff:
        messages.error(request, "Acceso denegado.")
        return redirect('gestion_academica:inicio_academico')

    # Obtenemos los grados de la institución del usuario
    grados_qs = Grado.objects.all()
    if not request.user.is_superuser and user_inst:
        grados_qs = grados_qs.filter(institucion=user_inst)
    
    grados_qs = grados_qs.order_by('nombre')

    if request.method == 'POST':
        # Procesamos el formulario enviado
        with transaction.atomic():
            for grado in grados_qs:
                siguiente_grado_id = request.POST.get(f'siguiente_grado_{grado.pk}')
                if siguiente_grado_id:
                    # Usamos .get() para asegurarnos de que el grado seleccionado exista
                    siguiente_grado = get_object_or_404(Grado, pk=siguiente_grado_id)
                    grado.siguiente_grado = siguiente_grado
                else:
                    # Si se selecciona "Ninguno", se guarda como nulo
                    grado.siguiente_grado = None
                grado.save()
        
        messages.success(request, "¡La secuencia de promoción de grados ha sido actualizada exitosamente!")
        return redirect('gestion_academica:gestionar_promocion_grados')

    context = {
        'titulo_pagina': "Configurar Secuencia de Promoción",
        'grados': grados_qs,
    }
    return render(request, 'gestion_academica/gestionar_promocion_grados.html', context)         

# Este es el "cerebro" de nuestra primera vista de API
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def perfil_usuario_api_view(request):
    """
    Devuelve la información esencial del usuario que ha iniciado sesión.
    """
    user = request.user

    # --- INICIO DE LOS PRINTS DE DEPURACIÓN ---
    print("=============================================")
    print(f"API de Perfil llamada por: {user.username}")
    print(f"Nombre completo: {user.get_full_name()}")
    print(f"Email: {user.email}")
    print(f"Rol en la base de datos: {user.rol}")
    print("=============================================")
    # --- FIN DE LOS PRINTS DE DEPURACIÓN ---

    # ... (el resto de la lógica de la foto y la data no cambia)
    foto_url = None
    if hasattr(user, 'usuario') and hasattr(user.usuario, 'foto_perfil') and user.usuario.foto_perfil:
        foto_url = request.build_absolute_uri(user.usuario.foto_perfil.url)

    data = {
        'id': user.id,
        'nombre_completo': user.get_full_name(),
        'email': user.email,
        'rol': user.rol,
        'foto_perfil_url': foto_url,
    }

    return Response(data)   

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mis_cursos_api_view(request):
    """
    Devuelve una lista de cursos.
    - Si el usuario es DOCENTE, devuelve solo sus cursos asignados.
    - Si el usuario es COORDINADOR, devuelve TODOS los cursos del periodo activo.
    """
    user = request.user
    
    try:
        # Buscamos el periodo académico activo de la institución del usuario
        institucion = user.institucion_asociada
        if not institucion:
            return Response({'error': 'El usuario no está asociado a ninguna institución.'}, status=404)
            
        periodo_activo = PeriodoAcademico.objects.filter(institucion=institucion, activo=True).first()
        if not periodo_activo:
            return Response({'error': 'No hay un periodo académico activo configurado.'}, status=404)

        cursos_a_devolver = []

        # --- LÓGICA DE ROLES ---
        if user.rol == 'docente':
            # Si es docente, busca sus cursos específicos
            cursos_a_devolver = Curso.objects.filter(
                docentes_asignados=user.docente, 
                periodo_academico=periodo_activo
            )
        elif user.rol == 'coordinador':
            # Si es coordinador, busca TODOS los cursos del periodo
            cursos_a_devolver = Curso.objects.filter(
                periodo_academico=periodo_activo
            )
        else:
            # Si es otro rol (estudiante, etc.), se le niega el acceso
            return Response({'error': 'Acceso denegado. Recurso solo para docentes y coordinadores.'}, status=403)
        # --- FIN LÓGICA DE ROLES ---

        # Preparamos los datos para la app
        data = []
        for curso in cursos_a_devolver.select_related('materia', 'grado').order_by('grado__nombre'):
            data.append({
                'id_curso': curso.id,
                'nombre_grado': curso.grado.nombre,
                'nombre_materia': curso.materia.nombre_materia,
            })
            
        return Response(data)

    except Exception as e:
        return Response({'error': f'Ocurrió un error inesperado: {str(e)}'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_docente_api_view(request):
    """
    Recopila y devuelve TODA la información para el dashboard del docente,
    incluyendo la lista COMPLETA de acciones de gestión y widgets.
    """
    user = request.user
    
    if not hasattr(user, 'docente'):
        return Response({'error': 'Acceso denegado. Este recurso es solo para docentes.'}, status=403)

    try:
        docente = user.docente
        periodo_activo = PeriodoAcademico.objects.filter(institucion=docente.institucion, activo=True).first()

        if not periodo_activo:
            return Response({'error': 'No hay un periodo académico activo configurado.'}, status=404)

        # --- Lógica de la vista (cálculos y queries) ---
        cursos_del_docente = Curso.objects.filter(docentes_asignados=docente, periodo_academico=periodo_activo)
        
        hoy_inicio = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
        fin_rango = hoy_inicio + timedelta(days=7)
        citas_count = CitaReunion.objects.filter(
            docente=docente, estado__in=['PENDIENTE', 'CONFIRMADA'],
            fecha_hora_inicio__range=(hoy_inicio, fin_rango)
        ).count()
        
        direccion_grupo = DirectorCurso.objects.filter(docente=docente, periodo_academico=periodo_activo).first()
        es_director = bool(direccion_grupo)
        cantidad_riesgo = 0
        if es_director:
            estudiantes_grupo = Estudiante.objects.filter(
                grado_actual=direccion_grupo.grado,
                institucion=docente.institucion,
            )
            cursos_grado = Curso.objects.filter(
                grado=direccion_grupo.grado,
                periodo_academico=periodo_activo,
            )
            cantidad_riesgo = contar_pares_estudiante_curso_en_riesgo_academico(
                estudiantes_grupo,
                cursos_grado,
            )

        # --- INICIO DE LA LISTA DE ACCIONES COMPLETA (AHORA SÍ) ---
        acciones_gestion = [
            {
                "id": "asistencia",
                "titulo": "Pasar Lista / Asistir Curso",
                "descripcion": "Registra la asistencia con QR o manual, incluso en cursos que no son tuyos.",
                "icono": "bi-person-check-fill",
                "url": reverse('gestion_academica:seleccionar_curso_asistencia')
            },
            {
                "id": "libro_notas",
                "titulo": "Libro de Notas",
                "descripcion": "Consulta y registra las calificaciones de tus estudiantes.",
                "icono": "bi-journal-richtext",
                "url": reverse('gestion_academica:docente_seleccionar_curso_libro_notas')
            },
            {
                "id": "gestionar_actividades",
                "titulo": "Gestionar Actividades",
                "descripcion": "Crea y edita las actividades calificables para tus cursos.",
                "icono": "bi-card-list",
                "url": reverse('gestion_academica:docente_lista_actividades')
            },
            {
                "id": "gestionar_categorias",
                "titulo": "Gestionar Categorías",
                "descripcion": "Define las categorías de evaluación y sus porcentajes.",
                "icono": "bi-tags-fill",
                "url": reverse('gestion_academica:docente_lista_tipos_actividad')
            },
            {
                "id": "gestionar_descriptores",
                "titulo": "Gestionar Descriptores",
                "descripcion": "Administra los logros e indicadores para los boletines.",
                "icono": "bi-patch-check-fill",
                "url": reverse('gestion_academica:docente_lista_descriptores')
            },
            {
                "id": "material_apoyo",
                "titulo": "Archivos y Materiales",
                "descripcion": "Sube y gestiona guías, talleres y otros recursos para tus clases.",
                "icono": "bi-folder-plus",
                "url": reverse('gestion_academica:docente_lista_materiales')
            },
            {
                "id": "registrar_leccion",
                "titulo": "Registrar Lección Diaria",
                "descripcion": "Lleva un registro de los temas y actividades de cada clase.",
                "icono": "bi-journal-plus",
                "url": reverse('gestion_academica:seleccionar_curso_para_leccion')
            },
            {
                "id": "observador_estudiante",
                "titulo": "Observador del Estudiante",
                "descripcion": "Realiza anotaciones académicas o de convivencia.",
                "icono": "bi-eye-fill",
                "url": reverse('gestion_academica:seleccionar_estudiante_observador')
            },
            {
                "id": "gestionar_disponibilidad",
                "titulo": "Mi Disponibilidad para Citas",
                "descripcion": "Define tus horarios para reuniones con acudientes.",
                "icono": "bi-clock-history",
                "url": reverse('gestion_academica:gestionar_disponibilidad')
            },
            {
                "id": "menciones_honores",
                "titulo": "Menciones y Honores",
                "descripcion": "Genera reconocimientos para tus estudiantes destacados.",
                "icono": "bi-award-fill",
                "url": reverse('gestion_academica:docente_lista_menciones')
            },
            {
                "id": "reporte_nota_minima",
                "titulo": "Reporte de Nota Mínima",
                "descripcion": "Calcula la nota que necesitan tus estudiantes para aprobar.",
                "icono": "bi-graph-down-arrow",
                "url": reverse('gestion_academica:seleccionar_curso_reporte_nota_minima')
            }
        ]
        
        if es_director:
            acciones_gestion.append({
                "id": "panel_director", "titulo": "Panel de Director de Grupo",
                "descripcion": "Supervisa el rendimiento general de tu grupo asignado.",
                "icono": "bi-person-video3",
                "url_endpoint": reverse('gestion_academica:api_seleccionar_curso_calificaciones')
            })
        # --- FIN DE LA LISTA DE ACCIONES ---

        data = {
            'docente': { 'nombre_completo': docente.usuario.get_full_name() },
            'widgets': {
                'cursos_asignados': cursos_del_docente.count(),
                'tareas_por_calificar': EntregaDeber.objects.filter(deber__curso__in=cursos_del_docente, calificacion_obtenida__isnull=True).count(),
                'notificaciones_nuevas': Notificacion.objects.filter(destinatario=user, leido=False).count(),
                'citas_proximas': citas_count,
                'estudiantes_riesgo_grupo': cantidad_riesgo if es_director else None,
            },
            'acciones_gestion': acciones_gestion
        }

        return Response(data)

    except Exception as e:
        return Response({'error': f'Ocurrió un error inesperado: {str(e)}'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_estudiante_api_view(request):
    """
    Recopila y devuelve TODA la información necesaria para construir
    el dashboard completo de un estudiante en la app móvil.
    """
    user = request.user
    
    if not hasattr(user, 'estudiante'):
        return Response({'error': 'Acceso denegado. Este recurso es solo para estudiantes.'}, status=403)

    try:
        estudiante = user.estudiante
        institucion = estudiante.institucion
        periodo_activo = PeriodoAcademico.objects.filter(institucion=institucion, activo=True).first()

        if not periodo_activo:
            return Response({'error': 'No hay un periodo académico activo configurado.'}, status=404)

        # Bloqueo por mora — la app móvil debe respetar el mismo contrato
        # del portal web (Fase C). Si está bloqueado, devolvemos el mínimo
        # necesario para mostrar la pantalla "Pagar para desbloquear".
        portal_bloqueado_por_mora = not estudiante.esta_al_dia()

        # --- RECOPILACIÓN DE DATOS COMPLETA ---
        
        # 1. Widgets y Alertas
        notificaciones_sin_leer = Notificacion.objects.filter(
            destinatario=user, leido=False, institucion=institucion
        )
        cuentas_vencidas = CuentaPorCobrarEstudiante.objects.filter(estudiante=estudiante, estado='VENCIDO')
        inasistencias_periodo = RegistroAsistencia.objects.filter(estudiante=estudiante, curso__periodo_academico=periodo_activo, estado='AUSENTE').count()

        # 2. Agenda Próxima (Tareas y Actividades)
        eventos_agenda = []
        hoy = timezone.now().date()
        limite_agenda = hoy + timedelta(days=14)
        cursos_del_estudiante = Curso.objects.filter(
            grado=estudiante.grado_actual,
            periodo_academico=periodo_activo,
            institucion=institucion,
        )
        
        deberes = Deber.objects.filter(curso__in=cursos_del_estudiante, fecha_entrega__range=(hoy, limite_agenda))
        for deber in deberes:
            eventos_agenda.append({'tipo': 'Tarea', 'titulo': deber.titulo, 'fecha_entrega': deber.fecha_entrega})

        actividades = ActividadCalificable.objects.filter(curso__in=cursos_del_estudiante, fecha_entrega_limite__range=(hoy, limite_agenda))
        for actividad in actividades:
            eventos_agenda.append({'tipo': 'Evaluación', 'titulo': actividad.titulo, 'fecha_entrega': actividad.fecha_entrega_limite})

        # 3. Calificaciones Recientes por Materia
        calificaciones_recientes = Calificacion.objects.filter(estudiante=estudiante, actividad_calificable__curso__in=cursos_del_estudiante).select_related('actividad_calificable__curso__materia').order_by('-fecha_registro')[:10]
        materias_con_calificaciones = defaultdict(list)
        for cal in calificaciones_recientes:
            materias_con_calificaciones[cal.actividad_calificable.curso.materia.nombre_materia].append({
                'actividad': cal.actividad_calificable.titulo,
                'nota': cal.valor_numerico
            })

        # 4. Horario Semanal
        horario_semanal = defaultdict(list)
        dias_semana_map = {0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves', 4: 'Viernes'}
        bloques_horario = BloqueHorario.objects.filter(curso__in=cursos_del_estudiante).select_related('curso__materia', 'aula').order_by('hora_inicio')
        for bloque in bloques_horario:
            dia_nombre = dias_semana_map.get(bloque.dia_semana, 'Día no definido')
            horario_semanal[dia_nombre].append({
                'materia': bloque.curso.materia.nombre_materia,
                'hora_inicio': bloque.hora_inicio.strftime('%I:%M %p'),
                'hora_fin': bloque.hora_fin.strftime('%I:%M %p'),
                'aula': bloque.aula.nombre if bloque.aula else ''
            })

        # 5. Últimas Noticias
        ultimas_noticias = Noticia.objects.filter(institucion=institucion).order_by('-fecha_publicacion')[:3]

        # 6. Construcción del objeto JSON final
        data = {
            'estudiante_info': {
                'nombre_completo': estudiante.usuario.get_full_name(),
                'grado': estudiante.grado_actual.nombre if estudiante.grado_actual else "Sin Grado"
            },
            'alertas_principales': {
                'notificaciones_nuevas': notificaciones_sin_leer.count(),
                'esta_en_mora': cuentas_vencidas.exists(),
                'portal_bloqueado_por_mora': portal_bloqueado_por_mora,
                'dias_atraso_max': estudiante.dias_de_atraso_max,
                'mensajes_notificaciones': [n.mensaje for n in notificaciones_sin_leer[:3]]
            },
            'acciones_principales': {
                'mi_boletin_url': reverse('gestion_academica:mi_boletin_periodo_actual'),
                'mi_perfil_url': reverse('gestion_academica:ver_mi_perfil'),
                'estado_cartera_url': reverse('finanzas:mi_estado_de_cuenta'),
                'mis_asignaturas_url': '#acordeonAsignaturas', # Esto se manejaría en la app
                'mi_horario_url': '#seccion-horario' # Esto se manejaría en la app
            },
            'widgets_resumen': {
                'inasistencias_periodo': inasistencias_periodo,
                'cuentas_vencidas_count': cuentas_vencidas.count(),
                'calificaciones_recientes': [{ 'materia': k, 'cantidad': len(v) } for k, v in materias_con_calificaciones.items()],
            },
            'agenda_proxima': sorted(eventos_agenda, key=lambda e: e['fecha_entrega'])[:5],
            'horario_semanal': dict(horario_semanal),
            'ultimas_noticias': [
                {'id': n.id, 'titulo': n.titulo, 'fecha': n.fecha_publicacion.strftime('%d %b, %Y')} for n in ultimas_noticias
            ]
        }
        return Response(data)

    except Exception as e:
        return Response({'error': f'Ocurrió un error inesperado: {str(e)}'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def portal_familiar_api_view(request):
    """
    Devuelve la lista de estudiantes asociados a un familiar
    y todas las acciones disponibles para cada uno.
    """
    user = request.user
    if not hasattr(user, 'familiar'):
        return Response({'error': 'Acceso denegado. Este recurso es solo para familiares.'}, status=403)

    try:
        familiar = user.familiar
        estudiantes_a_cargo = familiar.estudiantes_asociados.select_related('usuario', 'grado_actual')

        # Preparamos los datos para la app
        data = {
            'familiar_info': {
                'nombre_completo': familiar.usuario.get_full_name(),
            },
            'acciones_generales': [
                 {
                    'id': 'agendar_cita', 'titulo': 'Agendar Reunión con Docente', 'icono': 'bi-calendar-plus',
                    'url': reverse('gestion_academica:familiar_seleccionar_docente')
                },
                {
                    'id': 'perfil', 'titulo': 'Mi Perfil', 'icono': 'bi-person-circle',
                    'url': reverse('gestion_academica:ver_mi_perfil')
                }
            ],
            'estudiantes': [
                {
                    'id_estudiante': est.pk,
                    'nombre_completo': est.usuario.get_full_name(),
                    'grado': est.grado_actual.nombre if est.grado_actual else "Sin grado",
                    'resumen_portada': _resumen_por_estudiante_portal_familiar(est),
                    'estado_cuenta_url': reverse(
                        'finanzas:familiar_estado_cuenta_estudiante',
                        args=[est.pk],
                    ),
                    'acciones_especificas': {
                        'calificaciones_url': reverse('gestion_academica:familiar_ver_calificaciones_estudiante', args=[est.pk]),
                        'boletin_url': reverse('gestion_academica:familiar_ver_boletin_estudiante', args=[est.pk]),
                        'deberes_url': reverse('gestion_academica:familiar_ver_deberes_estudiante', args=[est.pk]),
                    }
                } for est in estudiantes_a_cargo
            ]
        }
        return Response(data)

    except Exception as e:
        return Response({'error': f'Ocurrió un error inesperado: {str(e)}'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_coordinador_api_view(request):
    """
    Recopila y devuelve TODA la información para el dashboard del coordinador,
    incluyendo widgets y la lista completa de acciones según sus permisos.
    """
    user = request.user
    if not (
        user.is_superuser
        or (user.is_staff and getattr(user, 'rol', None) in ['coordinador', 'administrador'])
    ):
        return Response({'error': 'Acceso denegado. Recurso solo para personal directivo.'}, status=403)

    try:
        user_inst = user.institucion_asociada
        
        # --- CÁLCULO DE WIDGETS ---
        alertas_bienestar = 0
        alertas_riesgo = 0
        if user_inst:
            alertas_bienestar = AnotacionObservador.objects.filter(estudiante__institucion=user_inst, requiere_revision=True).count()
            ultimo_analisis = AnalisisRiesgo.objects.filter(periodo_academico__institucion=user_inst).order_by('-fecha_analisis').first()
            if ultimo_analisis:
                alertas_riesgo = ultimo_analisis.predicciones.filter(nivel_riesgo__in=['ALTO', 'MEDIO']).count()

        # --- CONSTRUCCIÓN DE LA LISTA DE ACCIONES DISPONIBLES ---
        acciones_disponibles = []

        # 1. Panel de Coordinación (siempre visible para este rol)
        acciones_disponibles.append({
            "id": "panel_coordinacion", "titulo": "Panel de Coordinación", "icono": "bi-person-video3",
            "sub_acciones": [
                {"id": "control_asistencia", "titulo": "Control de Asistencia", "url": reverse('gestion_academica:admin_asistencia_diaria')},
                {"id": "alertas_bienestar", "titulo": "Alertas de Bienestar", "url": reverse('gestion_academica:dashboard_bienestar')},
                {"id": "supervisar_citas", "titulo": "Supervisión de Citas", "url": reverse('gestion_academica:supervisar_citas')},
                {"id": "reporte_riesgo_academico", "titulo": "HALU Sentinel - Riesgo Académico", "url": reverse('gestion_academica:reporte_riesgo_academico')},
            ]
        })

        # 2. Módulos de Gestión (con verificación de permisos)
        if user.has_perm('gestion_academica.acceso_modulo_academico'):
            acciones_disponibles.append({
                "id": "gestion_academica", "titulo": "Gestión Académica", "icono": "bi-buildings-fill",
                "sub_acciones": [
                    {"id": "gestion_personas", "titulo": "Gestión de Personas", "url": reverse('gestion_academica:lista_estudiantes')}, # Enlace genérico a estudiantes
                    {"id": "generar_certificados", "titulo": "Generar Certificados", "url": reverse('gestion_academica:seleccionar_estudiante_certificado')},
                ]
            })
            acciones_disponibles.append({
                "id": "configuracion_academica", "titulo": "Configuración Académica", "icono": "bi-sliders",
                "sub_acciones": [
                    {"id": "secuencia_promocion", "titulo": "Secuencia de Promoción", "url": reverse('gestion_academica:gestionar_promocion_grados')},
                ]
            })
        
        # 3. Módulos Externos y Admin Avanzado
        if user.has_perm('admisiones.view_aspirante'):
             acciones_disponibles.append({"id": "admisiones", "titulo": "Módulo de Admisiones", "icono": "bi-person-plus-fill", "url": reverse('admisiones:lista_aspirantes')})
        
        if user.is_superuser:
            acciones_disponibles.append({"id": "admin_avanzado", "titulo": "Admin Avanzado Django", "icono": "bi-speedometer2", "url": reverse('admin:index')})

        # --- CONSTRUCCIÓN DEL OBJETO JSON FINAL ---
        data = {
            'coordinador_info': { 'nombre_completo': user.get_full_name() },
            'widgets': {
                'alertas_bienestar': alertas_bienestar,
                'alertas_riesgo_academico': alertas_riesgo,
            },
            'acciones_disponibles': acciones_disponibles
        }
        return Response(data)

    except Exception as e:
        return Response({'error': f'Ocurrió un error inesperado: {str(e)}'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_admin_api_view(request):
    """
    Recopila y devuelve una lista completa de TODAS las secciones de gestión
    disponibles para un administrador, que por su rol, tiene acceso total.
    """
    user = request.user
    
    # Verificamos que el usuario tenga el rol correcto para ver este panel
    if not (user.is_staff and user.rol == 'administrador'):
        return Response({'error': 'Acceso denegado. Este recurso es solo para administradores.'}, status=403)

    try:
        # --- LÓGICA SIMPLIFICADA: Si es admin, tiene acceso a todo ---
        
        modulos_disponibles = [
            # 1. Módulo de Coordinación
            {
                "id": "panel_coordinacion", "titulo": "Panel de Coordinación", "icono": "bi-person-video3",
                "sub_acciones": [
                    {"id": "control_asistencia", "titulo": "Control de Asistencia", "url": reverse('gestion_academica:admin_asistencia_diaria')},
                    {"id": "alertas_bienestar", "titulo": "Alertas de Bienestar", "url": reverse('gestion_academica:dashboard_bienestar')},
                    {"id": "supervisar_citas", "titulo": "Supervisión de Citas", "url": reverse('gestion_academica:supervisar_citas')},
                    {"id": "reporte_riesgo_academico", "titulo": "HALU Sentinel - Riesgo Académico", "url": reverse('gestion_academica:reporte_riesgo_academico')},
                ]
            },
            # 2. Módulo de Admisiones
            {
                "id": "admisiones", "titulo": "Módulo de Admisiones", "icono": "bi-person-plus-fill", 
                "url": reverse('admisiones:lista_aspirantes')
            },
            # 3. Módulo Financiero
            {
                "id": "finanzas", "titulo": "Módulo Financiero", "icono": "bi-currency-dollar", 
                "url": reverse('finanzas:dashboard_financiero')
            },
            # 4. Módulo de Gestión Académica
            {
                "id": "gestion_academica", "titulo": "Gestión General", "icono": "bi-buildings-fill",
                "sub_acciones": [
                    {"id": "gestion_estudiantes", "titulo": "Gestionar Estudiantes", "url": reverse('gestion_academica:lista_estudiantes')},
                    {"id": "gestion_docentes", "titulo": "Gestionar Docentes", "url": reverse('gestion_academica:lista_docentes')},
                    {"id": "generar_certificados", "titulo": "Generar Certificados", "url": reverse('gestion_academica:seleccionar_estudiante_certificado')},
                ]
            },
            # 5. Módulo de Configuración Académica
            {
                "id": "configuracion_academica", "titulo": "Configuración Académica", "icono": "bi-sliders",
                "sub_acciones": [
                    {"id": "config_grados", "titulo": "Grados y Secciones", "url": reverse('gestion_academica:lista_grados')},
                    {"id": "config_materias", "titulo": "Materias", "url": reverse('gestion_academica:lista_materias')},
                    {"id": "config_periodos", "titulo": "Periodos Académicos", "url": reverse('gestion_academica:lista_periodos')},
                    {"id": "secuencia_promocion", "titulo": "Secuencia de Promoción", "url": reverse('gestion_academica:gestionar_promocion_grados')},
                ]
            },
            # 6. Panel de Admin Avanzado
            {
                "id": "admin_avanzado", "titulo": "Admin Avanzado Django", "icono": "bi-speedometer2",
                "url": reverse('admin:index')
            }
        ]

        # --- CONSTRUCCIÓN DEL OBJETO JSON FINAL ---
        data = {
            'admin_info': {
                'nombre_completo': user.get_full_name(),
                'rol': user.get_rol_display(),
            },
            'modulos_gestion': modulos_disponibles
        }

        return Response(data)

    except Exception as e:
        return Response({'error': f'Ocurrió un error inesperado: {str(e)}'}, status=500)                               

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_seleccionar_curso_calificaciones(request):
    """
    API que devuelve la lista de cursos de un docente para que la app
    pueda mostrar un selector antes de ir al libro de notas.
    """
    user = request.user
    if not hasattr(user, 'docente'):
        return Response({'error': 'Acceso denegado.'}, status=403)

    try:
        docente = user.docente
        periodo_activo = PeriodoAcademico.objects.filter(institucion=docente.institucion, activo=True).first()
        if not periodo_activo:
            return Response({'error': 'No hay un periodo académico activo.'}, status=404)

        cursos = Curso.objects.filter(
            docentes_asignados=docente,
            periodo_academico=periodo_activo
        ).select_related('materia', 'grado').order_by('grado__nombre', 'materia__nombre_materia')

        # Preparamos la lista de cursos para la app
        data = [
            {
                "id_curso": curso.id,
                "nombre_completo_curso": f"{curso.materia.nombre_materia} - {curso.grado.nombre}",
            } for curso in cursos
        ]
        
        return Response(data)

    except Exception as e:
        return Response({'error': f'Ocurrió un error: {str(e)}'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def libro_notas_api_view(request, curso_pk):
    """
    API que devuelve toda la estructura y datos de un libro de notas
    para un curso específico: estudiantes, actividades y calificaciones.
    """
    user = request.user
    if not hasattr(user, 'docente'):
        return Response({'error': 'Acceso denegado.'}, status=403)

    try:
        # 1. Obtenemos el curso y verificamos que el docente tenga permiso
        curso = get_object_or_404(Curso, pk=curso_pk)
        if not curso.docentes_asignados.filter(pk=user.docente.pk).exists() and not user.is_superuser:
            return Response({'error': 'No tienes permiso para ver este libro de notas.'}, status=403)

        # 2. Obtenemos los componentes del libro de notas
        estudiantes = Estudiante.objects.filter(grado_actual=curso.grado).select_related('usuario').order_by('usuario__last_name')
        actividades = ActividadCalificable.objects.filter(curso=curso).order_by('fecha_publicacion')
        calificaciones = Calificacion.objects.filter(actividad_calificable__in=actividades)

        # 3. Estructuramos los datos para que sean fáciles de usar para la app
        
        # Un mapa para acceder a las notas rápidamente: (id_estudiante, id_actividad) -> nota
        mapa_calificaciones = {
            (cal.estudiante_id, cal.actividad_calificable_id): cal.valor_numerico 
            for cal in calificaciones
        }

        # 4. Construimos la respuesta JSON
        data = {
            'curso_info': {
                'id': curso.id,
                'nombre_materia': curso.materia.nombre_materia,
                'nombre_grado': curso.grado.nombre,
            },
            'actividades_header': [ # La cabecera de la tabla
                {'id': act.id, 'titulo': act.titulo} for act in actividades
            ],
            'estudiantes_rows': [ # Las filas de la tabla
                {
                    'id_estudiante': est.pk,
                    'nombre_completo': est.usuario.get_full_name(),
                    'calificaciones': { # Un diccionario de notas para este estudiante
                        act.id: f"{mapa_calificaciones.get((est.pk, act.id)):.2f}" if mapa_calificaciones.get((est.pk, act.id)) is not None else None
                        for act in actividades
                    }
                } for est in estudiantes
            ]
        }
        
        return Response(data)

    except Exception as e:
        return Response({'error': f'Ocurrió un error inesperado: {str(e)}'}, status=500)

@api_view(['POST']) # <-- IMPORTANTE: Solo acepta peticiones POST
@permission_classes([IsAuthenticated])
def guardar_libro_notas_api_view(request, curso_pk):
    """
    API para recibir y guardar un lote de calificaciones enviadas
    desde la aplicación móvil.
    """
    user = request.user
    if not hasattr(user, 'docente'):
        return Response({'error': 'Acceso denegado.'}, status=403)

    try:
        curso = get_object_or_404(Curso, pk=curso_pk)
        if not curso.docentes_asignados.filter(pk=user.docente.pk).exists():
            return Response({'error': 'No tienes permiso para guardar notas en este curso.'}, status=403)

        # request.data contiene el JSON que envía la app
        calificaciones_data = request.data 
        if not isinstance(calificaciones_data, list):
            return Response({'error': 'El formato de los datos debe ser una lista de calificaciones.'}, status=400)

        calificaciones_procesadas_ids = []

        with transaction.atomic():
            for item in calificaciones_data:
                estudiante_id = item.get('estudiante_id')
                actividad_id = item.get('actividad_id')
                nota_str = item.get('nota')

                if nota_str is not None and nota_str.strip() != '':
                    valor_nota = Decimal(str(nota_str).replace(',', '.'))
                    
                    # Usamos update_or_create para eficiencia
                    calificacion_obj, created = Calificacion.objects.update_or_create(
                        estudiante_id=estudiante_id,
                        actividad_calificable_id=actividad_id,
                        defaults={
                            'valor_numerico': valor_nota,
                            'registrada_por': user.docente,
                            'institucion': curso.institucion
                        }
                    )
                    calificaciones_procesadas_ids.append(calificacion_obj.id)
        
        # --- Disparamos la lógica de IA manualmente después de guardar ---
        if calificaciones_procesadas_ids:
            # (Aquí va la misma lógica que ya tienes en la vista web para llamar a la IA)
            # ...
            pass

        return Response({'status': 'success', 'message': f'{len(calificaciones_procesadas_ids)} calificaciones guardadas exitosamente.'})

    except Exception as e:
        return Response({'error': f'Ocurrió un error inesperado: {str(e)}'}, status=500)    

@api_view(['GET'])
@permission_classes([IsAuthenticated, EstaAlDiaPermission])
def mis_deberes_api_view(request):
    """
    API que devuelve la lista de deberes asignados a un estudiante,
    indicando el estado de entrega de cada uno.
    """
    user = request.user
    if not hasattr(user, 'estudiante'):
        return Response({'error': 'Acceso denegado.'}, status=403)

    try:
        estudiante_actual = user.estudiante
        periodo_activo = PeriodoAcademico.objects.filter(institucion=estudiante_actual.institucion, activo=True).first()

        if not (periodo_activo and estudiante_actual.grado_actual):
            return Response([]) # Devuelve una lista vacía si no hay periodo o grado

        # Buscamos los cursos y luego los deberes de esos cursos
        cursos_del_estudiante = Curso.objects.filter(grado=estudiante_actual.grado_actual, periodo_academico=periodo_activo)
        deberes_asignados = Deber.objects.filter(curso__in=cursos_del_estudiante).select_related('curso__materia').order_by('-fecha_entrega')

        # Verificamos cuáles deberes ya han sido entregados por el estudiante
        entregas_realizadas_ids = EntregaDeber.objects.filter(
            estudiante=estudiante_actual,
            deber__in=deberes_asignados
        ).values_list('deber_id', flat=True)

        # Preparamos la respuesta JSON
        data = []
        for deber in deberes_asignados:
            data.append({
                "id_deber": deber.pk,
                "titulo": deber.titulo,
                "materia": deber.curso.materia.nombre_materia,
                "descripcion": deber.descripcion,
                "fecha_entrega": deber.fecha_entrega,
                "entregado": deber.pk in entregas_realizadas_ids,
                "url_entrega": reverse('gestion_academica:realizar_entrega_deber', args=[deber.pk])
            })

        return Response(data)

    except Exception as e:
        return Response({'error': f'Ocurrió un error inesperado: {str(e)}'}, status=500)                    

@api_view(['GET'])
@permission_classes([IsAuthenticated, EstaAlDiaPermission])
def mi_boletin_api_view(request):
    """
    API que devuelve los datos completos para construir el boletín
    de un estudiante en la aplicación móvil.
    """
    user = request.user
    if not hasattr(user, 'estudiante'):
        return Response({'error': 'Acceso denegado.'}, status=403)

    try:
        estudiante_actual = user.estudiante
        institucion = estudiante_actual.institucion
        periodo_activo = PeriodoAcademico.objects.filter(institucion=institucion, activo=True).first()

        if not (periodo_activo and estudiante_actual.grado_actual):
            return Response({'error': 'No se pueden mostrar las calificaciones. No hay un periodo activo o no estás asignado a un grado.'}, status=404)

        # --- Lógica de cálculo del boletín ---
        cursos_del_estudiante = Curso.objects.filter(
            grado=estudiante_actual.grado_actual,
            periodo_academico=periodo_activo
        ).select_related('materia')
        
        materias_data = []
        suma_notas_finales = Decimal('0.0')
        materias_validas_para_promedio = 0

        for curso in cursos_del_estudiante:
            # Reutilizamos la lógica de cálculo que ya tienes
            estado_academico = calcular_estado_academico_curso(curso, estudiante_actual)
            nota_final = estado_academico.get('nota_final_ponderada')
            
            materias_data.append({
                "nombre_materia": curso.materia.nombre_materia,
                "nota_final": f"{nota_final:.2f}" if nota_final is not None else "N/C", # N/C = No Calculada
                "desempeno": obtener_desempeno(nota_final, institucion) if nota_final is not None else ""
            })

            if nota_final is not None:
                suma_notas_finales += nota_final
                materias_validas_para_promedio += 1
        
        # Calculamos el promedio general del periodo
        promedio_general = (suma_notas_finales / materias_validas_para_promedio) if materias_validas_para_promedio > 0 else None
        
        # --- Construimos la respuesta JSON ---
        data = {
            'info_boletin': {
                'titulo_pagina': f"Mi Boletín - {periodo_activo.nombre}",
                'nombre_estudiante': estudiante_actual.usuario.get_full_name(),
                'nombre_grado': estudiante_actual.grado_actual.nombre,
                'periodo_id': periodo_activo.pk,  # ✅ AGREGA ESTA LÍNEA
            },
            'resumen_academico': {
                'promedio_general': f"{promedio_general:.2f}" if promedio_general is not None else "N/C",
                'desempeno_general': obtener_desempeno(promedio_general, institucion) if promedio_general is not None else ""
            },
            'calificaciones_por_materia': sorted(materias_data, key=lambda x: x['nombre_materia'])
        }

        return Response(data)

    except Exception as e:
        return Response({'error': f'Ocurrió un error inesperado: {str(e)}'}, status=500)        

@api_view(['GET'])
@permission_classes([IsAuthenticated, EstaAlDiaPermission])
def detalle_calificaciones_materia_api_view(request, materia_pk):
    """ API para la PANTALLA DE DETALLE de calificaciones de UNA materia. """
    user = request.user
    if not hasattr(user, 'estudiante'):
        return Response({'error': 'Acceso denegado.'}, status=403)
    
    try:
        estudiante = user.estudiante
        materia = get_object_or_404(Materia, pk=materia_pk)
        periodo_activo = PeriodoAcademico.objects.filter(institucion=estudiante.institucion, activo=True).first()

        if not periodo_activo:
            return Response({'error': 'No hay periodo activo.'}, status=404)

        calificaciones = Calificacion.objects.filter(
            estudiante=estudiante,
            actividad_calificable__curso__materia=materia,
            actividad_calificable__curso__periodo_academico=periodo_activo
        ).select_related('actividad_calificable').order_by('-fecha_registro')
        
        data = {
            'materia_info': {'nombre': materia.nombre_materia},
            'calificaciones': [
                {
                    'actividad': cal.actividad_calificable.titulo,
                    'nota': f"{cal.valor_numerico:.2f}" if cal.valor_numerico is not None else "N/A",
                    'fecha': cal.fecha_registro.strftime('%d %b, %Y')
                } for cal in calificaciones
            ]
        }
        return Response(data)
    except Exception as e:
        return Response({'error': f'Ocurrió un error: {str(e)}'}, status=500)


@api_view(['GET'])
def mis_asignaturas_api_view(request):
    """ API para la PANTALLA DE DETALLE de las asignaturas del estudiante. """
    user = request.user
    if not hasattr(user, 'estudiante'): return Response({'error': 'Acceso denegado.'}, status=403)
    try:
        estudiante = user.estudiante
        periodo_activo = PeriodoAcademico.objects.filter(institucion=estudiante.institucion, activo=True).first()
        if not (periodo_activo and estudiante.grado_actual): return Response([])
        cursos = Curso.objects.filter(grado=estudiante.grado_actual, periodo_academico=periodo_activo).select_related('materia', 'grado').order_by('materia__nombre_materia')
        data = [{'id_materia': c.materia.pk, 'nombre_materia': c.materia.nombre_materia, 'nombre_grado': c.grado.nombre} for c in cursos]
        return Response(data)
    except Exception as e: return Response({'error': f'Ocurrió un error: {str(e)}'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated, EstaAlDiaPermission])
def mi_horario_api_view(request):
    """ API para la PANTALLA DE DETALLE del horario semanal del estudiante. """
    user = request.user
    if not hasattr(user, 'estudiante'): return Response({'error': 'Acceso denegado.'}, status=403)
    try:
        estudiante = user.estudiante
        periodo_activo = PeriodoAcademico.objects.filter(institucion=estudiante.institucion, activo=True).first()
        if not (periodo_activo and estudiante.grado_actual): return Response({})
        cursos = Curso.objects.filter(grado=estudiante.grado_actual, periodo_academico=periodo_activo)
        horario = defaultdict(list)
        dias_map = {0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves', 4: 'Viernes'}
        bloques = BloqueHorario.objects.filter(curso__in=cursos).select_related('curso__materia', 'aula').order_by('hora_inicio')
        for b in bloques:
            horario[dias_map.get(b.dia_semana)].append({'materia': b.curso.materia.nombre_materia, 'hora_inicio': b.hora_inicio.strftime('%I:%M %p'), 'hora_fin': b.hora_fin.strftime('%I:%M %p'), 'aula': b.aula.nombre if b.aula else 'N/A'})
        return Response(dict(horario))
    except Exception as e: return Response({'error': f'Ocurrió un error: {str(e)}'}, status=500)

@api_view(['GET'])
def mi_estado_cartera_api_view(request):
    """ API para la PANTALLA DE DETALLE del estado de cartera del estudiante. """
    user = request.user
    if not hasattr(user, 'estudiante'): return Response({'error': 'Acceso denegado.'}, status=403)
    try:
        estudiante = user.estudiante
        cuentas = CuentaPorCobrarEstudiante.objects.filter(estudiante=estudiante).select_related('concepto_pago').order_by('-fecha_vencimiento')
        data = [{'id_cuenta': c.pk, 'concepto': c.concepto_pago.nombre, 'monto': f"{c.monto_asignado:,.0f}", 'estado': c.get_estado_display(), 'fecha_vencimiento': c.fecha_vencimiento} for c in cuentas]
        return Response(data)
    except Exception as e: return Response({'error': f'Ocurrió un error: {str(e)}'}, status=500)

@api_view(['GET'])
def lista_noticias_api_view(request):
    """ API para la PANTALLA DE DETALLE de noticias y anuncios. """
    user = request.user
    if not hasattr(user, 'estudiante'): return Response({'error': 'Acceso denegado.'}, status=403)
    try:
        noticias = Noticia.objects.filter(institucion=user.estudiante.institucion).order_by('-fecha_publicacion')
        data = [{'id_noticia': n.id, 'titulo': n.titulo, 'resumen': n.resumen, 'fecha': n.fecha_publicacion.strftime('%d %b, %Y')} for n in noticias]
        return Response(data)
    except Exception as e: return Response({'error': f'Ocurrió un error: {str(e)}'}, status=500)

# ===================================================================
# NUEVAS APIs PARA COORDINACIÓN - CONECTAR CON APP MÓVIL
# ===================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_dashboard_coordinador_data(request):
    """
    API que devuelve los datos del dashboard de coordinación para la app móvil
    """
    user = request.user
    if not (user.is_staff and user.rol in ['coordinador', 'administrador']):
        return Response({'error': 'Acceso denegado.'}, status=403)

    try:
        user_inst = getattr(user, 'institucion_asociada', None)
        
        # Estadísticas generales
        total_estudiantes = Estudiante.objects.filter(institucion=user_inst).count() if user_inst else 0
        total_docentes = Docente.objects.filter(institucion=user_inst).count() if user_inst else 0
        
        # Asistencia de hoy
        hoy = timezone.localdate()
        registros_hoy = RegistroAsistencia.objects.filter(
            estudiante__institucion=user_inst,
            fecha__date=hoy
        ) if user_inst else RegistroAsistencia.objects.none()
        
        presentes = registros_hoy.filter(estado='PRESENTE').count()
        ausentes = registros_hoy.filter(estado='AUSENTE').count()
        tardanzas = registros_hoy.filter(estado='TARDANZA').count()
        
        # Alertas de bienestar
        alertas_bienestar = AnotacionObservador.objects.filter(
            estudiante__institucion=user_inst,
            requiere_revision=True
        ).count() if user_inst else 0
        
        # Citas programadas próximas
        fin_semana = hoy + timedelta(days=7)
        citas_proximas = CitaReunion.objects.filter(
            docente__institucion=user_inst,
            fecha_hora_inicio__date__range=(hoy, fin_semana),
            estado__in=['PENDIENTE', 'CONFIRMADA']
        ).count() if user_inst else 0

        data = {
            'estadisticas_generales': {
                'total_estudiantes': total_estudiantes,
                'total_docentes': total_docentes,
                'alertas_activas': alertas_bienestar
            },
            'asistencia_hoy': {
                'presentes': presentes,
                'ausentes': ausentes,
                'tardanzas': tardanzas,
                'porcentaje': round((presentes / total_estudiantes * 100), 1) if total_estudiantes > 0 else 0
            },
            'alertas_bienestar': alertas_bienestar,
            'citas_programadas': citas_proximas
        }
        
        return Response(data)
        
    except Exception as e:
        return Response({'error': f'Error inesperado: {str(e)}'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_reportes_data(request):
    """
    API que devuelve datos para generar reportes
    """
    user = request.user
    if not (user.is_staff and user.rol in ['coordinador', 'administrador']):
        return Response({'error': 'Acceso denegado.'}, status=403)

    try:
        user_inst = getattr(user, 'institucion_asociada', None)
        periodo_activo = PeriodoAcademico.objects.filter(
            institucion=user_inst, 
            activo=True
        ).first() if user_inst else None

        if not periodo_activo:
            return Response({'error': 'No hay periodo académico activo'}, status=404)

        # Datos para reportes
        grados = Grado.objects.filter(institucion=user_inst).order_by('nombre')
        
        reportes_disponibles = [
            {
                'id': 'asistencia',
                'titulo': 'Reporte de Asistencia',
                'descripcion': 'Estadísticas de asistencia por curso y período',
                'icono': 'checkmark-done-outline'
            },
            {
                'id': 'academico', 
                'titulo': 'Reporte Académico',
                'descripcion': 'Rendimiento académico por materia y estudiante',
                'icono': 'school-outline'
            },
            {
                'id': 'disciplinario',
                'titulo': 'Reporte Disciplinario', 
                'descripcion': 'Incidentes y observaciones de comportamiento',
                'icono': 'alert-circle-outline'
            },
            {
                'id': 'bienestar',
                'titulo': 'Reporte de Bienestar',
                'descripcion': 'Alertas y seguimiento psicosocial',
                'icono': 'heart-outline'
            }
        ]

        data = {
            'periodo_activo': {
                'id': periodo_activo.id,
                'nombre': periodo_activo.nombre
            },
            'grados_disponibles': [
                {'id': g.id, 'nombre': g.nombre} for g in grados
            ],
            'reportes_disponibles': reportes_disponibles
        }
        
        return Response(data)
        
    except Exception as e:
        return Response({'error': f'Error inesperado: {str(e)}'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_asistencia_diaria_data(request):
    """
    API que devuelve datos detallados de asistencia diaria
    """
    user = request.user
    if not (user.is_staff and user.rol in ['coordinador', 'administrador']):
        return Response({'error': 'Acceso denegado.'}, status=403)

    try:
        user_inst = getattr(user, 'institucion_asociada', None)
        fecha_str = request.GET.get('fecha', timezone.localdate().strftime('%Y-%m-%d'))
        
        try:
            fecha_consulta = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            fecha_consulta = timezone.localdate()

        # Registros de asistencia del día
        registros_dia = RegistroAsistencia.objects.filter(
            estudiante__institucion=user_inst,
            fecha__date=fecha_consulta
        ).select_related('estudiante__usuario', 'estudiante__grado_actual', 'curso__materia')

        # Agrupar por grado
        asistencia_por_grado = defaultdict(lambda: {
            'total_estudiantes': 0,
            'presentes': 0,
            'ausentes': 0, 
            'tardanzas': 0,
            'estudiantes_detalle': []
        })

        # Obtener todos los estudiantes activos
        estudiantes_activos = Estudiante.objects.filter(
            institucion=user_inst,
            usuario__is_active=True
        ).select_related('usuario', 'grado_actual')

        # Crear mapa de registros por estudiante
        registros_map = {r.estudiante.id: r for r in registros_dia}

        for estudiante in estudiantes_activos:
            if estudiante.grado_actual:
                grado_nombre = estudiante.grado_actual.nombre
                asistencia_por_grado[grado_nombre]['total_estudiantes'] += 1
                
                registro = registros_map.get(estudiante.id)
                if registro:
                    if registro.estado == 'PRESENTE':
                        asistencia_por_grado[grado_nombre]['presentes'] += 1
                    elif registro.estado == 'AUSENTE':
                        asistencia_por_grado[grado_nombre]['ausentes'] += 1
                    elif registro.estado == 'TARDANZA':
                        asistencia_por_grado[grado_nombre]['tardanzas'] += 1
                    
                    estado_display = registro.get_estado_display()
                else:
                    estado_display = 'Sin registro'
                
                asistencia_por_grado[grado_nombre]['estudiantes_detalle'].append({
                    'id': estudiante.id,
                    'nombre': estudiante.usuario.get_full_name(),
                    'estado': estado_display,
                    'hora_registro': registro.fecha.strftime('%H:%M') if registro else None
                })

        # Calcular porcentajes
        for grado_data in asistencia_por_grado.values():
            total = grado_data['total_estudiantes']
            if total > 0:
                grado_data['porcentaje_asistencia'] = round(
                    (grado_data['presentes'] / total) * 100, 1
                )
            else:
                grado_data['porcentaje_asistencia'] = 0

        data = {
            'fecha_consulta': fecha_consulta.strftime('%Y-%m-%d'),
            'fecha_display': fecha_consulta.strftime('%d de %B de %Y'),
            'resumen_general': {
                'total_estudiantes': sum(g['total_estudiantes'] for g in asistencia_por_grado.values()),
                'total_presentes': sum(g['presentes'] for g in asistencia_por_grado.values()),
                'total_ausentes': sum(g['ausentes'] for g in asistencia_por_grado.values()),
                'total_tardanzas': sum(g['tardanzas'] for g in asistencia_por_grado.values())
            },
            'asistencia_por_grado': dict(asistencia_por_grado)
        }
        
        return Response(data)
        
    except Exception as e:
        return Response({'error': f'Error inesperado: {str(e)}'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_alertas_bienestar_data(request):
    """
    API que devuelve las alertas de bienestar para revisión
    """
    user = request.user
    if not (user.is_staff and user.rol in ['coordinador', 'administrador', 'psicologo']):
        return Response({'error': 'Acceso denegado.'}, status=403)

    try:
        user_inst = getattr(user, 'institucion_asociada', None)
        
        # Filtros opcionales
        categoria = request.GET.get('categoria', 'todas')
        
        alertas_query = AnotacionObservador.objects.filter(
            requiere_revision=True
        ).select_related('estudiante__usuario', 'estudiante__grado_actual', 'registrado_por')
        
        if user_inst:
            alertas_query = alertas_query.filter(estudiante__institucion=user_inst)
        
        # Aplicar filtro de categoría si se especifica
        if categoria != 'todas':
            alertas_query = alertas_query.filter(tipo_situacion_ia=categoria.upper())

        alertas = alertas_query.order_by('-fecha_hora')[:50]  # Limitar a 50 más recientes

        # Estadísticas por categoría
        estadisticas = {
            'total': alertas_query.count(),
            'criticas': alertas_query.filter(tipo_situacion_ia='CRITICA').count(),
            'activas': alertas_query.filter(requiere_revision=True).count(),
        }

        alertas_data = []
        for alerta in alertas:
            alertas_data.append({
                'id': alerta.id,
                'estudiante': {
                    'id': alerta.estudiante.id,
                    'nombre': alerta.estudiante.usuario.get_full_name(),
                    'grado': alerta.estudiante.grado_actual.nombre if alerta.estudiante.grado_actual else 'Sin grado'
                },
                'descripcion': alerta.descripcion,
                'tipo_situacion': alerta.tipo_situacion_ia or 'NINGUNO',
                'fecha': alerta.fecha_hora.strftime('%d/%m/%Y %H:%M'),
                'registrado_por': alerta.registrado_por.get_full_name() if alerta.registrado_por else 'Sistema'
            })

        data = {
            'estadisticas': estadisticas,
            'alertas': alertas_data,
            'categorias_disponibles': [
                {'id': 'todas', 'label': 'Todas'},
                {'id': 'academico', 'label': 'Académico'},
                {'id': 'comportamiento', 'label': 'Comportamiento'},
                {'id': 'emocional', 'label': 'Emocional'},
                {'id': 'familiar', 'label': 'Familiar'}
            ]
        }
        
        return Response(data)
        
    except Exception as e:
        return Response({'error': f'Error inesperado: {str(e)}'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_citas_supervision_data(request):
    """
    API que devuelve las citas programadas para supervisión
    """
    user = request.user
    if not (user.is_staff and user.rol in ['coordinador', 'administrador']):
        return Response({'error': 'Acceso denegado.'}, status=403)

    try:
        user_inst = getattr(user, 'institucion_asociada', None)
        
        citas_query = CitaReunion.objects.select_related(
            'docente__usuario', 
            'familiar__usuario', 
            'estudiante__usuario'
        )
        
        if user_inst:
            citas_query = citas_query.filter(docente__institucion=user_inst)
        
        # Filtros opcionales
        estado = request.GET.get('estado', 'todas')
        if estado != 'todas':
            citas_query = citas_query.filter(estado=estado.upper())

        citas = citas_query.order_by('-fecha_hora_inicio')[:100]  # Últimas 100 citas

        # Estadísticas
        estadisticas = {
            'total': citas_query.count(),
            'pendientes': citas_query.filter(estado='PENDIENTE').count(),
            'confirmadas': citas_query.filter(estado='CONFIRMADA').count(),
            'realizadas': citas_query.filter(estado='REALIZADA').count(),
            'canceladas': citas_query.filter(estado='CANCELADA').count()
        }

        citas_data = []
        for cita in citas:
            citas_data.append({
                'id': cita.id,
                'docente': cita.docente.usuario.get_full_name(),
                'familiar': cita.familiar.usuario.get_full_name(),
                'estudiante': cita.estudiante.usuario.get_full_name(),
                'fecha_hora': cita.fecha_hora_inicio.strftime('%d/%m/%Y %H:%M'),
                'asunto': cita.asunto,
                'estado': cita.get_estado_display(),
                'estado_codigo': cita.estado,
                'notas': cita.notas_reunion or '',
                'acuerdos': cita.acuerdos_establecidos or ''
            })

        data = {
            'estadisticas': estadisticas,
            'citas': citas_data,
            'estados_disponibles': [
                {'id': 'todas', 'label': 'Todas'},
                {'id': 'pendiente', 'label': 'Pendientes'},
                {'id': 'confirmada', 'label': 'Confirmadas'},
                {'id': 'realizada', 'label': 'Realizadas'},
                {'id': 'cancelada', 'label': 'Canceladas'}
            ]
        }
        
        return Response(data)
        
    except Exception as e:
        return Response({'error': f'Error inesperado: {str(e)}'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_noticias_gestion_data(request):
    """
    API que devuelve las noticias para gestión por coordinadores
    """
    user = request.user
    if not (user.is_staff and user.rol in ['coordinador', 'administrador']):
        return Response({'error': 'Acceso denegado.'}, status=403)

    try:
        user_inst = getattr(user, 'institucion_asociada', None)
        
        noticias_query = Noticia.objects.select_related('publicado_por')
        
        if user_inst:
            noticias_query = noticias_query.filter(institucion=user_inst)
        
        noticias = noticias_query.order_by('-fecha_publicacion')[:50]

        noticias_data = []
        for noticia in noticias:
            noticias_data.append({
                'id': noticia.id,
                'titulo': noticia.titulo,
                'resumen': noticia.resumen or noticia.contenido[:100] + '...',
                'fecha_publicacion': noticia.fecha_publicacion.strftime('%d/%m/%Y'),
                'publicado_por': noticia.publicado_por.get_full_name() if noticia.publicado_por else 'Sistema',
                'activa': noticia.activa if hasattr(noticia, 'activa') else True
            })

        data = {
            'noticias': noticias_data,
            'total_noticias': noticias_query.count()
        }
        
        return Response(data)
        
    except Exception as e:
        return Response({'error': f'Error inesperado: {str(e)}'}, status=500)

@api_view(['GET'])
def mis_menciones_api_view(request):
    """ API para la PANTALLA DE DETALLE de menciones y honores del estudiante. """
    user = request.user
    if not hasattr(user, 'estudiante'): return Response({'error': 'Acceso denegado.'}, status=403)
    try:
        menciones = MencionReconocimiento.objects.filter(estudiante=user.estudiante).order_by('-fecha_emision')
        data = [{'id_mencion': m.pk, 'titulo': m.titulo, 'descripcion': m.descripcion, 'fecha': m.fecha_emision.strftime('%d %b, %Y')} for m in menciones]
        return Response(data)
    except Exception as e: return Response({'error': f'Ocurrió un error: {str(e)}'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated, EstaAlDiaPermission])
def mi_historial_asistencia_api_view(request):
    """ API para la PANTALLA DE DETALLE del historial de asistencia. """
    user = request.user
    if not hasattr(user, 'estudiante'): return Response({'error': 'Acceso denegado.'}, status=403)
    try:
        estudiante = user.estudiante
        periodo_activo = PeriodoAcademico.objects.filter(institucion=estudiante.institucion, activo=True).first()
        if not periodo_activo: return Response([])
        historial = RegistroAsistencia.objects.filter(estudiante=estudiante, curso__periodo_academico=periodo_activo).select_related('curso__materia').order_by('-fecha')
        data = [{'fecha': r.fecha.strftime('%d/%m/%Y'), 'materia': r.curso.materia.nombre_materia if r.curso else 'N/A', 'estado': r.get_estado_display()} for r in historial]
        return Response(data)
    except Exception as e: return Response({'error': f'Ocurrió un error: {str(e)}'}, status=500)     

def get_cursos_docente_periodo_activo(user):
    docente = user.docente
    periodo_activo = PeriodoAcademico.objects.filter(institucion=docente.institucion, activo=True).first()
    if not periodo_activo: return Curso.objects.none()
    return Curso.objects.filter(docentes_asignados=docente, periodo_academico=periodo_activo).select_related('materia', 'grado')

@api_view(['GET'])
def api_seleccionar_curso_asistencia(request):
    """ API que devuelve TODOS los cursos de la institución para asistir. """
    user = request.user
    if not hasattr(user, 'docente'): return Response({'error': 'Acceso denegado.'}, status=403)
    periodo_activo = PeriodoAcademico.objects.filter(institucion=user.docente.institucion, activo=True).first()
    cursos = Curso.objects.filter(periodo_academico=periodo_activo).select_related('materia', 'grado')
    data = [{'id_curso': c.id, 'nombre_curso': f"{c.materia.nombre_materia} - {c.grado.nombre}"} for c in cursos]
    return Response(data)

@api_view(['GET'])
def api_seleccionar_curso_libro_notas(request):
    """ API que devuelve los cursos del docente para el Libro de Notas. """
    user = request.user
    if not hasattr(user, 'docente'): return Response({'error': 'Acceso denegado.'}, status=403)
    cursos = get_cursos_docente_periodo_activo(user)
    data = [{'id_curso': c.id, 'nombre_curso': f"{c.materia.nombre_materia} - {c.grado.nombre}"} for c in cursos]
    return Response(data)

@api_view(['GET'])
def api_docente_lista_actividades(request):
    """ API que devuelve la lista de actividades creadas por el docente. """
    user = request.user
    if not hasattr(user, 'docente'): return Response({'error': 'Acceso denegado.'}, status=403)
    cursos_docente = get_cursos_docente_periodo_activo(user)
    actividades = ActividadCalificable.objects.filter(curso__in=cursos_docente).select_related('curso__materia', 'tipo_actividad')
    data = [{'id': a.id, 'titulo': a.titulo, 'curso': a.curso.materia.nombre_materia, 'tipo': a.tipo_actividad.nombre} for a in actividades]
    return Response(data)

@api_view(['GET'])
def api_docente_lista_categorias(request):
    """ API para la lista de categorías del docente. """
    user = request.user
    if not hasattr(user, 'docente'): return Response({'error': 'Acceso denegado.'}, status=403)
    categorias = TipoActividad.objects.filter(institucion=user.docente.institucion)
    data = [{'id': c.id, 'nombre': c.nombre, 'porcentaje': c.porcentaje} for c in categorias]
    return Response(data)

@api_view(['GET'])
def api_docente_lista_descriptores(request):
    """ API para la lista de descriptores del docente. """
    user = request.user
    if not hasattr(user, 'docente'): return Response({'error': 'Acceso denegado.'}, status=403)
    descriptores = DescriptorLogro.objects.filter(creado_por=user).select_related('materia')
    data = [{'id': d.id, 'descripcion': d.descripcion, 'materia': d.materia.nombre_materia} for d in descriptores]
    return Response(data)

@api_view(['GET'])
def api_docente_lista_materiales(request):
    """ API para la lista de materiales del docente. """
    user = request.user
    if not hasattr(user, 'docente'): return Response({'error': 'Acceso denegado.'}, status=403)
    materiales = ArchivoPlanAcademico.objects.filter(subido_por=user)
    data = [{'id': m.id, 'nombre': m.nombre_archivo_descriptivo, 'url': request.build_absolute_uri(m.archivo.url)} for m in materiales if m.archivo]
    return Response(data)

@api_view(['GET'])
def api_seleccionar_curso_leccion(request):
    """ API que devuelve los cursos del docente para registrar lecciones. """
    return api_seleccionar_curso_libro_notas(request) # Reutiliza la misma lógica y vista

@api_view(['GET'])
def api_seleccionar_estudiante_observador(request):
    """ API que devuelve la lista de estudiantes del docente para el observador. """
    user = request.user
    if not hasattr(user, 'docente'): return Response({'error': 'Acceso denegado.'}, status=403)
    cursos_docente = get_cursos_docente_periodo_activo(user)
    grados_ids = cursos_docente.values_list('grado_id', flat=True).distinct()
    estudiantes = Estudiante.objects.filter(grado_actual_id__in=grados_ids).select_related('usuario')
    data = [{'id_estudiante': e.pk, 'nombre_completo': e.usuario.get_full_name()} for e in estudiantes]
    return Response(data)

@api_view(['GET'])
def api_gestionar_disponibilidad(request):
    """ API para la lista de disponibilidades del docente. """
    user = request.user
    if not hasattr(user, 'docente'): return Response({'error': 'Acceso denegado.'}, status=403)
    disponibilidades = DisponibilidadDocente.objects.filter(docente=user.docente)
    data = [{'id': d.id, 'dia': d.get_dia_semana_display(), 'inicio': d.hora_inicio, 'fin': d.hora_fin} for d in disponibilidades]
    return Response(data)

@api_view(['GET'])
def api_docente_lista_menciones(request):
    """ API para la lista de menciones creadas por el docente. """
    user = request.user
    if not hasattr(user, 'docente'): return Response({'error': 'Acceso denegado.'}, status=403)
    menciones = MencionReconocimiento.objects.filter(otorgado_por=user.docente).select_related('estudiante__usuario')
    data = [{'id': m.pk, 'tipo': m.tipo, 'estudiante': m.estudiante.usuario.get_full_name(), 'fecha': m.fecha_otorgamiento} for m in menciones]
    return Response(data)
    
@api_view(['GET'])
def api_seleccionar_curso_reporte_minima(request):
    """ API que devuelve los cursos del docente para el reporte de nota mínima. """
    return api_seleccionar_curso_libro_notas(request) # Reutiliza la misma lógica

@api_view(['GET'])
def api_panel_director_grupo(request):
    """ API que devuelve los datos del panel del director de grupo. """
    user = request.user
    if not hasattr(user, 'docente'): return Response({'error': 'Acceso denegado.'}, status=403)
    # ... (lógica para calcular y devolver los datos del panel)
    return Response({"mensaje": "Datos del panel del director de grupo"})               

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_admin_asistencia_diaria(request):
    """ API que devuelve un resumen de la asistencia del día. """
    user = request.user
    if not user.is_staff: return Response({'error': 'Acceso denegado.'}, status=403)
    
    user_inst = getattr(user, 'institucion_asociada', None)
    if not user_inst: return Response({'error': 'Usuario no asociado a una institución.'}, status=400)

    hoy = timezone.localdate()
    registros_hoy = RegistroAsistencia.objects.filter(estudiante__institucion=user_inst, fecha__date=hoy)
    
    total_estudiantes = Estudiante.objects.filter(institucion=user_inst, usuario__is_active=True).count()
    presentes = registros_hoy.filter(estado='PRESENTE').count()
    ausentes = registros_hoy.filter(estado='AUSENTE').count()
    tardanzas = registros_hoy.filter(estado='TARDANZA').count()

    data = {
        'fecha': hoy.strftime('%d de %B de %Y'),
        'total_estudiantes': total_estudiantes,
        'presentes': presentes,
        'ausentes': ausentes,
        'tardanzas': tardanzas,
        'sin_registro': total_estudiantes - (presentes + ausentes + tardanzas)
    }
    return Response(data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_dashboard_bienestar(request):
    """ API que devuelve la lista de alertas de bienestar. """
    user = request.user
    if not user.is_staff: return Response({'error': 'Acceso denegado.'}, status=403)
    
    user_inst = getattr(user, 'institucion_asociada', None)
    alertas = AnotacionObservador.objects.filter(requiere_revision=True)
    if user_inst:
        alertas = alertas.filter(estudiante__institucion=user_inst)
    
    data = [{
        'id': a.id, 
        'estudiante': a.estudiante.usuario.get_full_name(), 
        'fecha': a.fecha_hora, 
        'descripcion_corta': a.descripcion[:70] + '...'
    } for a in alertas.select_related('estudiante__usuario')]
    return Response(data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_supervisar_citas(request):
    """ API que devuelve la lista de todas las citas agendadas. """
    user = request.user
    if not user.is_staff: return Response({'error': 'Acceso denegado.'}, status=403)

    user_inst = getattr(user, 'institucion_asociada', None)
    citas = CitaReunion.objects.all()
    if user_inst:
        citas = citas.filter(docente__institucion=user_inst)
        
    data = [{
        'id': c.id, 
        'docente': c.docente.usuario.get_full_name(), 
        'familiar': c.familiar.usuario.get_full_name(),
        'estudiante': c.estudiante.usuario.get_full_name(),
        'fecha_hora': c.fecha_hora_inicio,
        'estado': c.get_estado_display()
    } for c in citas.select_related('docente__usuario', 'familiar__usuario', 'estudiante__usuario')]
    return Response(data)    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_admin_asistencia_diaria(request):
    user = request.user
    if not user.is_staff: return Response({'error': 'Acceso denegado.'}, status=403)
    user_inst = getattr(user, 'institucion_asociada', None)
    if not user_inst: return Response({'error': 'Usuario no asociado a una institución.'}, status=400)
    
    hoy = timezone.localdate()
    registros_hoy = RegistroAsistencia.objects.filter(estudiante__institucion=user_inst, fecha__date=hoy)
    total_estudiantes = Estudiante.objects.filter(institucion=user_inst, usuario__is_active=True).count()
    presentes = registros_hoy.filter(estado='PRESENTE').count()
    
    data = {
        'fecha': hoy.strftime('%d de %B de %Y'),
        'total_estudiantes': total_estudiantes,
        'presentes': presentes,
        'ausentes': registros_hoy.filter(estado='AUSENTE').count(),
        'tardanzas': registros_hoy.filter(estado='TARDANZA').count(),
    }
    return Response(data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_dashboard_bienestar(request):
    user = request.user
    if not user.is_staff: return Response({'error': 'Acceso denegado.'}, status=403)
    user_inst = getattr(user, 'institucion_asociada', None)
    alertas = AnotacionObservador.objects.filter(requiere_revision=True)
    if user_inst:
        alertas = alertas.filter(estudiante__institucion=user_inst)
    
    data = [{
        'id': a.id, 
        'estudiante': a.estudiante.usuario.get_full_name(), 
        'fecha': a.fecha_hora.strftime('%d/%m/%Y'), 
        'descripcion_corta': a.descripcion[:70] + '...'
    } for a in alertas.select_related('estudiante__usuario')]
    return Response(data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_supervisar_citas(request):
    user = request.user
    if not user.is_staff: return Response({'error': 'Acceso denegado.'}, status=403)
    user_inst = getattr(user, 'institucion_asociada', None)
    citas = CitaReunion.objects.all()
    if user_inst:
        citas = citas.filter(docente__institucion=user_inst)
        
    data = [{
        'id': c.id, 
        'docente': c.docente.usuario.get_full_name(), 
        'familiar': c.familiar.usuario.get_full_name(),
        'estudiante': c.estudiante.usuario.get_full_name(),
        'fecha_hora': c.fecha_hora_inicio.strftime('%d/%m/%Y %I:%M %p'),
        'estado': c.get_estado_display()
    } for c in citas.select_related('docente__usuario', 'familiar__usuario', 'estudiante__usuario')]
    return Response(data)    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_familiar_calificaciones_view(request, estudiante_pk):
    """ API que devuelve los cursos de un estudiante para que el familiar vea las calificaciones. """
    user = request.user
    if not hasattr(user, 'familiar'): return Response({'error': 'Acceso denegado.'}, status=403)
    try:
        estudiante = user.familiar.estudiantes_asociados.get(pk=estudiante_pk)
        periodo_activo = PeriodoAcademico.objects.filter(institucion=estudiante.institucion, activo=True).first()
        if not (periodo_activo and estudiante.grado_actual): return Response([])
        
        cursos = Curso.objects.filter(grado=estudiante.grado_actual, periodo_academico=periodo_activo).select_related('materia')
        data = [{'id_materia': c.materia.pk, 'nombre_materia': c.materia.nombre_materia} for c in cursos]
        return Response(data)
    except Estudiante.DoesNotExist:
        return Response({'error': 'No tienes permiso para ver este estudiante.'}, status=403)
    except Exception as e: return Response({'error': f'Ocurrió un error: {str(e)}'}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_familiar_boletin_view(request, estudiante_pk):
    """ API que devuelve los datos del boletín de un estudiante para su familiar. """
    user = request.user
    if not hasattr(user, 'familiar'): return Response({'error': 'Acceso denegado.'}, status=403)
    try:
        estudiante = user.familiar.estudiantes_asociados.get(pk=estudiante_pk)
        # (Aquí iría la misma lógica de cálculo de la vista `mi_boletin_api_view` para devolver el JSON del boletín)
        # Por simplicidad, devolvemos un mensaje de éxito por ahora.
        return Response({'mensaje': f"Datos del boletín para {estudiante.usuario.get_full_name()}"})
    except Estudiante.DoesNotExist:
        return Response({'error': 'No tienes permiso para ver este estudiante.'}, status=403)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_familiar_deberes_view(request, estudiante_pk):
    """ API que devuelve los deberes de un estudiante para su familiar. """
    user = request.user
    if not hasattr(user, 'familiar'): return Response({'error': 'Acceso denegado.'}, status=403)
    try:
        estudiante = user.familiar.estudiantes_asociados.get(pk=estudiante_pk)
        # (Aquí iría la misma lógica de la vista `mis_deberes_api_view` para devolver la lista de deberes)
        return Response({'mensaje': f"Lista de deberes para {estudiante.usuario.get_full_name()}"})
    except Estudiante.DoesNotExist:
        return Response({'error': 'No tienes permiso para ver este estudiante.'}, status=403)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_familiar_seleccionar_docente(request):
    """ API que devuelve la lista de docentes con los que un familiar puede agendar cita. """
    user = request.user
    if not hasattr(user, 'familiar'): return Response({'error': 'Acceso denegado.'}, status=403)
    try:
        estudiantes_asociados = user.familiar.estudiantes_asociados.all()
        cursos_hijos = Curso.objects.filter(grado__in=[e.grado_actual for e in estudiantes_asociados])
        docentes = Docente.objects.filter(cursos_impartidos__in=cursos_hijos).distinct().select_related('usuario')
        data = [{'id_docente': d.pk, 'nombre_completo': d.usuario.get_full_name()} for d in docentes]
        return Response(data)
    except Exception as e: return Response({'error': f'Ocurrió un error: {str(e)}'}, status=500)    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def detalle_noticia_api_view(request, noticia_pk):
    """
    Devuelve el detalle completo de una noticia específica.
    VERSIÓN CORREGIDA CON LOS NOMBRES DE CAMPO REALES.
    """
    noticia = get_object_or_404(Noticia, pk=noticia_pk)
    
    data = {
        'id': noticia.pk,
        'titulo': noticia.titulo,
        'cuerpo': noticia.contenido,  # <-- CORRECCIÓN: Usa 'contenido' en lugar de 'cuerpo'
        'fecha': noticia.fecha_publicacion.strftime('%d de %B de %Y'),
        'imagen_url': request.build_absolute_uri(noticia.imagen_destacada.url) if noticia.imagen_destacada else None # <-- CORRECCIÓN: Usa 'imagen_destacada'
    }
    return Response(data)

@login_required
def historial_convivencia_view(request):
    if not (request.user.is_superuser or (request.user.is_staff and request.user.rol in ['coordinador', 'psicologo', 'rector'])):
        messages.error(request, "Acceso denegado.")
        return redirect('gestion_academica:inicio_academico')

    # Excluir las situaciones sin clasificación
    anotaciones_convivencia = AnotacionObservador.objects.exclude(
        Q(tipo_situacion_ia='NINGUNO') | Q(tipo_situacion_ia__isnull=True)
    ).select_related(
        'estudiante__usuario', 'estudiante__grado_actual'
    ).order_by('estudiante__usuario__last_name', '-fecha_hora')

    # Agrupar por ID de estudiante para evitar errores de hash
    estudiantes_con_casos = OrderedDict()
    estudiantes_info = {}

    for anotacion in anotaciones_convivencia:
        estudiante = anotacion.estudiante
        if estudiante:
            pk = estudiante.pk
            if pk not in estudiantes_con_casos:
                estudiantes_con_casos[pk] = []
                estudiantes_info[pk] = estudiante
            estudiantes_con_casos[pk].append(anotacion)

    context = {
        'titulo_pagina': "Reporte de Ruta de Convivencia Escolar",
        'estudiantes_con_casos': estudiantes_con_casos,
        'estudiantes_info': estudiantes_info,
    }
    return render(request, 'gestion_academica/historial_convivencia.html', context)

@login_required
def votar_view(request, eleccion_id):
    eleccion = get_object_or_404(Eleccion, pk=eleccion_id)

    if Voto.objects.filter(eleccion=eleccion, votante=request.user.estudiante).exists():
        messages.info(request, "Ya has votado en esta elección.")
        return redirect('gestion_academica:dashboard_estudiante')

    candidatos = Candidato.objects.filter(eleccion=eleccion)

    if request.method == 'POST':
        candidato_id = request.POST.get('candidato')
        candidato = get_object_or_404(Candidato, pk=candidato_id)
        Voto.objects.create(eleccion=eleccion, votante=request.user.estudiante, candidato=candidato)
        messages.success(request, "Tu voto ha sido registrado exitosamente.")
        return redirect('gestion_academica:dashboard_estudiante')

    return render(request, 'gestion_academica/votar.html', {
        'eleccion': eleccion,
        'candidatos': candidatos
    })    

@login_required
@permission_required('gestion_academica.view_eleccion')
def dashboard_eleccion_ia(request, eleccion_id):
    """
    Muestra el análisis de resultados de una elección.
    VERSIÓN CORREGIDA: Obtiene el nombre del candidato correctamente.
    """
    user_inst = getattr(request.user, 'institucion_asociada', None)
    if user_inst:
        eleccion = Eleccion.objects.filter(pk=eleccion_id, institucion=user_inst).first()
    else:
        eleccion = Eleccion.objects.filter(pk=eleccion_id).first()
    if eleccion is None:
        messages.warning(
            request,
            "No hay ninguna elección registrada con ese identificador en su institución, "
            "o aún no ha creado un proceso electoral. Use «Gestión de Elecciones» para crear una.",
        )
        rol = getattr(request.user, 'rol', None)
        if request.user.is_superuser or rol in ('coordinador', 'administrador'):
            return redirect('gestion_academica:dashboard_coordinador')
        return redirect('gestion_academica:inicio_academico')
    
    candidatos = (
        Candidato.objects
        .filter(eleccion=eleccion)
        .select_related('estudiante__usuario', 'estudiante__grado_actual')
        .annotate(total_votos=Count('votos_recibidos'))
        .order_by('-total_votos')
    )

    labels = [c.estudiante.usuario.get_full_name() for c in candidatos]
    votos  = [c.total_votos for c in candidatos]

    # Totales de participación
    total_votos      = Voto.objects.filter(eleccion=eleccion).count()
    total_elegibles  = Estudiante.objects.filter(
        institucion=eleccion.institucion, activo=True
    ).count()
    porcentaje_part  = round(total_votos / total_elegibles * 100) if total_elegibles else 0

    # Votos por grado para cada candidato
    votos_por_grado_qs = (
        Voto.objects
        .filter(eleccion=eleccion)
        .values('candidato_id', 'votante__grado_actual__nombre')
        .annotate(total=Count('id'))
        .order_by('votante__grado_actual__nombre')
    )
    votos_por_grado = {}
    for row in votos_por_grado_qs:
        cid   = row['candidato_id']
        grado = row['votante__grado_actual__nombre'] or 'Sin grado'
        votos_por_grado.setdefault(cid, {})[grado] = row['total']

    # Lista de grados participantes
    grados_set = set()
    for d in votos_por_grado.values():
        grados_set.update(d.keys())
    grados_participantes = sorted(grados_set)

    context = {
        'eleccion':              eleccion,
        'candidatos':            candidatos,
        'labels':                json.dumps(labels),
        'votos':                 json.dumps(votos),
        'total_votos':           total_votos,
        'total_elegibles':       total_elegibles,
        'porcentaje_part':       porcentaje_part,
        'votos_por_grado':       votos_por_grado,
        'grados_participantes':  grados_participantes,
        'titulo_pagina':         f"Tablero Electoral: {eleccion.nombre}",
    }
    return render(request, 'gestion_academica/dashboard_eleccion_ia.html', context)

@login_required
@permission_required('gestion_academica.view_eleccion')
def acta_eleccion_view(request, eleccion_id):
    eleccion = get_object_or_404(Eleccion, pk=eleccion_id)
    candidatos = Candidato.objects.filter(eleccion=eleccion).annotate(
        total_votos=Count('voto')
    ).order_by('-total_votos')

    total_votantes = Voto.objects.filter(eleccion=eleccion).count()

    return render(request, 'gestion_academica/acta_eleccion.html', {
        'eleccion': eleccion,
        'candidatos': candidatos,
        'total_votantes': total_votantes,
    })  

@login_required
@require_POST
@permission_required('gestion_academica.change_candidato', raise_exception=True)
def analizar_propuestas_ia_view(request, eleccion_id):
    """
    Inicia las tareas de Celery para analizar las propuestas de todos los
    candidatos de una elección que aún no tengan un análisis.
    """
    eleccion = get_object_or_404(Eleccion, pk=eleccion_id)
    
    # Buscamos solo los candidatos que no tienen un análisis previo
    candidatos_a_analizar = Candidato.objects.filter(eleccion=eleccion, analisis_ia__isnull=True)
    
    if not candidatos_a_analizar.exists():
        messages.info(request, "Todas las propuestas de esta elección ya han sido analizadas.")
    else:
        for candidato in candidatos_a_analizar:
            analizar_propuesta_candidato_task.delay(candidato.id)
        messages.success(request, f"Se ha iniciado el análisis con IA para {candidatos_a_analizar.count()} propuesta(s). Los resultados aparecerán en breve.")
        
    return redirect('gestion_academica:dashboard_eleccion_ia', eleccion_id=eleccion.id)     


@login_required
@permission_required('gestion_academica.add_eleccion', raise_exception=True)
def gestionar_elecciones_view(request):
    """
    Permite a los coordinadores y administradores crear y ver las elecciones.
    VERSIÓN CORREGIDA: Asigna correctamente la institución a la nueva elección.
    """
    institucion = request.user.institucion_asociada
    
    if request.method == 'POST':
        form = EleccionForm(request.POST)
        if form.is_valid():
            # --- INICIO DE LA CORRECCIÓN CLAVE ---
            # 1. Creamos el objeto en memoria sin guardarlo aún en la base de datos.
            nueva_eleccion = form.save(commit=False)
            
            # 2. Asignamos la institución del usuario actual al nuevo objeto de elección.
            nueva_eleccion.institucion = institucion
            
            # 3. Ahora sí, guardamos el objeto completo en la base de datos.
            nueva_eleccion.save()
            # --- FIN DE LA CORRECCIÓN CLAVE ---
            
            messages.success(request, f"La elección '{nueva_eleccion.nombre}' ha sido creada exitosamente.")
            return redirect('gestion_academica:gestionar_elecciones')
    else:
        form = EleccionForm()

    elecciones = Eleccion.objects.filter(institucion=institucion).order_by('-fecha_inicio')
    
    context = {
        'titulo_pagina': "Gestión de Elecciones",
        'form': form,
        'elecciones': elecciones
    }
    return render(request, 'gestion_academica/gestionar_elecciones.html', context) 

@login_required
@permission_required('gestion_academica.add_candidato', raise_exception=True)
def gestionar_candidatos_view(request, eleccion_id):
    """
    Permite añadir y ver los candidatos para una elección específica.
    """
    eleccion = get_object_or_404(Eleccion, pk=eleccion_id, institucion=request.user.institucion_asociada)
    
    if request.method == 'POST':
        # Pasamos la institución al formulario para que filtre correctamente a los estudiantes
        form = CandidatoForm(request.POST, request.FILES, institucion=request.user.institucion_asociada)
        if form.is_valid():
            candidato = form.save(commit=False)
            candidato.eleccion = eleccion
            # La institución se hereda del modelo Eleccion, pero la añadimos por consistencia
            candidato.institucion = eleccion.institucion
            candidato.save()
            messages.success(request, f"Candidato '{candidato.estudiante.usuario.get_full_name()}' añadido a la elección.")
            return redirect('gestion_academica:gestionar_candidatos', eleccion_id=eleccion.id)
    else:
        form = CandidatoForm(institucion=request.user.institucion_asociada)

    candidatos = (
        Candidato.objects
        .filter(eleccion=eleccion)
        .select_related('estudiante__usuario', 'estudiante__grado_actual')
        .annotate(total_votos=Count('votos_recibidos'))
        .order_by('estudiante__grado_actual__nombre', 'estudiante__usuario__last_name')
    )

    context = {
        'titulo_pagina': f"Gestionar Candidatos: {eleccion.nombre}",
        'form': form,
        'eleccion': eleccion,
        'candidatos': candidatos,
    }
    return render(request, 'gestion_academica/gestionar_candidatos.html', context)


@login_required
@permission_required('gestion_academica.view_eleccion', raise_exception=True)
def detalle_candidato_view(request, candidato_id):
    """
    Perfil completo de un candidato: propuesta, análisis IA y desglose de votos por grado.
    """
    candidato = get_object_or_404(
        Candidato.objects.select_related(
            'estudiante__usuario', 'estudiante__grado_actual', 'eleccion'
        ).annotate(total_votos=Count('votos_recibidos')),
        pk=candidato_id,
    )
    eleccion = candidato.eleccion

    # Votos totales de la elección
    total_votos_eleccion = Voto.objects.filter(eleccion=eleccion).count()
    total_elegibles = Estudiante.objects.filter(
        institucion=eleccion.institucion, activo=True
    ).count()

    # Desglose por grado del votante
    votos_por_grado = (
        Voto.objects
        .filter(candidato=candidato)
        .values('votante__grado_actual__nombre')
        .annotate(total=Count('id'))
        .order_by('-total')
    )

    # Posición en el ranking
    candidatos_ranking = (
        Candidato.objects
        .filter(eleccion=eleccion)
        .annotate(total_votos=Count('votos_recibidos'))
        .order_by('-total_votos')
    )
    posicion = next(
        (i + 1 for i, c in enumerate(candidatos_ranking) if c.pk == candidato.pk),
        None
    )

    context = {
        'titulo_pagina':          f"Candidato: {candidato.estudiante.usuario.get_full_name()}",
        'candidato':              candidato,
        'eleccion':               eleccion,
        'total_votos_eleccion':   total_votos_eleccion,
        'total_elegibles':        total_elegibles,
        'votos_por_grado':        votos_por_grado,
        'posicion':               posicion,
        'total_candidatos':       candidatos_ranking.count(),
    }
    return render(request, 'gestion_academica/detalle_candidato.html', context)


class PreguntaCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """ Vista para crear una nueva Pregunta junto con sus Opciones. """
    model = Pregunta
    form_class = PreguntaForm
    template_name = 'gestion_academica/pregunta_formulario.html'
    permission_required = 'gestion_academica.add_pregunta'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        self.actividad = get_object_or_404(ActividadCalificable, pk=self.kwargs['actividad_pk'])
        context['actividad'] = self.actividad
        context['titulo_formulario'] = "Crear Nueva Pregunta"
        if self.request.POST:
            context['opciones_formset'] = OpcionFormSet(self.request.POST, prefix='opciones')
        else:
            context['opciones_formset'] = OpcionFormSet(prefix='opciones')
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        opciones_formset = context['opciones_formset']
        
        with transaction.atomic():
            self.actividad = get_object_or_404(ActividadCalificable, pk=self.kwargs['actividad_pk'])
            form.instance.actividad = self.actividad
            form.instance.institucion = self.request.user.institucion_asociada
            self.object = form.save()

            if opciones_formset.is_valid():
                opciones_formset.instance = self.object
                opciones_formset.save()
            else:
                return self.form_invalid(form)

        messages.success(self.request, "Pregunta creada exitosamente.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('gestion_academica:gestionar_preguntas_actividad', kwargs={'actividad_pk': self.kwargs['actividad_pk']})


class PreguntaUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """ Vista para editar una Pregunta existente y sus Opciones. """
    model = Pregunta
    form_class = PreguntaForm
    template_name = 'gestion_academica/pregunta_formulario.html'
    permission_required = 'gestion_academica.change_pregunta'

    def get_queryset(self):
        return Pregunta.objects.filter(institucion=self.request.user.institucion_asociada)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['actividad'] = self.object.actividad
        context['titulo_formulario'] = "Editar Pregunta"
        if self.request.POST:
            context['opciones_formset'] = OpcionFormSet(self.request.POST, instance=self.object, prefix='opciones')
        else:
            context['opciones_formset'] = OpcionFormSet(instance=self.object, prefix='opciones')
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        opciones_formset = context['opciones_formset']
        
        if opciones_formset.is_valid():
            with transaction.atomic():
                self.object = form.save()
                opciones_formset.instance = self.object
                opciones_formset.save()
        else:
            return self.form_invalid(form)

        messages.success(self.request, "Pregunta actualizada exitosamente.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('gestion_academica:gestionar_preguntas_actividad', kwargs={'actividad_pk': self.object.actividad.pk})


class PreguntaDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Pregunta
    template_name = 'gestion_academica/pregunta_confirmar_eliminar.html'
    context_object_name = 'pregunta'
    permission_required = 'gestion_academica.delete_pregunta'

    def get_queryset(self):
        return Pregunta.objects.filter(institucion=self.request.user.institucion_asociada)

    def get_success_url(self):
        messages.success(self.request, "La pregunta ha sido eliminada.")
        return reverse('gestion_academica:gestionar_preguntas_actividad', kwargs={'actividad_pk': self.object.actividad.pk})

class GestionarPreguntasActividadView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = ActividadCalificable
    template_name = 'gestion_academica/gestionar_preguntas.html'
    context_object_name = 'actividad'
    pk_url_kwarg = 'actividad_pk'
    permission_required = 'gestion_academica.change_actividadcalificable'

    def get_queryset(self):
        if hasattr(self.request.user, 'docente'):
            return ActividadCalificable.objects.filter(curso__docentes_asignados=self.request.user.docente)
        return ActividadCalificable.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Gestionar: {self.object.titulo}"
        context['preguntas'] = self.object.preguntas.order_by('orden')
        
        # Pasamos el formulario de configuración a la plantilla
        if 'config_form' not in context:
            context['config_form'] = ActividadConfigForm(instance=self.object)
            
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = ActividadConfigForm(request.POST, instance=self.object)

        if form.is_valid():
            form.save()
            messages.success(request, "La configuración de la actividad ha sido actualizada.")
            return redirect(self.request.path_info)
        else:
            messages.error(request, "Hubo un error al guardar la configuración.")
            context = self.get_context_data()
            context['config_form'] = form
            return self.render_to_response(context) 
             

@login_required
@permission_required('gestion_academica.add_docente')
def descargar_plantilla_docentes(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Docentes"

    encabezados = [
        "username", "first_name", "last_name", "email",
        "documento", "codigo_docente", "especialidad",
        "rol"
    ]
    ws.append(encabezados)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    # -------- ROL DESPLEGABLE --------
    # Extraer todos los roles únicos del modelo Usuario
    roles_disponibles = [rol[0] for rol in Usuario.ROLES]

    # Crear hoja oculta con roles válidos
    ws_roles = wb.create_sheet(title="Roles")
    for i, rol in enumerate(roles_disponibles, start=1):
        ws_roles[f"A{i}"] = rol
    ws_roles.sheet_state = 'hidden'

    # Crear validación con referencia absoluta a la hoja "Roles"
    formula = f"'Roles'!$A$1:$A${len(roles_disponibles)}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=False)
    dv.error = "Selecciona un valor válido de la lista."
    dv.errorTitle = "Valor inválido"
    dv.prompt = "Selecciona un rol desde la lista desplegable."
    dv.promptTitle = "Rol requerido"
    ws.add_data_validation(dv)

    # Aplicar validación a la columna 'rol' (columna G = 7)
    for row in range(2, 101):  # Desde la fila 2 a la 100
        dv.add(ws.cell(row=row, column=8))  # Columna 8 (índice base 1)

    # -------- FIN VALIDACIÓN --------

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_docentes.xlsx'
    wb.save(response)
    return response


@login_required
@permission_required('gestion_academica.add_docente')
def importar_docentes_excel(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            wb = load_workbook(filename=request.FILES['file'])
            sheet = wb.active
            errores = []
            creados = 0
            correos_enviados = 0
            sin_correo = []
            institucion = get_current_institution(request.user)

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Tomar solo las primeras 8 columnas (ignorar extras como contraseñas de carga anterior)
                    row8 = list(row)[:8]
                    if len(row8) < 8:
                        row8 += [None] * (8 - len(row8))
                    username, first_name, last_name, email, doc_id, cod_doc, especialidad, rol = row8

                    if not username or not doc_id:
                        errores.append(f"Fila {i}: Falta username o documento.")
                        continue

                    user, creado = Usuario.objects.get_or_create(username=username, defaults={
                        "first_name": first_name,
                        "last_name": last_name,
                        "email": email,
                        "rol": rol,
                        "institucion_asociada": institucion
                    })

                    if creado:
                        creados += 1
                        temp_password = get_random_string(
                            length=16,
                            allowed_chars='abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789',
                        )
                        user.set_password(temp_password)
                        user.save()

                        nombre_docente = f"{first_name or ''} {last_name or ''}".strip() or username
                        if email:
                            try:
                                from admisiones.utils import enviar_correo_dinamico
                                html = f"""
                                <div style="font-family:Arial,sans-serif;max-width:520px;margin:auto;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;">
                                  <div style="background:linear-gradient(135deg,#4f46e5,#7c3aed);padding:28px 32px;text-align:center;">
                                    <h2 style="color:#fff;margin:0;font-size:22px;">Bienvenido/a a {institucion.nombre}</h2>
                                    <p style="color:#c7d2fe;margin:6px 0 0;">Tu cuenta docente ha sido creada</p>
                                  </div>
                                  <div style="padding:28px 32px;background:#fff;">
                                    <p style="color:#374151;">Hola <strong>{nombre_docente}</strong>,</p>
                                    <p style="color:#374151;">Tu cuenta en <strong>Halu Plataforma Escolar</strong> está lista. Estas son tus credenciales de acceso:</p>
                                    <div style="background:#f3f4f6;border-radius:8px;padding:16px 20px;margin:20px 0;">
                                      <p style="margin:0 0 8px;color:#6b7280;font-size:13px;">USUARIO</p>
                                      <p style="margin:0 0 16px;font-size:18px;font-weight:700;color:#1f2937;letter-spacing:1px;">{username}</p>
                                      <p style="margin:0 0 8px;color:#6b7280;font-size:13px;">CONTRASEÑA TEMPORAL</p>
                                      <p style="margin:0;font-size:18px;font-weight:700;color:#4f46e5;letter-spacing:2px;">{temp_password}</p>
                                    </div>
                                    <p style="color:#ef4444;font-size:13px;">⚠️ Por seguridad, cambia tu contraseña la primera vez que inicies sesión.</p>
                                    <p style="color:#6b7280;font-size:12px;margin-top:24px;">Si tienes preguntas, contacta al administrador de tu institución.</p>
                                  </div>
                                </div>
                                """
                                enviar_correo_dinamico(
                                    institucion=institucion,
                                    asunto=f"Bienvenido/a a {institucion.nombre} — Tus credenciales de acceso",
                                    destinatarios=[email],
                                    html_content=html,
                                )
                                correos_enviados += 1
                            except Exception as email_exc:
                                sin_correo.append(f"{username} ({email_exc})")
                        else:
                            sin_correo.append(f"{username} (sin correo registrado)")

                    Docente.objects.get_or_create(usuario=user, defaults={
                        "documento_identidad": doc_id,
                        "codigo_docente": cod_doc,
                        "especialidad": especialidad,
                        "institucion": institucion
                    })

                except Exception as e:
                    errores.append(f"Fila {i}: {str(e)}")

            # ── Resumen ──
            if creados:
                if correos_enviados == creados:
                    messages.success(request, f"✅ {creados} docente(s) creado(s). Se enviaron las credenciales por correo a cada uno.")
                elif correos_enviados > 0:
                    messages.success(request, f"✅ {creados} docente(s) creado(s). Correos enviados: {correos_enviados}.")
                    for u in sin_correo:
                        messages.warning(request, f"No se pudo enviar correo a: {u}")
                else:
                    messages.warning(request, f"{creados} docente(s) creado(s), pero no se pudo enviar ningún correo.")
                    for u in sin_correo:
                        messages.warning(request, f"Sin correo: {u}")
            elif not errores:
                messages.info(request, "No se crearon docentes nuevos (ya existían todos los usuarios del archivo).")

            if errores:
                messages.warning(request, f"Se encontraron {len(errores)} error(es) en el archivo:")
                for e in errores:
                    messages.warning(request, e)

            return redirect('gestion_academica:lista_docentes')
    else:
        form = UploadFileForm()

    return render(request, 'gestion_academica/docente_importar_excel.html', {'form': form})   

@login_required
def generar_carnet_docente(request, pk):
    """
    Genera la vista HTML del carnet digital del docente con código QR (basado en UUID seguro).
    """

    docente = get_object_or_404(Docente.objects.select_related('usuario', 'institucion'), pk=pk)

    # Seguridad multiinstitucional
    if not request.user.is_superuser and docente.institucion != request.user.institucion_asociada:
        return HttpResponseForbidden("No tienes permiso para ver este carnet.")

    # Generar código QR con el identificador único del docente
    qr_data = str(docente.qr_identifier)
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    img_str = base64.b64encode(buffer.getvalue()).decode()
    qr_code_url = f"data:image/png;base64,{img_str}"

    context = {
        'docente': docente,
        'qr_code_url': qr_code_url,
        'titulo_pagina': f"Carnet de {docente.usuario.get_full_name()}",
    }

    return render(request, 'gestion_academica/carnet_docente.html', context)

@require_POST
@login_required
def registrar_asistencia_docente_api(request):
    try:
        data = json.loads(request.body)
        qr_identifier = data.get('qr_identifier')

        if not qr_identifier:
            return JsonResponse({'status': 'error', 'message': 'Falta el identificador del QR'}, status=400)

        docente = Docente.objects.select_related('usuario', 'institucion').get(qr_identifier=qr_identifier)

        if not request.user.is_superuser and docente.institucion != request.user.institucion_asociada:
            return JsonResponse({'status': 'error', 'message': 'Docente de otra institución'}, status=403)

        now = timezone.now()
        dia = timezone.localdate(now)
        Est = RegistroAsistenciaDocente.Estado

        with transaction.atomic():
            reg = (
                RegistroAsistenciaDocente.objects.select_for_update()
                .filter(docente=docente, dia=dia)
                .first()
            )
            if reg is None:
                RegistroAsistenciaDocente.objects.create(
                    docente=docente,
                    dia=dia,
                    estado=Est.PRESENTE,
                    hora_entrada=now,
                    hora_salida=None,
                    registrado_por=request.user,
                    institucion=docente.institucion,
                )
                message = "Entrada registrada."
            else:
                if reg.estado == Est.AUSENTE:
                    reg.estado = Est.PRESENTE
                    reg.hora_entrada = now
                    reg.hora_salida = None
                    reg.registrado_por = request.user
                    reg.save()
                    message = "Entrada registrada (se corrige ausencia del día)."
                elif reg.hora_entrada is None:
                    reg.hora_entrada = now
                    reg.registrado_por = request.user
                    reg.save()
                    message = "Entrada registrada."
                elif reg.hora_salida is None:
                    if reg.hora_entrada and now <= reg.hora_entrada:
                        return JsonResponse(
                            {
                                "status": "error",
                                "message": "La salida no puede ser anterior o igual a la entrada.",
                            },
                            status=400,
                        )
                    reg.hora_salida = now
                    reg.registrado_por = request.user
                    reg.save()
                    message = "Salida registrada. Jornada cerrada."
                else:
                    reg.hora_salida = now
                    reg.registrado_por = request.user
                    reg.save()
                    message = "Salida actualizada (corrección de hora de salida)."

        return JsonResponse({'status': 'success', 'message': message})

    except Docente.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Docente no encontrado'}, status=404)
    except Exception as e:
        logger.error(f"Error en asistencia de docente: {e}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': 'Error interno del servidor'}, status=500)

@login_required
def asistencias_docentes_hoy_api(request):
    hoy = localdate()
    institucion = getattr(request.user, 'institucion_asociada', None)

    registros = RegistroAsistenciaDocente.objects.filter(dia=hoy).select_related('docente__usuario')
    if not request.user.is_superuser:
        if institucion:
            registros = registros.filter(institucion=institucion)
        else:
            registros = registros.none()

    registros = registros.order_by('-hora_entrada', '-hora_salida')

    html = render_to_string('gestion_academica/_tabla_asistencias_hoy.html', {
        'asistencias_hoy': registros
    }, request=request)

    return JsonResponse({'html': html})       

@login_required
@permission_required('gestion_academica.add_registroasistenciadocente')
def escaner_asistencia_docente(request):
    # ⚠️ Asegura que el usuario tenga una institución
    institucion = None
    if request.user.is_superuser:
        institucion = None  # Puede ver todo o se puede elegir una
    elif hasattr(request.user, 'institucion'):
        institucion = request.user.institucion
    elif hasattr(request.user, 'docente'):
        institucion = request.user.docente.institucion
    else:
        messages.error(request, "No tienes una institución asociada.")
        return redirect('inicio')

    if not (request.user.is_superuser or request.user.rol in ['coordinador', 'secretaria']):
        messages.error(request, "No tienes permisos para registrar asistencia de docentes.")
        return redirect('inicio')

    # 👉 Aquí podrías cargar la lista de docentes si se requiere en la vista
    # docentes = Docente.objects.filter(institucion=institucion)

    context = {
        'titulo_pagina': "Tomar Asistencia de Docentes",
        'institucion': institucion.nombre if institucion else 'Todas',
        # 'docentes': docentes,
    }
    return render(request, 'gestion_academica/escaner_asistencia_docente.html', context)


class RegistroAsistenciaDocenteListView(ListView):
    model = RegistroAsistenciaDocente
    template_name = 'gestion_academica/asistencias_docentes_list.html'
    context_object_name = 'asistencias'
    paginate_by = 25

    def get_queryset(self):
        queryset = RegistroAsistenciaDocente.objects.select_related('docente__usuario', 'institucion', 'registrado_por')

        # Filtro por institución si el usuario no es superuser
        if not self.request.user.is_superuser and self.request.user.institucion_asociada:
            institucion = self.request.user.institucion_asociada

            # Registrar automáticamente inasistencias si no marcaron hoy
            registrar_inasistencias_docentes(institucion)

            queryset = queryset.filter(institucion=institucion)
            

        # --- Filtros GET ---
        docente_id = self.request.GET.get('docente')
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')

        if docente_id:
            queryset = queryset.filter(docente_id=docente_id)

        if fecha_inicio:
            try:
                d1 = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
                queryset = queryset.filter(dia__gte=d1)
            except ValueError:
                pass  # Fecha inválida

        if fecha_fin:
            try:
                d2 = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
                queryset = queryset.filter(dia__lte=d2)
            except ValueError:
                pass  # Fecha inválida

        return queryset.order_by('-dia', '-hora_entrada')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        docentes_qs = Docente.objects.select_related('usuario')
        if not self.request.user.is_superuser and self.request.user.institucion_asociada:
            docentes_qs = docentes_qs.filter(institucion=self.request.user.institucion_asociada)

        context.update({
            'titulo_pagina': 'Historial de Asistencia Docente',
            'docentes': docentes_qs,
        })
        return context  

@login_required
@permission_required('gestion_academica.view_registroasistenciadocente')
def exportar_asistencia_docentes_excel(request):
    from decimal import Decimal

    queryset = RegistroAsistenciaDocente.objects.select_related(
        'docente__usuario', 'institucion', 'docente', 'registrado_por'
    )

    if not request.user.is_superuser and request.user.institucion_asociada:
        queryset = queryset.filter(institucion=request.user.institucion_asociada)

    docente_id = request.GET.get('docente')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if docente_id:
        queryset = queryset.filter(docente_id=docente_id)

    if fecha_inicio:
        try:
            d1 = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            queryset = queryset.filter(dia__gte=d1)
        except ValueError:
            pass

    if fecha_fin:
        try:
            d2 = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
            queryset = queryset.filter(dia__lte=d2)
        except ValueError:
            pass

    queryset = list(queryset.order_by('dia', 'docente__usuario__last_name', 'docente__usuario__first_name'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle"

    encabezados = [
        'Día',
        'Docente',
        'Documento',
        'Modalidad liq.',
        'Valor hora ref.',
        'Entrada',
        'Salida',
        'Horas',
        'Estado',
        'Institución',
        'Registrado por',
    ]
    ws.append(encabezados)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    resumen_horas = defaultdict(lambda: Decimal('0'))
    resumen_dias_jornada = defaultdict(int)

    for asistencia in queryset:
        h = asistencia.horas_en_institucion
        dpk = asistencia.docente_id
        if h is not None:
            resumen_horas[dpk] += Decimal(str(h))
            resumen_dias_jornada[dpk] += 1
        ws.append([
            asistencia.dia.isoformat() if asistencia.dia else '',
            asistencia.docente.usuario.get_full_name(),
            asistencia.docente.documento_identidad or '',
            asistencia.docente.get_modalidad_liquidacion_display(),
            float(asistencia.docente.valor_hora_docencia) if asistencia.docente.valor_hora_docencia is not None else '',
            asistencia.hora_entrada.strftime('%Y-%m-%d %H:%M') if asistencia.hora_entrada else '',
            asistencia.hora_salida.strftime('%Y-%m-%d %H:%M') if asistencia.hora_salida else '',
            h if h is not None else '',
            asistencia.estado,
            asistencia.institucion.nombre,
            asistencia.registrado_por.get_full_name() if asistencia.registrado_por else 'Automático',
        ])

    ws2 = wb.create_sheet('Resumen periodo')
    ws2.append([
        'Docente',
        'Documento',
        'Modalidad',
        'Valor hora ref.',
        'Días con entrada y salida',
        'Total horas',
        'Estimado (horas × valor hora)',
        'Nota',
    ])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    docentes_por_id = {}
    for a in queryset:
        docentes_por_id[a.docente_id] = a.docente

    for d in sorted(
        docentes_por_id.values(),
        key=lambda x: ((x.usuario.last_name or '').lower(), (x.usuario.first_name or '').lower()),
    ):
        th = resumen_horas.get(d.pk, Decimal('0'))
        djm = resumen_dias_jornada.get(d.pk, 0)
        if d.modalidad_liquidacion == Docente.ModalidadLiquidacion.POR_HORA and d.valor_hora_docencia:
            estimado = (th * d.valor_hora_docencia).quantize(Decimal('0.01'))
            estimado_out = float(estimado)
        else:
            estimado_out = ''
        nota = (
            'Salario fijo: las horas son solo control; el pago mensual es fijo.'
            if d.modalidad_liquidacion == Docente.ModalidadLiquidacion.SALARIO_FIJO
            else 'Referencia; validar con nómina oficial.'
        )
        ws2.append([
            d.usuario.get_full_name(),
            d.documento_identidad or '',
            d.get_modalidad_liquidacion_display(),
            float(d.valor_hora_docencia) if d.valor_hora_docencia is not None else '',
            djm,
            float(th) if th else 0,
            estimado_out,
            nota,
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=asistencia_docentes_nomina.xlsx'
    wb.save(response)
    return response   

class ListaAsistenciasDocenteView(ListView):
    model = RegistroAsistenciaDocente
    template_name = 'gestion_academica/lista_asistencias_docentes.html'
    context_object_name = 'asistencias'
    paginate_by = 20  # opcional: para paginación

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user

        if user.is_superuser:
            return queryset.select_related(
                'docente__usuario', 'institucion', 'registrado_por'
            ).order_by('-dia', '-hora_entrada')

        if hasattr(user, 'docente'):
            queryset = queryset.filter(docente=user.docente)
        elif getattr(user, 'institucion_asociada', None):
            queryset = queryset.filter(institucion=user.institucion_asociada)
        else:
            queryset = queryset.none()

        return queryset.select_related(
            'docente__usuario', 'institucion', 'registrado_por'
        ).order_by('-dia', '-hora_entrada')       

@login_required
def crear_area_academica(request):
    if request.method == 'POST':
        form = AreaAcademicaForm(request.POST, request=request)
        if form.is_valid():
            area = form.save(commit=False)
            # Asignar institución si no es superuser
            if not request.user.is_superuser:
                area.institucion = request.user.institucion_asociada
            area.save()
            form.save_m2m()
            return redirect('gestion_academica:listar_areas_academicas')
    else:
        form = AreaAcademicaForm(request=request)

    # Filtrar materias por institución
    if request.user.is_superuser:
        materias_disponibles = Materia.objects.all()
    else:
        materias_disponibles = Materia.objects.filter(institucion=request.user.institucion_asociada)

    return render(request, 'gestion_academica/areaacademica_form.html', {
        'form': form,
        'titulo_pagina': 'Crear Área Académica',
        'materias_disponibles': materias_disponibles,
    })

@login_required
def listar_areas_academicas(request):
    institucion = request.user.institucion_asociada if not request.user.is_superuser else None
    areas = AreaAcademica.objects.filter(institucion=institucion) if institucion else AreaAcademica.objects.all()
    return render(request, 'gestion_academica/areaacademica_list.html', {
        'areas': areas,
        'titulo_pagina': 'Áreas Académicas',
    })    

@login_required
def editar_area_academica(request, pk):
    area = get_object_or_404(AreaAcademica, pk=pk)

    # Verificación de acceso
    if not request.user.is_superuser and area.institucion != request.user.institucion_asociada:
        return redirect('gestion_academica:listar_areas_academicas')

    if request.method == 'POST':
        form = AreaAcademicaForm(request.POST, instance=area, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Área actualizada correctamente.')
            return redirect('gestion_academica:listar_areas_academicas')
    else:
        form = AreaAcademicaForm(instance=area, request=request)

    # Obtener materias que aún no están asociadas al área
    if request.user.is_superuser:
        materias_disponibles = Materia.objects.exclude(areaacademica=area)
    else:
        materias_disponibles = Materia.objects.filter(
            institucion=request.user.institucion_asociada
        ).exclude(areacademica=area)

    return render(request, 'gestion_academica/areaacademica_form.html', {
        'form': form,
        'titulo_pagina': 'Editar Área Académica',
        'materias_disponibles': materias_disponibles,
    })


@login_required
def eliminar_area_academica(request, pk):
    area = get_object_or_404(AreaAcademica, pk=pk)
    if not request.user.is_superuser and area.institucion != request.user.institucion_asociada:
        return redirect('gestion_academica:listar_areas_academicas')
    
    area.delete()
    messages.success(request, 'Área eliminada correctamente.')
    return redirect('gestion_academica:listar_areas_academicas')    

@login_required
# Te recomiendo crear un permiso específico para esta acción tan delicada
# @permission_required('gestion_academica.can_graduate_students', raise_exception=True)
def proceso_graduacion_view(request):
    institucion = request.user.institucion_asociada
    
    # Busca el último grado de la institución (ej: "Once", "11°", etc.)
    ultimo_grado = Grado.objects.filter(institucion=institucion).order_by('-orden').first()
    
    if not ultimo_grado:
        messages.error(request, "No se ha configurado un grado final en el sistema.")
        return redirect('gestion_academica:inicio_academico') # Redirige a tu dashboard principal

    estudiantes_a_graduar = Estudiante.objects.filter(
        institucion=institucion,
        grado_actual=ultimo_grado,
        activo=True
    )

    if request.method == 'POST':
        # --- LÓGICA ASÍNCRONA (RECOMENDADO PARA PRODUCCIÓN) ---
        # iniciar_proceso_graduacion_task.delay(institucion.id, ultimo_grado.id)
        # messages.success(request, "El proceso de graduación ha comenzado. Recibirás una notificación cuando termine.")
        
        # --- LÓGICA SÍNCRONA (PARA PROBAR AHORA CON POCOS ESTUDIANTES) ---
        # Advertencia: Esto puede causar un 'timeout' si hay muchos estudiantes.
        try:
            graduar_estudiantes(institucion.id, ultimo_grado.id)
            messages.success(request, f"Proceso completado. Se han graduado {estudiantes_a_graduar.count()} estudiantes.")
        except Exception as e:
            messages.error(request, f"Ocurrió un error durante el proceso: {e}")
        
        return redirect('gestion_academica:proceso_graduacion')

    context = {
        'titulo_pagina': 'Proceso de Graduación Anual',
        'ultimo_grado': ultimo_grado,
        'conteo_estudiantes': estudiantes_a_graduar.count(),
    }
    return render(request, 'gestion_academica/proceso_graduacion.html', context)


# Esta es la función que hace el trabajo pesado.
# La separamos para poder convertirla en una tarea de Celery fácilmente en el futuro.
def graduar_estudiantes(institucion_id, ultimo_grado_id):
    estudiantes_a_graduar = Estudiante.objects.filter(
        institucion_id=institucion_id,
        grado_actual_id=ultimo_grado_id,
        activo=True
    )
    
    año_actual = date.today().year

    for estudiante in estudiantes_a_graduar:
        with transaction.atomic():
            egresado, created = Egresado.objects.get_or_create(
                estudiante=estudiante,
                defaults={
                    'año_graduacion': año_actual,
                    'fecha_egreso': date.today(),
                }
            )

            # --- ✅ LÓGICA DE TRANSFERENCIA DE BOLETÍN ---
            try:
                # 1. Llama a tu lógica existente para generar el PDF en memoria
                pdf_buffer = generar_boletin_pdf_en_memoria(estudiante, año_actual)
                
                # 2. Crea un nombre de archivo único
                nombre_archivo = f"boletin_final_{estudiante.pk}_{año_actual}.pdf"

                # 3. Guarda el PDF en el Archivo Histórico del Egresado
                ArchivoHistorico.objects.create(
                    egresado=egresado,
                    tipo_documento='BOL_FINAL',
                    año_academico=año_actual,
                    archivo_pdf=ContentFile(pdf_buffer.read(), name=nombre_archivo)
                )
            except Exception as e:
                # Si la generación del PDF falla, la transacción se revierte
                raise e

            # --- El resto de la lógica no cambia ---
            estudiante.activo = False
            estudiante.save(update_fields=['activo'])
            
            if estudiante.usuario:
                usuario = estudiante.usuario
                usuario.rol = 'egresado'
                usuario.save(update_fields=['rol'])

def tu_vista_actual_de_boletin(request, estudiante_id, año):
    estudiante = get_object_or_404(Estudiante, pk=estudiante_id)
    
    try:
        # Llama a la función centralizada para obtener el PDF
        pdf_buffer = generar_boletin_pdf_en_memoria(estudiante, año)
        
        # Devuelve el PDF en la respuesta HTTP
        return HttpResponse(pdf_buffer, content_type='application/pdf')
    
    except Exception as e:
        return HttpResponse(f"Error al generar el boletín: {e}", status=500)

@login_required
def portal_egresado_view(request):
    try:
        egresado = request.user.perfil_estudiante.perfil_egresado
    except ObjectDoesNotExist:
        messages.error(request, "No tienes acceso al portal de egresados.")
        return redirect('gestion_academica:inicio_academico')

    # Lógica para crear una nueva solicitud de documento
    if request.method == 'POST':
        tipo_documento = request.POST.get('tipo_documento')
        # Busca el Concepto de Pago correspondiente a este certificado
        concepto_pago = get_object_or_404(ConceptoPago, nombre_concepto=tipo_documento, institucion=request.user.institucion_asociada)

        with transaction.atomic():
            # 1. Crear la solicitud
            solicitud = SolicitudDocumento.objects.create(
                egresado=egresado,
                tipo_documento_solicitado=tipo_documento
            )
            
            # 2. Crear la cuenta por cobrar
            cuenta = CuentaPorCobrarEstudiante.objects.create(
                estudiante=egresado.estudiante,
                concepto_pago=concepto_pago,
                monto_asignado=concepto_pago.valor,
                fecha_vencimiento_especifica=date.today() + timedelta(days=15),
                institucion=egresado.estudiante.institucion
            )
            
            # 3. Vincular la solicitud con la cuenta por cobrar
            solicitud.cuenta_por_cobrar = cuenta
            solicitud.save()

        messages.success(request, f"Solicitud para '{tipo_documento}' creada. Serás redirigido para realizar el pago.")
        # Redirige a la vista que genera la preferencia de Mercado Pago
        return redirect('finanzas:crear_preferencia_pago', cuenta_id=cuenta.id)

    # Lógica para mostrar la información del portal (peticiones GET)
    archivos_historicos = egresado.archivos.all().order_by('-año_academico')
    solicitudes_activas = egresado.solicitudes.all().order_by('-fecha_solicitud')
    
    # Obtiene la lista de certificados que se pueden solicitar y pagar
    certificados_solicitables = ConceptoPago.objects.filter(
    institucion=request.user.institucion_asociada,
    es_solicitable_por_egresado=True
    )

    context = {
        'titulo_pagina': 'Portal del Egresado',
        'egresado': egresado,
        'archivos': archivos_historicos,
        'solicitudes': solicitudes_activas,
        'certificados_solicitables': certificados_solicitables,
    }
    return render(request, 'gestion_academica/portal_egresado.html', context)       

@login_required
def evaluar_logros_curso(request, curso_pk):
    curso = get_object_or_404(Curso.objects.select_related('grado', 'materia', 'periodo_academico'), pk=curso_pk)
    
    # Tu lógica de permisos se mantiene, es correcta.
    if not (request.user.is_staff or (hasattr(request.user, 'docente') and request.user.docente in curso.docentes_asignados.all())):
        raise PermissionDenied
        
    if curso.grado.tipo_evaluacion != 'CUALITATIVO':
        messages.error(request, "Este curso se gestiona con el libro de notas numérico.")
        return redirect('gestion_academica:dashboard_docente')

    # --- INICIO DE LA CORRECCIÓN PRINCIPAL ---
    # 1. Buscamos en el modelo correcto: LogroPreescolar.
    logros = LogroPreescolar.objects.filter(materia=curso.materia, periodo=curso.periodo_academico)
    # --- FIN DE LA CORRECCIÓN PRINCIPAL ---

    estudiantes = Estudiante.objects.filter(grado_actual=curso.grado, activo=True).select_related('usuario')
    
    if not logros.exists():
        messages.warning(request, f"No se encontraron logros para la materia '{curso.materia.nombre_materia}' en el '{curso.periodo_academico.nombre}'. Por favor, créalos primero en la sección de gestión de logros de Preescolar.")

    escala_cualitativa = EscalaCualitativa.objects.filter(institucion=curso.institucion).order_by('orden')

    # Tu lógica para guardar los datos (POST) se mantiene, pero ahora usará los modelos correctos.
    if request.method == 'POST':
        for estudiante in estudiantes:
            for logro in logros:
                estado_id = request.POST.get(f'eval-E{estudiante.pk}-L{logro.pk}')
                if estado_id:
                    estado_obj = get_object_or_404(EscalaCualitativa, pk=estado_id)
                    # 2. Guardamos en el modelo correcto: EvaluacionLogroPreescolar
                    EvaluacionLogroPreescolar.objects.update_or_create(
                        estudiante=estudiante,
                        logro=logro, # 'logro' ahora es una instancia de LogroPreescolar
                        defaults={
                            'estado': estado_obj,
                            'registrado_por': request.user.docente,
                            'institucion': curso.institucion
                        }
                    )
        messages.success(request, "Evaluaciones de logros guardadas exitosamente.")
        return redirect('gestion_academica:evaluar_logros_curso', curso_pk=curso.pk)

    # 3. Obtenemos las evaluaciones del modelo correcto: EvaluacionLogroPreescolar
    evaluaciones_existentes = EvaluacionLogroPreescolar.objects.filter(
        estudiante__in=estudiantes, logro__in=logros
    )
    
    evaluaciones_map = defaultdict(dict)
    for ev in evaluaciones_existentes:
        evaluaciones_map[ev.estudiante_id][ev.logro_id] = ev.estado_id

    context = {
        'titulo_pagina': f"Evaluar Logros: {curso}",
        'curso': curso,
        'logros': logros,
        'estudiantes': estudiantes,
        'evaluaciones_map': dict(evaluaciones_map),
        'estados_posibles': escala_cualitativa
    }
    return render(request, 'gestion_academica/evaluar_logros_curso.html', context)

def generar_boletin_descriptivo_pdf(request, estudiante_pk, periodo_pk):
    estudiante = get_object_or_404(Estudiante, pk=estudiante_pk)
    periodo = get_object_or_404(PeriodoAcademico, pk=periodo_pk)

    # Lógica de seguridad (puedes adaptarla de tus otras vistas de boletín)
    # ...

    cursos = Curso.objects.filter(grado=estudiante.grado_actual, periodo_academico=periodo).select_related('materia')
    
    materias_con_logros = []
    for curso in cursos:
        logros = Logro.objects.filter(materia=curso.materia, periodo=periodo)
        evaluaciones = EvaluacionLogroPreescolar.objects.filter(estudiante=estudiante, logro__in=logros)
        evaluaciones_map = {ev.logro_id: ev for ev in evaluaciones}
        
        materias_con_logros.append({
            'materia': curso.materia,
            'logros_evaluados': [
                {'logro': logro, 'evaluacion': evaluaciones_map.get(logro.id)}
                for logro in logros
            ]
        })

    context = {
        'estudiante': estudiante,
        'periodo': periodo,
        'institucion': estudiante.institucion,
        'materias_con_logros': materias_con_logros,
        'observacion_general': ObservacionBoletin.objects.filter(estudiante=estudiante, periodo=periodo).first(),
        'es_bilingue': getattr(estudiante.institucion, 'es_bilingue', False),
    }

    template = get_template('gestion_academica/pdfs/boletin_descriptivo.html')
    html = template.render(context)
    
    # Lógica para generar y devolver el PDF (puedes copiarla de tu vista de boletín numérico)
    # ...
    response = HttpResponse(content_type='application/pdf')
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF.')
    return response



class LogroListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Logro
    template_name = 'gestion_academica/logro_lista.html'
    context_object_name = 'logros'
    permission_required = 'gestion_academica.view_logro' # Asegúrate de que los docentes tengan este permiso

    def get_queryset(self):
        # El docente solo ve los logros de su institución
        return Logro.objects.filter(institucion=self.request.user.institucion_asociada).select_related('materia', 'periodo')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Mis Logros de Aprendizaje (Preescolar)"
        return context


class LogroCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Logro
    form_class = LogroForm
    template_name = 'gestion_academica/logro_formulario.html'
    success_url = reverse_lazy('gestion_academica:logro_lista')
    permission_required = 'gestion_academica.add_logro'

    def get_form_kwargs(self):
        # Pasamos el usuario al formulario para que filtre los desplegables
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        # Asignamos la institución automáticamente
        form.instance.institucion = self.request.user.institucion_asociada
        messages.success(self.request, "Logro creado exitosamente.")
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Crear Nuevo Logro"
        return context


class LogroUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Logro
    form_class = LogroForm
    template_name = 'gestion_academica/logro_formulario.html'
    success_url = reverse_lazy('gestion_academica:logro_lista')
    permission_required = 'gestion_academica.change_logro'

    def get_queryset(self):
        # Seguridad: El docente solo puede editar logros de su institución
        return Logro.objects.filter(institucion=self.request.user.institucion_asociada)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, "Logro actualizado exitosamente.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Editar Logro"
        return context


class LogroDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Logro
    template_name = 'gestion_academica/logro_confirmar_eliminar.html'
    success_url = reverse_lazy('gestion_academica:logro_lista')
    permission_required = 'gestion_academica.delete_logro'
    context_object_name = 'logro'

    def get_queryset(self):
        return Logro.objects.filter(institucion=self.request.user.institucion_asociada)

    def form_valid(self, form):
        messages.success(self.request, f"El logro '{self.object.descripcion[:30]}...' ha sido eliminado.")
        return super().form_valid(form)

     
@require_POST
@login_required
def asistente_halu_api(request):
    """
    API para el asistente HALU, con uso de herramientas y manejo
    seguro de múltiples instituciones. VERSIÓN FINAL.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Esta vista solo acepta peticiones POST.'}, status=405)

    try:
        data = json.loads(request.body)
        pregunta = data.get('pregunta', '').strip()
        historial_previo = data.get('historial', []) # 2. HISTORIAL: Recibimos la memoria de la conversación enviada por el frontend
        
        if not pregunta:
            return JsonResponse({'respuesta': 'Por favor, escribe una pregunta.'})

        institucion = getattr(request.user, 'institucion_asociada', None)
        if not institucion:
            return JsonResponse({'respuesta': "Error: Tu usuario no está asociado a ninguna institución."}, status=403)

        api_key = institucion_google_api_key(institucion)
        if not api_key:
            return JsonResponse(
                {'respuesta': "Error de configuración: la institución no tiene configurada la API key de Google (Gemini)."},
                status=500,
            )

        genai.configure(api_key=api_key)

        tools_disponibles = {}
        instrucciones_sistema = ""
        user = request.user

        if user.is_superuser or (hasattr(user, 'rol') and user.rol in ['administrador', 'coordinador']):
            tools_disponibles = {
                'obtener_promedio_materia_por_grado': obtener_promedio_materia_por_grado,
                'obtener_conteo_estudiantes_por_grado': obtener_conteo_estudiantes_por_grado,
                'get_absent_students_by_grade': get_absent_students_by_grade,
                'obtener_resumen_financiero_estudiantes': obtener_resumen_financiero_estudiantes,
                'get_top_student_in_school': get_top_student_in_school,
                'get_observation_count_for_student': get_observation_count_for_student,
            }
            instrucciones_sistema = f"""
            Eres HALU, un asistente virtual experto en analítica escolar.
            Estás trabajando para un directivo de la institución '{institucion.nombre}'.
            El ID de esta institución es {institucion.id}, úsalo silenciosamente si una herramienta lo requiere.
            Tienes acceso a datos globales de la institución. Responde de manera profesional, proactiva y amigable.
            """
        elif hasattr(user, 'rol') and user.rol == 'docente':
            from ..utils import obtener_resumen_cursos_docente, obtener_estudiantes_riesgo_docente
            tools_disponibles = {
                'obtener_resumen_cursos_docente': obtener_resumen_cursos_docente,
                'obtener_estudiantes_riesgo_docente': obtener_estudiantes_riesgo_docente,
            }
            instrucciones_sistema = f"""
            Eres HALU, un asistente virtual pedagógico.
            Trabajas apoyando al docente {user.get_full_name()} de la institución '{institucion.nombre}'.
            Ayúdale a organizar sus clases, analizar el rendimiento de sus estudiantes y redactar comunicaciones.
            Responde de forma amable, profesional y motivadora.
            """
        elif hasattr(user, 'rol') and user.rol == 'estudiante':
            from ..utils import obtener_tareas_pendientes_estudiante, obtener_resumen_notas_estudiante
            tools_disponibles = {
                'obtener_tareas_pendientes_estudiante': obtener_tareas_pendientes_estudiante,
                'obtener_resumen_notas_estudiante': obtener_resumen_notas_estudiante,
            }
            instrucciones_sistema = f"""
            Eres HALU, un tutor virtual y compañero de estudio.
            Hablas con el estudiante {user.get_full_name()} de la institución '{institucion.nombre}'.
            Tu objetivo es motivarlo, recordarle sus tareas, explicarle conceptos y darle resúmenes de sus notas de forma amigable y alentadora.
            No hagas su tarea, guíalo para que aprenda.
            """
        elif hasattr(user, 'rol') and user.rol == 'familiar':
            from ..utils import obtener_resumen_hijos_familiar
            tools_disponibles = {
                'obtener_resumen_hijos_familiar': obtener_resumen_hijos_familiar,
            }
            instrucciones_sistema = f"""
            Eres HALU, un asistente de apoyo para padres de familia y acudientes.
            Estás ayudando a {user.get_full_name()}, familiar de estudiantes en la institución '{institucion.nombre}'.
            Da respuestas cordiales, claras y orientadas a involucrar al acudiente en la educación de sus hijos.
            """
        else:
            instrucciones_sistema = f"Eres HALU, el asistente virtual amigable de la plataforma escolar en '{institucion.nombre}'."
        
        # Quitamos system_instruction de los parámetros para evitar crash en versiones antiguas de la librería
        model_kwargs = {'model_name': 'gemini-2.5-flash'}
        if tools_disponibles:
            model_kwargs['tools'] = list(tools_disponibles.values())
            
        model = genai.GenerativeModel(**model_kwargs)
        
        chat = model.start_chat(history=historial_previo)
        
        # Inyectamos la instrucción al principio de forma manual si es el primer mensaje
        mensaje_enviar = pregunta
        if not historial_previo:
            mensaje_enviar = f"{instrucciones_sistema}\n\nPregunta del usuario: {pregunta}"
            
        response = chat.send_message(mensaje_enviar)
        
        # Verificación segura de Tool Calls para evitar IndexError o AttributeError
        part = response.candidates[0].content.parts[0] if response.candidates and response.candidates[0].content.parts else None
        
        while part and hasattr(part, 'function_call') and part.function_call:
            function_call = part.function_call
            tool_name = function_call.name
            
            # Extracción segura de argumentos (el objeto de Google no siempre soporta .items())
            tool_args = {}
            if hasattr(function_call, 'args'):
                for key in function_call.args:
                    tool_args[key] = function_call.args[key]
            
            if tool_name in tools_disponibles:
                tool_function = tools_disponibles[tool_name]

                if 'institucion_id' in tool_function.__code__.co_varnames and 'institucion_id' not in tool_args:
                    tool_args['institucion_id'] = institucion.id
                if 'docente_usuario_id' in tool_function.__code__.co_varnames and 'docente_usuario_id' not in tool_args:
                    tool_args['docente_usuario_id'] = request.user.id
                if 'estudiante_usuario_id' in tool_function.__code__.co_varnames and 'estudiante_usuario_id' not in tool_args:
                    tool_args['estudiante_usuario_id'] = request.user.id
                if 'familiar_usuario_id' in tool_function.__code__.co_varnames and 'familiar_usuario_id' not in tool_args:
                    tool_args['familiar_usuario_id'] = request.user.id

                try:
                    tool_response = tool_function(**tool_args)
                except Exception as e:
                    tool_response = f"Ocurrió un error al buscar los datos: {str(e)}"
                
                # Usamos glm.Part que es el estándar oficial más robusto para evitar errores 500
                response = chat.send_message(
                    glm.Part(
                        function_response=glm.FunctionResponse(
                            name=tool_name,
                            response={'resultado': str(tool_response)}
                        )
                    )
                )
                part = response.candidates[0].content.parts[0] if response.candidates and response.candidates[0].content.parts else None
            else:
                break 

        if not response.text:
            response = chat.send_message("Ok, ahora dame un resumen amigable con los resultados obtenidos de la base de datos.")

        # Extraemos el historial seguro (solo texto) para la memoria del frontend
        nuevo_historial = []
        for message in chat.history:
            textos = [p.text for p in message.parts if hasattr(p, 'text') and p.text]
            if textos:
                nuevo_historial.append({
                    "role": message.role,
                    "parts": [{"text": "\n".join(textos)}]
                })

        return JsonResponse({'respuesta': response.text, 'historial': nuevo_historial})

    except Exception as e:
        logger.error(f"Error inesperado en asistente_halu_api: {e}", exc_info=True)
        return JsonResponse({'respuesta': f"Ocurrió un error interno en la IA: {str(e)}. Por favor, avisa a soporte."}, status=500)

@login_required
def redirigir_a_libro_de_notas(request, curso_pk):
    """
    Vista inteligente que actúa como un despachador.
    Revisa el tipo de evaluación del curso y redirige al docente
    a la vista de calificación correcta (cuantitativa o cualitativa).
    """
    # 1. Obtenemos el curso y su grado para verificar el tipo de evaluación
    curso = get_object_or_404(
        Curso.objects.select_related('grado'), 
        pk=curso_pk
    )

    # 2. Lógica de seguridad: docente del curso, coordinador/admin (is_staff) o superusuario
    es_coord = request.user.is_staff and getattr(request.user, 'rol', None) in ['administrador', 'coordinador']
    if not (request.user.is_superuser or es_coord or (hasattr(request.user, 'docente') and request.user.docente in curso.docentes_asignados.all())):
        messages.error(request, "No tienes permiso para acceder a este curso.")
        return redirect('gestion_academica:dashboard_docente')

    # Preservar parámetros de URL (ej. ?tab=anual)
    extra_params = ''
    tab = request.GET.get('tab')
    if tab:
        extra_params = f'?tab={tab}'

    # 3. ¡LA LÓGICA CLAVE! Leemos el campo del modelo Grado
    if curso.grado.tipo_evaluacion == 'CUALITATIVO':
        # Si es cualitativo, redirigimos a la vista de evaluación de logros
        return redirect('gestion_academica:evaluar_logros_curso', curso_pk=curso.pk)
    else:
        # Por defecto, o si es cuantitativo, redirigimos al libro de notas numérico
        url = reverse('gestion_academica:docente_libro_de_notas_por_curso', kwargs={'curso_pk': curso.pk})
        return redirect(url + extra_params)

@login_required
def gestionar_curso_cualitativo(request, curso_pk):
    curso = get_object_or_404(Curso.objects.select_related('grado'), pk=curso_pk)
    
    # Doble verificación de seguridad
    if curso.grado.tipo_evaluacion != 'CUALITATIVO':
        messages.error(request, "Esta sección es solo para cursos de evaluación cualitativa.")
        return redirect('gestion_academica:dashboard_docente')

    context = {
        'titulo_pagina': f"Gestionar Curso: {curso.materia.nombre_materia} ({curso.grado.nombre})",
        'curso': curso
    }
    return render(request, 'gestion_academica/gestion_curso_cualitativo.html', context)

class DimensionListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = DimensionDesarrollo
    template_name = 'gestion_academica/dimension_lista.html'
    context_object_name = 'dimensiones'
    permission_required = 'gestion_academica.view_dimensiondesarrollo' # Necesitarás crear este permiso

    def get_queryset(self):
        return get_filtered_queryset(self.model, self.request.user).order_by('orden')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Gestionar Dimensiones de Desarrollo (Preescolar)"
        return context

class DimensionCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = DimensionDesarrollo
    form_class = DimensionDesarrolloForm
    template_name = 'gestion_academica/dimension_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_dimensiones')
    permission_required = 'gestion_academica.add_dimensiondesarrollo'

    def form_valid(self, form):
        form.instance.institucion = self.request.user.institucion_asociada
        messages.success(self.request, "Dimensión creada exitosamente.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Crear Nueva Dimensión"
        return context

class DimensionUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = DimensionDesarrollo
    form_class = DimensionDesarrolloForm
    template_name = 'gestion_academica/dimension_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_dimensiones')
    permission_required = 'gestion_academica.change_dimensiondesarrollo'

    def get_queryset(self):
        return get_filtered_queryset(self.model, self.request.user)

    def form_valid(self, form):
        messages.success(self.request, "Dimensión actualizada exitosamente.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = f"Editar Dimensión: {self.object.nombre}"
        return context

class DimensionDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = DimensionDesarrollo
    template_name = 'gestion_academica/confirmar_eliminar_generico.html'
    success_url = reverse_lazy('gestion_academica:lista_dimensiones')
    permission_required = 'gestion_academica.delete_dimensiondesarrollo'
    context_object_name = 'object'

    def get_queryset(self):
        return get_filtered_queryset(self.model, self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, f"La dimensión '{self.object.nombre}' ha sido eliminada.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Dimensión"
        context['mensaje_confirmacion'] = f"¿Estás seguro de que deseas eliminar la dimensión '{self.object.nombre}'? Todos los logros asociados perderán esta categoría."
        context['url_cancelar'] = reverse_lazy('gestion_academica:lista_dimensiones')
        return context    

class EscalaCualitativaListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = EscalaCualitativa
    template_name = 'gestion_academica/escala_cualitativa_lista.html'
    context_object_name = 'escalas'
    permission_required = 'gestion_academica.view_escalacualitativa'

    def get_queryset(self):
        return get_filtered_queryset(self.model, self.request.user).order_by('orden')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Gestionar Escala Cualitativa (Preescolar)"
        return context

class EscalaCualitativaCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = EscalaCualitativa
    form_class = EscalaCualitativaForm
    template_name = 'gestion_academica/escala_cualitativa_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_escala_cualitativa')
    permission_required = 'gestion_academica.add_escalacualitativa'

    def form_valid(self, form):
        form.instance.institucion = self.request.user.institucion_asociada
        messages.success(self.request, "Nuevo nivel de escala creado exitosamente.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Crear Nuevo Nivel de Escala"
        return context

class EscalaCualitativaUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = EscalaCualitativa
    form_class = EscalaCualitativaForm
    template_name = 'gestion_academica/escala_cualitativa_formulario.html'
    success_url = reverse_lazy('gestion_academica:lista_escala_cualitativa')
    permission_required = 'gestion_academica.change_escalacualitativa'

    def get_queryset(self):
        return get_filtered_queryset(self.model, self.request.user)

    def form_valid(self, form):
        messages.success(self.request, "Nivel de escala actualizado exitosamente.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = f"Editar Nivel: {self.object.nombre_escala}"
        return context

class EscalaCualitativaDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = EscalaCualitativa
    template_name = 'gestion_academica/confirmar_eliminar_generico.html'
    success_url = reverse_lazy('gestion_academica:lista_escala_cualitativa')
    permission_required = 'gestion_academica.delete_escalacualitativa'
    context_object_name = 'object'

    def get_queryset(self):
        return get_filtered_queryset(self.model, self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, f"El nivel de escala '{self.object.nombre_escala}' ha sido eliminado.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Nivel de Escala"
        context['mensaje_confirmacion'] = f"¿Estás seguro de que deseas eliminar el nivel '{self.object.nombre_escala}'?"
        context['url_cancelar'] = reverse_lazy('gestion_academica:lista_escala_cualitativa')
        return context        

class LogroListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = LogroPreescolar
    template_name = 'gestion_academica/logro_lista.html'
    context_object_name = 'logros'
    permission_required = 'gestion_academica.view_logropreescolar' # <-- CORREGIDO

    def get_queryset(self):
        return get_filtered_queryset(self.model, self.request.user).select_related('materia', 'periodo').order_by('-periodo__año_escolar', 'materia__nombre_materia')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Mis Logros de Aprendizaje (Preescolar)"
        return context

class LogroCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = LogroPreescolar
    form_class = LogroPreescolarForm
    template_name = 'gestion_academica/logro_formulario.html'
    success_url = reverse_lazy('gestion_academica:logro_lista')
    permission_required = 'gestion_academica.add_logropreescolar' # <-- CORREGIDO

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.institucion = self.request.user.institucion_asociada
        messages.success(self.request, "Logro de Preescolar creado exitosamente.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Crear Nuevo Logro de Preescolar"
        return context

class LogroUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = LogroPreescolar
    form_class = LogroPreescolarForm
    template_name = 'gestion_academica/logro_formulario.html'
    success_url = reverse_lazy('gestion_academica:logro_lista')
    permission_required = 'gestion_academica.change_logropreescolar' # <-- CORREGIDO

    def get_queryset(self):
        return get_filtered_queryset(self.model, self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, "Logro de Preescolar actualizado exitosamente.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Logro de Preescolar"
        return context

class LogroDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = LogroPreescolar
    template_name = 'gestion_academica/confirmar_eliminar_generico.html'
    success_url = reverse_lazy('gestion_academica:logro_lista')
    permission_required = 'gestion_academica.delete_logropreescolar' # <-- CORREGIDO
    context_object_name = 'object'

    def get_queryset(self):
        return get_filtered_queryset(self.model, self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, "El logro de Preescolar ha sido eliminado.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Logro"
        context['mensaje_confirmacion'] = f"¿Estás seguro de que deseas eliminar este logro?"
        context['url_cancelar'] = reverse_lazy('gestion_academica:logro_lista')
        return context       

@login_required
def boletin_descriptivo_preescolar_pdf(request, estudiante_pk, periodo_pk):
    """
    Genera un boletín en PDF CUALITATIVO para un estudiante de Preescolar.
    VERSIÓN FINAL: La vista prepara todos los datos para que la plantilla sea simple.
    """
    # 1. Obtención de datos y seguridad (se mantiene tu lógica)
    try:
        estudiante = Estudiante.objects.select_related(
            'usuario', 'grado_actual', 'institucion'
        ).get(pk=estudiante_pk)
        periodo = PeriodoAcademico.objects.get(
            pk=periodo_pk, institucion=estudiante.institucion
        )
    except (Estudiante.DoesNotExist, PeriodoAcademico.DoesNotExist):
        messages.error(request, "El estudiante o periodo solicitado no es válido.")
        return redirect('gestion_academica:inicio_academico')

    # (Tu lógica de permisos se mantiene)
    if not (request.user.pk == estudiante.usuario.pk or request.user.is_staff or (hasattr(request.user, 'familiar') and request.user.familiar.estudiantes_asociados.filter(pk=estudiante_pk).exists())):
        messages.error(request, "No tienes permiso para ver este boletín.")
        return redirect('gestion_academica:inicio_academico')

    # --- INICIO DE LA LÓGICA DE CONSULTA CORREGIDA ---
    
    # 2. Obtenemos las evaluaciones del estudiante para este periodo
    evaluaciones_del_estudiante = EvaluacionLogroPreescolar.objects.filter(
        estudiante=estudiante,
        logro__periodo=periodo
    ).select_related('estado')
    
    # Creamos un mapa para buscar rápidamente la evaluación de cada logro
    evaluaciones_map = {ev.logro_id: ev.estado for ev in evaluaciones_del_estudiante}

    # 3. Obtenemos las Dimensiones y, para cada una, sus logros correspondientes
    dimensiones_con_logros = DimensionDesarrollo.objects.filter(
        institucion=estudiante.institucion
    ).prefetch_related(
        Prefetch(
            'logros_preescolar',
            queryset=LogroPreescolar.objects.filter(
                periodo=periodo,
                materia__cursos__grado=estudiante.grado_actual
            ).distinct(),
            to_attr='logros_de_la_dimension'
        )
    ).order_by('orden')
    
    # 4. *** EL PASO CLAVE ***
    #    Añadimos la evaluación directamente a cada objeto 'logro'
    for dimension in dimensiones_con_logros:
        for logro in dimension.logros_de_la_dimension:
            # .get() devuelve la evaluación si existe, o None si no existe (sin dar error)
            logro.evaluacion = evaluaciones_map.get(logro.id)

    # --- FIN DE LA LÓGICA DE CONSULTA CORREGIDA ---

    # 5. Obtenemos datos adicionales (tu lógica original)
    escala_cualitativa = EscalaCualitativa.objects.filter(institucion=estudiante.institucion).order_by('orden')
    observacion_obj = ObservacionBoletin.objects.filter(estudiante=estudiante, periodo=periodo).first()
    director_de_grupo = DirectorCurso.objects.select_related('docente__usuario').filter(
        grado=estudiante.grado_actual, periodo_academico=periodo
    ).first()

    # 6. Construimos el contexto final
    context = {
        'institucion': estudiante.institucion,
        'periodo': periodo,
        'estudiante': estudiante,
        'dimensiones_data': dimensiones_con_logros,
        'escala_cualitativa': escala_cualitativa,
        'director_de_grupo': director_de_grupo,
        'observaciones': observacion_obj.observacion if observacion_obj else "No hay observaciones registradas.",
        'fecha_emision': timezone.now(),
    }

    # 7. Renderizamos el PDF
    template_path = 'gestion_academica/boletin_descriptivo_preescolar.html'
    template = get_template(template_path)
    html = template.render(context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Boletin_Descriptivo_{estudiante.usuario.username}.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
    
    if pisa_status.err:
        logger.error("Error al generar PDF con xhtml2pdf (pisa.err=%s)", pisa_status.err)
        return HttpResponse('Error al generar el PDF. Por favor, inténtelo de nuevo.', status=500)
        
    return response

@login_required
# Aquí podrías añadir un user_passes_test para limitar a coordinadores/directivos
def reportes_dashboard(request):
    """
    Muestra un dashboard central con enlaces a todos los reportes académicos disponibles.
    """
    context = {
        'titulo_pagina': "Dashboard de Reportes Académicos"
    }
    return render(request, 'gestion_academica/reportes/dashboard.html', context)    


    
@login_required
def generar_boletin_dispatcher(request, estudiante_pk, periodo_pk):
    """
    Revisa el tipo de evaluación del grado del estudiante y redirige
    a la vista de generación de PDF correcta (cuantitativa o cualitativa).
    """
    estudiante = get_object_or_404(Estudiante.objects.select_related('grado_actual'), pk=estudiante_pk)
    
    # (Aquí puedes añadir la lógica de seguridad que usas en tus otras vistas de boletín
    # para verificar que el usuario (estudiante, familiar, staff) tiene permiso)

    if estudiante.grado_actual and estudiante.grado_actual.tipo_evaluacion == 'CUALITATIVO':
        # Si el grado es Cualitativo (Preescolar), redirigimos a la vista del boletín descriptivo.
        return redirect('gestion_academica:boletin_descriptivo_preescolar', estudiante_pk=estudiante.pk, periodo_pk=periodo_pk)
    else:
        # Para todos los demás casos (Cuantitativo), redirigimos al boletín numérico tradicional.
        return redirect('gestion_academica:boletin_imprimible', estudiante_pk=estudiante.pk, periodo_pk=periodo_pk)       

@login_required
def reporte_rendimiento_estudiante(request):
    """
    Vista MEJORADA para generar reportes de rendimiento.
    Distingue entre evaluación Cuantitativa (con gráfica) y Cualitativa.
    """
    grados = Grado.objects.all()
    periodos = PeriodoAcademico.objects.all().order_by('-año_escolar', '-fecha_inicio')
    
    grado_id = request.GET.get('grado')
    estudiante_id = request.GET.get('estudiante')
    periodo_id = request.GET.get('periodo')
    
    estudiantes_del_grado = Estudiante.objects.none()
    estudiante_seleccionado = None
    periodo_seleccionado = None
    contexto_reporte = {}

    if grado_id:
        estudiantes_del_grado = Estudiante.objects.filter(grado_actual__id=grado_id).select_related('usuario').order_by('usuario__last_name')

    if estudiante_id and periodo_id:
        estudiante_seleccionado = get_object_or_404(Estudiante.objects.select_related('grado_actual'), pk=estudiante_id)
        periodo_seleccionado = get_object_or_404(PeriodoAcademico, pk=periodo_id)
        
        # --- LÓGICA DE SELECCIÓN DE REPORTE ---
        if estudiante_seleccionado.grado_actual.tipo_evaluacion == 'CUALITATIVO':
            # --- LÓGICA PARA PREESCOLAR ---
            dimensiones = DimensionDesarrollo.objects.filter(institucion=estudiante_seleccionado.institucion).prefetch_related(
                Prefetch('logros_preescolar', queryset=LogroPreescolar.objects.filter(periodo=periodo_seleccionado, materia__cursos__grado=estudiante_seleccionado.grado_actual))
            )
            evaluaciones = EvaluacionLogroPreescolar.objects.filter(estudiante=estudiante_seleccionado, logro__periodo=periodo_seleccionado).select_related('estado')
            evaluaciones_map = {ev.logro_id: ev.estado for ev in evaluaciones}

            for dim in dimensiones:
                for logro in dim.logros_preescolar.all():
                    logro.evaluacion = evaluaciones_map.get(logro.id)

            contexto_reporte = {
                'tipo_reporte': 'CUALITATIVO',
                'dimensiones_data': dimensiones
            }
        else:
            # --- LÓGICA PARA PRIMARIA/SECUNDARIA (CUANTITATIVO) ---
            cursos_del_estudiante = Curso.objects.filter(
                grado=estudiante_seleccionado.grado_actual,
                periodo_academico=periodo_seleccionado
            ).select_related('materia').order_by('materia__nombre_materia')

            cursos_con_detalle = []
            promedio_numerador = 0
            promedio_denominador = 0
            
            # Datos para la gráfica
            chart_labels = []
            chart_data = []

            for curso in cursos_del_estudiante:
                nota_final_curso = Calificacion.objects.filter(
                    estudiante=estudiante_seleccionado,
                    actividad_calificable__curso=curso,
                    valor_numerico__isnull=False
                ).aggregate(promedio=Avg('valor_numerico'))['promedio']
                
                if nota_final_curso is not None:
                    promedio_numerador += nota_final_curso
                    promedio_denominador += 1
                    chart_labels.append(curso.materia.nombre_materia)
                    chart_data.append(float(nota_final_curso))
                
                cursos_con_detalle.append({
                    'curso': curso,
                    'nota_final_curso': nota_final_curso
                })

            promedio_general_calculado = (promedio_numerador / promedio_denominador) if promedio_denominador > 0 else None

            contexto_reporte = {
                'tipo_reporte': 'CUANTITATIVO',
                'cursos_con_detalle': cursos_con_detalle,
                'promedio_general_periodo': promedio_general_calculado,
                'chart_labels': json.dumps(chart_labels),
                'chart_data': json.dumps(chart_data)
            }

    context = {
        'titulo_pagina': "Reporte de Rendimiento Académico",
        'grados': grados,
        'estudiantes_del_grado': estudiantes_del_grado,
        'periodos': periodos,
        'grado_seleccionado_id': grado_id,
        'estudiante_seleccionado_id': estudiante_id,
        'periodo_seleccionado_id': periodo_id,
        'estudiante_seleccionado': estudiante_seleccionado,
        'periodo_seleccionado': periodo_seleccionado,
        'contexto_reporte': contexto_reporte,
    }
    
    return render(request, 'gestion_academica/reportes/rendimiento_estudiante.html', context)   


@login_required
def reporte_acumulado_periodo(request):
    """
    Muestra el rendimiento de un estudiante a lo largo de todos los periodos de un año.
    Distingue entre reportes cuantitativos y cualitativos, AMBOS CON GRÁFICOS.
    """
    grados = Grado.objects.all()
    años_escolares = PeriodoAcademico.objects.values_list('año_escolar', flat=True).distinct().order_by('-año_escolar')

    grado_id = request.GET.get('grado')
    estudiante_id = request.GET.get('estudiante')
    año_seleccionado_str = request.GET.get('año')
    
    año_seleccionado = int(año_seleccionado_str) if año_seleccionado_str else (años_escolares.first() or timezone.now().year)

    estudiantes_del_grado = Estudiante.objects.none()
    estudiante_seleccionado = None
    reporte_data = {}

    if grado_id:
        estudiantes_del_grado = Estudiante.objects.filter(grado_actual__id=grado_id).select_related('usuario').order_by('usuario__last_name')

    if estudiante_id:
        estudiante_seleccionado = get_object_or_404(Estudiante.objects.select_related('grado_actual'), pk=estudiante_id)
        periodos_del_año = PeriodoAcademico.objects.filter(año_escolar=año_seleccionado, institucion=estudiante_seleccionado.institucion).order_by('fecha_inicio')
        periodos_header = [p.nombre for p in periodos_del_año]

        if estudiante_seleccionado.grado_actual and estudiante_seleccionado.grado_actual.tipo_evaluacion == 'CUALITATIVO':
            # --- LÓGICA PARA REPORTE CUALITATIVO + GRÁFICO ---
            logros = LogroPreescolar.objects.filter(materia__cursos__grado=estudiante_seleccionado.grado_actual, periodo__in=periodos_del_año).select_related('dimension', 'materia').order_by('dimension__orden', 'orden')
            evaluaciones = EvaluacionLogroPreescolar.objects.filter(estudiante=estudiante_seleccionado, logro__in=logros).select_related('logro', 'estado')
            evaluaciones_map = {(ev.logro.periodo_id, ev.logro_id): ev.estado for ev in evaluaciones}
            
            logros_por_dimension = OrderedDict()
            unique_logros = set()
            for logro in logros:
                if logro.pk not in unique_logros:
                    dimension_nombre = logro.dimension.nombre
                    if dimension_nombre not in logros_por_dimension:
                        logros_por_dimension[dimension_nombre] = []
                    
                    evaluaciones_logro = []
                    for periodo in periodos_del_año:
                        evaluaciones_logro.append(evaluaciones_map.get((periodo.id, logro.id)))

                    logros_por_dimension[dimension_nombre].append({'logro': logro, 'evaluaciones': evaluaciones_logro})
                    unique_logros.add(logro.pk)

            # Preparar datos para el gráfico de barras apiladas
            escala = EscalaCualitativa.objects.filter(institucion=estudiante_seleccionado.institucion).order_by('orden')
            chart_datasets_cualitativo = []
            colores = ['rgba(75, 192, 192, 0.7)', 'rgba(255, 206, 86, 0.7)', 'rgba(255, 99, 132, 0.7)', 'rgba(153, 102, 255, 0.7)']
            for i, nivel_escala in enumerate(escala):
                data = []
                for periodo in periodos_del_año:
                    count = sum(1 for ev in evaluaciones if ev.logro.periodo == periodo and ev.estado == nivel_escala)
                    data.append(count)
                chart_datasets_cualitativo.append({'label': nivel_escala.nombre_escala, 'data': data, 'backgroundColor': colores[i % len(colores)]})

            reporte_data = {
                'tipo_reporte': 'CUALITATIVO', 'periodos_header': periodos_header,
                'logros_por_dimension': logros_por_dimension,
                'chart_labels': json.dumps(periodos_header),
                'chart_datasets': json.dumps(chart_datasets_cualitativo)
            }

        else:
            # --- LÓGICA PARA REPORTE CUANTITATIVO + GRÁFICO ---
            materias = Materia.objects.filter(cursos__grado=estudiante_seleccionado.grado_actual, cursos__periodo_academico__in=periodos_del_año).distinct().order_by('nombre_materia')
            notas_por_materia = OrderedDict()
            chart_datasets_cuantitativo = []

            for materia in materias:
                notas_periodos = []
                for periodo in periodos_del_año:
                    curso = Curso.objects.filter(materia=materia, periodo_academico=periodo, grado=estudiante_seleccionado.grado_actual).first()
                    nota_final = calcular_estado_academico_curso(curso, estudiante_seleccionado).get('nota_final_ponderada') if curso else None
                    notas_periodos.append(nota_final)
                notas_por_materia[materia.nombre_materia] = notas_periodos
                
                # Preparamos dataset para el gráfico de líneas
                color = f'rgb({random.randint(0, 255)}, {random.randint(0, 255)}, {random.randint(0, 255)})'
                chart_datasets_cuantitativo.append({
                    'label': materia.nombre_materia,
                    'data': [float(n) if n is not None else None for n in notas_periodos],
                    'borderColor': color,
                    'backgroundColor': color,
                    'fill': False,
                    'tension': 0.1
                })

            reporte_data = {
                'tipo_reporte': 'CUANTITATIVO', 'periodos_header': periodos_header,
                'notas_por_materia': notas_por_materia,
                'chart_labels': json.dumps(periodos_header),
                'chart_datasets': json.dumps(chart_datasets_cuantitativo)
            }

    context = {
        'titulo_pagina': "Reporte Acumulado por Periodo",
        'grados': grados, 'años_escolares': años_escolares, 'estudiantes_del_grado': estudiantes_del_grado,
        'grado_seleccionado_id': grado_id, 'estudiante_seleccionado_id': estudiante_id,
        'año_seleccionado': año_seleccionado, 'estudiante_seleccionado': estudiante_seleccionado,
        'reporte_data': reporte_data,
    }
    return render(request, 'gestion_academica/reportes/reporte_acumulado.html', context)

@login_required
def reporte_promedio_general_grado(request):
    """
    Muestra un ranking de estudiantes de un grado basado en su promedio
    general para un periodo específico. Incluye una gráfica comparativa.
    Solo funciona para grados con evaluación CUANTITATIVA.
    """
    grados = Grado.objects.filter(tipo_evaluacion='CUANTITATIVO') # Solo mostramos grados cuantitativos
    periodos = PeriodoAcademico.objects.all().order_by('-año_escolar', '-fecha_inicio')

    grado_id = request.GET.get('grado')
    periodo_id = request.GET.get('periodo')

    reporte_data = []
    grado_seleccionado = None
    periodo_seleccionado = None
    chart_labels = []
    chart_data = []

    if grado_id and periodo_id:
        grado_seleccionado = get_object_or_404(Grado, pk=grado_id)
        periodo_seleccionado = get_object_or_404(PeriodoAcademico, pk=periodo_id)
        
        estudiantes_del_grado = Estudiante.objects.filter(grado_actual=grado_seleccionado, activo=True)
        
        # Calculamos el promedio para cada estudiante
        for estudiante in estudiantes_del_grado:
            cursos = Curso.objects.filter(grado=estudiante.grado_actual, periodo_academico=periodo_seleccionado)
            
            total_puntos_ponderados = Decimal('0.0')
            total_ihs = 0
            
            for curso in cursos:
                estado_academico = calcular_estado_academico_curso(curso, estudiante)
                nota_final = estado_academico.get('nota_final_ponderada')
                ihs = curso.materia.intensidad_horaria_semanal

                if nota_final is not None and ihs > 0:
                    total_puntos_ponderados += nota_final * ihs
                    total_ihs += ihs
            
            promedio_general = total_puntos_ponderados / total_ihs if total_ihs > 0 else None
            
            if promedio_general is not None:
                reporte_data.append({'estudiante': estudiante, 'promedio': promedio_general})

        # Ordenamos la lista de mayor a menor promedio
        reporte_data = sorted(reporte_data, key=lambda x: x['promedio'], reverse=True)

        # Preparamos los datos para el gráfico de barras
        chart_labels = [item['estudiante'].usuario.get_full_name() for item in reporte_data]
        chart_data = [float(item['promedio']) for item in reporte_data]

    context = {
        'titulo_pagina': "Ranking de Estudiantes por Grado",
        'grados': grados,
        'periodos': periodos,
        'grado_seleccionado': grado_seleccionado,
        'periodo_seleccionado': periodo_seleccionado,
        'reporte_data': reporte_data,
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
    }
    return render(request, 'gestion_academica/reportes/reporte_promedio_general.html', context)


@login_required
def reporte_estudiante_dashboard(request):
    """
    Muestra un dashboard consolidado con toda la información relevante de un
    único estudiante, adaptado para evaluación cuantitativa y cualitativa.
    """
    grados = Grado.objects.all()
    estudiantes_del_grado = Estudiante.objects.none()
    
    grado_id = request.GET.get('grado')
    estudiante_id = request.GET.get('estudiante')
    
    contexto_reporte = {}
    estudiante_seleccionado = None

    if grado_id:
        estudiantes_del_grado = Estudiante.objects.filter(grado_actual__id=grado_id).select_related('usuario').order_by('usuario__last_name')

    if estudiante_id:
        estudiante_seleccionado = get_object_or_404(Estudiante.objects.select_related('grado_actual', 'institucion'), pk=estudiante_id)
        institucion = estudiante_seleccionado.institucion
        periodo_activo = PeriodoAcademico.objects.filter(institucion=institucion, activo=True).first()
        
        # --- DATOS COMUNES ---
        anotaciones_recientes = AnotacionObservador.objects.filter(estudiante=estudiante_seleccionado).order_by('-fecha_hora')[:5]
        esta_al_dia = not CuentaPorCobrarEstudiante.objects.filter(estudiante=estudiante_seleccionado, estado='VENCIDO').exists()

        if estudiante_seleccionado.grado_actual and estudiante_seleccionado.grado_actual.tipo_evaluacion == 'CUALITATIVO':
            # --- LÓGICA PARA DASHBOARD CUALITATIVO (PREESCOLAR) ---
            logros_alcanzados = 0
            logros_en_proceso = 0
            if periodo_activo:
                evaluaciones = EvaluacionLogroPreescolar.objects.filter(estudiante=estudiante_seleccionado, logro__periodo=periodo_activo).select_related('estado')
                for ev in evaluaciones:
                    if "alcanzado" in ev.estado.nombre_escala.lower():
                        logros_alcanzados += 1
                    elif "proceso" in ev.estado.nombre_escala.lower():
                        logros_en_proceso += 1

            contexto_reporte = {
                'tipo_reporte': 'CUALITATIVO',
                'logros_alcanzados': logros_alcanzados,
                'logros_en_proceso': logros_en_proceso,
                'anotaciones_recientes': anotaciones_recientes,
                'esta_al_dia': esta_al_dia,
                # El gráfico para cualitativo podría ser un resumen total, lo añadiremos en un futuro reporte.
            }

        else:
            # --- LÓGICA PARA DASHBOARD CUANTITATIVO ---
            promedio_periodo_actual = None
            materias_en_riesgo = 0
            inasistencias = 0
            
            if periodo_activo:
                cursos = Curso.objects.filter(grado=estudiante_seleccionado.grado_actual, periodo_academico=periodo_activo)
                nota_minima = institucion.nota_minima_aprobacion if institucion else Decimal('3.0')
                
                total_puntos = Decimal('0.0')
                total_ihs = 0
                for curso in cursos:
                    estado = calcular_estado_academico_curso(curso, estudiante_seleccionado)
                    nota_final = estado.get('nota_final_ponderada')
                    if nota_final is not None:
                        if nota_final < nota_minima:
                            materias_en_riesgo += 1
                        if curso.materia.intensidad_horaria_semanal > 0:
                            total_puntos += nota_final * curso.materia.intensidad_horaria_semanal
                            total_ihs += curso.materia.intensidad_horaria_semanal
                
                promedio_periodo_actual = total_puntos / total_ihs if total_ihs > 0 else None
                inasistencias = RegistroAsistencia.objects.filter(estudiante=estudiante_seleccionado, curso__periodo_academico=periodo_activo, estado='AUSENTE').count()

            # Datos para el gráfico de evolución
            periodos_año = PeriodoAcademico.objects.filter(año_escolar=periodo_activo.año_escolar, institucion=institucion).order_by('fecha_inicio') if periodo_activo else []
            chart_labels = [p.nombre for p in periodos_año]
            chart_data = []
            for p in periodos_año:
                # Lógica similar para calcular el promedio en cada periodo del año
                cursos_p = Curso.objects.filter(grado=estudiante_seleccionado.grado_actual, periodo_academico=p)
                total_puntos_p = Decimal('0.0')
                total_ihs_p = 0
                for c in cursos_p:
                    estado_p = calcular_estado_academico_curso(c, estudiante_seleccionado)
                    nota_p = estado_p.get('nota_final_ponderada')
                    if nota_p and c.materia.intensidad_horaria_semanal > 0:
                        total_puntos_p += nota_p * c.materia.intensidad_horaria_semanal
                        total_ihs_p += c.materia.intensidad_horaria_semanal
                promedio_p = total_puntos_p / total_ihs_p if total_ihs_p > 0 else None
                chart_data.append(float(promedio_p) if promedio_p else None)

            contexto_reporte = {
                'tipo_reporte': 'CUANTITATIVO',
                'promedio_actual': promedio_periodo_actual,
                'materias_riesgo': materias_en_riesgo,
                'inasistencias': inasistencias,
                'esta_al_dia': esta_al_dia,
                'anotaciones_recientes': anotaciones_recientes,
                'chart_labels': json.dumps(chart_labels),
                'chart_data': json.dumps(chart_data),
            }

    context = {
        'titulo_pagina': "Dashboard del Estudiante",
        'grados': grados,
        'estudiantes_del_grado': estudiantes_del_grado,
        'grado_seleccionado_id': grado_id,
        'estudiante_seleccionado_id': estudiante_id,
        'estudiante_seleccionado': estudiante_seleccionado,
        'contexto_reporte': contexto_reporte,
    }
    return render(request, 'gestion_academica/reportes/reporte_estudiante_dashboard.html', context)

@login_required
def reporte_rendimiento_por_grado(request):
    """
    Muestra el promedio por materia (cuantitativo) o un resumen de logros (cualitativo)
    para un grado y periodo específicos. Incluye gráficos para ambos casos.
    VERSIÓN CORREGIDA PARA PREESCOLAR.
    """
    grados = Grado.objects.all().order_by('orden', 'nombre')
    periodos = PeriodoAcademico.objects.all().order_by('-año_escolar', '-fecha_inicio')

    grado_id = request.GET.get('grado')
    periodo_id = request.GET.get('periodo')

    reporte_data = {}
    grado_seleccionado = None
    periodo_seleccionado = None

    if grado_id and periodo_id:
        grado_seleccionado = get_object_or_404(Grado, pk=grado_id)
        periodo_seleccionado = get_object_or_404(PeriodoAcademico, pk=periodo_id)
        
        if grado_seleccionado.tipo_evaluacion == 'CUALITATIVO':
            # --- LÓGICA MEJORADA PARA REPORTE CUALITATIVO POR GRADO ---
            # 1. Obtenemos todas las escalas posibles para usarlas como base.
            escala_completa = EscalaCualitativa.objects.filter(institucion=grado_seleccionado.institucion).order_by('orden')
            
            # 2. Inicializamos nuestro contador con todas las escalas en cero.
            conteo_por_escala = OrderedDict((escala.nombre_escala, 0) for escala in escala_completa)

            # 3. Buscamos las evaluaciones existentes.
            estudiantes = Estudiante.objects.filter(grado_actual=grado_seleccionado, activo=True)
            evaluaciones = EvaluacionLogroPreescolar.objects.filter(
                estudiante__in=estudiantes,
                logro__periodo=periodo_seleccionado
            ).select_related('estado')

            # 4. Actualizamos el conteo con los datos reales.
            for ev in evaluaciones:
                if ev.estado and ev.estado.nombre_escala in conteo_por_escala:
                    conteo_por_escala[ev.estado.nombre_escala] += 1
            
            reporte_data = {
                'tipo_reporte': 'CUALITATIVO',
                'conteo_logros': conteo_por_escala,
                'total_evaluaciones': sum(conteo_por_escala.values())
            }
            # 5. Preparamos los datos para el gráfico (ahora nunca estarán vacíos).
            chart_labels = list(conteo_por_escala.keys())
            chart_data = list(conteo_por_escala.values())

        else:
            # --- LÓGICA CUANTITATIVA (SIN CAMBIOS) ---
            cursos_del_grado = Curso.objects.filter(grado=grado_seleccionado, periodo_academico=periodo_seleccionado)
            datos_cuantitativos = []
            
            for curso in cursos_del_grado:
                promedio_curso = Calificacion.objects.filter(
                    actividad_calificable__curso=curso, valor_numerico__isnull=False
                ).aggregate(promedio=Avg('valor_numerico'))['promedio']

                if promedio_curso is not None:
                    datos_cuantitativos.append({'materia': curso.materia.nombre_materia, 'promedio': promedio_curso})
            
            datos_cuantitativos = sorted(datos_cuantitativos, key=lambda x: x['promedio'], reverse=True)
            
            reporte_data = {
                'tipo_reporte': 'CUANTITATIVO',
                'datos_tabla': datos_cuantitativos,
            }
            chart_labels = [item['materia'] for item in datos_cuantitativos]
            chart_data = [float(item['promedio']) for item in datos_cuantitativos]

        reporte_data['chart_labels'] = json.dumps(chart_labels)
        reporte_data['chart_data'] = json.dumps(chart_data)

    context = {
        'titulo_pagina': "Rendimiento General por Grado",
        'grados': grados, 'periodos': periodos,
        'grado_seleccionado': grado_seleccionado, 'periodo_seleccionado': periodo_seleccionado,
        'reporte_data': reporte_data,
    }
    return render(request, 'gestion_academica/reportes/reporte_rendimiento_grado.html', context)  

@login_required
def reporte_promedio_por_area(request):
    """
    Muestra el promedio por Área (cuantitativo) o un resumen de logros por
    Dimensión (cualitativo). VERSIÓN CORREGIDA PARA INCLUIR PREESCOLAR.
    """
    grados = Grado.objects.all().order_by('orden', 'nombre')
    periodos = PeriodoAcademico.objects.all().order_by('-año_escolar', '-fecha_inicio')

    grado_id = request.GET.get('grado')
    periodo_id = request.GET.get('periodo')

    reporte_data = {}
    grado_seleccionado = None
    periodo_seleccionado = None

    if grado_id and periodo_id:
        grado_seleccionado = get_object_or_404(Grado, pk=grado_id)
        periodo_seleccionado = get_object_or_404(PeriodoAcademico, pk=periodo_id)
        institucion = grado_seleccionado.institucion

        if grado_seleccionado.tipo_evaluacion == 'CUALITATIVO':
            # --- LÓGICA PARA REPORTE CUALITATIVO POR DIMENSIÓN ---
            dimensiones = DimensionDesarrollo.objects.filter(institucion=institucion).order_by('orden')
            estudiantes = Estudiante.objects.filter(grado_actual=grado_seleccionado, activo=True)
            
            datos_cualitativos = []
            conteo_general_por_escala = OrderedDict((escala.nombre_escala, 0) for escala in EscalaCualitativa.objects.filter(institucion=institucion))

            for dimension in dimensiones:
                logros_dimension = LogroPreescolar.objects.filter(dimension=dimension, periodo=periodo_seleccionado)
                evaluaciones = EvaluacionLogroPreescolar.objects.filter(
                    estudiante__in=estudiantes, logro__in=logros_dimension
                ).select_related('estado')
                
                conteo_dimension = defaultdict(int)
                for ev in evaluaciones:
                    if ev.estado:
                        conteo_dimension[ev.estado.nombre_escala] += 1
                        if ev.estado.nombre_escala in conteo_general_por_escala:
                             conteo_general_por_escala[ev.estado.nombre_escala] +=1

                datos_cualitativos.append({'dimension': dimension.nombre, 'conteo': dict(conteo_dimension)})

            reporte_data = {
                'tipo_reporte': 'CUALITATIVO',
                'datos_tabla': datos_cualitativos,
                'chart_labels': json.dumps(list(conteo_general_por_escala.keys())),
                'chart_data': json.dumps(list(conteo_general_por_escala.values()))
            }
        else:
            # --- LÓGICA CUANTITATIVA ---
            areas_academicas = AreaAcademica.objects.filter(institucion=institucion)
            datos_cuantitativos = []
            for area in areas_academicas:
                materias_del_area = area.materias.all()
                promedio_area = Calificacion.objects.filter(
                    actividad_calificable__curso__grado=grado_seleccionado,
                    actividad_calificable__curso__periodo_academico=periodo_seleccionado,
                    actividad_calificable__curso__materia__in=materias_del_area,
                    valor_numerico__isnull=False
                ).aggregate(promedio=Avg('valor_numerico'))['promedio']
                if promedio_area is not None:
                    datos_cuantitativos.append({'area': area.nombre, 'promedio': promedio_area})
            
            datos_cuantitativos = sorted(datos_cuantitativos, key=lambda x: x['promedio'], reverse=True)
            reporte_data = {
                'tipo_reporte': 'CUANTITATIVO',
                'datos_tabla': datos_cuantitativos,
                'chart_labels': json.dumps([item['area'] for item in datos_cuantitativos]),
                'chart_data': json.dumps([float(item['promedio']) for item in datos_cuantitativos])
            }
            
    context = {
        'titulo_pagina': "Rendimiento por Áreas Académicas",
        'grados': grados, 'periodos': periodos,
        'grado_seleccionado': grado_seleccionado, 'periodo_seleccionado': periodo_seleccionado,
        'reporte_data': reporte_data,
    }
    return render(request, 'gestion_academica/reportes/reporte_promedio_area.html', context) 

@login_required
def reporte_final_reprobacion(request):
    """
    Genera un informe de fin de año con los estudiantes que reprobaron una o
    más materias. Se adapta a evaluaciones cuantitativas y cualitativas.
    """
    grados = Grado.objects.all().order_by('orden', 'nombre')
    años_escolares = PeriodoAcademico.objects.values_list('año_escolar', flat=True).distinct().order_by('-año_escolar')
    
    grado_id = request.GET.get('grado')
    año_seleccionado_str = request.GET.get('año')
    año_seleccionado = int(año_seleccionado_str) if año_seleccionado_str else (años_escolares.first() or timezone.now().year)

    reporte_data = {}
    grado_seleccionado = None

    if grado_id:
        grado_seleccionado = get_object_or_404(Grado, pk=grado_id)
        estudiantes_del_grado = Estudiante.objects.filter(grado_actual=grado_seleccionado, activo=True)
        periodos_del_año = PeriodoAcademico.objects.filter(año_escolar=año_seleccionado, institucion=grado_seleccionado.institucion)
        
        estudiantes_reprobados = defaultdict(list)
        conteo_reprobados_por_materia = defaultdict(int)

        if grado_seleccionado.tipo_evaluacion == 'CUALITATIVO':
            # --- LÓGICA PARA REPROBACIÓN CUALITATIVA ---
            escala_reprobado = EscalaCualitativa.objects.filter(institucion=grado_seleccionado.institucion, es_reprobatoria=True).first()
            if escala_reprobado:
                materias = Materia.objects.filter(cursos__grado=grado_seleccionado, cursos__periodo_academico__in=periodos_del_año).distinct()
                for estudiante in estudiantes_del_grado:
                    for materia in materias:
                        # Contamos si en el último periodo tuvo una evaluación reprobatoria en algún logro de esa materia
                        logros_reprobados = EvaluacionLogroPreescolar.objects.filter(
                            estudiante=estudiante,
                            logro__materia=materia,
                            logro__periodo=periodos_del_año.order_by('-fecha_fin').first(),
                            estado=escala_reprobado
                        ).exists()
                        if logros_reprobados:
                            estudiantes_reprobados[estudiante].append(materia.nombre_materia)
                            conteo_reprobados_por_materia[materia.nombre_materia] += 1
        
        else:
            # --- LÓGICA PARA REPROBACIÓN CUANTITATIVA ---
            nota_minima = grado_seleccionado.institucion.nota_minima_aprobacion
            cursos_del_grado = Curso.objects.filter(grado=grado_seleccionado, periodo_academico__in=periodos_del_año).select_related('materia')
            materias_del_grado = Materia.objects.filter(cursos__in=cursos_del_grado).distinct()

            for estudiante in estudiantes_del_grado:
                promedio_anual_por_materia = {}
                for materia in materias_del_grado:
                    cursos_materia_año = cursos_del_grado.filter(materia=materia)
                    notas_periodos = []
                    for curso in cursos_materia_año:
                        estado = calcular_estado_academico_curso(curso, estudiante)
                        nota_final = estado.get('nota_final_ponderada')
                        if nota_final is not None:
                            notas_periodos.append(nota_final)
                    
                    if notas_periodos:
                        promedio_anual_materia = sum(notas_periodos) / len(notas_periodos)
                        if promedio_anual_materia < nota_minima:
                            estudiantes_reprobados[estudiante].append(f"{materia.nombre_materia} ({promedio_anual_materia:.2f})")
                            conteo_reprobados_por_materia[materia.nombre_materia] += 1

        # Preparar datos para el gráfico
        chart_labels = list(conteo_reprobados_por_materia.keys())
        chart_data = list(conteo_reprobados_por_materia.values())

        reporte_data = {
            'estudiantes_reprobados': dict(estudiantes_reprobados),
            'chart_labels': json.dumps(chart_labels),
            'chart_data': json.dumps(chart_data)
        }

    context = {
        'titulo_pagina': "Informe Final de Reprobación",
        'grados': grados,
        'años_escolares': años_escolares,
        'grado_seleccionado': grado_seleccionado,
        'año_seleccionado': año_seleccionado,
        'reporte_data': reporte_data,
    }
    return render(request, 'gestion_academica/reportes/reporte_reprobacion.html', context)    

@login_required
def reporte_consolidado_materia(request):
    """
    Muestra una planilla de notas detallada (consolidado).
    VERSIÓN FINAL: Asegura que el gráfico cualitativo siempre se muestre.
    """
    grados = Grado.objects.all().order_by('orden', 'nombre')
    periodos = PeriodoAcademico.objects.all().order_by('-año_escolar', '-fecha_inicio')
    materias_del_grado = Materia.objects.none()

    grado_id = request.GET.get('grado')
    periodo_id = request.GET.get('periodo')
    materia_id = request.GET.get('materia')

    reporte_data = {}
    grado_seleccionado = None
    periodo_seleccionado = None
    materia_seleccionada = None

    if grado_id:
        grado_seleccionado = get_object_or_404(Grado, pk=grado_id)
        materias_del_grado = Materia.objects.filter(cursos__grado=grado_seleccionado).distinct().order_by('nombre_materia')
    if periodo_id:
        periodo_seleccionado = get_object_or_404(PeriodoAcademico, pk=periodo_id)

    if grado_seleccionado and periodo_seleccionado:
        estudiantes = Estudiante.objects.filter(grado_actual=grado_seleccionado, activo=True).select_related('usuario').order_by('usuario__last_name')
        institucion = grado_seleccionado.institucion

        if grado_seleccionado.tipo_evaluacion == 'CUALITATIVO':
            # --- LÓGICA DE GRÁFICO CORREGIDA PARA CUALITATIVO ---
            logros = LogroPreescolar.objects.filter(periodo=periodo_seleccionado, materia__cursos__grado=grado_seleccionado).distinct().select_related('dimension').order_by('dimension__orden', 'orden')
            evaluaciones = EvaluacionLogroPreescolar.objects.filter(estudiante__in=estudiantes, logro__in=logros)
            evaluaciones_map = {(ev.estudiante_id, ev.logro_id): ev.estado for ev in evaluaciones}
            
            logros_agrupados_por_dimension = OrderedDict()
            for logro in logros:
                if logro.dimension not in logros_agrupados_por_dimension:
                    logros_agrupados_por_dimension[logro.dimension] = []
                logros_agrupados_por_dimension[logro.dimension].append(logro)

            datos_tabla_cualitativa = []
            for est in estudiantes:
                evaluaciones_ordenadas = [evaluaciones_map.get((est.pk, logro.pk)) for logro in logros]
                datos_tabla_cualitativa.append({'estudiante': est, 'evaluaciones': evaluaciones_ordenadas})
            
            # 1. Obtenemos todas las escalas de la institución
            escala_completa = EscalaCualitativa.objects.filter(institucion=institucion).order_by('orden')
            # 2. Inicializamos el contador con todas las escalas en CERO
            conteo_desempenos = OrderedDict((escala.nombre_escala, 0) for escala in escala_completa)
            # 3. Actualizamos el conteo con las evaluaciones que sí existen
            for ev in evaluaciones:
                if ev.estado and ev.estado.nombre_escala in conteo_desempenos:
                    conteo_desempenos[ev.estado.nombre_escala] += 1
            
            reporte_data = {
                'tipo_reporte': 'CUALITATIVO', 'logros_agrupados': logros_agrupados_por_dimension,
                'datos_tabla': datos_tabla_cualitativa,
                'chart_labels': json.dumps(list(conteo_desempenos.keys())), # Esta lista nunca estará vacía
                'chart_data': json.dumps(list(conteo_desempenos.values())) # Podrá tener ceros, pero existirá
            }
        
        elif materia_id:
            # Lógica cuantitativa (sin cambios)
            materia_seleccionada = get_object_or_404(Materia, pk=materia_id)
            curso = Curso.objects.filter(grado=grado_seleccionado, periodo_academico=periodo_seleccionado, materia=materia_seleccionada).first()
            datos_tabla_cuantitativa = []
            notas_finales_para_grafico = []
            if curso:
                actividades = ActividadCalificable.objects.filter(curso=curso).order_by('fecha_publicacion')
                calificaciones = Calificacion.objects.filter(actividad_calificable__in=actividades)
                calificaciones_map = {(cal.estudiante_id, cal.actividad_calificable_id): cal.valor_numerico for cal in calificaciones}
                for est in estudiantes:
                    # ... (resto de la lógica cuantitativa sin cambios) ...
                    estado = calcular_estado_academico_curso(curso, est)
                    nota_final = estado.get('nota_final_ponderada')
                    if nota_final is not None:
                        notas_finales_para_grafico.append(nota_final)
                    datos_tabla_cuantitativa.append({'estudiante': est, 'calificaciones': [calificaciones_map.get((est.pk, act.pk)) for act in actividades], 'nota_final': nota_final})
                
                rangos = {"Reprobado": 0, "Básico": 0, "Alto": 0, "Superior": 0}
                nota_minima = institucion.nota_minima_aprobacion
                for nota in notas_finales_para_grafico:
                    if nota < nota_minima: rangos["Reprobado"] += 1
                    elif nota < 4.0: rangos["Básico"] += 1
                    elif nota < 4.6: rangos["Alto"] += 1
                    else: rangos["Superior"] += 1
                
                reporte_data = {
                    'tipo_reporte': 'CUANTITATIVO', 'actividades_header': actividades,
                    'datos_tabla': datos_tabla_cuantitativa,
                    'chart_labels': json.dumps(list(rangos.keys())),
                    'chart_data': json.dumps(list(rangos.values()))
                }
            
    context = {
        'titulo_pagina': "Consolidado de Notas por Materia",
        'grados': grados, 'periodos': periodos, 'materias_del_grado': materias_del_grado,
        'grado_seleccionado': grado_seleccionado, 'periodo_seleccionado': periodo_seleccionado,
        'materia_seleccionada': materia_seleccionada, 'reporte_data': reporte_data,
    }
    return render(request, 'gestion_academica/reportes/reporte_consolidado_materia.html', context)


@login_required
def reporte_consolidado_areas(request):
    """
    Muestra un consolidado por Área (cuantitativo) o por Dimensión (cualitativo).
    VERSIÓN CORREGIDA Y DEFINITIVA PARA INCLUIR PREESCOLAR.
    """
    # CORRECCIÓN: Quitamos el filtro inicial para mostrar TODOS los grados
    grados = Grado.objects.all().order_by('orden', 'nombre')
    periodos = PeriodoAcademico.objects.all().order_by('-año_escolar', '-fecha_inicio')

    grado_id = request.GET.get('grado')
    periodo_id = request.GET.get('periodo')

    reporte_data = {}
    grado_seleccionado = None
    periodo_seleccionado = None

    if grado_id and periodo_id:
        grado_seleccionado = get_object_or_404(Grado, pk=grado_id)
        periodo_seleccionado = get_object_or_404(PeriodoAcademico, pk=periodo_id)
        institucion = grado_seleccionado.institucion
        estudiantes = Estudiante.objects.filter(grado_actual=grado_seleccionado, activo=True).select_related('usuario')

        if grado_seleccionado.tipo_evaluacion == 'CUALITATIVO':
            # --- LÓGICA PARA CONSOLIDADO CUALITATIVO POR DIMENSIÓN ---
            dimensiones = DimensionDesarrollo.objects.filter(institucion=institucion).order_by('orden')
            datos_tabla = []
            conteo_general_para_grafico = defaultdict(int)

            for estudiante in estudiantes:
                resumen_por_dimension = OrderedDict()
                for dimension in dimensiones:
                    evaluaciones = EvaluacionLogroPreescolar.objects.filter(
                        estudiante=estudiante, logro__dimension=dimension, logro__periodo=periodo_seleccionado
                    ).select_related('estado')
                    
                    resumen_dimension = defaultdict(int)
                    for ev in evaluaciones:
                        if ev.estado:
                            resumen_dimension[ev.estado.abreviatura] += 1
                            conteo_general_para_grafico[ev.estado.nombre_escala] += 1
                    
                    resumen_por_dimension[dimension.nombre] = dict(resumen_dimension)
                
                datos_tabla.append({'estudiante': estudiante, 'resumen_por_dimension': resumen_por_dimension})

            reporte_data = {
                'tipo_reporte': 'CUALITATIVO', 'areas_header': dimensiones,
                'datos_tabla': datos_tabla,
                'chart_labels': json.dumps(list(conteo_general_para_grafico.keys())),
                'chart_data': json.dumps(list(conteo_general_para_grafico.values()))
            }
        
        else:
            # --- LÓGICA PARA CONSOLIDADO CUANTITATIVO POR ÁREA ---
            areas = AreaAcademica.objects.filter(institucion=institucion).prefetch_related('materias')
            datos_tabla = []
            conteo_desempenos_por_area = OrderedDict((area.nombre, defaultdict(int)) for area in areas)

            for estudiante in estudiantes:
                promedios_por_area = OrderedDict()
                notas_para_promedio_general = []

                for area in areas:
                    materias_del_area = area.materias.all()
                    cursos_del_area = Curso.objects.filter(grado=grado_seleccionado, periodo_academico=periodo_seleccionado, materia__in=materias_del_area)
                    notas_finales_area = [estado.get('nota_final_ponderada') for curso in cursos_del_area if (estado := calcular_estado_academico_curso(curso, estudiante)) and estado.get('nota_final_ponderada') is not None]
                    promedio_area = sum(notas_finales_area) / len(notas_finales_area) if notas_finales_area else None
                    promedios_por_area[area.nombre] = promedio_area
                    if promedio_area:
                        notas_para_promedio_general.append(promedio_area)
                        desempeno = obtener_desempeno(promedio_area, institucion)
                        if desempeno: conteo_desempenos_por_area[area.nombre][desempeno] += 1
                
                promedio_general_estudiante = sum(notas_para_promedio_general) / len(notas_para_promedio_general) if notas_para_promedio_general else None
                datos_tabla.append({'estudiante': estudiante, 'promedios_por_area': promedios_por_area, 'promedio_general': promedio_general_estudiante})
            
            escalas = EscalaValorativa.objects.filter(institucion=institucion).order_by('orden')
            chart_labels = list(conteo_desempenos_por_area.keys())
            chart_datasets = [{'label': esc.nombre_desempeno, 'data': [conteo_desempenos_por_area[area][esc.abreviatura] for area in chart_labels], 'backgroundColor': ['rgba(220, 53, 69, 0.7)','rgba(255, 193, 7, 0.7)','rgba(25, 135, 84, 0.7)','rgba(13, 110, 253, 0.7)'][i % 4]} for i, esc in enumerate(escalas)]
            
            reporte_data = {
                'tipo_reporte': 'CUANTITATIVO', 'areas_header': areas,
                'datos_tabla': datos_tabla,
                'chart_labels': json.dumps(chart_labels),
                'chart_datasets': json.dumps(chart_datasets)
            }

    context = {
        'titulo_pagina': "Consolidado por Áreas Académicas",
        'grados': grados, 'periodos': periodos,
        'grado_seleccionado': grado_seleccionado, 'periodo_seleccionado': periodo_seleccionado,
        'reporte_data': reporte_data,
    }
    return render(request, 'gestion_academica/reportes/reporte_consolidado_areas.html', context)

@login_required
def reporte_ranking_institucion(request):
    """
    Vista que INICIA la tarea de Celery para el ranking y muestra
    una página para esperar y ver los resultados.
    """
    periodos = PeriodoAcademico.objects.all().order_by('-año_escolar', '-fecha_inicio')
    periodo_id = request.GET.get('periodo')
    task_id = None

    if periodo_id:
        # En lugar de calcular aquí, llamamos a la tarea con .delay()
        # Esto envía la tarea a Redis y devuelve inmediatamente un ID.
        task = generar_ranking_institucional_task.delay(periodo_id)
        task_id = task.id

    context = {
        'titulo_pagina': "Ranking General de la Institución",
        'periodos': periodos,
        'periodo_seleccionado_id': periodo_id,
        'task_id': task_id, # Pasamos el ID de la tarea a la plantilla
    }
    return render(request, 'gestion_academica/reportes/reporte_ranking_institucion.html', context)

@login_required
def reporte_promedio_cualitativo(request):
    """
    Genera un resumen estadístico y un gráfico de pastel con la distribución
    de los desempeños cualitativos para un grado de preescolar en un periodo.
    """
    # Filtramos para que en el selector solo aparezcan grados cualitativos
    grados = Grado.objects.filter(tipo_evaluacion='CUALITATIVO').order_by('orden', 'nombre')
    periodos = PeriodoAcademico.objects.all().order_by('-año_escolar', '-fecha_inicio')

    grado_id = request.GET.get('grado')
    periodo_id = request.GET.get('periodo')

    reporte_data = {}
    grado_seleccionado = None
    periodo_seleccionado = None

    if grado_id and periodo_id:
        grado_seleccionado = get_object_or_404(Grado, pk=grado_id)
        periodo_seleccionado = get_object_or_404(PeriodoAcademico, pk=periodo_id)
        institucion = grado_seleccionado.institucion

        # 1. Obtenemos todas las escalas posibles de la institución para usarlas como base
        escala_completa = EscalaCualitativa.objects.filter(institucion=institucion).order_by('orden')
        
        # 2. Inicializamos el contador con todas las escalas en cero
        conteo_desempenos = OrderedDict((escala.nombre_escala, 0) for escala in escala_completa)

        # 3. Buscamos todas las evaluaciones existentes para ese grado y periodo
        estudiantes = Estudiante.objects.filter(grado_actual=grado_seleccionado, activo=True)
        evaluaciones = EvaluacionLogroPreescolar.objects.filter(
            estudiante__in=estudiantes,
            logro__periodo=periodo_seleccionado
        ).select_related('estado')

        # 4. Actualizamos el conteo con los datos reales
        for ev in evaluaciones:
            if ev.estado and ev.estado.nombre_escala in conteo_desempenos:
                conteo_desempenos[ev.estado.nombre_escala] += 1
        
        total_evaluaciones = sum(conteo_desempenos.values())
        
        # 5. Preparamos los datos para la tabla y el gráfico
        datos_tabla = []
        for nombre, cantidad in conteo_desempenos.items():
            porcentaje = (cantidad / total_evaluaciones * 100) if total_evaluaciones > 0 else 0
            datos_tabla.append({
                'escala': nombre,
                'cantidad': cantidad,
                'porcentaje': porcentaje
            })

        reporte_data = {
            'datos_tabla': datos_tabla,
            'total_evaluaciones': total_evaluaciones,
            'chart_labels': json.dumps(list(conteo_desempenos.keys())),
            'chart_data': json.dumps(list(conteo_desempenos.values()))
        }

    context = {
        'titulo_pagina': "Resumen de Desempeño Cualitativo por Grado",
        'grados': grados,
        'periodos': periodos,
        'grado_seleccionado': grado_seleccionado,
        'periodo_seleccionado': periodo_seleccionado,
        'reporte_data': reporte_data,
    }
    return render(request, 'gestion_academica/reportes/reporte_promedio_cualitativo.html', context) 


@login_required
def reporte_promedio_por_materia(request):
    """
    Muestra el rendimiento promedio de una materia específica a través de
    todos los grados en los que se imparte durante un periodo.
    """
    # Para los filtros, mostramos todas las materias y periodos de la institución
    materias = Materia.objects.all().order_by('nombre_materia')
    periodos = PeriodoAcademico.objects.all().order_by('-año_escolar', '-fecha_inicio')

    materia_id = request.GET.get('materia')
    periodo_id = request.GET.get('periodo')

    reporte_data = {}
    materia_seleccionada = None
    periodo_seleccionado = None

    if materia_id and periodo_id:
        materia_seleccionada = get_object_or_404(Materia, pk=materia_id)
        periodo_seleccionado = get_object_or_404(PeriodoAcademico, pk=periodo_id)
        
        # 1. Encontramos todos los cursos de esa materia en ese periodo
        cursos_de_la_materia = Curso.objects.filter(
            materia=materia_seleccionada,
            periodo_academico=periodo_seleccionado,
            grado__tipo_evaluacion='CUANTITATIVO' # Solo en grados cuantitativos
        ).select_related('grado')

        datos_tabla = []
        for curso in cursos_de_la_materia:
            # 2. Para cada curso (es decir, para cada grado), calculamos el promedio general
            promedio_grado_en_materia = Calificacion.objects.filter(
                actividad_calificable__curso=curso,
                valor_numerico__isnull=False
            ).aggregate(
                promedio=Avg('valor_numerico')
            )['promedio']

            if promedio_grado_en_materia is not None:
                datos_tabla.append({
                    'grado': curso.grado,
                    'promedio': promedio_grado_en_materia
                })

        # 3. Preparamos los datos para la tabla y el gráfico
        datos_tabla = sorted(datos_tabla, key=lambda x: x['grado'].orden)
        
        reporte_data = {
            'datos_tabla': datos_tabla,
            'chart_labels': json.dumps([item['grado'].nombre for item in datos_tabla]),
            'chart_data': json.dumps([float(item['promedio']) for item in datos_tabla])
        }

    context = {
        'titulo_pagina': "Rendimiento Comparativo por Materia",
        'materias': materias,
        'periodos': periodos,
        'materia_seleccionada': materia_seleccionada,
        'periodo_seleccionado': periodo_seleccionado,
        'reporte_data': reporte_data,
    }
    return render(request, 'gestion_academica/reportes/reporte_promedio_materia.html', context)

          
@login_required
def cuadro_honor_grado(request):
    """
    Muestra un ranking de estudiantes DENTRO de un grado específico.
    VERSIÓN CORREGIDA: Muestra todos los grados y maneja la selección.
    """
    # CORRECCIÓN: Quitamos el filtro para mostrar TODOS los grados
    grados = Grado.objects.all().order_by('orden', 'nombre')
    periodos = PeriodoAcademico.objects.all().order_by('-año_escolar', '-fecha_inicio')

    grado_id = request.GET.get('grado')
    periodo_id = request.GET.get('periodo')

    reporte_data = []
    grado_seleccionado = None
    periodo_seleccionado = None
    
    # Esta nueva variable nos dirá si el reporte no aplica
    reporte_no_aplica = False

    if grado_id and periodo_id:
        grado_seleccionado = get_object_or_404(Grado, pk=grado_id)
        periodo_seleccionado = get_object_or_404(PeriodoAcademico, pk=periodo_id)
        
        # VERIFICAMOS EL TIPO DE EVALUACIÓN DESPUÉS DE SELECCIONAR
        if grado_seleccionado.tipo_evaluacion == 'CUANTITATIVO':
            estudiantes_del_grado = Estudiante.objects.filter(grado_actual=grado_seleccionado, activo=True)
            
            for estudiante in estudiantes_del_grado:
                cursos = Curso.objects.filter(grado=estudiante.grado_actual, periodo_academico=periodo_seleccionado)
                total_puntos_ponderados = Decimal('0.0')
                total_ihs = 0
                for curso in cursos:
                    estado = calcular_estado_academico_curso(curso, estudiante)
                    nota_final = estado.get('nota_final_ponderada')
                    ihs = curso.materia.intensidad_horaria_semanal
                    if nota_final is not None and ihs > 0:
                        total_puntos_ponderados += nota_final * ihs
                        total_ihs += ihs
                promedio_general = total_puntos_ponderados / total_ihs if total_ihs > 0 else None
                if promedio_general is not None:
                    reporte_data.append({'estudiante': estudiante, 'promedio': promedio_general})

            reporte_data = sorted(reporte_data, key=lambda x: x['promedio'], reverse=True)
        else:
            # Si el grado es CUALITATIVO, activamos nuestra bandera
            reporte_no_aplica = True

    context = {
        'titulo_pagina': "Cuadro de Honor por Grado",
        'grados': grados,
        'periodos': periodos,
        'grado_seleccionado': grado_seleccionado,
        'periodo_seleccionado': periodo_seleccionado,
        'reporte_data': reporte_data,
        'reporte_no_aplica': reporte_no_aplica, # Pasamos la bandera a la plantilla
        'chart_labels': json.dumps([item['estudiante'].usuario.get_full_name() for item in reporte_data]),
        'chart_data': json.dumps([float(item['promedio']) for item in reporte_data]),
    }
    return render(request, 'gestion_academica/reportes/cuadro_honor_grado.html', context)        


@login_required
def reporte_estadistica_asistencia_diaria(request):
    """
    Muestra un resumen estadístico de la asistencia (presentes, ausentes, etc.)
    para una fecha específica, con un gráfico de pastel.
    """
    # Lógica de filtros
    fecha_str = request.GET.get('fecha', timezone.localdate().strftime('%Y-%m-%d'))
    try:
        fecha_seleccionada = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        fecha_seleccionada = timezone.localdate()

    institucion = request.user.institucion_asociada
    reporte_data = {}
    
    if institucion:
        # 1. Obtenemos el total de estudiantes activos
        total_estudiantes = Estudiante.objects.filter(institucion=institucion, activo=True).count()

        # 2. Contamos los registros de asistencia para la fecha seleccionada
        conteo_estados = RegistroAsistencia.objects.filter(
            fecha__date=fecha_seleccionada,
            institucion=institucion
        ).values('estado').annotate(total=Count('id'))
        
        # 3. Procesamos los conteos en un diccionario limpio
        resumen = {
            'Presentes': 0,
            'Ausentes': 0,
            'Tardanzas': 0,
            'Justificados': 0
        }
        for item in conteo_estados:
            # Mapeamos los valores del modelo a nombres más amigables
            if item['estado'] == 'PRESENTE': resumen['Presentes'] = item['total']
            elif item['estado'] == 'AUSENTE': resumen['Ausentes'] = item['total']
            elif item['estado'] == 'TARDANZA': resumen['Tardanzas'] = item['total']
            elif item['estado'] == 'JUSTIFICADO': resumen['Justificados'] = item['total']
        
        # 4. Calculamos los estudiantes sin registro
        total_registrados = sum(resumen.values())
        resumen['Sin Registro'] = total_estudiantes - total_registrados
        
        reporte_data = {
            'total_estudiantes': total_estudiantes,
            'resumen': resumen,
            'chart_labels': json.dumps(list(resumen.keys())),
            'chart_data': json.dumps(list(resumen.values())),
        }

    context = {
        'titulo_pagina': "Estadística de Asistencia Diaria",
        'fecha_seleccionada': fecha_seleccionada,
        'reporte_data': reporte_data,
    }
    return render(request, 'gestion_academica/reportes/reporte_asistencia_diaria.html', context)


@login_required
@permission_required('gestion_academica.view_registroasistencia') # Permiso adecuado
def reporte_asistencia_materia(request):
    """
    Muestra un reporte detallado de asistencia por materia, listando a cada
    estudiante y su conteo de presentes, ausentes y tardanzas.
    """
    institucion = request.user.institucion_asociada
    
    # Filtramos por la institución del usuario para seguridad
    grados = Grado.objects.filter(institucion=institucion).order_by('orden')
    periodos = PeriodoAcademico.objects.filter(institucion=institucion).order_by('-año_escolar', '-fecha_inicio')
    materias_del_grado = Materia.objects.none()

    grado_id = request.GET.get('grado')
    periodo_id = request.GET.get('periodo')
    materia_id = request.GET.get('materia')

    reporte_data = {}
    curso_seleccionado = None

    if grado_id:
        materias_del_grado = Materia.objects.filter(
            cursos__grado_id=grado_id, 
            cursos__institucion=institucion
        ).distinct().order_by('nombre_materia')

    if grado_id and periodo_id and materia_id:
        curso_seleccionado = get_object_or_404(
            Curso.objects.select_related('grado', 'periodo_academico', 'materia'),
            grado_id=grado_id, 
            periodo_academico_id=periodo_id, 
            materia_id=materia_id,
            institucion=institucion
        )
        
        # Obtenemos los estudiantes y anotamos sus conteos de asistencia para este curso
        estudiantes_con_asistencia = Estudiante.objects.filter(
            grado_actual=curso_seleccionado.grado, 
            activo=True,
            institucion=institucion
        ).annotate(
            total_presente=Count('asistencias', filter=Q(asistencias__curso=curso_seleccionado, asistencias__estado='PRESENTE')),
            total_ausente=Count('asistencias', filter=Q(asistencias__curso=curso_seleccionado, asistencias__estado='AUSENTE')),
            total_tardanza=Count('asistencias', filter=Q(asistencias__curso=curso_seleccionado, asistencias__estado='TARDANZA')),
            total_justificado=Count('asistencias', filter=Q(asistencias__curso=curso_seleccionado, asistencias__estado='JUSTIFICADO')),
        ).select_related('usuario').order_by('usuario__last_name')

        # Preparamos los datos para el gráfico
        total_general_presentes = sum(e.total_presente for e in estudiantes_con_asistencia)
        total_general_ausentes = sum(e.total_ausente for e in estudiantes_con_asistencia)
        total_general_tardanzas = sum(e.total_tardanza for e in estudiantes_con_asistencia)
        total_general_justificados = sum(e.total_justificado for e in estudiantes_con_asistencia)

        reporte_data = {
            'estudiantes_data': estudiantes_con_asistencia,
            'chart_labels': json.dumps(['Presentes', 'Ausentes', 'Tardanzas', 'Justificados']),
            'chart_data': json.dumps([total_general_presentes, total_general_ausentes, total_general_tardanzas, total_general_justificados])
        }

    context = {
        'titulo_pagina': "Reporte de Asistencia por Materia",
        'grados': grados, 'periodos': periodos, 'materias_del_grado': materias_del_grado,
        'curso_seleccionado': curso_seleccionado,
        'reporte_data': reporte_data
    }
    return render(request, 'gestion_academica/reportes/reporte_asistencia_materia.html', context)


@login_required
def reporte_incidencias_estudiante(request):
    """
    Muestra un historial detallado de todas las anotaciones en el observador
    para un estudiante específico, con un gráfico resumen.
    """
    grados = Grado.objects.all().order_by('orden', 'nombre')
    estudiantes_del_grado = Estudiante.objects.none()

    grado_id = request.GET.get('grado')
    estudiante_id = request.GET.get('estudiante')

    reporte_data = {}
    estudiante_seleccionado = None

    if grado_id:
        estudiantes_del_grado = Estudiante.objects.filter(grado_actual__id=grado_id).select_related('usuario').order_by('usuario__last_name')

    if estudiante_id:
        estudiante_seleccionado = get_object_or_404(Estudiante, pk=estudiante_id)
        
        # Obtenemos todas las anotaciones del estudiante
        anotaciones = AnotacionObservador.objects.filter(
            estudiante=estudiante_seleccionado
        ).select_related('registrado_por').order_by('-fecha_hora')

        # Preparamos los datos para el gráfico, contando por tipo de anotación
        conteo_por_tipo = defaultdict(int)
        for an in anotaciones:
            conteo_por_tipo[an.get_tipo_display()] += 1
        
        reporte_data = {
            'anotaciones': anotaciones,
            'chart_labels': json.dumps(list(conteo_por_tipo.keys())),
            'chart_data': json.dumps(list(conteo_por_tipo.values()))
        }

    context = {
        'titulo_pagina': "Reporte de Incidencias y Observador",
        'grados': grados,
        'estudiantes_del_grado': estudiantes_del_grado,
        'grado_seleccionado_id': grado_id,
        'estudiante_seleccionado_id': estudiante_id,
        'estudiante_seleccionado': estudiante_seleccionado,
        'reporte_data': reporte_data
    }
    return render(request, 'gestion_academica/reportes/reporte_incidencias.html', context)

@login_required
def reporte_consolidado_convivencia(request):
    """
    Muestra un consolidado de todas las anotaciones de convivencia (Halu Sentinel)
    clasificadas por la IA para toda la institución o un grado específico.
    """
    grados = Grado.objects.all().order_by('orden', 'nombre')
    periodos = PeriodoAcademico.objects.all().order_by('-año_escolar', '-fecha_inicio')

    grado_id = request.GET.get('grado')
    periodo_id = request.GET.get('periodo')

    reporte_data = {}
    grado_seleccionado = None
    periodo_seleccionado = None

    # El coordinador puede ver el consolidado sin necesidad de filtrar
    institucion = request.user.institucion_asociada
    
    # 1. Obtenemos todas las anotaciones que han sido clasificadas por la IA
    anotaciones_qs = AnotacionObservador.objects.filter(
        institucion=institucion
    ).exclude(
        Q(tipo_situacion_ia='NINGUNO') | Q(tipo_situacion_ia__isnull=True)
    ).select_related('estudiante__usuario', 'estudiante__grado_actual')

    # 2. Aplicamos los filtros si existen
    if grado_id:
        grado_seleccionado = get_object_or_404(Grado, pk=grado_id)
        anotaciones_qs = anotaciones_qs.filter(estudiante__grado_actual=grado_seleccionado)
    if periodo_id:
        periodo_seleccionado = get_object_or_404(PeriodoAcademico, pk=periodo_id)
        anotaciones_qs = anotaciones_qs.filter(fecha_hora__range=(periodo_seleccionado.fecha_inicio, periodo_seleccionado.fecha_fin))

    # 3. Agrupamos las anotaciones por estudiante
    casos_por_estudiante = defaultdict(list)
    for anotacion in anotaciones_qs:
        casos_por_estudiante[anotacion.estudiante].append(anotacion)

    # 4. Preparamos datos para el gráfico (conteo por tipo de situación)
    conteo_por_tipo = anotaciones_qs.values('tipo_situacion_ia').annotate(
        total=Count('id')
    ).order_by('tipo_situacion_ia')

    reporte_data = {
        'casos_por_estudiante': sorted(casos_por_estudiante.items(), key=lambda item: item[0].usuario.last_name),
        'chart_labels': json.dumps([item['tipo_situacion_ia'] for item in conteo_por_tipo]),
        'chart_data': json.dumps([item['total'] for item in conteo_por_tipo])
    }

    context = {
        'titulo_pagina': "Consolidado de Convivencia Escolar (Halu Sentinel)",
        'grados': grados,
        'periodos': periodos,
        'grado_seleccionado': grado_seleccionado,
        'periodo_seleccionado': periodo_seleccionado,
        'reporte_data': reporte_data,
    }
    return render(request, 'gestion_academica/reportes/reporte_consolidado_convivencia.html', context)   

@login_required
def espera_activacion(request):
    """
    Página que se muestra a los usuarios que se han registrado pero
    aún no tienen un rol (Estudiante, Docente, etc.) asignado.
    """
    context = {
        'titulo_pagina': "Cuenta Creada - Esperando Activación",
        'usuario_actual': request.user
    }
    return render(request, 'gestion_academica/espera_activacion.html', context)  
           

@login_required
def google_calendar_callback(request):
    """
    Se ejecuta después de que el usuario autoriza el acceso a su calendario.
    Guarda el ID del calendario principal del usuario y lanza la primera
    sincronización en segundo plano.
    """
    try:
        # 1. Obtenemos el token de acceso que allauth guardó en la base de datos
        #    para el usuario actual y el proveedor 'google'.
        social_token = SocialToken.objects.get(account__user=request.user, account__provider='google')
        
        # 2. Creamos las credenciales de Google a partir del token guardado.
        #    Estas credenciales nos dan el poder de actuar en nombre del usuario.
        credentials = Credentials(
            token=social_token.token,
            refresh_token=social_token.token_secret,
            token_uri='https://oauth2.googleapis.com/token',
            client_id=social_token.app.client_id,
            client_secret=social_token.app.secret,
            scopes=['https://www.googleapis.com/auth/calendar']
        )

        # 3. Usamos las credenciales para conectarnos al servicio de la API de Google Calendar.
        service = build('calendar', 'v3', credentials=credentials)

        # 4. Le pedimos a la API que nos dé la información del calendario "principal" del usuario.
        #    Normalmente, este es el calendario asociado a su dirección de correo.
        calendar_info = service.calendars().get(calendarId='primary').execute()
        calendar_id = calendar_info['id']

        # 5. Guardamos este ID único en el perfil de nuestro usuario en la base de datos.
        user = request.user
        user.google_calendar_id = calendar_id
        user.save(update_fields=['google_calendar_id'])

        # 6. Despachamos la tarea de sincronización a Celery para que se ejecute en segundo plano.
        #    Le pasamos el ID del usuario para que la tarea sepa a quién sincronizar.
        sincronizar_horario_google_calendar_task.delay(user.id)

        messages.success(request, "¡Tu calendario ha sido conectado! La sincronización de tu horario ha comenzado y aparecerá en unos minutos.")

    except SocialToken.DoesNotExist:
        messages.error(request, "No se encontró la conexión con tu cuenta de Google. Por favor, intenta conectar tu cuenta desde tu perfil primero.")
    except Exception as e:
        logger.error(f"Error al conectar con Google Calendar para el usuario {request.user.id}: {e}", exc_info=True)
        messages.error(request, f"Ocurrió un error inesperado al conectar tu calendario: {e}")

    # 7. Finalmente, redirigimos al usuario de vuelta a su página de perfil.
    return redirect('gestion_academica:ver_mi_perfil') 

@login_required
def planeacion_clases_view(request):
    """
    Vista principal para el módulo de planeación de clases.
    AHORA: Crea la planeación, la pone a generar y redirige al detalle.
    """
    try:
        docente = request.user.docente
    except AttributeError:
        messages.error(request, "Acceso denegado. Esta sección es solo para docentes.")
        return redirect('gestion_academica:inicio_academico')

    if request.method == 'POST':
        form = PlaneacionClaseForm(request.POST, user=request.user)
        if form.is_valid():
            # 1. Crea la instancia pero no la guardes aún en la base de datos
            planeacion = form.save(commit=False)
            planeacion.docente = docente
            
            # --- INICIO DE LA CORRECCIÓN CLAVE ---
            # 2. Establece el estado a 'GENERANDO' inmediatamente
            planeacion.estado_generacion = PlaneacionClase.EstadoGeneracion.GENERANDO
            
            # 3. Guarda la planeación en la base de datos. Ahora se crea con el estado correcto.
            planeacion.save()

            # 4. Llama a la tarea de Celery para que comience a trabajar
            from ..tasks import generar_contenido_planeacion_task
            generar_contenido_planeacion_task.delay(planeacion.pk)
            
            # 5. Redirige al usuario DIRECTAMENTE a la página de detalle
            messages.info(request, "La generación de la planeación ha comenzado. La página se actualizará automáticamente.")
            return redirect('gestion_academica:planeacion_detalle', pk=planeacion.pk) 
            # --- FIN DE LA CORRECCIÓN CLAVE ---

    else:
        form = PlaneacionClaseForm(user=request.user)

    planeaciones_existentes = PlaneacionClase.objects.filter(docente=docente).select_related('curso__materia', 'curso__grado')

    context = {
        'titulo_pagina': "Planeador de Clases con IA",
        'form': form,
        'planeaciones': planeaciones_existentes
    }
    return render(request, 'gestion_academica/planeacion_clases.html', context)
  

@login_required
def planeacion_detalle_view(request, pk):
    """
    Muestra el detalle de una planeación y lanza la tarea de Celery
    para generar el contenido con IA.
    """
    planeacion = get_object_or_404(PlaneacionClase, pk=pk, docente=request.user.docente)

    # Si la petición es POST, significa que el usuario hizo clic en "Generar" o "Reintentar"
    if request.method == 'POST':
        # Cambiamos el estado a 'GENERANDO' inmediatamente para que el usuario vea el feedback
        planeacion.estado_generacion = PlaneacionClase.EstadoGeneracion.GENERANDO
        planeacion.error_generacion = None
        planeacion.save()

        # Despachamos la tarea a Celery. Esto es instantáneo.
        generar_contenido_planeacion_task.delay(planeacion.id)
        
        messages.info(request, "La IA ha comenzado a generar tu planeación. La página se actualizará automáticamente cuando esté lista.")
        # Redirigimos de vuelta a la misma página para mostrar el estado "Generando..."
        return redirect('gestion_academica:planeacion_detalle', pk=pk)

    context = {
        'titulo_pagina': f"Planeación: {planeacion.titulo}",
        'planeacion': planeacion,
    }
    return render(request, 'gestion_academica/planeacion_detalle.html', context)


@login_required
@require_POST # Esta vista solo acepta peticiones POST para mayor seguridad
def cancelar_generacion_planeacion(request, pk):
    """
    Permite al docente cancelar una tarea de generación de planeación
    que se ha quedado atascada en el estado 'GENERANDO'.
    """
    planeacion = get_object_or_404(PlaneacionClase, pk=pk, docente=request.user.docente)

    # Solo actuamos si la planeación está efectivamente "generando"
    if planeacion.estado_generacion == PlaneacionClase.EstadoGeneracion.GENERANDO:
        planeacion.estado_generacion = PlaneacionClase.EstadoGeneracion.FALLIDO
        planeacion.error_generacion = "El proceso de generación fue cancelado manualmente por el usuario."
        planeacion.save()
        messages.warning(request, "El proceso de generación de la planeación ha sido detenido.")
    else:
        messages.info(request, "Esta planeación no se estaba generando activamente.")

    return redirect('gestion_academica:planeacion_detalle', pk=planeacion.pk)    

@login_required
@never_cache
def get_planeacion_status_api(request, pk):
    """Endpoint de polling para el modal del planeador IA.

    Devuelve estado + datos suficientes para que el frontend muestre un toast
    informativo al completarse (o un mensaje de error detallado al fallar) sin
    necesidad de recargar antes para "descubrir" qué pasó.
    """
    try:
        planeacion = PlaneacionClase.objects.get(pk=pk)

        if not request.user.is_superuser:
            if not hasattr(request.user, 'docente') or planeacion.docente_id != request.user.docente.pk:
                return JsonResponse({'status': 'FORBIDDEN'}, status=403)

        estado = str(planeacion.estado_generacion or "").strip().upper()
        detalles_count = planeacion.detalles_clase.count()

        data = {
            'status': estado,
            'detalles_count': detalles_count,
            'error_generacion': planeacion.error_generacion or "",
            'titulo': planeacion.titulo,
            'duracion_clases': planeacion.duracion_clases,
        }

        # Mensaje listo para mostrarse como toast en el frontend
        if estado == "COMPLETADO":
            data['mensaje'] = (
                f"¡Planeación generada con éxito! Se crearon "
                f"{detalles_count} clase{'s' if detalles_count != 1 else ''} para '{planeacion.titulo}'."
            )
        elif estado == "FALLIDO":
            data['mensaje'] = (
                f"La generación falló: {planeacion.error_generacion or 'error desconocido'}."
            )
        elif estado == "GENERANDO":
            data['mensaje'] = "La IA sigue procesando tu planeación. Por favor espera…"
        else:
            data['mensaje'] = "Pendiente de iniciar la generación."

        return JsonResponse(data)
    except PlaneacionClase.DoesNotExist:
        return JsonResponse({'status': 'NOT_FOUND'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'ERROR', 'message': str(e)}, status=500)

@login_required
def generar_planeacion_pdf(request, pk):
    """
    Genera un documento PDF para una planeación de clase específica.
    """
    try:
        # Buscamos la planeación, asegurándonos de que pertenezca al docente logueado
        planeacion = get_object_or_404(PlaneacionClase, pk=pk, docente=request.user.docente)
        
        # Obtenemos los detalles de la clase asociados
        detalles = planeacion.detalles_clase.all()

        # Preparamos el contexto para la plantilla del PDF
        context = {
            'planeacion': planeacion,
            'detalles': detalles,
            'institucion': planeacion.institucion,
        }

        template_path = 'gestion_academica/pdfs/planeacion_pdf.html'
        template = get_template(template_path)
        html = template.render(context)

        # Creamos el PDF en memoria
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="planeacion_{planeacion.titulo}.pdf"'
        
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html, dest=pdf_buffer)

        if pisa_status.err:
            logger.error("Error al generar PDF con xhtml2pdf (pisa.err=%s)", pisa_status.err)
        return HttpResponse('Error al generar el PDF. Por favor, inténtelo de nuevo.', status=500)

        response.write(pdf_buffer.getvalue())
        return response

    except Exception as e:
        logger.error(f"Error al generar el PDF de la planeación {pk}: {e}", exc_info=True)
        messages.error(request, "No se pudo generar el PDF de la planeación.")
        return redirect('gestion_academica:planeacion_detalle', pk=pk)



@login_required
@require_POST # Esta vista solo se puede llamar con un método POST para seguridad
def anadir_planeacion_a_lecciones(request, pk):
    """
    Activa el proceso de crear los registros de LeccionDiaria a partir de una planeación.
    """
    planeacion = get_object_or_404(PlaneacionClase, pk=pk, docente=request.user.docente)
    
    # Llamamos a la función del utils.py que hace todo el trabajo
    lecciones_creadas, mensaje = crear_lecciones_diarias_desde_planeacion(planeacion.id)
    
    if lecciones_creadas > 0:
        messages.success(request, mensaje)
    else:
        messages.error(request, mensaje)
        
    return redirect('gestion_academica:planeacion_detalle', pk=planeacion.pk)


@login_required
@permission_required('gestion_academica.view_bloquehorario', raise_exception=True)
def gestion_horarios_view(request):
    """
    Muestra una vista organizada de todos los horarios de la institución,
    agrupados por Nivel de Escolaridad y Grado, en una parrilla correcta.
    """
    institucion = request.user.institucion_asociada
    
    # Obtenemos los niveles y grados para la navegación/filtros
    niveles = NivelEscolaridad.objects.filter(institucion=institucion).order_by('pk')
    
    # Lógica de filtrado
    grado_seleccionado = None
    grado_id = request.GET.get('grado_id')
    if grado_id:
        grado_seleccionado = get_object_or_404(Grado, pk=grado_id, institucion=institucion)

    horario_grid = {}
    horas_del_dia = []
    
    if grado_seleccionado:
        bloques = BloqueHorario.objects.filter(
            curso__grado=grado_seleccionado
        ).select_related('curso__materia', 'aula').order_by('hora_inicio', 'dia_semana')

        # --- INICIO DE LA LÓGICA CLAVE ---
        # 1. Obtenemos todas las franjas horarias únicas del día
        horas_del_dia = sorted(list(set(b.hora_inicio for b in bloques)))
        
        # 2. Creamos un mapa para acceso rápido: {(hora, dia): bloque}
        bloques_map = {(b.hora_inicio, b.dia_semana): b for b in bloques}
        
        # 3. Construimos la parrilla (grid) que la plantilla usará
        for hora in horas_del_dia:
            horario_grid[hora] = {}
            for dia in range(5): # 0=Lunes, 1=Martes, ..., 4=Viernes
                horario_grid[hora][dia] = bloques_map.get((hora, dia))
        # --- FIN DE LA LÓGICA CLAVE ---

    context = {
        'titulo_pagina': "Gestión de Horarios por Grado",
        'niveles_escolaridad': niveles,
        'grados': Grado.objects.filter(institucion=institucion).order_by('orden'),
        'grado_seleccionado': grado_seleccionado,
        'horario_grid': horario_grid,
        'horas_del_dia': horas_del_dia,
        'dias_semana': range(5) # Pasamos un rango de 0 a 4
    }
    return render(request, 'gestion_academica/gestion_horarios.html', context)

@login_required
def seleccionar_curso_para_lecciones(request):
    """
    Muestra al docente una lista de sus cursos para que elija
    de cuál quiere ver el historial de lecciones.
    """
    try:
        docente = request.user.docente
        periodo_activo = PeriodoAcademico.objects.filter(institucion=docente.institucion, activo=True).first()
        if periodo_activo:
            cursos = Curso.objects.filter(docentes_asignados=docente, periodo_academico=periodo_activo)
        else:
            cursos = Curso.objects.none()
    except AttributeError:
        messages.error(request, "Acceso denegado.")
        return redirect('gestion_academica:inicio_academico')

    context = {
        'titulo_pagina': "Seleccionar Curso para Ver Lecciones",
        'cursos': cursos
    }
    return render(request, 'gestion_academica/seleccionar_curso_lecciones.html', context)


@login_required
def lista_lecciones_diarias(request, curso_pk):
    """
    Muestra todas las lecciones diarias registradas para un curso específico.
    """
    curso = get_object_or_404(Curso, pk=curso_pk)
    # Validamos que el docente que solicita tenga permiso sobre este curso
    if not curso.docentes_asignados.filter(pk=request.user.docente.pk).exists():
        messages.error(request, "No tienes permiso para ver las lecciones de este curso.")
        return redirect('gestion_academica:seleccionar_curso_para_lecciones')

    lecciones = LeccionDiaria.objects.filter(curso=curso).order_by('fecha')

    context = {
        'titulo_pagina': f"Historial de Lecciones para {curso}",
        'curso': curso,
        'lecciones': lecciones
    }
    return render(request, 'gestion_academica/lista_lecciones_diarias.html', context)    


@login_required
def detalle_leccion(request, leccion_pk):
    """
    Muestra el detalle de una lección diaria con lógica de permisos y depuración.
    """
    leccion = get_object_or_404(LeccionDiaria, pk=leccion_pk)
    curso = leccion.curso
    user = request.user

    # --- DEPURACIÓN ---
    print("======== DETALLE LECCIÓN DEBUG ========")
    print("Usuario:", user.username)
    print("Es superusuario:", user.is_superuser)
    print("Rol:", getattr(user, 'rol', 'No definido'))
    print("Curso:", curso)
    print("Curso.grado:", curso.grado)

    if hasattr(user, 'estudiante'):
        print("Usuario tiene perfil estudiante.")
        print("Estudiante.grado_actual:", user.estudiante.grado_actual)
    else:
        print("Usuario NO tiene perfil estudiante.")

    if hasattr(user, 'docente'):
        print("Usuario tiene perfil docente.")
    else:
        print("Usuario NO tiene perfil docente.")

    # --- LÓGICA DE PERMISOS ---
    tiene_permiso = False

    if user.is_superuser:
        tiene_permiso = True
        print("Permiso concedido: Superusuario.")
    elif getattr(user, 'rol', None) == 'docente':
        tiene_permiso = True
        print("Permiso concedido: Rol docente.")
    elif getattr(user, 'rol', None) == 'estudiante' and hasattr(user, 'estudiante'):
        if user.estudiante.grado_actual == curso.grado:
            tiene_permiso = True
            print("Permiso concedido: Estudiante del mismo grado.")
        else:
            print("Permiso denegado: Estudiante de otro grado.")
    else:
        print("Permiso denegado: No cumple ninguna condición.")

    if not tiene_permiso:
        messages.error(request, "No tienes permiso para acceder a esta lección.")
        print("Redirigiendo a 'inicio_academico'")
        return redirect('gestion_academica:inicio_academico')

    # --- FORMULARIO ---
    if getattr(user, 'rol', None) == 'docente' and request.method == 'POST':
        form = LeccionDiariaIaForm(request.POST, request.FILES, instance=leccion)
        if form.is_valid():
            form.save()
            messages.success(request, "Lección actualizada exitosamente.")
            return redirect('gestion_academica:lista_lecciones_diarias', curso_pk=curso.pk)
    else:
        form = LeccionDiariaIaForm(instance=leccion)

    docente = getattr(leccion, 'docente', None)

    context = {
        'titulo_pagina': "Detalle de Lección",
        'form': form,
        'leccion': leccion,
        'curso': curso,
        'docente': docente,
    }
    return render(request, 'gestion_academica/detalle_leccion_ia.html', context)


@login_required
@permission_required('gestion_academica.view_usuario', raise_exception=True) # Permiso para ver usuarios
def lista_usuarios_view(request):
    """
    Muestra una lista centralizada de todos los usuarios de la institución,
    con filtros por rol y un buscador.
    """
    institucion = request.user.institucion_asociada
    
    # Obtenemos los parámetros de filtrado y búsqueda desde la URL
    rol_filtrado = request.GET.get('rol', '')
    query = request.GET.get('q', '')

    # Empezamos con todos los usuarios de la institución
    usuarios_list = Usuario.objects.filter(institucion_asociada=institucion)

    # Aplicamos el filtro por rol si se ha seleccionado uno
    if rol_filtrado:
        usuarios_list = usuarios_list.filter(rol=rol_filtrado)

    # Aplicamos el filtro de búsqueda si se ha escrito algo
    if query:
        usuarios_list = usuarios_list.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(username__icontains=query) |
            Q(email__icontains=query)
        )

    context = {
        'titulo_pagina': "Gestión de Usuarios",
        'usuarios': usuarios_list.order_by('last_name', 'first_name'),
        'roles': Usuario.ROLES, # Pasamos la lista de roles para los botones de filtro
        'rol_actual': rol_filtrado,
        'query_actual': query,
    }
    return render(request, 'gestion_academica/lista_usuarios.html', context)


@login_required
@permission_required('gestion_academica.change_usuario', raise_exception=True)
def editar_usuario_view(request, user_pk):
    """
    Permite a un administrador editar los detalles y la contraseña de un usuario.
    VERSIÓN MEJORADA: También obtiene el perfil del estudiante si aplica.
    """
    user_to_edit = get_object_or_404(Usuario, pk=user_pk)
    
    # --- INICIO DE LA LÓGICA AÑADIDA ---
    # Buscamos el perfil de estudiante solo si el rol del usuario es 'estudiante'
    estudiante_profile = None
    if user_to_edit.rol == 'estudiante':
        try:
            # La relación OneToOne desde Usuario a Estudiante es 'estudiante'
            estudiante_profile = user_to_edit.estudiante 
        except Estudiante.DoesNotExist:
            # En caso de que haya alguna inconsistencia de datos
            estudiante_profile = None
    # --- FIN DE LA LÓGICA AÑADIDA ---
    
    if request.method == 'POST':
        # Verificamos qué formulario se envió
        if 'update_details' in request.POST:
            user_form = UserEditForm(request.POST, instance=user_to_edit)
            password_form = UserPasswordChangeForm()
            if user_form.is_valid():
                user_form.save()
                messages.success(request, f"Los datos de '{user_to_edit.get_full_name()}' han sido actualizados.")
                return redirect('gestion_academica:lista_usuarios')
        
        elif 'change_password' in request.POST:
            password_form = UserPasswordChangeForm(request.POST)
            user_form = UserEditForm(instance=user_to_edit)
            if password_form.is_valid():
                new_password = password_form.cleaned_data['new_password1']
                user_to_edit.set_password(new_password)
                user_to_edit.save()
                messages.success(request, f"La contraseña de '{user_to_edit.get_full_name()}' ha sido cambiada exitosamente.")
                return redirect('gestion_academica:lista_usuarios')
    else:
        user_form = UserEditForm(instance=user_to_edit)
        password_form = UserPasswordChangeForm()

    context = {
        'titulo_pagina': f"Editando Usuario: {user_to_edit.get_full_name()}",
        'user_to_edit': user_to_edit,
        'user_form': user_form,
        'password_form': password_form,
        'estudiante_profile': estudiante_profile, # Pasamos el perfil a la plantilla
    }
    return render(request, 'gestion_academica/editar_usuario.html', context)

@login_required
def pasar_lista_view(request, curso_pk):
    """
    Muestra la lista de estudiantes de un curso y permite al docente
    actualizar su estado de asistencia para la fecha actual.
    VERSIÓN CORREGIDA: Refuerza la seguridad para sistemas multi-institución.
    """
    institucion = request.user.institucion_asociada
    
    # --- INICIO DE LA CORRECCIÓN DE SEGURIDAD ---
    # 1. Filtramos el curso por la institución del usuario desde el principio.
    #    Esto previene que un usuario de una institución acceda a datos de otra.
    curso_qs = Curso.objects.filter(institucion=institucion)
    curso = get_object_or_404(curso_qs, pk=curso_pk)
    # --- FIN DE LA CORRECCIÓN ---

    hoy = timezone.now().date()
    
    # 2. La lógica de permisos ahora es una segunda capa de seguridad.
    if not (request.user.is_superuser or (hasattr(request.user, 'docente') and request.user.docente in curso.docentes_asignados.all())):
        messages.error(request, "No tienes permiso para pasar lista en este curso.")
        return redirect('gestion_academica:dashboard_docente')

    if request.method == 'POST':
        for key, value in request.POST.items():
            if key.startswith('estado_'):
                asistencia_id = int(key.split('_')[1])
                try:
                    # Nos aseguramos de que el registro de asistencia también pertenezca a la institución.
                    asistencia = RegistroAsistencia.objects.get(pk=asistencia_id, institucion=institucion)
                    if asistencia.estado != value:
                        asistencia.estado = value
                        asistencia.registrado_por = request.user
                        asistencia.save()
                except RegistroAsistencia.DoesNotExist:
                    continue
        messages.success(request, "La lista de asistencia ha sido actualizada.")
        return redirect('gestion_academica:pasar_lista', curso_pk=curso.pk)

    # 3. Aseguramos que la lista de estudiantes también esté filtrada por la institución.
    estudiantes = Estudiante.objects.filter(grado_actual=curso.grado, activo=True, institucion=institucion)
    
    asistencias_hoy = []
    for est in estudiantes:
        registro, created = RegistroAsistencia.objects.get_or_create(
            estudiante=est,
            curso=curso,
            fecha__date=hoy,
            defaults={
                'estado': 'AUSENTE',
                'institucion': curso.institucion,
                'registrado_por': request.user
            }
        )
        asistencias_hoy.append(registro)

    context = {
        'titulo_pagina': f"Pasar Lista: {curso}",
        'curso': curso,
        'asistencias': asistencias_hoy,
        'fecha': hoy,
        'estados': RegistroAsistencia.ESTADOS
    }
    return render(request, 'gestion_academica/pasar_lista.html', context)    


class MaterialRefuerzoView(LoginRequiredMixin, View):
    template_name = 'gestion_academica/material_refuerzo.html'

    def get(self, request, actividad_pk):
        try:
            estudiante = request.user.estudiante
        except Estudiante.DoesNotExist:
            raise PermissionDenied

        actividad = get_object_or_404(ActividadCalificable, pk=actividad_pk, institucion=estudiante.institucion)
        
        # Lógica de búsqueda (similar a la señal)
        recursos_sugeridos = ArchivoPlanAcademico.objects.filter(
            institucion=estudiante.institucion,
            temas_relacionados__icontains=actividad.titulo
        ).distinct()

        context = {
            'titulo_pagina': f"Material de Refuerzo para {actividad.titulo}",
            'actividad': actividad,
            'recursos': recursos_sugeridos
        }
        return render(request, self.template_name, context)

class HistorialEntregasView(LoginRequiredMixin, View):
    template_name = 'gestion_academica/historial_entregas.html'

    def get(self, request, *args, **kwargs):
        docente = request.user.docente
        institucion = request.user.institucion_asociada
        periodo_activo = PeriodoAcademico.objects.filter(activo=True, institucion=institucion).first()

        todas_las_entregas = EntregaDeber.objects.none()
        todos_los_intentos = IntentoCuestionario.objects.none()

        if periodo_activo:
            cursos_del_docente = Curso.objects.filter(docentes_asignados=docente, periodo_academico=periodo_activo)
            
            # 1. Obtenemos TODAS las entregas de deberes de los cursos del docente
            todas_las_entregas = EntregaDeber.objects.filter(
                deber__curso__in=cursos_del_docente,
            ).select_related('deber__curso', 'estudiante__usuario').order_by('-fecha_entrega_real')

            # 2. Obtenemos TODOS los intentos de cuestionario FINALIZADOS
            todos_los_intentos = IntentoCuestionario.objects.filter(
                cuestionario__actividad_calificable__curso__in=cursos_del_docente,
                estado='FINALIZADO',
            ).select_related(
                'estudiante__usuario', 
                'cuestionario__actividad_calificable__curso'
            ).distinct().order_by('-fecha_fin')

        context = {
            'titulo_pagina': 'Historial de Entregas y Evaluaciones',
            'entregas': todas_las_entregas,
            'intentos': todos_los_intentos
        }
        
        return render(request, self.template_name, context)

class SincronizarPermisosView(LoginRequiredMixin, View):
    template_name = 'gestion_academica/sincronizar_permisos.html'

    def get(self, request, *args, **kwargs):
        # El método GET no necesita cambios, ya funciona bien.
        if not request.user.is_staff:
            raise PermissionDenied

        grupos_gestionables = Group.objects.all().order_by('name')
        
        context = {
            'titulo_pagina': 'Sincronización de Permisos por Rol',
            'grupos': grupos_gestionables
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        if not request.user.is_staff:
            raise PermissionDenied
            
        grupo_id = request.POST.get('grupo_id')
        if not grupo_id:
            messages.error(request, "No se especificó un grupo para sincronizar.")
            return redirect('gestion_academica:sincronizar_permisos')

        try:
            grupo = Group.objects.get(pk=grupo_id)
            institucion = request.user.institucion_asociada

            # --- INICIO DE LA CORRECCIÓN CLAVE ---
            # 1. Usamos un diccionario para mapear de forma segura el nombre del grupo al rol.
            #    ¡Ajusta las claves si tus grupos se llaman diferente!
            rol_map = {
                'estudiantes': 'estudiante',
                'docentes': 'docente',
                'coordinadores': 'coordinador',
                'administradores': 'administrador',
                'familiares': 'familiar',
                # Añade más mapeos si tienes otros grupos
            }
            
            # 2. Buscamos el rol correspondiente en nuestro mapa.
            rol_a_buscar = rol_map.get(grupo.name.lower())

            if not rol_a_buscar:
                messages.error(request, f"No se pudo determinar el rol para el grupo '{grupo.name}'. Asegúrate de que esté configurado en la vista.")
                return redirect('gestion_academica:sincronizar_permisos')
            
            # 3. Buscamos a todos los usuarios con ese rol que aún no están en el grupo.
            usuarios_a_sincronizar = Usuario.objects.filter(
                institucion_asociada=institucion,
                rol=rol_a_buscar
            ).exclude(groups=grupo)

            # 4. Añadimos los usuarios encontrados al grupo.
            if usuarios_a_sincronizar.exists():
                # El método .add() puede recibir múltiples objetos a la vez.
                grupo.user_set.add(*usuarios_a_sincronizar)
                messages.success(request, f"¡Sincronización completa! Se añadieron {usuarios_a_sincronizar.count()} usuario(s) al grupo '{grupo.name}'.")
            else:
                messages.info(request, f"No se encontraron nuevos usuarios con el rol '{rol_a_buscar}' para añadir al grupo '{grupo.name}'. Ya están todos sincronizados.")
            
            # --- FIN DE LA CORRECCIÓN CLAVE ---

        except Group.DoesNotExist:
            messages.error(request, "El grupo especificado no existe.")
        except Exception as e:
            messages.error(request, f"Ocurrió un error inesperado: {e}")
        
        return redirect('gestion_academica:sincronizar_permisos')      


class MarcarNotificacionLeidaView(LoginRequiredMixin, APIView):
    
    def post(self, request, *args, **kwargs):
        notificacion_id = request.data.get('notificacion_id')

        if not notificacion_id:
            return JsonResponse({'status': 'error', 'message': 'Falta el ID de la notificación.'}, status=400)

        try:
            # Buscamos la notificación, asegurándonos de que pertenezca al usuario logueado por seguridad.
            notificacion = Notificacion.objects.get(pk=notificacion_id, destinatario=request.user)
            
            # Actualizamos el estado si no ha sido leída.
            if not notificacion.leido:
                notificacion.leido = True
                notificacion.fecha_leido = timezone.now()
                notificacion.save()

            return JsonResponse({'status': 'success'})

        except Notificacion.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Notificación no encontrada o no tienes permiso.'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


class EjecutarAnalisisComportamientoView(LoginRequiredMixin, APIView): # <-- Cambiado a APIView
    def post(self, request, *args, **kwargs):
        if not request.user.is_staff:
            raise PermissionDenied

        # Llamamos a la tarea y OBTENEMOS SU ID
        task = analizar_comportamiento_task.delay(request.user.id)
        
        # Devolvemos el ID de la tarea al frontend
        return JsonResponse({'task_id': task.id}, status=202)


# Añade esta nueva vista completa
class AnalisisStatusView(LoginRequiredMixin, APIView):
    def get(self, request, *args, **kwargs):
        task_id = request.GET.get('task_id')
        if not task_id:
            return JsonResponse({'status': 'error', 'message': 'No se proporcionó un ID de tarea.'}, status=400)

        task_result = AsyncResult(task_id)

        return JsonResponse({
            'task_id': task_id,
            'status': task_result.status,
            'result': task_result.result if task_result.ready() else None
        })


class GenerarResumenEstudianteIAView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, estudiante_pk, periodo_pk):
        estudiante = get_object_or_404(Estudiante, pk=estudiante_pk)
        periodo = get_object_or_404(PeriodoAcademico, pk=periodo_pk)

        # --- INICIO DE LA CORRECCIÓN DE PERMISOS ---
        user = request.user
        tiene_permiso = False

        # Un superusuario o un staff de la misma institución siempre tienen permiso
        if user.is_superuser or (user.is_staff and user.institucion_asociada == estudiante.institucion):
            tiene_permiso = True
        
        # Un docente tiene permiso si es el director de ese grupo en ese periodo
        elif hasattr(user, 'docente'):
            if DirectorCurso.objects.filter(
                docente=user.docente, 
                grado=estudiante.grado_actual, 
                periodo_academico=periodo
            ).exists():
                tiene_permiso = True

        if not tiene_permiso:
            return Response({'error': 'No tienes permiso para generar un resumen para este estudiante.'}, status=403)
        # --- FIN DE LA CORRECCIÓN DE PERMISOS ---
            
        estudiante = get_object_or_404(Estudiante, pk=estudiante_pk, institucion=request.user.institucion_asociada)
        periodo = get_object_or_404(PeriodoAcademico, pk=periodo_pk, institucion=request.user.institucion_asociada)
        
        # 2. Recopilación de Datos del Estudiante
        
        # a) Rendimiento Académico
        cursos = Curso.objects.filter(grado=estudiante.grado_actual, periodo_academico=periodo).select_related('materia')
        notas_texto = ""
        materias_rendimiento_bajo = []
        materias_rendimiento_alto = []
        nota_minima = estudiante.institucion.nota_minima_aprobacion or Decimal('3.0')

        for curso in cursos:
            estado = calcular_estado_academico_curso(curso, estudiante)
            nota_final = estado.get('nota_final_ponderada')
            if nota_final is not None:
                notas_texto += f"- {curso.materia.nombre_materia}: {nota_final:.2f}\n"
                if nota_final < nota_minima:
                    materias_rendimiento_bajo.append(curso.materia.nombre_materia)
                elif nota_final >= 4.5:
                    materias_rendimiento_alto.append(curso.materia.nombre_materia)

        # b) Asistencia
        inasistencias = RegistroAsistencia.objects.filter(
            estudiante=estudiante, curso__periodo_academico=periodo, estado='AUSENTE'
        ).count()

        # c) Comportamiento (Observador)
        anotaciones = AnotacionObservador.objects.filter(
            estudiante=estudiante, fecha_hora__range=(periodo.fecha_inicio, periodo.fecha_fin)
        ).order_by('-fecha_hora')[:5]
        anotaciones_texto = "\n".join([f"- {a.descripcion}" for a in anotaciones])

        # 3. Construcción del Prompt para la IA
        prompt = f"""
        Actúa como un experimentado director de grupo llamado HALU. Estás redactando una observación para el boletín de un estudiante.
        Sé profesional, constructivo y empático. Usa un lenguaje formal pero cercano.

        Aquí tienes los datos del estudiante '{estudiante.usuario.get_full_name()}' para el '{periodo.nombre}':

        Rendimiento Académico (Notas finales por materia):
        {notas_texto if notas_texto else "No hay calificaciones registradas."}

        Materias con rendimiento destacado: {', '.join(materias_rendimiento_alto) or 'Ninguna'}.
        Materias con rendimiento bajo (por debajo de {nota_minima}): {', '.join(materias_rendimiento_bajo) or 'Ninguna'}.

        Registro de Asistencia:
        - Total de inasistencias en el periodo: {inasistencias}.

        Observaciones de Comportamiento Recientes:
        {anotaciones_texto if anotaciones_texto else "No hay anotaciones recientes."}

        Basado en TODOS estos datos, redacta un párrafo de resumen (máximo 120 palabras) que describa de forma integral el desempeño del estudiante. Destaca sus fortalezas, menciona sus áreas de mejora de forma constructiva y, si aplica, conecta su comportamiento o asistencia con su rendimiento académico. Finaliza con una nota de motivación.
        """

        try:
            # Llamada a la IA y respuesta
            api_key = institucion_google_api_key(estudiante.institucion)
            if not api_key:
                return Response(
                    {'status': 'error', 'message': 'La institución no tiene configurada la API key de Google (Gemini).'},
                    status=500,
                )
            genai.configure(api_key=api_key)
            
            # --- CORRECCIÓN: De GenerativaModel a GenerativeModel ---
            model = genai.GenerativeModel('gemini-2.5-flash')
            # --- FIN DE LA CORRECCIÓN ---
            
            response = model.generate_content(prompt) # El prompt se construye con la lógica anterior
            
            return Response({'status': 'success', 'resumen': response.text})
        except Exception as e:
            return Response({'status': 'error', 'message': f"Error al contactar la IA: {e}"}, status=500)


class OptimizadorHorariosView(LoginRequiredMixin, View):
    template_name = 'gestion_academica/optimizador_horarios.html'

    def get(self, request, *args, **kwargs):
        # El método GET se mantiene exactamente igual
        if not request.user.is_staff or request.user.rol not in ['coordinador', 'administrador']:
            raise PermissionDenied("No tienes permiso para acceder a esta herramienta.")

        institucion = request.user.institucion_asociada
        periodo_activo = PeriodoAcademico.objects.filter(institucion=institucion, activo=True).first()
        
        grados = Grado.objects.filter(institucion=institucion).order_by('orden')
        
        context = {
            'titulo_pagina': 'Optimizador de Horarios con IA',
            'periodo_activo': periodo_activo,
            'grados': grados,
            'grado_seleccionado': None,
            'docentes_disponibles': [],
            'cursos_a_programar': [],
            'aulas_disponibles': []
        }

        grado_id = request.GET.get('grado_id')
        if grado_id:
            grado_seleccionado = get_object_or_404(Grado, pk=grado_id, institucion=institucion)
            context['grado_seleccionado'] = grado_seleccionado
            
            if periodo_activo:
                cursos_del_grado = Curso.objects.filter(periodo_academico=periodo_activo, grado=grado_seleccionado)
                context['cursos_a_programar'] = cursos_del_grado.select_related('materia', 'grado').order_by('materia__nombre_materia')
                docentes_ids = cursos_del_grado.values_list('docentes_asignados', flat=True)
                context['docentes_disponibles'] = Docente.objects.filter(pk__in=set(docentes_ids)).prefetch_related('disponibilidades').order_by('usuario__last_name')
                context['aulas_disponibles'] = Aula.objects.filter(institucion=institucion).order_by('nombre')

        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        periodo_activo = PeriodoAcademico.objects.filter(institucion=request.user.institucion_asociada, activo=True).first()
        
        # --- INICIO DE LA CORRECCIÓN ---
        # Leemos el cuerpo de la petición y lo decodificamos como JSON
        try:
            data = json.loads(request.body)
            grado_pk = data.get('grado_pk')
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Cuerpo de la petición inválido.'}, status=400)
        # --- FIN DE LA CORRECCIÓN ---

        if not periodo_activo or not grado_pk:
            return JsonResponse({'status': 'error', 'message': 'Falta el periodo activo o el grado.'}, status=400)

        task = generar_propuesta_horario_task.delay(
            periodo_pk=periodo_activo.pk, 
            institucion_id=request.user.institucion_asociada.pk,
            grado_pk=grado_pk
        )
        
        return JsonResponse({'task_id': task.id}, status=202)
               

class GuardarHorarioView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        if not request.user.is_staff:
            raise PermissionDenied

        try:
            horario_data = request.data.get('horario')
            periodo_pk = request.data.get('periodo_pk')
            # --- NUEVO: Obtenemos el grado para un borrado más seguro ---
            grado_pk = request.data.get('grado_pk')

            if not all([horario_data, periodo_pk, grado_pk]):
                return Response({'status': 'error', 'message': 'Faltan datos (horario, periodo o grado).'}, status=400)

            with transaction.atomic():
                # Borramos solo los bloques del grado y periodo específicos
                cursos_del_grado = Curso.objects.filter(periodo_academico_id=periodo_pk, grado_id=grado_pk)
                BloqueHorario.objects.filter(curso__in=cursos_del_grado).delete()

                # --- INICIO DE LA CORRECCIÓN CLAVE ---
                for evento in horario_data:
                    # Creamos los nuevos bloques SIN el campo 'docente'
                    BloqueHorario.objects.create(
                        curso=get_object_or_404(Curso, pk=evento['curso_id']),
                        aula=get_object_or_404(Aula, pk=evento['aula_id']),
                        dia_semana=int(evento['dia_semana']),
                        hora_inicio=evento['hora_inicio'],
                        hora_fin=evento['hora_fin'],
                        institucion=request.user.institucion_asociada
                        # El campo 'docente' se elimina de aquí
                    )
                # --- FIN DE LA CORRECCIÓN CLAVE ---
            
            return Response({'status': 'success', 'message': '¡Horario guardado y publicado exitosamente!'})

        except Exception as e:
            return Response({'status': 'error', 'message': str(e)}, status=500)     


@login_required
def cursos_por_institucion_api(request, institucion_id):
    # Superusuario puede consultar cualquier institución.
    # El resto (docentes, coordinadores, admin_institucion) solo pueden
    # consultar su propia institución — nunca la de otra.
    if not request.user.is_superuser:
        institucion_usuario = getattr(request.user, 'institucion_asociada', None)
        if not institucion_usuario or institucion_usuario.pk != int(institucion_id):
            return JsonResponse({'error': 'Acceso denegado'}, status=403)

    cursos = Curso.objects.filter(
        institucion_id=institucion_id
    ).select_related('grado', 'materia').order_by('grado__orden', 'materia__nombre_materia')

    data = [{'id': curso.pk, 'text': str(curso)} for curso in cursos]
    return JsonResponse(data, safe=False)

class GenerarCorreoAcudienteIAView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, estudiante_pk, periodo_pk):
        estudiante = get_object_or_404(Estudiante, pk=estudiante_pk)
        periodo = get_object_or_404(PeriodoAcademico, pk=periodo_pk)
        
        # --- INICIO: Lógica de permisos y recopilación de datos (completa) ---
        user = request.user
        tiene_permiso = False

        if user.is_superuser or (user.is_staff and user.institucion_asociada == estudiante.institucion):
            tiene_permiso = True
        elif hasattr(user, 'docente'):
            if DirectorCurso.objects.filter(docente=user.docente, grado=estudiante.grado_actual, periodo_academico=periodo).exists():
                tiene_permiso = True

        if not tiene_permiso:
            return Response({'error': 'No tienes permiso para generar una comunicación para este estudiante.'}, status=403)

        cursos = Curso.objects.filter(grado=estudiante.grado_actual, periodo_academico=periodo).select_related('materia')
        notas_texto = ""
        materias_rendimiento_bajo = []
        materias_rendimiento_alto = []
        nota_minima = estudiante.institucion.nota_minima_aprobacion or Decimal('3.0')

        for curso in cursos:
            estado = calcular_estado_academico_curso(curso, estudiante)
            nota_final = estado.get('nota_final_ponderada')
            if nota_final is not None:
                notas_texto += f"- {curso.materia.nombre_materia}: {nota_final:.2f}\n"
                if nota_final < nota_minima:
                    materias_rendimiento_bajo.append(curso.materia.nombre_materia)
                elif nota_final >= 4.5:
                    materias_rendimiento_alto.append(curso.materia.nombre_materia)

        inasistencias = RegistroAsistencia.objects.filter(estudiante=estudiante, curso__periodo_academico=periodo, estado='AUSENTE').count()
        anotaciones = AnotacionObservador.objects.filter(estudiante=estudiante, fecha_hora__range=(periodo.fecha_inicio, periodo.fecha_fin)).order_by('-fecha_hora')[:5]
        anotaciones_texto = "\n".join([f"- {a.descripcion}" for a in anotaciones])
        # --- FIN: Lógica de permisos y recopilación de datos ---

        # --- Prompt específico para el correo electrónico ---
        prompt = f"""
        Actúa como un docente profesional y empático llamado HALU. Estás escribiendo un borrador de correo electrónico para el acudiente del estudiante '{estudiante.usuario.get_full_name()}'.
        El correo debe tener un tono formal pero cercano y estar estructurado con un saludo, un cuerpo y una despedida.

        Usa los siguientes datos para redactar el cuerpo del correo:
        - Rendimiento Académico: {notas_texto if notas_texto else "Aún no hay calificaciones consolidadas."}
        - Materias Destacadas: {', '.join(materias_rendimiento_alto) or 'Ninguna'}.
        - Materias a Reforzar: {', '.join(materias_rendimiento_bajo) or 'Ninguna'}.
        - Inasistencias: {inasistencias}.
        - Comportamiento (notas recientes): {anotaciones_texto if anotaciones_texto else "Sin novedades."}

        TAREA:
        Redacta el correo. Si el rendimiento general es bueno, enfócate en felicitar y motivar. Si hay áreas de mejora (notas bajas o inasistencias), menciónalas constructivamente y sugiere agendar una reunión para hablar sobre estrategias de apoyo. Finaliza con un saludo cordial.
        """

        try:
            api_key = institucion_google_api_key(estudiante.institucion)
            if not api_key:
                return Response(
                    {'status': 'error', 'message': 'La institución no tiene configurada la API key de Google (Gemini).'},
                    status=500,
                )
            genai.configure(api_key=api_key)
            model = genai.GenerativeModel('gemini-2.5-flash')
            response = model.generate_content(prompt)
            return Response({'status': 'success', 'correo': response.text})
        except Exception as e:
            logger.exception("GenerarCorreoIA error: %s", e)
            return Response({'status': 'error', 'message': f"Error al contactar la IA: {e}"}, status=500)