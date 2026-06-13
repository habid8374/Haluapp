"""Vistas del módulo Simulacros Saber."""
import json
import logging

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Avg, Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from .models import (
    BancoPregunta, IntentoSimulacro, OpcionPregunta,
    PreguntaSimulacro, RespuestaSimulacro, Simulacro,
)

logger = logging.getLogger(__name__)


def _get_institucion(request):
    return getattr(request.user, 'institucion_asociada', None)


def _es_docente_o_coordinador(user):
    rol = getattr(user, 'rol', '') or ''
    return rol in ('docente', 'coordinador', 'admin_institucion') or user.is_superuser


def _es_estudiante(user):
    return (getattr(user, 'rol', '') or '') == 'estudiante'


# ──────────────────────────────────────────────────────────────────────────────
# BANCO DE PREGUNTAS — vistas del docente/coordinador
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def banco_preguntas(request):
    if not _es_docente_o_coordinador(request.user):
        messages.error(request, "Acceso restringido.")
        return redirect('gestion_academica:inicio_academico')

    institucion = _get_institucion(request)

    # Preguntas públicas + privadas de la institución
    qs = BancoPregunta.objects.prefetch_related('opciones')
    if not request.user.is_superuser:
        qs = qs.filter(Q(es_publica=True) | Q(institucion=institucion))

    # Filtros
    grado   = request.GET.get('grado', '')
    area    = request.GET.get('area', '')
    q       = (request.GET.get('q') or '').strip()
    solo_mias = request.GET.get('solo_mias', '')

    if grado:
        qs = qs.filter(grado_nivel=grado)
    if area:
        qs = qs.filter(area=area)
    if q:
        qs = qs.filter(enunciado__icontains=q)
    if solo_mias:
        qs = qs.filter(institucion=institucion, es_publica=False)

    paginator = Paginator(qs.order_by('grado_nivel', 'area', 'pk'), 30)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'simulacros/banco_preguntas.html', {
        'preguntas': page,
        'page_obj': page,
        'grados': BancoPregunta.GradoNivel.choices,
        'areas': BancoPregunta.Area.choices,
        'filtros': {'grado': grado, 'area': area, 'q': q, 'solo_mias': solo_mias},
        'total': qs.count(),
        'titulo_pagina': 'Banco de Preguntas Saber',
    })


@login_required
def crear_pregunta(request):
    if not _es_docente_o_coordinador(request.user):
        return redirect('gestion_academica:inicio_academico')

    if request.method == 'POST':
        return _guardar_pregunta(request, None)

    return render(request, 'simulacros/form_pregunta.html', {
        'grados': BancoPregunta.GradoNivel.choices,
        'areas': BancoPregunta.Area.choices,
        'dificultades': BancoPregunta.Dificultad.choices,
        'titulo_pagina': 'Nueva Pregunta',
        'accion': 'Crear',
    })


@login_required
def editar_pregunta(request, pk):
    if not _es_docente_o_coordinador(request.user):
        return redirect('gestion_academica:inicio_academico')

    institucion = _get_institucion(request)
    pregunta = get_object_or_404(
        BancoPregunta,
        pk=pk,
        **({} if request.user.is_superuser else {'institucion': institucion}),
    )

    if request.method == 'POST':
        return _guardar_pregunta(request, pregunta)

    opciones = {o.letra: o.texto for o in pregunta.opciones.all()}
    correcta = pregunta.opciones.filter(es_correcta=True).values_list('letra', flat=True).first()

    return render(request, 'simulacros/form_pregunta.html', {
        'pregunta': pregunta,
        'opciones': opciones,
        'correcta': correcta,
        'grados': BancoPregunta.GradoNivel.choices,
        'areas': BancoPregunta.Area.choices,
        'dificultades': BancoPregunta.Dificultad.choices,
        'titulo_pagina': 'Editar Pregunta',
        'accion': 'Guardar',
    })


def _guardar_pregunta(request, pregunta_existente):
    """Crea o actualiza una BancoPregunta con sus 4 opciones."""
    institucion = _get_institucion(request)
    p = request.POST

    enunciado = (p.get('enunciado') or '').strip()
    if not enunciado:
        messages.error(request, "El enunciado no puede estar vacío.")
        return redirect(request.path)

    correcta = p.get('correcta', 'A')
    opciones_texto = {l: (p.get(f'opcion_{l}') or '').strip() for l in 'ABCD'}
    if not all(opciones_texto.values()):
        messages.error(request, "Debes completar las 4 opciones.")
        return redirect(request.path)

    if pregunta_existente is None:
        pregunta_existente = BancoPregunta(
            institucion=institucion,
            es_publica=False,
            creado_por=request.user,
        )

    # A03 — validar que imagen_url solo permita http/https
    from urllib.parse import urlparse
    imagen_url = (p.get('imagen_url') or '').strip()
    if imagen_url:
        parsed_img = urlparse(imagen_url)
        if parsed_img.scheme not in ('http', 'https'):
            messages.error(request, "La URL de imagen debe comenzar con http:// o https://")
            return redirect(request.path)
    else:
        imagen_url = ''

    pregunta_existente.enunciado        = enunciado
    pregunta_existente.imagen_url       = imagen_url
    pregunta_existente.grado_nivel      = p.get('grado_nivel', BancoPregunta.GradoNivel.GRADO_11)
    pregunta_existente.area             = p.get('area', BancoPregunta.Area.MATEMATICAS)
    pregunta_existente.competencia      = (p.get('competencia') or '').strip()
    pregunta_existente.componente       = (p.get('componente') or '').strip()
    pregunta_existente.nivel_dificultad = p.get('nivel_dificultad', BancoPregunta.Dificultad.MEDIO)
    pregunta_existente.explicacion      = (p.get('explicacion') or '').strip()
    pregunta_existente.fuente           = (p.get('fuente') or '').strip()
    pregunta_existente.save()

    # Recrear opciones
    pregunta_existente.opciones.all().delete()
    for letra in 'ABCD':
        OpcionPregunta.objects.create(
            pregunta=pregunta_existente,
            letra=letra,
            texto=opciones_texto[letra],
            es_correcta=(letra == correcta),
        )

    messages.success(request, "Pregunta guardada correctamente.")
    return redirect('simulacros:banco_preguntas')


@login_required
@require_POST
def eliminar_pregunta(request, pk):
    if not _es_docente_o_coordinador(request.user):
        return redirect('gestion_academica:inicio_academico')
    institucion = _get_institucion(request)
    pregunta = get_object_or_404(
        BancoPregunta, pk=pk,
        **({} if request.user.is_superuser else {'institucion': institucion, 'es_publica': False}),
    )
    pregunta.delete()
    messages.success(request, "Pregunta eliminada.")
    return redirect('simulacros:banco_preguntas')


# ──────────────────────────────────────────────────────────────────────────────
# IMPORTAR DESDE EXCEL
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def importar_preguntas(request):
    if not _es_docente_o_coordinador(request.user):
        return redirect('gestion_academica:inicio_academico')

    if request.method == 'GET':
        return render(request, 'simulacros/importar_preguntas.html', {
            'grados': BancoPregunta.GradoNivel.choices,
            'areas': BancoPregunta.Area.choices,
            'titulo_pagina': 'Importar Preguntas desde Excel',
        })

    archivo = request.FILES.get('archivo')
    if not archivo:
        messages.error(request, "Selecciona un archivo Excel (.xlsx).")
        return redirect('simulacros:importar_preguntas')

    # A08 — validación de tipo, extensión y tamaño antes de procesar
    ALLOWED_CONTENT_TYPES = {
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/octet-stream',  # algunos navegadores envían este tipo para .xlsx
    }
    MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB
    if not archivo.name.lower().endswith('.xlsx'):
        messages.error(request, "Solo se aceptan archivos Excel con extensión .xlsx")
        return redirect('simulacros:importar_preguntas')
    if archivo.content_type not in ALLOWED_CONTENT_TYPES:
        messages.error(request, "Tipo de archivo no permitido. Sube un archivo .xlsx válido.")
        return redirect('simulacros:importar_preguntas')
    if archivo.size > MAX_UPLOAD_BYTES:
        messages.error(request, "El archivo supera el límite de 5 MB.")
        return redirect('simulacros:importar_preguntas')

    try:
        import openpyxl
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        institucion = _get_institucion(request)
        creadas = 0
        errores = []

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            try:
                enunciado, grado, area, comp, comp2, dif, op_a, op_b, op_c, op_d, correcta, expl, fuente = (
                    (row[j] or '') for j in range(13)
                )
                if not enunciado or not op_a or str(correcta).upper() not in 'ABCD':
                    errores.append(f"Fila {i}: datos incompletos o respuesta inválida.")
                    continue

                pregunta = BancoPregunta.objects.create(
                    institucion=institucion,
                    es_publica=False,
                    enunciado=str(enunciado).strip(),
                    grado_nivel=str(grado).strip() or BancoPregunta.GradoNivel.GRADO_11,
                    area=str(area).strip() or BancoPregunta.Area.MATEMATICAS,
                    competencia=str(comp).strip(),
                    componente=str(comp2).strip(),
                    nivel_dificultad=str(dif).strip() or BancoPregunta.Dificultad.MEDIO,
                    explicacion=str(expl).strip(),
                    fuente=str(fuente).strip(),
                    creado_por=request.user,
                )
                for letra, texto in zip('ABCD', [op_a, op_b, op_c, op_d]):
                    OpcionPregunta.objects.create(
                        pregunta=pregunta,
                        letra=letra,
                        texto=str(texto).strip(),
                        es_correcta=(letra == str(correcta).strip().upper()),
                    )
                creadas += 1
            except Exception as exc:
                logger.warning("importar_preguntas fila %s: %s", i, exc)
                errores.append(f"Fila {i}: datos incorrectos o incompletos.")

        if creadas:
            messages.success(request, f"✅ {creadas} pregunta(s) importada(s) correctamente.")
        if errores:
            for e in errores[:5]:
                messages.warning(request, e)

    except Exception as exc:
        logger.error("importar_preguntas error leyendo archivo: %s", exc, exc_info=True)
        messages.error(request, "El archivo no pudo procesarse. Verifica que sea un .xlsx válido.")

    return redirect('simulacros:banco_preguntas')


@login_required
def descargar_plantilla_excel(request):
    """Genera y descarga la plantilla Excel para importar preguntas."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Preguntas"

    headers = [
        'enunciado', 'grado_nivel', 'area', 'competencia', 'componente',
        'nivel_dificultad', 'opcion_A', 'opcion_B', 'opcion_C', 'opcion_D',
        'correcta (A/B/C/D)', 'explicacion', 'fuente',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='4f46e5')
        ws.column_dimensions[cell.column_letter].width = 20

    # Fila de ejemplo
    ws.append([
        '¿Cuál es el resultado de 2x + 3 = 11?',
        'GRADO_11', 'MATEMATICAS', 'Razonamiento y argumentación', 'Numérico-variacional',
        'BASICO', 'x = 4', 'x = 3', 'x = 5', 'x = 2', 'A',
        'Despejando: 2x = 8, x = 4', 'ICFES Saber 11 2019-1',
    ])

    # Hoja de referencia de valores válidos
    ws2 = wb.create_sheet("Valores válidos")
    ws2.append(['grado_nivel', 'area', 'nivel_dificultad'])
    for g, a, d in zip(
        [c[0] for c in BancoPregunta.GradoNivel.choices],
        [c[0] for c in BancoPregunta.Area.choices],
        [c[0] for c in BancoPregunta.Dificultad.choices],
    ):
        ws2.append([g, a, d])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_preguntas_saber.xlsx"'
    wb.save(response)
    return response


# ──────────────────────────────────────────────────────────────────────────────
# GENERADOR IA
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@require_POST
def generar_preguntas_ia(request):
    if not _es_docente_o_coordinador(request.user):
        return JsonResponse({'ok': False, 'error': 'Sin permiso.'}, status=403)

    grado    = request.POST.get('grado', 'GRADO_11')
    area     = request.POST.get('area', 'MATEMATICAS')
    cantidad = min(int(request.POST.get('cantidad', 5)), 10)
    dificultad = request.POST.get('dificultad', 'MEDIO')

    grado_label = dict(BancoPregunta.GradoNivel.choices).get(grado, grado)
    area_label  = dict(BancoPregunta.Area.choices).get(area, area)
    dif_label   = dict(BancoPregunta.Dificultad.choices).get(dificultad, dificultad)

    prompt = f"""Eres un experto en las Pruebas Saber del ICFES de Colombia.
Genera exactamente {cantidad} preguntas de opción múltiple con única respuesta para la prueba {grado_label}, área {area_label}, nivel de dificultad {dif_label}.

Formato estricto de respuesta — JSON puro, sin markdown, sin texto antes ni después:
[
  {{
    "enunciado": "texto completo de la pregunta",
    "competencia": "nombre de la competencia ICFES",
    "componente": "nombre del componente",
    "opciones": {{"A": "texto A", "B": "texto B", "C": "texto C", "D": "texto D"}},
    "correcta": "A",
    "explicacion": "explicación breve de por qué es correcta"
  }}
]

Reglas:
- Las preguntas deben seguir exactamente el formato y el estilo de las Pruebas Saber del ICFES.
- Las opciones incorrectas (distractores) deben ser plausibles.
- La respuesta correcta debe ser inequívoca.
- No uses LaTeX ni HTML, solo texto plano con símbolos matemáticos cuando sea necesario.
"""

    try:
        import google.generativeai as genai
        from django.conf import settings as dj_settings
        genai.configure(api_key=dj_settings.GEMINI_API_KEY)
        model = genai.GenerativeModel('gemini-1.5-flash')
        resp = model.generate_content(prompt)
        raw = resp.text.strip()

        # Limpiar posible markdown
        if raw.startswith('```'):
            raw = raw.split('```')[1]
            if raw.startswith('json'):
                raw = raw[4:]
        preguntas_data = json.loads(raw)

        # M5 — validar estructura básica antes de enviar al cliente
        if not isinstance(preguntas_data, list) or not preguntas_data:
            raise ValueError("La IA no devolvió una lista de preguntas válida.")
        for p in preguntas_data:
            if not isinstance(p, dict) or not p.get('enunciado') or not isinstance(p.get('opciones'), dict):
                raise ValueError("Estructura de pregunta inválida en la respuesta de la IA.")

        return JsonResponse({'ok': True, 'preguntas': preguntas_data, 'grado': grado, 'area': area, 'dificultad': dificultad})
    except Exception as exc:
        logger.error("generar_preguntas_ia error: %s", exc, exc_info=True)
        return JsonResponse({'ok': False, 'error': 'Error al generar preguntas. Intenta de nuevo.'}, status=500)


@login_required
@require_POST
def guardar_preguntas_ia(request):
    """Guarda en el banco las preguntas generadas por IA tras revisión del docente."""
    if not _es_docente_o_coordinador(request.user):
        return JsonResponse({'ok': False, 'error': 'Sin permiso.'}, status=403)

    institucion = _get_institucion(request)
    try:
        data = json.loads(request.body)
        preguntas_raw = data.get('preguntas', [])
        if not isinstance(preguntas_raw, list) or not preguntas_raw:
            return JsonResponse({'ok': False, 'error': 'Datos inválidos.'}, status=400)

        # A02/A08 — validar valores de grado/area/dificultad contra choices permitidos
        grado_validos = {v for v, _ in BancoPregunta.GradoNivel.choices}
        area_validas  = {v for v, _ in BancoPregunta.Area.choices}
        dif_validas   = {v for v, _ in BancoPregunta.Dificultad.choices}

        grado      = data.get('grado', 'GRADO_11')
        area       = data.get('area', 'MATEMATICAS')
        dificultad = data.get('dificultad', 'MEDIO')

        if grado not in grado_validos or area not in area_validas or dificultad not in dif_validas:
            return JsonResponse({'ok': False, 'error': 'Parámetros inválidos.'}, status=400)

        creadas = 0
        for p in preguntas_raw[:10]:  # máximo 10 preguntas por llamada
            enunciado = str(p.get('enunciado', '')).strip()[:5000]
            if not enunciado:
                continue
            opciones = p.get('opciones', {})
            if not isinstance(opciones, dict):
                continue
            correcta = str(p.get('correcta', 'A')).strip().upper()
            if correcta not in 'ABCD':
                correcta = 'A'

            pregunta = BancoPregunta.objects.create(
                institucion=institucion,
                es_publica=False,
                enunciado=enunciado,
                grado_nivel=grado,
                area=area,
                competencia=str(p.get('competencia', ''))[:120],
                componente=str(p.get('componente', ''))[:120],
                nivel_dificultad=dificultad,
                explicacion=str(p.get('explicacion', ''))[:2000],
                fuente='Generada con IA (Gemini)',
                creado_por=request.user,
            )
            for letra in 'ABCD':
                texto_opcion = str(opciones.get(letra, '')).strip()[:2000]
                OpcionPregunta.objects.create(
                    pregunta=pregunta,
                    letra=letra,
                    texto=texto_opcion or f'Opción {letra}',
                    es_correcta=(letra == correcta),
                )
            creadas += 1

        return JsonResponse({'ok': True, 'creadas': creadas})
    except Exception as exc:
        logger.error("guardar_preguntas_ia error: %s", exc, exc_info=True)
        return JsonResponse({'ok': False, 'error': 'Error al guardar preguntas.'}, status=500)


# ──────────────────────────────────────────────────────────────────────────────
# SIMULACROS — gestión del docente
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def lista_simulacros(request):
    if not _es_docente_o_coordinador(request.user):
        return redirect('gestion_academica:inicio_academico')

    institucion = _get_institucion(request)
    simulacros = Simulacro.objects.filter(institucion=institucion).annotate(
        num_preguntas=Count('preguntas'),
        num_intentos=Count('intentos'),
    ).order_by('-fecha_creacion')

    return render(request, 'simulacros/lista_simulacros.html', {
        'simulacros': simulacros,
        'titulo_pagina': 'Simulacros Saber',
    })


@login_required
def crear_simulacro(request):
    if not _es_docente_o_coordinador(request.user):
        return redirect('gestion_academica:inicio_academico')

    if request.method == 'POST':
        return _guardar_simulacro(request, None)

    institucion = _get_institucion(request)
    preguntas_banco = BancoPregunta.objects.filter(
        Q(es_publica=True) | Q(institucion=institucion)
    ).prefetch_related('opciones').order_by('grado_nivel', 'area', 'pk')

    return render(request, 'simulacros/form_simulacro.html', {
        'grados': BancoPregunta.GradoNivel.choices,
        'preguntas_banco': preguntas_banco,
        'areas': BancoPregunta.Area.choices,
        'titulo_pagina': 'Nuevo Simulacro',
        'accion': 'Crear',
    })


@login_required
def editar_simulacro(request, pk):
    if not _es_docente_o_coordinador(request.user):
        return redirect('gestion_academica:inicio_academico')

    institucion = _get_institucion(request)
    simulacro = get_object_or_404(Simulacro, pk=pk, institucion=institucion)

    if request.method == 'POST':
        return _guardar_simulacro(request, simulacro)

    preguntas_banco = BancoPregunta.objects.filter(
        Q(es_publica=True) | Q(institucion=institucion)
    ).prefetch_related('opciones').order_by('grado_nivel', 'area', 'pk')
    seleccionadas = set(simulacro.preguntas.values_list('pk', flat=True))

    return render(request, 'simulacros/form_simulacro.html', {
        'simulacro': simulacro,
        'grados': BancoPregunta.GradoNivel.choices,
        'preguntas_banco': preguntas_banco,
        'seleccionadas': seleccionadas,
        'areas': BancoPregunta.Area.choices,
        'titulo_pagina': 'Editar Simulacro',
        'accion': 'Guardar',
    })


def _guardar_simulacro(request, simulacro_existente):
    p = request.POST
    institucion = _get_institucion(request)

    titulo = (p.get('titulo') or '').strip()
    if not titulo:
        messages.error(request, "El título es obligatorio.")
        return redirect(request.path)

    from django.utils.dateparse import parse_datetime
    fecha_inicio = parse_datetime(p.get('fecha_inicio', ''))
    fecha_cierre = parse_datetime(p.get('fecha_cierre', ''))
    if not fecha_inicio or not fecha_cierre:
        messages.error(request, "Las fechas son obligatorias.")
        return redirect(request.path)

    if simulacro_existente is None:
        simulacro_existente = Simulacro(institucion=institucion, docente=request.user)

    simulacro_existente.titulo             = titulo
    simulacro_existente.descripcion        = (p.get('descripcion') or '').strip()
    simulacro_existente.grado_nivel        = p.get('grado_nivel', BancoPregunta.GradoNivel.GRADO_11)
    simulacro_existente.tiempo_minutos     = int(p.get('tiempo_minutos', 60))
    simulacro_existente.fecha_inicio       = fecha_inicio
    simulacro_existente.fecha_cierre       = fecha_cierre
    simulacro_existente.mostrar_respuestas = 'mostrar_respuestas' in p
    simulacro_existente.save()

    # Actualizar preguntas seleccionadas
    ids_seleccionados = list(map(int, p.getlist('preguntas')))
    PreguntaSimulacro.objects.filter(simulacro=simulacro_existente).delete()
    for orden, pid in enumerate(ids_seleccionados):
        pregunta = BancoPregunta.objects.filter(pk=pid).first()
        if pregunta:
            PreguntaSimulacro.objects.create(
                simulacro=simulacro_existente, pregunta=pregunta, orden=orden
            )

    messages.success(request, f"Simulacro '{titulo}' guardado con {len(ids_seleccionados)} pregunta(s).")
    return redirect('simulacros:lista_simulacros')


@login_required
@require_POST
def cambiar_estado_simulacro(request, pk):
    if not _es_docente_o_coordinador(request.user):
        return redirect('gestion_academica:inicio_academico')
    institucion = _get_institucion(request)
    simulacro = get_object_or_404(Simulacro, pk=pk, institucion=institucion)
    nuevo_estado = request.POST.get('estado', Simulacro.Estado.PUBLICADO)
    if nuevo_estado in dict(Simulacro.Estado.choices):
        simulacro.estado = nuevo_estado
        simulacro.save(update_fields=['estado'])
        messages.success(request, f"Simulacro '{simulacro.titulo}' ahora está {simulacro.get_estado_display()}.")
    return redirect('simulacros:lista_simulacros')


@login_required
@require_POST
def eliminar_simulacro(request, pk):
    if not _es_docente_o_coordinador(request.user):
        return redirect('gestion_academica:inicio_academico')
    institucion = _get_institucion(request)
    simulacro = get_object_or_404(Simulacro, pk=pk, institucion=institucion)
    simulacro.delete()
    messages.success(request, "Simulacro eliminado.")
    return redirect('simulacros:lista_simulacros')


@login_required
def resultados_simulacro(request, pk):
    if not _es_docente_o_coordinador(request.user):
        return redirect('gestion_academica:inicio_academico')

    institucion = _get_institucion(request)
    simulacro = get_object_or_404(Simulacro, pk=pk, institucion=institucion)

    intentos = (
        IntentoSimulacro.objects
        .filter(simulacro=simulacro)
        .select_related('estudiante__usuario')
        .order_by('-puntaje')
    )

    # Estadísticas por área/competencia
    stats_preguntas = []
    for pq in simulacro.preguntasimulacro_set.select_related('pregunta').order_by('orden'):
        total_resp = RespuestaSimulacro.objects.filter(pregunta=pq.pregunta, intento__simulacro=simulacro).count()
        correctas  = RespuestaSimulacro.objects.filter(pregunta=pq.pregunta, intento__simulacro=simulacro, es_correcta=True).count()
        stats_preguntas.append({
            'pregunta': pq.pregunta,
            'orden': pq.orden + 1,
            'total': total_resp,
            'correctas': correctas,
            'pct': round(correctas / total_resp * 100, 1) if total_resp else 0,
        })

    puntaje_promedio = intentos.filter(completado=True).aggregate(avg=Avg('puntaje'))['avg']

    return render(request, 'simulacros/resultados_simulacro.html', {
        'simulacro': simulacro,
        'intentos': intentos,
        'stats_preguntas': stats_preguntas,
        'puntaje_promedio': puntaje_promedio,
        'titulo_pagina': f'Resultados — {simulacro.titulo}',
    })


# ──────────────────────────────────────────────────────────────────────────────
# VISTAS DEL ESTUDIANTE
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def simulacros_estudiante(request):
    if not _es_estudiante(request.user):
        messages.error(request, "Esta sección es solo para estudiantes.")
        return redirect('gestion_academica:inicio_academico')

    institucion = _get_institucion(request)
    estudiante  = getattr(request.user, 'estudiante', None)
    ahora       = timezone.now()

    simulacros_qs = Simulacro.objects.filter(
        institucion=institucion,
        estado=Simulacro.Estado.PUBLICADO,
    ).annotate(num_preguntas=Count('preguntas')).order_by('-fecha_inicio')

    # Enriquecer con estado del intento del estudiante
    simulacros_info = []
    for s in simulacros_qs:
        intento = None
        if estudiante:
            intento = IntentoSimulacro.objects.filter(simulacro=s, estudiante=estudiante).first()
        simulacros_info.append({
            'simulacro': s,
            'intento': intento,
            'disponible': s.fecha_inicio <= ahora <= s.fecha_cierre,
            'cerrado': ahora > s.fecha_cierre,
        })

    return render(request, 'simulacros/simulacros_estudiante.html', {
        'simulacros_info': simulacros_info,
        'titulo_pagina': 'Simulacros Saber',
    })


@login_required
def resolver_simulacro(request, pk):
    if not _es_estudiante(request.user):
        return redirect('gestion_academica:inicio_academico')

    institucion = _get_institucion(request)
    estudiante  = getattr(request.user, 'estudiante', None)
    if not estudiante:
        messages.error(request, "No tienes perfil de estudiante.")
        return redirect('gestion_academica:inicio_academico')

    simulacro = get_object_or_404(
        Simulacro,
        pk=pk,
        institucion=institucion,
        estado=Simulacro.Estado.PUBLICADO,
    )

    if not simulacro.esta_disponible():
        messages.warning(request, "Este simulacro no está disponible en este momento.")
        return redirect('simulacros:simulacros_estudiante')

    # Obtener o crear intento
    intento, creado = IntentoSimulacro.objects.get_or_create(
        simulacro=simulacro,
        estudiante=estudiante,
        defaults={'institucion': institucion},
    )

    if intento.completado:
        return redirect('simulacros:resultado_intento', pk=intento.pk)

    if request.method == 'POST':
        # Guardar respuestas y calcular puntaje
        preguntas_qs = simulacro.preguntasimulacro_set.select_related('pregunta').order_by('orden')
        for pq in preguntas_qs:
            opcion_id = request.POST.get(f'pregunta_{pq.pregunta.pk}')
            opcion = OpcionPregunta.objects.filter(pk=opcion_id, pregunta=pq.pregunta).first() if opcion_id else None
            RespuestaSimulacro.objects.update_or_create(
                intento=intento,
                pregunta=pq.pregunta,
                defaults={
                    'opcion_elegida': opcion,
                    'es_correcta': opcion.es_correcta if opcion else False,
                },
            )
        intento.calcular_y_guardar_puntaje()
        return redirect('simulacros:resultado_intento', pk=intento.pk)

    preguntas_qs = simulacro.preguntasimulacro_set.select_related('pregunta').prefetch_related('pregunta__opciones').order_by('orden')

    return render(request, 'simulacros/resolver_simulacro.html', {
        'simulacro': simulacro,
        'intento': intento,
        'preguntas': preguntas_qs,
        'tiempo_segundos': simulacro.tiempo_minutos * 60,
        'titulo_pagina': simulacro.titulo,
    })


@login_required
def resultado_intento(request, pk):
    estudiante = getattr(request.user, 'estudiante', None)
    if not estudiante and not _es_docente_o_coordinador(request.user):
        return redirect('gestion_academica:inicio_academico')

    institucion = _get_institucion(request)
    intento = get_object_or_404(IntentoSimulacro, pk=pk, institucion=institucion)

    # Verificar acceso: el estudiante solo ve su propio intento
    if _es_estudiante(request.user) and (not estudiante or intento.estudiante != estudiante):
        messages.error(request, "No tienes acceso a este resultado.")
        return redirect('simulacros:simulacros_estudiante')

    respuestas = intento.respuestas.select_related(
        'pregunta', 'opcion_elegida', 'pregunta__opcion_correcta'
    ).prefetch_related('pregunta__opciones').order_by('pregunta__pk')

    return render(request, 'simulacros/resultado_intento.html', {
        'intento': intento,
        'simulacro': intento.simulacro,
        'respuestas': respuestas,
        'titulo_pagina': f'Resultado — {intento.simulacro.titulo}',
    })
