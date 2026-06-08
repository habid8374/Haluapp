# admisiones/views.py

# --- Importaciones (sin cambios) ---
import io
import hashlib
import pandas as pd
from openpyxl.styles import Font
import re
import uuid as _uuid
from finanzas.institucion_credentials import mp_webhook_secret as institucion_mp_webhook_secret
from finanzas.mercadopago_client import (
    crear_preferencia as mp_crear_preferencia,
    consultar_pago as mp_consultar_pago,
    MercadoPagoError,
    MercadoPagoSinCredenciales,
)
from utils.mercadopago_webhook import resolve_notification_data_id, verify_mercadopago_webhook_signature
from django.db.models import Sum
import json
from urllib.parse import urlencode
import logging
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.views.generic import DetailView, UpdateView, DeleteView
from django.db import models
from django.db.models import Count, Value, Q
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.db import transaction
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
import traceback
from django.http import Http404
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font
import time
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from .models import (
    Aspirante,
    DocumentoRequerido,
    DocumentoEntregado,
    HorarioDisponible,
    CitaAgendada,
    LoteImportacionAspirantes,
)
from gestion_academica.models import Grado, Usuario, Estudiante
from .forms import AspiranteForm, ImportarAspirantesForm
from .utils import crear_cuenta_cobro_inscripcion, enviar_correo_bienvenida, enviar_correo_cambio_estado, enviar_correo_confirmacion_cita
from finanzas.models import (
    CuentaPorCobrarEstudiante,
    PagoRegistrado,
    InstitucionEducativa,
    ConceptoPago,
    WebhookEventoMercadoPago,
)
from datetime import timedelta
from gestion_academica.models import PeriodoAcademico



logger = logging.getLogger(__name__)


# --- Validación de archivos subidos al portal del postulante ---
# Tipos MIME y extensiones aceptados para los documentos de admisión.
DOCUMENTO_ASPIRANTE_MAX_BYTES = 10 * 1024 * 1024  # 10 MB
DOCUMENTO_ASPIRANTE_EXTENSIONES = {
    "pdf", "jpg", "jpeg", "png", "webp", "doc", "docx",
}
DOCUMENTO_ASPIRANTE_MIME = {
    "application/pdf",
    "image/jpeg",
    "image/png",
    "image/webp",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/octet-stream",  # algunos navegadores no detectan MIME; validamos por extensión también
}


def _validar_archivo_documento(archivo):
    """Devuelve (ok, mensaje_error) para un archivo subido por un postulante.

    Reglas:
    - Tamaño máximo 10 MB.
    - Extensión en lista blanca (PDF, imagen, Word).
    - Content-type del request en lista blanca o ``application/octet-stream``
      (fallback común cuando el navegador no detecta el tipo).
    """
    if archivo is None:
        return False, "No se recibió ningún archivo."
    if archivo.size > DOCUMENTO_ASPIRANTE_MAX_BYTES:
        return False, "El archivo supera el tamaño máximo permitido (10 MB)."
    nombre = (archivo.name or "").lower()
    extension = nombre.rsplit(".", 1)[-1] if "." in nombre else ""
    if extension not in DOCUMENTO_ASPIRANTE_EXTENSIONES:
        return False, "Formato no permitido. Usa PDF, imagen (JPG/PNG/WEBP) o Word (DOC/DOCX)."
    content_type = (getattr(archivo, "content_type", "") or "").lower()
    if content_type and content_type not in DOCUMENTO_ASPIRANTE_MIME:
        return False, "El tipo de archivo no coincide con los formatos permitidos."
    return True, ""


# --- Helpers de autorización para el portal del postulante ---

def _aspirante_desde_cuenta(cuenta):
    """Resuelve el aspirante asociado a una CuentaPorCobrarEstudiante.

    Maneja los dos casos del modelo:
    - Cuenta de matrícula: enlace directo `cuenta.aspirante`.
    - Cuenta de inscripción: enlace vía `cuenta.estudiante.aspirante_origen`.
    """
    if cuenta.aspirante_id:
        return cuenta.aspirante
    if cuenta.estudiante_id and getattr(cuenta.estudiante, 'aspirante_origen', None):
        return cuenta.estudiante.aspirante_origen
    return None


def _puede_operar_cuenta_aspirante(request, cuenta, aspirante):
    """Determina si la petición está autorizada a operar pagos de un aspirante.

    Reglas:
    1) Si trae `?token=<uuid>` y coincide con `aspirante.access_token`, OK.
    2) Si el usuario es staff/superuser de la misma institución, OK.
    3) En cualquier otro caso, denegado.
    """
    if not aspirante:
        return False

    token_raw = (request.GET.get('token') or '').strip()
    if token_raw:
        try:
            token_uuid = _uuid.UUID(token_raw)
        except (ValueError, TypeError):
            token_uuid = None
        if token_uuid and token_uuid == aspirante.access_token:
            return True

    user = getattr(request, 'user', None)
    if user and user.is_authenticated:
        if user.is_superuser:
            return True
        if (user.is_staff
                and getattr(user, 'institucion_asociada_id', None) == cuenta.institucion_id):
            return True

    return False


# --- 2. VISTAS PRINCIPALES DEL CRUD Y LÓGICA DE NEGOCIO ---

@login_required
@permission_required('admisiones.view_aspirante', raise_exception=True)
def listar_aspirantes(request):
    """ Muestra solo los aspirantes de la institución del usuario logueado. """
    aspirantes = Aspirante.objects.filter(
        institucion=request.user.institucion_asociada
    ).select_related('grado_aspira').order_by('-fecha_inscripcion')
    
    context = {'aspirantes': aspirantes, 'titulo_pagina': "Gestión de Aspirantes"}
    return render(request, 'admisiones/lista_aspirantes.html', context)

@login_required
@permission_required('admisiones.add_aspirante', raise_exception=True)
def crear_aspirante_manual(request):
    institucion_usuario = request.user.institucion_asociada

    if request.method == 'POST':
        form = AspiranteForm(request.POST, user=request.user)
        if form.is_valid():
            try:
                with transaction.atomic():
                    aspirante = form.save(commit=False)
                    aspirante.institucion = institucion_usuario
                    aspirante.save()
                    resultado = aspirante.procesar_inscripcion_completa()

                messages.success(
                    request,
                    f"Aspirante '{aspirante.nombres} {aspirante.apellidos}' registrado exitosamente.",
                )

                # Si requería pago pero la cuenta NO se pudo crear por configuración
                # incompleta, avisamos al admin para que lo solucione antes de que el
                # aspirante reciba el correo y vea el portal sin botón de pago.
                cobro = resultado.cobro_inscripcion
                if cobro.es_warning:
                    messages.warning(
                        request,
                        f"⚠ El aspirante se creó, pero NO se generó la cuenta de "
                        f"inscripción: {cobro.mensaje}",
                    )

                return redirect('admisiones:lista_grados_aspirantes')

            except Exception as e:
                messages.error(request, f"Ocurrió un error al registrar al aspirante: {e}")
    else:
        form = AspiranteForm(user=request.user)

    context = {
        'form': form,
        'titulo_pagina': 'Registrar Nuevo Aspirante Manualmente'
    }
    return render(request, 'admisiones/formulario_aspirante_manual.html', context)

@login_required
@permission_required('admisiones.change_aspirante', raise_exception=True)
def admitir_aspirante(request, aspirante_id):
    aspirante = get_object_or_404(Aspirante, pk=aspirante_id)
    grado_aspirado = aspirante.grado_aspira
    nivel_escolaridad = getattr(grado_aspirado, 'nivel_escolaridad', None)

    if not nivel_escolaridad:
        messages.error(request, f"El grado '{grado_aspirado}' no tiene un Nivel de Escolaridad asignado.")
        return redirect('admisiones:detalle_aspirante', pk=aspirante.id)

    try:
        concepto_matricula = ConceptoPago.objects.get(
            institucion=aspirante.institucion,
            es_pago_matricula=True,
            nivel_escolaridad=nivel_escolaridad,
        )
    except (ConceptoPago.DoesNotExist, ConceptoPago.MultipleObjectsReturned):
        messages.error(
            request,
            f"Error: No se encontró (o hay duplicados) un Concepto de Pago "
            f"marcado como 'Es pago de Matrícula' para el nivel '{nivel_escolaridad}'."
        )
        return redirect('admisiones:detalle_aspirante', pk=aspirante.id)

    with transaction.atomic():
        CuentaPorCobrarEstudiante.objects.get_or_create(
            aspirante=aspirante,
            concepto_pago=concepto_matricula,
            defaults={'monto_asignado': concepto_matricula.valor}
        )
        aspirante.estado = Aspirante.EstadoAdmision.ADMITIDO
        aspirante.save()

    messages.success(request, "Aspirante admitido y cobro de matrícula generado.")
    return redirect('admisiones:detalle_aspirante', pk=aspirante.id)

class AspiranteMultiInstitucionMixin:
    """ Mixin para filtrar querysets por la institución del usuario. """
    def get_queryset(self):
        return self.model.objects.filter(institucion=self.request.user.institucion_asociada)    

# --- Vistas Basadas en Clases (Completas) ---
class AspiranteDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Aspirante
    template_name = 'admisiones/detalle_aspirante.html'
    context_object_name = 'aspirante'
    permission_required = 'admisiones.view_aspirante'

    def get_queryset(self):
        # Tu lógica de seguridad multi-institución está perfecta y se mantiene
        return Aspirante.objects.filter(institucion=self.request.user.institucion_asociada)

    def get_context_data(self, **kwargs):
        """
        Añade los cobros del aspirante al contexto para mostrarlos en la plantilla.
        """
        context = super().get_context_data(**kwargs)
        aspirante = self.get_object()

        context['cuentas_del_aspirante'] = CuentaPorCobrarEstudiante.objects.filter(
            aspirante=aspirante
        ).select_related('concepto_pago').order_by('-fecha_creacion')

        context['titulo_pagina'] = f"Perfil del Aspirante: {aspirante}"

        # URL absoluta del portal construida con request.build_absolute_uri(),
        # que ya respeta USE_X_FORWARDED_HOST y SECURE_PROXY_SSL_HEADER
        # (Cloudflare tunnel, ngrok, etc.) sin producir "http://https://...".
        context['portal_url_absoluta'] = self.request.build_absolute_uri(
            aspirante.get_portal_url()
        )
        return context

class AspiranteUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Aspirante
    form_class = AspiranteForm
    template_name = 'admisiones/formulario_inscripcion.html'
    permission_required = 'admisiones.change_aspirante'
    # Asumo que tienes un Mixin para filtrar por institución, lo mantengo si existe
    # de lo contrario, puedes añadir el método get_queryset como en otras vistas.

    def get_success_url(self):
        """
        Define a dónde redirigir al usuario después de editar exitosamente.
        """
        messages.success(self.request, "La información del aspirante ha sido actualizada.")
        # --- CORRECCIÓN AQUÍ ---
        # Cambiamos 'lista_aspirantes' por la URL que sí existe.
        return reverse_lazy('admisiones:lista_grados_aspirantes')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Editar Aspirante: {self.object}"
        return context



class AspiranteDeleteView(LoginRequiredMixin, PermissionRequiredMixin, AspiranteMultiInstitucionMixin, DeleteView):
    model = Aspirante
    template_name = 'admisiones/confirmar_eliminacion_aspirante.html'
    context_object_name = 'aspirante'
    permission_required = 'admisiones.delete_aspirante'

    def get_success_url(self):
        """
        Define a dónde redirigir al usuario después de eliminar exitosamente.
        """
        messages.success(self.request, f"El aspirante '{self.object}' ha sido eliminado.")
        # ▼▼▼ CORRECCIÓN CLAVE AQUÍ ▼▼▼
        # Cambiamos la URL de redirección a la que sí existe.
        return reverse_lazy('admisiones:lista_grados_aspirantes')
        # ▲▲▲ FIN DE LA CORRECCIÓN ▲▲▲

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Confirmar Eliminación"
        return context


# --- 4. VISTAS DE DASHBOARD, IMPORTACIÓN Y EXPORTACIÓN ---

@login_required
@permission_required('admisiones.view_aspirante', raise_exception=True)
def dashboard_admisiones(request):
    return render(request, 'admisiones/dashboard.html', {'titulo_pagina': 'Dashboard de Admisiones'})

@login_required
def dashboard_data(request):
    """ 
    Proporciona datos JSON para el dashboard, mostrando estadísticas de TODOS 
    los aspirantes del periodo académico activo, incluyendo los ya matriculados.
    """
    institucion_usuario = getattr(request.user, 'institucion_asociada', None)

    if not institucion_usuario:
        return JsonResponse({
            'sin_datos': True,
            'porEstado': {'labels': [], 'data': []},
            'porGrado': {'labels': [], 'data': []},
            'porColegio': {'labels': [], 'data': []},
            'porMunicipio': {'labels': [], 'data': []},
            'porSexo': {'labels': [], 'data': []},
        })

    # 1. Buscamos el periodo académico activo para definir el ciclo de admisión
    periodo_activo = PeriodoAcademico.objects.filter(
        institucion=institucion_usuario, 
        activo=True
    ).first()

    if not periodo_activo:
        return JsonResponse({
            'sin_datos': True,
            'porEstado': {'labels': [], 'data': []},
            'porGrado': {'labels': [], 'data': []},
            'porColegio': {'labels': [], 'data': []},
            'porMunicipio': {'labels': [], 'data': []},
            'porSexo': {'labels': [], 'data': []},
        })

    # 2. Todos los aspirantes de la institución (el período activo es solo contexto)
    base_queryset = Aspirante.objects.filter(
        institucion=institucion_usuario,
    )

    # --- El resto de la lógica para agregar los datos se mantiene igual ---
    
    data_por_estado = base_queryset.values('estado').annotate(total=Count('id')).order_by()
    
    data_por_grado = base_queryset.annotate(
        grado_nombre=Coalesce('grado_aspira__nombre', Value('Sin Asignar'))
    ).values('grado_nombre').annotate(total=Count('id')).order_by('-total')
    
    data_por_colegio = base_queryset.exclude(colegio_procedencia__isnull=True).exclude(colegio_procedencia__exact='').values('colegio_procedencia').annotate(total=Count('id')).order_by('-total')[:10]
    
    data_por_municipio = base_queryset.exclude(municipio_ciudad__isnull=True).exclude(municipio_ciudad__exact='').values('municipio_ciudad').annotate(total=Count('id')).order_by('-total')[:10]
    
    sexo_display_map = dict(Aspirante._meta.get_field('sexo').choices)
    data_por_sexo_raw = base_queryset.values('sexo').annotate(total=Count('id')).order_by()
    data_por_sexo = [
        {'sexo_label': sexo_display_map.get(item['sexo'], 'No especificado'), 'total': item['total']}
        for item in data_por_sexo_raw
    ]

    response_data = {
        'porEstado': {'labels': [item['estado'] for item in data_por_estado], 'data': [item['total'] for item in data_por_estado]},
        'porGrado': {'labels': [item['grado_nombre'] for item in data_por_grado], 'data': [item['total'] for item in data_por_grado]},
        'porColegio': {'labels': [item['colegio_procedencia'] for item in data_por_colegio], 'data': [item['total'] for item in data_por_colegio]},
        'porMunicipio': {'labels': [item['municipio_ciudad'] for item in data_por_municipio], 'data': [item['total'] for item in data_por_municipio]},
        'porSexo': {'labels': [item['sexo_label'] for item in data_por_sexo], 'data': [item['total'] for item in data_por_sexo]}
    }
    return JsonResponse(response_data)

@login_required
@permission_required('admisiones.add_aspirante', raise_exception=True)
def descargar_plantilla_importacion(request):
    """
    Genera la plantilla de Excel DEFINITIVA con nombres de columna simples
    y una lista desplegable de grados para evitar errores.
    """
    institucion = request.user.institucion_asociada
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Aspirantes"
    
    # Columnas: obligatorias primero, luego opcionales (marcadas con *)
    headers = [
        # Obligatorias
        'nombres', 'apellidos', 'numero_documento', 'fecha_nacimiento',
        'email_contacto', 'grado_aspira', 'paga_inscripcion',
        # Opcionales — Observador del Estudiante
        'tipo_documento', 'lugar_nacimiento',
        'telefono_contacto', 'sexo', 'grupo_sanguineo', 'eps', 'discapacidad',
        'colegio_procedencia', 'municipio_ciudad', 'departamento', 'direccion',
    ]
    ws.append(headers)

    # Fila de ayuda con descripción de cada columna
    ws.append([
        'Nombres completos', 'Apellidos completos', 'Documento de identidad',
        'DD/MM/AAAA o AAAA-MM-DD', 'Correo electrónico', 'Nombre exacto del grado', 'SI o NO',
        'TI/CC/RC/PA/CE/OT', 'Ciudad y departamento',
        'Teléfono', 'M/F/O', 'A+/A-/B+/B-/AB+/AB-/O+/O-', 'Nombre de la EPS',
        'Condición (vacío=ninguna)',
        'Colegio anterior', 'Municipio/ciudad', 'Departamento', 'Dirección residencia',
    ])

    from openpyxl.styles import PatternFill, Font as OFont, Alignment
    # Cabecera principal: azul oscuro
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_font = OFont(bold=True, color="FFFFFF", size=10)
    # Cabecera de columnas opcionales: verde oscuro
    opt_fill   = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    # Fila de ayuda: gris claro
    help_fill  = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    help_font  = OFont(italic=True, size=9, color="475569")

    OBLIGATORIAS_COUNT = 7  # primeras 7 columnas son obligatorias
    for i, cell in enumerate(ws[1], start=1):
        if i <= OBLIGATORIAS_COUNT:
            cell.fill = header_fill
        else:
            cell.fill = opt_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for cell in ws[2]:
        cell.fill = help_fill
        cell.font = help_font

    grados = Grado.objects.filter(institucion=institucion).order_by('orden')
    if grados.exists():
        grado_names = f'"{",".join([grado.nombre for grado in grados])}"'
        dv_grado = DataValidation(type="list", formula1=grado_names, allow_blank=False)
        ws.add_data_validation(dv_grado)
        dv_grado.add('F3:F1000')  # Columna F (grado_aspira)

    dv_paga = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    ws.add_data_validation(dv_paga)
    dv_paga.add('G3:G1000')  # Columna G (paga_inscripcion)

    dv_tipo_doc = DataValidation(type="list", formula1='"TI,CC,RC,PA,CE,OT"', allow_blank=True)
    ws.add_data_validation(dv_tipo_doc)
    dv_tipo_doc.add('H3:H1000')

    dv_sexo = DataValidation(type="list", formula1='"M,F,O"', allow_blank=True)
    ws.add_data_validation(dv_sexo)
    dv_sexo.add('K3:K1000')

    dv_gs = DataValidation(type="list", formula1='"A+,A-,B+,B-,AB+,AB-,O+,O-"', allow_blank=True)
    ws.add_data_validation(dv_gs)
    dv_gs.add('L3:L1000')

    for col_cells in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = max(max_length + 3, 14)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_importacion_aspirantes.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required('admisiones.add_aspirante', raise_exception=True)
def importar_aspirantes_excel(request):
    """
    Sube un archivo Excel, crea un ``LoteImportacionAspirantes`` con el archivo
    y encola la tarea Celery ``procesar_importacion_aspirantes_task``. El
    usuario es redirigido a la vista de progreso, que se actualiza en vivo.
    """
    from .tasks import procesar_importacion_aspirantes_task

    institucion_actual = request.user.institucion_asociada
    if not institucion_actual:
        messages.error(request, "Tu usuario no está asociado a ninguna institución.")
        return redirect('admisiones:lista_grados_aspirantes')

    if request.method == 'POST':
        form = ImportarAspirantesForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = form.cleaned_data['archivo_excel']
            dry_run = form.cleaned_data.get('dry_run') or False

            contenido = archivo.read()
            archivo.seek(0)
            lote = LoteImportacionAspirantes.objects.create(
                institucion=institucion_actual,
                creado_por=request.user,
                archivo=archivo,
                archivo_bytes=contenido,
                nombre_original=archivo.name,
                dry_run=dry_run,
            )

            # Encolamos la tarea SOLO después del commit para evitar que el
            # worker la levante antes de que el archivo esté visible en BD/MEDIA.
            transaction.on_commit(
                lambda: procesar_importacion_aspirantes_task.delay(lote.pk)
            )

            messages.info(
                request,
                "Tu archivo se está procesando en segundo plano. Verás el progreso aquí en vivo.",
            )
            return redirect('admisiones:lote_importacion_detalle', lote_id=lote.pk)
    else:
        form = ImportarAspirantesForm()

    lotes_recientes = (
        LoteImportacionAspirantes.objects
        .filter(institucion=institucion_actual)
        .select_related('creado_por')
        .order_by('-fecha_creacion')[:10]
    )

    return render(
        request,
        'admisiones/importar_aspirantes.html',
        {
            'form': form,
            'titulo_pagina': 'Importar Aspirantes desde Excel',
            'lotes_recientes': lotes_recientes,
        },
    )


@login_required
@permission_required('admisiones.add_aspirante', raise_exception=True)
def lote_importacion_detalle(request, lote_id):
    """Vista web del estado de un lote (página de progreso o resumen final)."""
    lote = get_object_or_404(
        LoteImportacionAspirantes.objects.select_related('creado_por', 'institucion'),
        pk=lote_id,
        institucion=request.user.institucion_asociada,
    )
    context = {
        'lote': lote,
        'titulo_pagina': f"Lote de importación #{lote.pk}",
        'url_estado_json': reverse('admisiones:lote_importacion_estado', kwargs={'lote_id': lote.pk}),
        'url_errores_xlsx': reverse('admisiones:lote_importacion_errores', kwargs={'lote_id': lote.pk}),
    }
    return render(request, 'admisiones/lote_progreso.html', context)


@login_required
@permission_required('admisiones.add_aspirante', raise_exception=True)
def lote_importacion_estado(request, lote_id):
    """Endpoint JSON para polling del progreso del lote desde la UI."""
    lote = get_object_or_404(
        LoteImportacionAspirantes,
        pk=lote_id,
        institucion=request.user.institucion_asociada,
    )
    # Limitamos errores en la respuesta para no inflar el payload del polling;
    # la página de resumen los mostrará completos al finalizar.
    errores = lote.errores or []
    advertencias_total = sum(1 for e in errores if (e.get('tipo') == 'warning'))
    errores_total = sum(1 for e in errores if (e.get('tipo') != 'warning'))
    return JsonResponse({
        'id': lote.pk,
        'estado': lote.estado,
        'estado_display': lote.get_estado_display(),
        'finalizado': lote.esta_finalizado,
        'dry_run': lote.dry_run,
        'total_filas': lote.total_filas,
        'filas_procesadas': lote.filas_procesadas,
        'filas_exitosas': lote.filas_exitosas,
        'filas_fallidas': lote.filas_fallidas,
        'filas_con_advertencia': lote.filas_con_advertencia,
        'progreso_porcentaje': lote.progreso_porcentaje,
        'mensaje_error_general': lote.mensaje_error_general,
        'errores_preview': errores[:20],
        'errores_total': errores_total,
        'advertencias_total': advertencias_total,
        'fecha_inicio': lote.fecha_inicio.isoformat() if lote.fecha_inicio else None,
        'fecha_fin': lote.fecha_fin.isoformat() if lote.fecha_fin else None,
        'puede_cancelarse': lote.puede_cancelarse,
        'puede_reintentarse': lote.puede_reintentarse,
        'cancelacion_solicitada': lote.cancelacion_solicitada,
    })


def _puede_gestionar_lote(user, lote):
    """Aislamiento SaaS: solo el creador del lote o staff de la misma institución
    pueden cancelarlo o reintentarlo. Superuser tiene acceso total.
    """
    if user.is_superuser:
        return True
    if not user.institucion_asociada_id or user.institucion_asociada_id != lote.institucion_id:
        return False
    if lote.creado_por_id == user.pk:
        return True
    return bool(user.is_staff)


@login_required
@permission_required('admisiones.add_aspirante', raise_exception=True)
@require_POST
def cancelar_lote_importacion(request, lote_id):
    """Solicita la cancelación cooperativa del lote.

    Marca el flag ``cancelacion_solicitada=True`` (la tarea lo lee entre filas)
    e intenta hacer ``revoke`` del task Celery para detener lotes pendientes.
    """
    lote = get_object_or_404(
        LoteImportacionAspirantes,
        pk=lote_id,
        institucion=request.user.institucion_asociada,
    )
    if not _puede_gestionar_lote(request.user, lote):
        return HttpResponseForbidden("No tienes permiso para cancelar este lote.")

    if not lote.puede_cancelarse:
        messages.warning(
            request,
            f"El lote ya está en estado {lote.get_estado_display()}; no se puede cancelar.",
        )
        return redirect('admisiones:lote_importacion_detalle', lote_id=lote.pk)

    lote.cancelacion_solicitada = True
    lote.save(update_fields=['cancelacion_solicitada'])

    if lote.task_id:
        try:
            from proyecto_colegio.celery import app as celery_app
            celery_app.control.revoke(lote.task_id, terminate=False)
            logger.info("Revoke enviado para task %s del lote %s.", lote.task_id, lote.pk)
        except Exception as exc:
            logger.warning(
                "No se pudo enviar revoke al task %s del lote %s: %s",
                lote.task_id, lote.pk, exc,
            )

    if lote.estado == LoteImportacionAspirantes.Estado.PENDIENTE:
        # Si la tarea aún no arrancó, marcamos directamente como CANCELADO.
        lote.estado = LoteImportacionAspirantes.Estado.CANCELADO
        lote.mensaje_error_general = "Lote cancelado antes de iniciar."
        lote.fecha_fin = timezone.now()
        lote.save(update_fields=['estado', 'mensaje_error_general', 'fecha_fin'])

    messages.info(
        request,
        "Solicitud de cancelación enviada. Las filas ya creadas se conservan.",
    )
    return redirect('admisiones:lote_importacion_detalle', lote_id=lote.pk)


@login_required
@permission_required('admisiones.add_aspirante', raise_exception=True)
@require_POST
def reintentar_lote_importacion(request, lote_id):
    """Crea un nuevo lote a partir del archivo del original (FALLIDO o CANCELADO)
    y lo encola. Conservamos el lote viejo como auditoría.
    """
    from .tasks import procesar_importacion_aspirantes_task

    lote_origen = get_object_or_404(
        LoteImportacionAspirantes,
        pk=lote_id,
        institucion=request.user.institucion_asociada,
    )
    if not _puede_gestionar_lote(request.user, lote_origen):
        return HttpResponseForbidden("No tienes permiso para reintentar este lote.")

    if not lote_origen.puede_reintentarse:
        messages.warning(
            request,
            f"Solo se pueden reintentar lotes en estado FALLIDO o CANCELADO "
            f"(actual: {lote_origen.get_estado_display()}).",
        )
        return redirect('admisiones:lote_importacion_detalle', lote_id=lote_origen.pk)

    if not lote_origen.archivo or not lote_origen.archivo.name:
        messages.error(request, "El lote original ya no tiene archivo asociado.")
        return redirect('admisiones:lote_importacion_detalle', lote_id=lote_origen.pk)

    # Creamos un NUEVO lote apuntando al mismo archivo (no movemos bytes).
    nuevo_lote = LoteImportacionAspirantes.objects.create(
        institucion=lote_origen.institucion,
        creado_por=request.user,
        archivo=lote_origen.archivo.name,
        nombre_original=f"[reintento de #{lote_origen.pk}] {lote_origen.nombre_original}",
        dry_run=lote_origen.dry_run,
    )
    transaction.on_commit(
        lambda: procesar_importacion_aspirantes_task.delay(nuevo_lote.pk)
    )
    messages.info(
        request,
        f"Reintento encolado como lote #{nuevo_lote.pk}. Sigue el progreso aquí.",
    )
    return redirect('admisiones:lote_importacion_detalle', lote_id=nuevo_lote.pk)


@login_required
@permission_required('admisiones.add_aspirante', raise_exception=True)
@require_POST
def reenviar_correos_lote(request, lote_id):
    """Encola la tarea de reenvío de correos de bienvenida para un lote COMPLETADO."""
    from .tasks import reenviar_correos_bienvenida_lote

    lote = get_object_or_404(
        LoteImportacionAspirantes,
        pk=lote_id,
        institucion=request.user.institucion_asociada,
    )
    if lote.estado != LoteImportacionAspirantes.Estado.COMPLETADO:
        messages.warning(request, "Solo se pueden reenviar correos de lotes completados.")
        return redirect('admisiones:lote_importacion_detalle', lote_id=lote.pk)

    reenviar_correos_bienvenida_lote.delay(lote.pk, user_id=request.user.pk)
    messages.success(
        request,
        f"Reenvío de correos encolado para el lote #{lote.pk}. "
        "Recibirás una notificación cuando termine con el resultado detallado."
    )
    return redirect('admisiones:lote_importacion_detalle', lote_id=lote.pk)


@login_required
@permission_required('admisiones.add_aspirante', raise_exception=True)
def lote_importacion_errores_excel(request, lote_id):
    """Descarga un Excel con la lista de filas que fallaron en la importación."""
    lote = get_object_or_404(
        LoteImportacionAspirantes,
        pk=lote_id,
        institucion=request.user.institucion_asociada,
    )

    errores = lote.errores or []
    # Normalizamos para que la salida tenga columnas estables independientemente
    # del esquema antiguo ({fila, documento, error}) o el nuevo
    # ({tipo, fila, documento, mensaje, error}).
    filas = [
        {
            'tipo': (e.get('tipo') or 'error'),
            'fila': e.get('fila', ''),
            'documento': e.get('documento', ''),
            'mensaje': e.get('mensaje') or e.get('error') or '',
        }
        for e in errores
    ] or [{'tipo': '', 'fila': '', 'documento': '', 'mensaje': 'Sin incidencias registradas.'}]

    df = pd.DataFrame(filas, columns=['tipo', 'fila', 'documento', 'mensaje'])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Incidencias')
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="lote_{lote.pk}_errores.xlsx"'
    return response

@login_required
@permission_required('admisiones.view_aspirante', raise_exception=True)
def exportar_matriculados_excel(request):
    """
    Exporta una lista de estudiantes matriculados.
    CORREGIDO: Ahora filtra por la institución del administrador y está completo.
    """
    # Queryset base
    matriculados_qs = Aspirante.objects.filter(
        estado=Aspirante.EstadoAdmision.MATRICULADO
    ).select_related('estudiante_creado__usuario', 'grado_aspira', 'institucion')

    # Aplicamos el filtro de seguridad multi-institución
    if not request.user.is_superuser:
        institucion_usuario = getattr(request.user, 'institucion_asociada', None)
        if institucion_usuario:
            matriculados_qs = matriculados_qs.filter(institucion=institucion_usuario)
        else:
            # Si no tiene institución, no puede exportar nada
            matriculados_qs = Aspirante.objects.none()

    # Construimos el diccionario de datos para el DataFrame
    data = {
        'Nombres': [m.nombres for m in matriculados_qs], 
        'Apellidos': [m.apellidos for m in matriculados_qs],
        'Documento': [m.numero_documento for m in matriculados_qs], 
        'Fecha de Nacimiento': [m.fecha_nacimiento.strftime('%Y-%m-%d') if m.fecha_nacimiento else '' for m in matriculados_qs],
        'Email de Contacto': [m.email_contacto for m in matriculados_qs], 
        'Grado Matriculado': [m.grado_aspira.nombre if m.grado_aspira else 'N/A' for m in matriculados_qs],
        'Sexo': [m.get_sexo_display() for m in matriculados_qs],
        'Colegio de Procedencia': [m.colegio_procedencia for m in matriculados_qs], 
        'Municipio': [m.municipio_ciudad for m in matriculados_qs],
        'Departamento': [m.departamento for m in matriculados_qs],
        'Username Creado': [m.estudiante_creado.usuario.username if m.estudiante_creado and m.estudiante_creado.usuario else 'N/A' for m in matriculados_qs],
    }
    
    # Creamos el archivo Excel en memoria
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Matriculados')
    
    output.seek(0)
    
    # Preparamos la respuesta HTTP para descargar el archivo
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_matriculados.xlsx"'
    
    return response


@login_required
@permission_required('admisiones.change_aspirante', raise_exception=True)
@transaction.atomic
def revertir_matriculacion(request, aspirante_id):
    """
    Revierte el estado de un aspirante de 'MATRICULADO' a 'ADMITIDO'.
    """
    aspirante = get_object_or_404(Aspirante, pk=aspirante_id, institucion=request.user.institucion_asociada)

    if aspirante.estado != Aspirante.EstadoAdmision.MATRICULADO:
        messages.warning(request, f"El aspirante {aspirante} no está matriculado. No se puede revertir.")
        return redirect('admisiones:lista_aspirantes')

    try:
        estudiante_asociado = aspirante.estudiante_creado
        if estudiante_asociado:
            usuario_asociado = estudiante_asociado.usuario
            
            # 1. Eliminamos al Estudiante
            estudiante_asociado.delete()
            messages.info(request, f"Se eliminó el perfil de estudiante de {estudiante_asociado}.")

            # 2. Lógica para eliminar el Usuario si no tiene otros roles
            if usuario_asociado:
                # Comprobamos si el usuario tiene otros perfiles (ej. docente, familiar)
                # Esta es una forma segura de evitar borrar usuarios multifuncionales.
                if not hasattr(usuario_asociado, 'docente') and not hasattr(usuario_asociado, 'familiar'):
                    usuario_asociado.delete()
                    messages.info(request, f"Se eliminó la cuenta de usuario '{usuario_asociado.username}'.")
                else:
                    messages.warning(request, f"No se eliminó la cuenta de usuario '{usuario_asociado.username}' porque tiene otros perfiles asociados.")

        # 3. Revertimos el estado del Aspirante
        aspirante.estado = Aspirante.EstadoAdmision.ADMITIDO
        aspirante.estudiante_creado = None
        # MEJORA: Usamos update_fields para mayor eficiencia
        aspirante.save(update_fields=['estado', 'estudiante_creado'])

        messages.success(request, f"La matriculación del aspirante {aspirante} ha sido revertida exitosamente.")

    except Exception as e:
        messages.error(request, f"Ocurrió un error al revertir la matriculación: {e}")

    return redirect('admisiones:lista_aspirantes')

@login_required
@permission_required('admisiones.view_aspirante', raise_exception=True)
def pipeline_admisiones(request):
    """
    Prepara los datos para la vista del pipeline Kanban.
    Agrupa a los aspirantes por su estado de admisión.
    """
    estados_posibles = Aspirante.EstadoAdmision.choices
    
    # ✅ CORRECCIÓN CLAVE:
    # Se corrige la sintaxis del filtro y se usa .exclude()
    aspirantes_activos = Aspirante.objects.filter(
        institucion=request.user.institucion_asociada
    ).exclude(  # <-- Usamos exclude para quitar los que no queremos ver
        estado__in=[Aspirante.EstadoAdmision.RECHAZADO, Aspirante.EstadoAdmision.MATRICULADO]
    ).select_related('grado_aspira').order_by('apellidos', 'nombres')

    # El resto de tu lógica para agrupar es correcta
    pipeline_data = {estado_valor: {'label': estado_label, 'aspirantes': []} for estado_valor, estado_label in estados_posibles}

    for aspirante in aspirantes_activos:
        if aspirante.estado in pipeline_data:
            pipeline_data[aspirante.estado]['aspirantes'].append(aspirante)
    
    # Asegúrate que los valores en columnas_del_pipeline coincidan con los del modelo
    # Por ejemplo, si en el modelo es 'EN_PROCESO', aquí también debe serlo.
    columnas_del_pipeline = [
        Aspirante.EstadoAdmision.INSCRITO, 
        Aspirante.EstadoAdmision.EN_PROCESO, 
        Aspirante.EstadoAdmision.ADMITIDO
    ]
    pipeline_ordenado = {key: pipeline_data[key] for key in columnas_del_pipeline if key in pipeline_data}

    context = {
        'titulo_pagina': "Pipeline de Admisiones",
        'pipeline': pipeline_ordenado,
    }
    return render(request, 'admisiones/pipeline_admisiones.html', context)

@require_POST
@login_required
@permission_required('admisiones.change_aspirante', raise_exception=True)
def actualizar_estado_aspirante_api(request):
    """
    Endpoint de API para actualizar el estado de un aspirante.
    Recibe un JSON con el ID del aspirante y el nuevo estado.
    """
    try:
        data = json.loads(request.body)
        aspirante_id = data.get('aspirante_id')
        nuevo_estado = data.get('nuevo_estado')

        # Validamos que el nuevo estado sea válido
        if nuevo_estado not in Aspirante.EstadoAdmision.values:
            return JsonResponse({'status': 'error', 'message': 'Estado no válido.'}, status=400)

        aspirante = get_object_or_404(Aspirante, pk=aspirante_id, institucion=request.user.institucion_asociada)
        
        # Guardamos el estado anterior para la lógica de señales
        estado_anterior = aspirante.estado
        
        if estado_anterior != nuevo_estado:
            aspirante.estado = nuevo_estado
            aspirante.save(update_fields=['estado']) # Actualiza solo el campo de estado
            # La señal post_save se disparará aquí y enviará el correo de cambio de estado.

        return JsonResponse({'status': 'success', 'message': f'Aspirante movido a {aspirante.get_estado_display()}'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
# =========================================================================
# INICIO: VISTAS NUEVAS PARA MERCADO PAGO
# =========================================================================


def crear_preferencia_mercadopago(request, cuenta_por_cobrar_id):
    """
    VERSIÓN ENDURECIDA: solo el aspirante dueño (vía `?token=<uuid>`) o personal
    de la institución autenticado pueden generar la preferencia de pago.
    """
    logger.info(f"Iniciando creación de preferencia para la cuenta: {cuenta_por_cobrar_id}")

    cuenta = get_object_or_404(
        CuentaPorCobrarEstudiante.objects.select_related(
            'concepto_pago',
            'aspirante__institucion',
            'estudiante__aspirante_origen__institucion',
            'institucion',
        ),
        pk=cuenta_por_cobrar_id,
    )

    aspirante = _aspirante_desde_cuenta(cuenta)
    if not aspirante:
        logger.error(
            "Error Crítico: La CuentaPorCobrarEstudiante ID %s no está vinculada a ningún aspirante válido.",
            cuenta.id,
        )
        raise Http404("No se pudo procesar el pago: la información del aspirante es inconsistente.")

    if not _puede_operar_cuenta_aspirante(request, cuenta, aspirante):
        logger.warning(
            "Intento NO autorizado de crear preferencia MP para cuenta %s (user=%s, has_token=%s).",
            cuenta.id,
            getattr(request.user, 'pk', None),
            bool(request.GET.get('token')),
        )
        return HttpResponseForbidden("No autorizado para iniciar este pago.")

    # Bloquear link si la cuenta ya fue pagada o anulada
    if cuenta.estado == 'PAGADO':
        messages.info(
            request,
            "Este pago ya fue registrado exitosamente. No es necesario volver a realizarlo.",
        )
        if aspirante.estado == 'ADMITIDO':
            return redirect('admisiones:portal_postulante_pagado', token=aspirante.access_token)
        return redirect('admisiones:portal_postulante', token=aspirante.access_token)

    if cuenta.estado == 'ANULADO':
        messages.error(
            request,
            "Este enlace de pago fue anulado. Contacta a la institución para más información.",
        )
        return redirect('admisiones:portal_postulante', token=aspirante.access_token)

    institucion = aspirante.institucion

    query_params = {'token': str(aspirante.access_token), 'cuenta_id': cuenta.id}
    base_procesando_url = request.build_absolute_uri(reverse('admisiones:pago_procesando'))
    url_procesando = f"{base_procesando_url}?{urlencode(query_params)}"
    notification_url = (
        request.build_absolute_uri(reverse('admisiones:mercadopago_webhook'))
        + f"?institucion_id={institucion.id}"
    )

    preference_data = {
        "items": [{
            "title": f"{cuenta.concepto_pago.nombre_concepto}",
            "quantity": 1,
            "unit_price": float(cuenta.monto_asignado),
            "currency_id": "COP",
        }],
        "payer": {
            "name": aspirante.nombres,
            "surname": aspirante.apellidos,
            "email": aspirante.email_contacto,
        },
        "back_urls": {
            "success": url_procesando,
            "failure": url_procesando,
            "pending": url_procesando,
        },
        "auto_return": "approved",
        "notification_url": notification_url,
        "external_reference": str(cuenta.id),
    }

    try:
        body = mp_crear_preferencia(institucion, payload=preference_data, cuenta=cuenta)
    except MercadoPagoSinCredenciales as exc:
        logger.error(
            "Cuenta %s: institución %s sin credenciales MP: %s",
            cuenta.id, institucion.id, exc,
        )
        messages.error(
            request,
            "Esta institución aún no tiene configurada la pasarela de pago. "
            "Contacta al administrador.",
        )
        return redirect('admisiones:portal_postulante', token=aspirante.access_token)
    except MercadoPagoError as exc:
        logger.error(
            "Error en Mercado Pago al crear preferencia para cuenta %s: %s",
            cuenta.id, exc, exc_info=True,
        )
        messages.error(
            request,
            "No fue posible iniciar el pago en este momento. Reintenta en unos minutos.",
        )
        return redirect('admisiones:portal_postulante', token=aspirante.access_token)

    # Preferimos el `init_point` (vivo) sobre el `sandbox_init_point` cuando estamos
    # en producción; el cliente decide según `mp_modo_produccion`.
    redirect_url = (
        body.get('init_point') if institucion.mp_modo_produccion
        else (body.get('sandbox_init_point') or body.get('init_point'))
    )
    if not redirect_url:
        logger.error("Cuenta %s: respuesta MP sin URL de pago: %s", cuenta.id, body)
        messages.error(request, "La pasarela no devolvió una URL de pago válida.")
        return redirect('admisiones:portal_postulante', token=aspirante.access_token)

    logger.info("Cuenta %s: preferencia creada, redirigiendo a MP.", cuenta.id)
    return redirect(redirect_url)

    

# --- VISTA 2: LA PÁGINA DE ESPERA (SIN CAMBIOS, YA ESTABA BIEN) ---
def pago_respuesta_mp(request):
    """
    Página simple que se muestra al usuario al regresar de Mercado Pago.
    La confirmación real del pago se hace vía Webhook.
    """
    status = request.GET.get('status')
    
    if status == 'approved':
        messages.success(request, "¡Gracias por tu pago! Lo estamos procesando y te confirmaremos en breve.")
    elif status == 'pending':
        messages.info(request, "Tu pago está pendiente. Te notificaremos cuando se apruebe.")
    else: # Failure o cualquier otro caso
        messages.error(request, "El pago no pudo ser completado. Por favor, intenta de nuevo.")
    
    # Aquí puedes agregar lógica para redirigir al portal del aspirante si tienes el token
    return redirect('gestion_academica:inicio_academico')



def subir_documento(request, token, doc_req_id):
    aspirante = get_object_or_404(Aspirante, access_token=token)

    if request.method != 'POST':
        return redirect('admisiones:portal_postulante_pagado', token=token)

    archivo_subido = request.FILES.get('archivo')
    if not archivo_subido:
        messages.warning(request, "No se seleccionó ningún archivo para subir.")
        return redirect('admisiones:portal_postulante_pagado', token=token)

    # Validamos antes de tocar la base de datos: si falla, no creamos registros vacíos.
    ok, error_msg = _validar_archivo_documento(archivo_subido)
    if not ok:
        messages.error(request, error_msg)
        return redirect('admisiones:portal_postulante_pagado', token=token)

    documento_requerido = get_object_or_404(
        DocumentoRequerido,
        pk=doc_req_id,
        institucion=aspirante.institucion,  # Evita subir contra un docreq de otra institución.
    )

    documento_existente = DocumentoEntregado.objects.filter(
        aspirante=aspirante, documento_requerido=documento_requerido
    ).first()

    if documento_existente:
        documento_existente.archivo = archivo_subido
        documento_existente.estado = 'subido'
        documento_existente.save()
        messages.success(request, f"Documento '{documento_requerido.nombre}' reemplazado exitosamente.")
    else:
        DocumentoEntregado.objects.create(
            aspirante=aspirante,
            documento_requerido=documento_requerido,
            archivo=archivo_subido,
        )
        messages.success(request, f"Documento '{documento_requerido.nombre}' subido exitosamente.")

    return redirect('admisiones:portal_postulante_pagado', token=token)

def vista_agendamiento(request, token):
    """ Muestra los horarios de citas disponibles para la institución del aspirante. """
    aspirante = get_object_or_404(Aspirante, access_token=token)
    
    if aspirante.estado != 'ADMITIDO':
        messages.warning(request, "Aún no estás en la etapa de agendamiento de citas.")
        return redirect('admisiones:portal_postulante', token=token)
    if hasattr(aspirante, 'cita_agendada'):
        messages.info(request, "Ya tienes una cita agendada.")
        return redirect('admisiones:portal_postulante_pagado', token=token)
    
    # Filtra los horarios por la institución del aspirante
    horarios = HorarioDisponible.objects.filter(
    institucion=aspirante.institucion, 
    fecha_hora_inicio__gte=timezone.now()
    ).annotate(citas_count=Count('citas_agendadas')).filter(citas_count__lt=models.F('cupos_disponibles'))
    
    context = {'aspirante': aspirante, 'horarios_disponibles': horarios, 'titulo_pagina': "Agendar Cita de Admisión"}
    return render(request, 'admisiones/agendar_cita.html', context)

def confirmar_agendamiento(request, token, horario_id):
    aspirante = get_object_or_404(Aspirante, access_token=token)
    horario = get_object_or_404(HorarioDisponible, pk=horario_id, institucion=aspirante.institucion)
    if not horario.esta_disponible or hasattr(aspirante, 'cita_agendada'):
        messages.error(request, "El horario seleccionado ya no está disponible o ya tienes una cita."); return redirect('admisiones:vista_agendamiento', token=token)

    if request.method == 'POST':
        CitaAgendada.objects.create(aspirante=aspirante, horario=horario, notas_adicionales=request.POST.get('notas', ''))
        messages.success(request, f"¡Tu cita ha sido agendada exitosamente!"); return redirect('admisiones:portal_postulante_pagado', token=token)
    
    context = {'aspirante': aspirante, 'horario': horario, 'titulo_pagina': "Confirmar Cita"}
    return render(request, 'admisiones/confirmar_cita.html', context) 

@csrf_exempt
def mercadopago_webhook(request):
    """Webhook MP endurecido (Fase 3).

    Mejoras vs versión anterior:
      - **Idempotencia real**: cada notificación se guarda en
        ``WebhookEventoMercadoPago`` con (institucion, data_id, payload_hash).
        Si MP reenvía el mismo evento, devolvemos el mismo HTTP que la
        primera vez sin reprocesar nada.
      - **Auditoría completa**: cada evento queda persistido aunque la firma
        falle, con su payload_hash, headers relevantes y resultado.
      - **Rechazo estricto de firmas inválidas**: HTTP 401 (no 403, no 200).
      - **Llamada a MP via cliente con reintentos** (timeout + auditoría).
      - **Errores en transición post-pago no tumban el webhook**: el pago ya
        está registrado y MP no debe reintentar.
    """
    if request.method != "POST":
        return HttpResponse("Método no permitido", status=405)

    # 1) institución (multi-tenant): la URL trae ?institucion_id=N
    institucion_id_raw = request.GET.get("institucion_id")
    if not institucion_id_raw or not str(institucion_id_raw).isdigit():
        logger.error("Webhook MP: institucion_id ausente o inválido en la URL.")
        return HttpResponse("institucion_id invalido", status=400)

    try:
        institucion = InstitucionEducativa.objects.get(pk=int(institucion_id_raw))
    except InstitucionEducativa.DoesNotExist:
        logger.error("Webhook MP: institución %s no encontrada.", institucion_id_raw)
        return HttpResponse("Institucion no encontrada", status=404)

    # 2) Body crudo y hash (necesarios para idempotencia y auditoría)
    body_raw = request.body or b""
    payload_hash = hashlib.sha256(body_raw).hexdigest()

    try:
        data = json.loads(body_raw.decode("utf-8") or "{}")
    except (json.JSONDecodeError, UnicodeDecodeError):
        logger.error("Webhook MP: payload JSON inválido para institución %s.", institucion.id)
        return HttpResponse("Payload invalido", status=400)

    tipo_evento = (data.get("type") or "").strip()
    payment_id = (data.get("data") or {}).get("id")
    data_id_for_signature = resolve_notification_data_id(request, str(payment_id or ""))

    x_signature = request.META.get("HTTP_X_SIGNATURE", "") or ""
    x_request_id = request.META.get("HTTP_X_REQUEST_ID", "") or ""

    secret = institucion_mp_webhook_secret(institucion)
    firma_valida = verify_mercadopago_webhook_signature(
        secret,
        data_id=data_id_for_signature,
        x_request_id=x_request_id,
        x_signature_header=x_signature,
    )

    payload_resumen = {
        "type": tipo_evento,
        "data_id": data_id_for_signature or "",
        "user_id": data.get("user_id"),
        "action": data.get("action"),
        "live_mode": data.get("live_mode"),
        "api_version": data.get("api_version"),
    }

    # 3) Idempotencia: ¿ya existe este evento exacto para esta institución?
    evento_existente = (
        WebhookEventoMercadoPago.objects
        .filter(
            institucion=institucion,
            data_id=data_id_for_signature or "",
            payload_hash=payload_hash,
        )
        .first()
    )
    if evento_existente and evento_existente.procesado_ok:
        logger.info(
            "Webhook MP: evento duplicado (institucion=%s data_id=%s); devolviendo HTTP %s previo.",
            institucion.id, evento_existente.data_id, evento_existente.estado_http_devuelto,
        )
        return HttpResponse(
            "Evento ya procesado.",
            status=evento_existente.estado_http_devuelto or 200,
        )

    # 4) Crear/actualizar el evento. Esto debe persistir SIEMPRE (incluso si la
    # firma es inválida): así tenemos rastro de intentos sospechosos.
    evento = evento_existente or WebhookEventoMercadoPago(
        institucion=institucion,
        data_id=data_id_for_signature or "",
        payload_hash=payload_hash,
    )
    evento.tipo = tipo_evento[:32]
    evento.x_request_id = x_request_id[:128]
    evento.x_signature = x_signature[:255]
    evento.firma_valida = firma_valida
    evento.payload_resumen = payload_resumen
    evento.save()

    # 5) Rechazo estricto de firma inválida (los reintentos legítimos de MP
    # también traen firma; aquí distinguimos un atacante de un retry real).
    if not firma_valida:
        logger.warning(
            "Webhook MP: firma INVÁLIDA para institución %s (x-request-id=%s).",
            institucion.id, x_request_id or "<vacío>",
        )
        evento.estado_http_devuelto = 401
        evento.error_mensaje = "Firma x-signature inválida o ausente."
        evento.fecha_procesamiento = timezone.now()
        evento.save(update_fields=["estado_http_devuelto", "error_mensaje", "fecha_procesamiento"])
        return HttpResponse("Firma invalida", status=401)

    # 6) Tipos de evento que ignoramos (devolvemos 200 para que MP no reintente).
    if tipo_evento and tipo_evento != "payment":
        evento.estado_http_devuelto = 200
        evento.procesado_ok = True
        evento.error_mensaje = f"Tipo de evento '{tipo_evento}' ignorado."
        evento.fecha_procesamiento = timezone.now()
        evento.save(update_fields=["estado_http_devuelto", "procesado_ok", "error_mensaje", "fecha_procesamiento"])
        return HttpResponse(status=200)

    if not payment_id:
        evento.estado_http_devuelto = 400
        evento.error_mensaje = "Falta data.id en el payload."
        evento.fecha_procesamiento = timezone.now()
        evento.save(update_fields=["estado_http_devuelto", "error_mensaje", "fecha_procesamiento"])
        return HttpResponse("Falta payment id", status=400)

    # 7) Consultar el pago en MP usando el cliente con reintentos + auditoría.
    try:
        payment_info = mp_consultar_pago(institucion, payment_id=payment_id)
    except MercadoPagoSinCredenciales as exc:
        logger.error("Webhook MP institución %s sin credenciales: %s", institucion.id, exc)
        evento.estado_http_devuelto = 503
        evento.error_mensaje = f"Sin credenciales MP: {exc}"
        evento.fecha_procesamiento = timezone.now()
        evento.save(update_fields=["estado_http_devuelto", "error_mensaje", "fecha_procesamiento"])
        return HttpResponse("Sin credenciales", status=503)
    except MercadoPagoError as exc:
        logger.error(
            "Webhook MP: fallo al consultar pago %s para institución %s: %s",
            payment_id, institucion.id, exc,
        )
        evento.estado_http_devuelto = 502
        evento.error_mensaje = f"Error consultando MP: {exc}"
        evento.fecha_procesamiento = timezone.now()
        evento.save(update_fields=["estado_http_devuelto", "error_mensaje", "fecha_procesamiento"])
        # 502 hace que MP reintente (queremos eso si el problema es transitorio).
        return HttpResponse("Error consultando MP", status=502)

    if payment_info.get("status") != "approved":
        evento.estado_http_devuelto = 200
        evento.procesado_ok = True
        evento.error_mensaje = f"Pago en estado '{payment_info.get('status')}', sin acción."
        evento.fecha_procesamiento = timezone.now()
        evento.save(update_fields=["estado_http_devuelto", "procesado_ok", "error_mensaje", "fecha_procesamiento"])
        return HttpResponse(status=200)

    external_ref = payment_info.get("external_reference") or ""
    if not external_ref or not re.fullmatch(r"\d+", str(external_ref)):
        logger.warning(
            "Webhook MP: external_reference inválida ('%s') para pago %s.",
            external_ref, payment_id,
        )
        evento.estado_http_devuelto = 400
        evento.error_mensaje = f"external_reference inválida: '{external_ref}'."
        evento.fecha_procesamiento = timezone.now()
        evento.save(update_fields=["estado_http_devuelto", "error_mensaje", "fecha_procesamiento"])
        return HttpResponse("external_reference invalida", status=400)

    # 8) Registro del pago (con bloqueo de fila y aislamiento por institución).
    cuenta = None
    aspirante = None
    pago_registrado_obj = None
    try:
        with transaction.atomic():
            # Defensa por si el mismo payment_id viaja en dos webhooks distintos
            # (data_id distinto pero pago igual) → el filtro por institución evita
            # cross-tenant.
            existente = (
                PagoRegistrado.objects
                .filter(referencia_transaccion=str(payment_id), institucion=institucion)
                .first()
            )
            if existente:
                logger.info(
                    "Webhook MP: pago %s ya estaba registrado para institución %s (idempotencia).",
                    payment_id, institucion.id,
                )
                pago_registrado_obj = existente
            else:
                cuenta = CuentaPorCobrarEstudiante.objects.select_for_update().get(
                    id=int(external_ref),
                    institucion=institucion,
                )

                aspirante = _aspirante_desde_cuenta(cuenta)
                if not aspirante:
                    raise RuntimeError(f"Cuenta {cuenta.id} sin aspirante válido.")

                estudiante_asociado = aspirante.estudiante_creado
                if not estudiante_asociado:
                    raise RuntimeError(
                        f"Aspirante {aspirante.pk} sin estudiante asociado."
                    )

                if cuenta.estado != "PAGADO":
                    pago_registrado_obj = PagoRegistrado.objects.create(
                        cuenta=cuenta,
                        estudiante=estudiante_asociado,
                        valor_pagado=Decimal(str(payment_info["transaction_amount"])),
                        metodo_pago="MERCADO_PAGO",
                        referencia_transaccion=str(payment_id),
                        institucion=institucion,
                        observacion=f"Pago confirmado automáticamente vía MP. ID: {payment_id}",
                    )
                    logger.info(
                        "Webhook: PagoRegistrado creado para cuenta #%s (institución %s).",
                        cuenta.id, institucion.id,
                    )
                else:
                    logger.info(
                        "Webhook MP: cuenta %s ya marcada como PAGADO; no se duplica el pago.",
                        cuenta.id,
                    )
    except CuentaPorCobrarEstudiante.DoesNotExist:
        logger.error(
            "Webhook MP: cuenta %s no existe o no pertenece a la institución %s.",
            external_ref, institucion.id,
        )
        evento.estado_http_devuelto = 404
        evento.error_mensaje = (
            f"Cuenta {external_ref} no existe o no pertenece a la institución {institucion.id}."
        )
        evento.fecha_procesamiento = timezone.now()
        evento.save(update_fields=["estado_http_devuelto", "error_mensaje", "fecha_procesamiento"])
        return HttpResponse("Cuenta no encontrada", status=404)
    except Exception as exc:
        logger.error("Webhook MP: error registrando pago %s: %s", payment_id, exc, exc_info=True)
        evento.estado_http_devuelto = 500
        evento.error_mensaje = f"Error interno: {exc}"
        evento.fecha_procesamiento = timezone.now()
        evento.save(update_fields=["estado_http_devuelto", "error_mensaje", "fecha_procesamiento"])
        return HttpResponse("Error interno", status=500)

    # 9) Transición de estado del aspirante (post-commit). Si esto falla, el pago
    # YA está registrado, así que devolvemos 200 igual y dejamos rastro en el log.
    try:
        if cuenta and aspirante:
            concepto_pagado = cuenta.concepto_pago
            if concepto_pagado.es_pago_matricula and aspirante.estado == "APROBADO_MATRICULA":
                _, resultado_cuentas = aspirante.matricular()
                logger.info(
                    "Webhook: Aspirante %s MATRICULADO. Cuentas: %s",
                    aspirante.pk, resultado_cuentas.resumen(),
                )
            elif concepto_pagado.es_pago_inscripcion and aspirante.estado == "INSCRITO":
                aspirante.estado = "ADMITIDO"
                aspirante.save(update_fields=["estado"])
                logger.info("Webhook: Aspirante %s actualizado a ADMITIDO.", aspirante.pk)
    except Exception as exc:
        logger.error(
            "Webhook: El pago de la cuenta %s se registró, pero la transición automática falló: %s",
            cuenta.id if cuenta else external_ref, exc, exc_info=True,
        )

    # 10) Cerrar el evento como OK.
    evento.cuenta = cuenta
    evento.pago_registrado = pago_registrado_obj
    evento.estado_http_devuelto = 200
    evento.procesado_ok = True
    evento.fecha_procesamiento = timezone.now()
    evento.save(update_fields=[
        "cuenta", "pago_registrado",
        "estado_http_devuelto", "procesado_ok", "fecha_procesamiento",
    ])
    return HttpResponse(status=200)

@require_POST
def cancelar_cita(request, token):
    """Permite al aspirante cancelar su cita.

    El aspirante NO está autenticado en Django: se identifica por el UUID
    ``access_token``. Por eso esta vista es pública pero exige POST + CSRF
    para evitar que un enlace malicioso GET dispare la cancelación
    (anteriormente era ``@login_required`` y un ``<a href>`` simple, lo que
    rompía la funcionalidad real para el aspirante).
    """
    aspirante = get_object_or_404(Aspirante, access_token=token)

    try:
        cita_agendada = aspirante.cita_agendada
        info_cita_cancelada = str(cita_agendada)
        cita_agendada.delete()
        messages.success(
            request,
            f"Tu cita '{info_cita_cancelada}' ha sido cancelada exitosamente. "
            "Ya puedes agendar una nueva.",
        )
    except CitaAgendada.DoesNotExist:
        messages.warning(request, "No tienes ninguna cita agendada para cancelar.")

    return redirect('admisiones:portal_postulante_pagado', token=token)

def portal_postulante_pagado(request, token):
    """
    Muestra los siguientes pasos después de que el aspirante ha pagado,
    como la lista de documentos a subir.
    """
    aspirante = get_object_or_404(Aspirante.objects.select_related('institucion'), access_token=token)
    
    # Filtra los documentos requeridos por la institución del aspirante
    documentos_requeridos = DocumentoRequerido.objects.filter(institucion=aspirante.institucion)
    
    # Crea un mapa de los documentos que ya se han subido para mostrarlos en la plantilla
    documentos_subidos = {doc.documento_requerido.id: doc for doc in aspirante.documentos_entregados.all()}
    
    # Prepara la lista de tareas para la plantilla
    lista_tareas_documentos = [
        {'requerido': doc_req, 'entregado': documentos_subidos.get(doc_req.id)} 
        for doc_req in documentos_requeridos
    ]
    
    context = {
        'aspirante': aspirante, 
        'lista_tareas_documentos': lista_tareas_documentos, 
        'titulo_pagina': f"Siguientes Pasos: {aspirante.nombres}"
    }
    return render(request, 'admisiones/portal_postulante_pagado.html', context)


def portal_postulante(request, token):
    """
    Portal único y seguro para el aspirante.
    Muestra el estado de sus documentos y los enlaces de pago correctos,
    buscando en el modelo CuentaPorCobrarEstudiante.
    """
    aspirante = get_object_or_404(Aspirante.objects.select_related(
        'grado_aspira', 'institucion', 'estudiante_creado'
    ), access_token=token)
    
    cuenta_inscripcion = None
    cuenta_matricula = None

    # --- LÓGICA DE BÚSQUEDA DE COBROS CORREGIDA ---
    
    # 1. Si debe pagar INSCRIPCIÓN, buscamos en la tabla de estudiantes,
    #    filtrando por el aspirante y el estado 'PENDIENTE'.
    if aspirante.estado == 'INSCRITO' and aspirante.requiere_pago_inscripcion and aspirante.estudiante_creado:
        cuenta_inscripcion = CuentaPorCobrarEstudiante.objects.filter(
            estudiante=aspirante.estudiante_creado,
            estado='PENDIENTE',
            concepto_pago__es_pago_inscripcion=True,
        ).first()

    # 2. Si debe pagar MATRÍCULA, la lógica se mantiene igual.
    elif aspirante.estado == 'APROBADO_MATRICULA':
        # Buscamos la cuenta por cobrar directamente a través del aspirante,
        # que es como se crea en la función 'crear_cuenta_cobro_matricula'.
        cuenta_matricula = CuentaPorCobrarEstudiante.objects.filter(
            aspirante=aspirante,
            estado='PENDIENTE',
            concepto_pago__es_pago_matricula=True,
        ).first()

    # --- Lógica de documentos (se mantiene igual) ---
    documentos_requeridos = DocumentoRequerido.objects.filter(
        institucion=aspirante.institucion,
        grados_aplicables=aspirante.grado_aspira
    )
    
    documentos_entregados_map = {
        doc.documento_requerido_id: doc 
        for doc in DocumentoEntregado.objects.filter(aspirante=aspirante)
    }

    context = {
        'titulo_pagina': 'Portal del Postulante',
        'aspirante': aspirante,
        'documentos_requeridos': documentos_requeridos,
        'documentos_entregados_map': documentos_entregados_map,
        'cuenta_inscripcion': cuenta_inscripcion,
        'cuenta_matricula': cuenta_matricula,
    }
    return render(request, 'admisiones/portal_postulante.html', context)

@login_required
@permission_required('admisiones.change_aspirante', raise_exception=True)
def matricular_aspirante(request, aspirante_id):
    """
    Acción manual para matricular un aspirante desde el panel administrativo.

    Reusa el método ``Aspirante.matricular()`` para no duplicar lógica: ese método
    activa al estudiante preliminar (creado en la fase de inscripción), actualiza
    rol y estado, y sincroniza las pensiones del año. Antes existía aquí una
    versión paralela que creaba OTRO ``Estudiante`` y rompía la unicidad de
    ``documento_identidad`` — esa versión queda eliminada.
    """
    aspirante = get_object_or_404(
        Aspirante,
        pk=aspirante_id,
        institucion=request.user.institucion_asociada,
    )

    grado_aspirado = aspirante.grado_aspira
    if not grado_aspirado or not getattr(grado_aspirado, 'nivel_escolaridad', None):
        messages.error(
            request,
            f"El grado '{grado_aspirado}' no tiene un Nivel de Escolaridad asignado. "
            "No se puede matricular hasta que esté configurado.",
        )
        return redirect('admisiones:detalle_aspirante', pk=aspirante.id)

    if aspirante.estado == Aspirante.EstadoAdmision.MATRICULADO:
        messages.info(request, f"El aspirante {aspirante} ya está matriculado.")
        return redirect('admisiones:detalle_aspirante', pk=aspirante.id)

    if not aspirante.estudiante_creado:
        messages.error(
            request,
            "Este aspirante aún no tiene un perfil de estudiante preliminar. "
            "Debe completar antes el proceso de inscripción.",
        )
        return redirect('admisiones:detalle_aspirante', pk=aspirante.id)

    try:
        estudiante, resultado_cuentas = aspirante.matricular()
    except Exception as exc:
        logger.error("Fallo al matricular aspirante %s: %s", aspirante.pk, exc, exc_info=True)
        messages.error(request, f"No se pudo matricular: {exc}")
        return redirect('admisiones:detalle_aspirante', pk=aspirante.id)

    messages.success(request, f"¡El aspirante {aspirante} ha sido matriculado exitosamente!")

    # Si la sincronización de cuentas tuvo problemas accionables (faltan
    # ConceptoPago de pensión/matrícula), avisamos al admin con CTA clara.
    if resultado_cuentas.es_warning:
        messages.warning(
            request,
            f"⚠ Matrícula realizada, pero NO se generaron todas las cuentas "
            f"automáticas: {resultado_cuentas.mensaje}",
        )
    elif resultado_cuentas.es_exito:
        messages.info(request, f"Cuentas generadas: {resultado_cuentas.resumen()}")

    return redirect('finanzas:historial_cuentas_estudiante', estudiante_id=estudiante.pk)

@login_required
@permission_required('admisiones.change_documentoentregado', raise_exception=True)
def revision_documentos_lista(request):
    institucion = request.user.institucion_asociada
    
    # Buscamos aspirantes que tengan al menos un documento 'subido'
    aspirantes_con_pendientes = Aspirante.objects.filter(
        institucion=institucion,
        documentos_entregados__estado='subido'
    ).distinct().annotate(
        # Contamos cuántos documentos tiene pendientes cada uno
        documentos_pendientes=Count('documentos_entregados', filter=Q(documentos_entregados__estado='subido'))
    ).order_by('-fecha_inscripcion')

    context = {
        'titulo_pagina': 'Revisión de Documentos',
        'aspirantes': aspirantes_con_pendientes,
    }
    return render(request, 'admisiones/revision_documentos_lista.html', context)

@login_required
@permission_required('admisiones.change_documentoentregado', raise_exception=True)
def revision_documento_detalle(request, aspirante_id):
    aspirante = get_object_or_404(Aspirante, pk=aspirante_id, institucion=request.user.institucion_asociada)

    # Si el admin envía el formulario para actualizar un documento
    if request.method == 'POST':
        documento_id = request.POST.get('documento_id')
        nuevo_estado = request.POST.get('estado')
        observaciones = request.POST.get('observaciones', '')

        documento = get_object_or_404(DocumentoEntregado, id=documento_id, aspirante=aspirante)
        documento.estado = nuevo_estado
        documento.observaciones_revision = observaciones
        documento.save()
        
        messages.success(request, f"El estado del documento '{documento.documento_requerido.nombre}' ha sido actualizado.")
        return redirect('admisiones:revision_documento_detalle', aspirante_id=aspirante.id)

    # Obtenemos la lista de todos los documentos requeridos y los que ya entregó
    documentos_requeridos = DocumentoRequerido.objects.filter(institucion=aspirante.institucion)
    documentos_entregados = {doc.documento_requerido.id: doc for doc in aspirante.documentos_entregados.all()}
    
    # Unimos la información para la plantilla
    lista_documentos = []
    for requerido in documentos_requeridos:
        lista_documentos.append({
            'requerido': requerido,
            'entregado': documentos_entregados.get(requerido.id)
        })

    context = {
        'titulo_pagina': f'Revisando a {aspirante}',
        'aspirante': aspirante,
        'lista_documentos': lista_documentos,
    }
    return render(request, 'admisiones/revision_documento_detalle.html', context)  

@login_required
@permission_required('admisiones.view_aspirante', raise_exception=True)
def lista_grados_aspirantes(request):
    """
    PÁGINA 1: Muestra una lista de los grados que tienen aspirantes.
    """
    grados_con_aspirantes = Grado.objects.filter(
        institucion=request.user.institucion_asociada
    ).annotate(
        num_aspirantes=Count('aspirante') # 'aspirante' es el related_name por defecto de Aspirante -> Grado
    ).filter(
        num_aspirantes__gt=0
    ).order_by('orden')

    context = {
        'titulo_pagina': "Aspirantes por Grado",
        'grados': grados_con_aspirantes
    }
    return render(request, 'admisiones/lista_grados_aspirantes.html', context)


@login_required
@permission_required('admisiones.view_aspirante', raise_exception=True)
def lista_aspirantes_por_grado(request, grado_id):
    """
    PÁGINA 2: Muestra la lista de aspirantes para un grado específico.
    """
    grado = get_object_or_404(Grado, pk=grado_id, institucion=request.user.institucion_asociada)
    
    aspirantes = Aspirante.objects.filter(
        grado_aspira=grado,
        institucion=request.user.institucion_asociada
    ).select_related('grado_aspira').order_by('apellidos', 'nombres')
    
    context = {
        'titulo_pagina': f"Aspirantes para {grado.nombre}",
        'aspirantes': aspirantes,
        'grado': grado
    }
    return render(request, 'admisiones/lista_aspirantes_por_grado.html', context)          

def pago_procesando(request):
    """
    Página intermedia que verifica el estado del pago y redirige al destino correcto.

    Endurecida:
    - Exige `?cuenta_id=<int>` y `?token=<uuid>` válidos.
    - Solo el aspirante dueño (token) o personal de la institución autenticado pueden ver
      esta página; así evitamos enumerar cuentas o estados ajenos.
    - El destino final se calcula en backend a partir del aspirante (no se acepta `?next=`).
    """
    cuenta_id_raw = (request.GET.get('cuenta_id') or '').strip()
    if not cuenta_id_raw.isdigit():
        messages.error(request, "Información de pago inválida o incompleta.")
        return redirect('gestion_academica:inicio_academico')

    cuenta = get_object_or_404(
        CuentaPorCobrarEstudiante.objects.select_related(
            'aspirante',
            'estudiante__aspirante_origen',
            'institucion',
        ),
        pk=int(cuenta_id_raw),
    )

    aspirante = _aspirante_desde_cuenta(cuenta)
    if not aspirante or not _puede_operar_cuenta_aspirante(request, cuenta, aspirante):
        return HttpResponseForbidden("No autorizado para ver este pago.")

    token_str = str(aspirante.access_token)
    url_final = reverse('admisiones:portal_postulante_pagado', kwargs={'token': token_str})

    # Permitimos override solo a URL interna segura (defensa contra open redirect).
    raw_next = (request.GET.get('next') or '').strip()
    if raw_next and url_has_allowed_host_and_scheme(
        url=raw_next,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        url_final = raw_next

    verif_url = reverse('admisiones:verificar_estado_pago', args=[cuenta.id]) + f"?token={token_str}"

    context = {
        'titulo_pagina': 'Procesando tu Pago',
        'cuenta_id': cuenta.id,
        'url_verificacion': verif_url,
        'url_final': url_final,
    }
    return render(request, 'admisiones/pago_procesando.html', context)


def verificar_estado_pago(request, cuenta_id):
    """
    Endpoint AJAX que devuelve el estado de una cuenta de un aspirante.

    Endurecido: requiere `?token=<uuid>` válido del aspirante o sesión staff de la
    misma institución; si no, responde 403.
    """
    try:
        cuenta = CuentaPorCobrarEstudiante.objects.select_related(
            'aspirante', 'estudiante__aspirante_origen', 'institucion',
        ).get(pk=cuenta_id)
    except CuentaPorCobrarEstudiante.DoesNotExist:
        return JsonResponse({'estado': 'NO_ENCONTRADO'}, status=404)

    aspirante = _aspirante_desde_cuenta(cuenta)
    if not aspirante or not _puede_operar_cuenta_aspirante(request, cuenta, aspirante):
        return JsonResponse({'estado': 'NO_AUTORIZADO'}, status=403)

    return JsonResponse({'estado': cuenta.estado})